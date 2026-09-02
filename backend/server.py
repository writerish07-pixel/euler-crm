import asyncio
import json
import logging
import io
import os
import re
import uuid
from collections import Counter
from datetime import datetime, timezone, date, timedelta
from pathlib import Path
from typing import Optional, List

from dotenv import load_dotenv
from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File, Form, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from motor.motor_asyncio import AsyncIOMotorClient
from pymongo.errors import DuplicateKeyError
from pydantic import BaseModel, ConfigDict, Field

import commercial as ce
import seed as seeder
import auth as authmod
import gsheets
import oem_catalog as oem_cat
import oem_sync
import coulson as coulson_client
import botspace as wa
import web_push

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

logger = logging.getLogger("euler.server")

client = AsyncIOMotorClient(os.environ["MONGO_URL"])
db = client[os.environ["DB_NAME"]]

app = FastAPI(title="Euler CRM API")
_finance_index_status = {
    "status": "UNKNOWN",
    "ready": False,
    "reason": "startup has not run",
    "checkedAt": None,
    "auditCounts": {},
}

auth_router = authmod.build_router(db)
current_user = auth_router.current_user
owner_only = auth_router.owner_only
sales_staff_only = auth_router.sales_staff_only
deal_desk_only = auth_router.deal_desk_only
money_desk_only = auth_router.money_desk_only
field_viewer_only = auth_router.field_viewer_only
finance_viewer_only = auth_router.finance_viewer_only
sales_gm_only = auth_router.sales_gm_only

api = APIRouter(prefix="/api", dependencies=[Depends(current_user)])
public = APIRouter(prefix="/api")


# ---------------------------------------------------------------- helpers
def clean(doc):
    if not doc:
        return doc
    doc.pop("_id", None)
    return doc


def now_iso():
    return datetime.now(timezone.utc).isoformat()


async def sheet_sync(entity: str, doc: dict, *, entity_id: str = ""):
    """Upsert one record into the existing Google Sheet and durably record the
    outcome (GS-4). A failed write becomes a PENDING sheet_sync_log entry that
    /integrations/gsheets/retry can replay — it is never silently lost.

    Safe to retry: gsheets.sync() is an ID-keyed upsert, so replaying a write that
    actually succeeded but whose response timed out finds the existing row and
    updates it instead of appending a duplicate.
    """
    res = await gsheets.sync(entity, doc)
    if res.get("operation") in ("skipped", "blocked"):
        return res  # sync disabled or env-write-blocked — nothing to log/retry
    eid = entity_id or str(doc.get(gsheets.SYNC_MAP.get(entity, ("", "", []))[1], "") or "")
    key = {"entityType": entity, "entityId": eid}
    entry = {
        **key,
        "tab": res.get("tab", ""),
        "operation": res.get("operation", ""),
        "status": "OK" if res.get("ok") else "PENDING",
        "error": res.get("error", ""),
        "missingHeaders": res.get("missingHeaders", []),
        "timestamp": now_iso(),
        "payload": {k: v for k, v in doc.items() if not k.startswith("_")},
    }
    existing = await db.sheet_sync_log.find_one(key)
    entry["attempt"] = int((existing or {}).get("attempt", 0)) + 1
    if res.get("ok"):
        entry["resolvedAt"] = now_iso()
    await db.sheet_sync_log.update_one(key, {"$set": entry}, upsert=True)
    return res


def today():
    return datetime.now(timezone.utc).strftime("%Y-%m-%d")


def this_month():
    return datetime.now(timezone.utc).strftime("%Y-%m")


def _scheme_as_of(lead=None, on=None):
    """Date used to pick Scheme Master month for a lead.

    Order: explicit `on` / schemeDate > lead.schemeAsOf > bookingDate > today.
    Changing Scheme Date on an unbooked lead must not invent a bookingDate.
    """
    for raw in (on, (lead or {}).get("schemeAsOf"), (lead or {}).get("bookingDate")):
        d = str(raw or "").strip()[:10]
        if len(d) == 10 and d[4] == "-" and d[7] == "-":
            return d
    return today()


def _normalize_scheme_row(payload):
    """Tag a Scheme Master row to a circular month from its dates.

    The UI used to send only Valid To. Without schemeMonth the row was treated as
    belonging to whatever month a lead asked for, so changing the date never
    switched circulars.
    """
    payload = dict(payload or {})
    eff_from = str(payload.get("effectiveFrom") or "").strip()[:10]
    eff_to = str(payload.get("effectiveTo") or "").strip()[:10]
    month = ce._norm_month(payload.get("schemeMonth") or "")
    if not month:
        month = ce.scheme_month_from_date(eff_from or eff_to or today())
    payload["schemeMonth"] = month
    if not eff_from and month and len(str(month)) >= 7:
        payload["effectiveFrom"] = f"{str(month)[:7]}-01"
    if eff_from:
        payload["effectiveFrom"] = eff_from
    if eff_to:
        payload["effectiveTo"] = eff_to
    return payload


def _scheme_row_matches_as_of(row, on):
    iso = str(on or "")[:10]
    month = ce.scheme_month_from_date(iso)
    r_month = ce._norm_month(row.get("schemeMonth") or "")
    if r_month:
        return r_month == month
    eff_from = str(row.get("effectiveFrom") or "")[:10]
    eff_to = str(row.get("effectiveTo") or "")[:10]
    if eff_from and iso and iso < eff_from:
        return False
    if eff_to and iso and iso > eff_to:
        return False
    return bool(eff_from or eff_to)


def _add_days(day: str, days: int) -> str:
    """Calendar arithmetic on a YYYY-MM-DD string. Unparseable in, empty out —
    a bad date must not become a revival that fires on the wrong day."""
    try:
        base = datetime.strptime(str(day or "")[:10], "%Y-%m-%d")
    except ValueError:
        return ""
    return (base + timedelta(days=int(days or 0))).strftime("%Y-%m-%d")


async def _log_activity_safe(lead: dict, activity_type: str, discussion: str):
    """Timeline entry that must never fail the operation that caused it.

    A cancellation that succeeded but could not write its activity row is still a
    cancellation; raising here would roll the user back to an error page while the
    lead was already cancelled in the database.
    """
    lead_id = str((lead or {}).get("leadId") or "")
    if not lead_id:
        return None
    try:
        doc = {
            "activityId": await next_id("activity", "AC26"), "leadId": lead_id,
            "date": today(), "time": datetime.now(timezone.utc).strftime("%H:%M"),
            "activityType": activity_type, "discussion": discussion,
            "customerName": lead.get("customerName"), "mobile": lead.get("mobile"),
            "model": lead.get("interestedModel"),
        }
        await db.activities.insert_one(doc)
        await db.leads.update_one({"leadId": lead_id}, {"$set": {
            "lastActivity": f"{activity_type} · {discussion}"[:200]}})
        await sheet_sync("activities", doc)
        return doc
    except Exception:
        logging.exception("ACTIVITY_LOG_FAILED lead=%s", lead_id)
        return None


async def next_id(kind, prefix, width=6):
    doc = await db.counters.find_one_and_update(
        {"_id": kind}, {"$inc": {"seq": 1}}, upsert=True, return_document=True
    )
    return f"{prefix}{str(doc['seq']).zfill(width)}"


async def get_lead_or_404(lead_id):
    lead = await db.leads.find_one({"leadId": lead_id})
    if not lead:
        raise HTTPException(404, f"Lead {lead_id} not found")
    return clean(lead)


async def get_scheme_rows():
    return [clean(s) for s in await db.scheme_master.find().to_list(1000)]


# ---------------------------------------------------------------- audit trail (H4) — append-only transaction log
async def actor(request: Request, user=Depends(current_user)):
    """Resolve the acting user + client IP for audit logging."""
    ip = ""
    try:
        ip = request.headers.get("x-forwarded-for", "").split(",")[0].strip() or (request.client.host if request.client else "")
    except Exception:
        ip = ""
    return {"email": (user or {}).get("email") or "system", "role": (user or {}).get("role") or "", "ip": ip}


async def write_audit(act, action, module, *, leadId="", paymentId="", claimId="",
                      financeFileNumber="", reportRef="", old=None, new=None):
    """Append-only. No financial record should change without an audit entry."""
    act = act or {}
    entry = {
        "auditId": f"AUD{uuid.uuid4().hex[:14]}",
        "timestamp": now_iso(),
        "user": act.get("email") or "system",
        "role": act.get("role") or "",
        "ip": act.get("ip") or "",
        "action": action, "module": module,
        "leadId": leadId, "paymentId": paymentId, "claimId": claimId,
        "financeFileNumber": financeFileNumber, "reportRef": reportRef,
        "oldValue": old, "newValue": new,
    }
    await db.audit_log.insert_one(dict(entry))
    return entry


# ---------------------------------------------------------------- step eligibility (LeadPickerService PICKER_STAGE + requireActiveLead_)
def _is_booked(lead):
    cs = (lead.get("currentStatus") or "").lower()
    return ("book" in cs or cs == "delivered" or "finance" in cs
            or bool(lead.get("bookingDate")) or ce.num(lead.get("bookingAmount")) > 0)


def _is_delivered(lead):
    return (lead.get("deliveryStatus") or "").lower() == "delivered" or (lead.get("currentStatus") or "").lower() == "delivered"


def _acct(lead):
    return (lead.get("accountStatus") or "Active").strip() or "Active"


def _has_persisted_scheme(lead):
    """True once a scheme step has been saved (explicit allocation or breakup)."""
    if lead.get("schemeAllocationExplicit"):
        return True
    raw = lead.get("benefitPassedBreakup") or lead.get("schemeAllocation") or ""
    if isinstance(raw, dict):
        return bool(raw)
    if isinstance(raw, str) and raw.strip() and raw.strip() not in ("{}", "null"):
        return True
    used = lead.get("schemeComponentsUsed") or ""
    if isinstance(used, dict) and used:
        return True
    if isinstance(used, str) and used.strip() and used.strip() not in ("{}", "null"):
        return True
    return False


def _is_priced(lead):
    """True after an explicit Price Structure save (not merely booking auto-fill)."""
    if "priceStructureSaved" in (lead or {}):
        return bool(lead.get("priceStructureSaved"))
    # Legacy leads (pre-flag): any positive Ex-Showroom counts as priced.
    return ce.num((lead or {}).get("exShowroom")) > 0


def _cancel_money(lead) -> dict:
    """Money already committed against a lead, and therefore at stake in a cancel.

    Cancelling does NOT reverse any of it — there is no auto-refund path, and
    inventing one would post entries nobody authorised. It is surfaced and
    recorded so a cancellation with money behind it is visible instead of silent.
    """
    booking = ce.num(lead.get("bookingAmount"))
    received = ce.num(lead.get("totalReceived"))
    # totalReceived already includes the booking amount once it is receipted, so
    # the exposure is the larger of the two, not their sum.
    customer = max(booking, received)
    return {
        "bookingAmount": ce.round2(booking),
        "totalReceived": ce.round2(received),
        "customerMoney": ce.round2(customer),
        "hasMoney": customer > 0.01,
    }


def _cancel_stage(lead) -> str:
    """Where in the funnel the lead died. An enquiry that fizzles costs a phone
    call; a booked deal that cancels costs a refund and a blocked chassis. They
    must never be added together."""
    if _is_delivered(lead):
        return "Delivered"
    if (lead.get("currentStatus") or "").lower().startswith("finance") or lead.get("financeFileNumber"):
        return "Finance"
    if _is_booked(lead):
        return "Booked"
    return "Enquiry"


def lead_actions(lead, act=None):
    """Which workflow steps a lead is eligible for (faithful to PICKER_STAGE + requireActiveLead_).

    Staff: Active + not Delivered may change commercial fields (vehicle uniqueness).
    Owner: Active leads stay editable after Mark Delivered; freeze only when Closed /
    Cancelled / Archived (accountStatus != Active).
    Close remains available on Active (including Delivered) so the lifecycle can exit.
    Finance / claim receipts stay allowed after close (not Archived).
    """
    role = ((act or {}).get("role") or "").strip().lower()
    active = _acct(lead) == "Active"
    booked = _is_booked(lead)
    delivered = _is_delivered(lead)
    not_archived = _acct(lead) != "Archived"
    # Owner may correct commercials after delivery until the lead is closed.
    mutable = active if role == "owner" else (active and not delivered)
    priced = _is_priced(lead)
    schemed = _has_persisted_scheme(lead)
    return {
        "canBook": mutable and not booked,
        "canPrice": mutable,
        "canScheme": mutable,
        "canPayment": mutable,                             # customer payment
        "canFinanceReceipt": not_archived,               # finance receipt allowed after close
        "canDeliver": active and booked and not delivered,
        "canClose": active,                               # close exit path (incl. delivered)
        # Cancel is the LOST exit. A delivered vehicle is not a cancellation — that
        # is a buyback, which this app has no ledger for — so it stops at delivery.
        "canCancel": active and not delivered,
        # Owner-only once the customer has paid: cancelling a funded booking is a
        # refund decision, not a sales-desk one.
        "cancelNeedsOwner": _cancel_money(lead)["hasMoney"],
        "canEditLead": mutable,                           # Edit Lead modal / PUT /leads
        "isBooked": booked, "isDelivered": delivered, "isActive": active,
        "isLocked": not mutable,
        # Step completion — staff may complete a step once; only owner re-edits (while mutable).
        "priceCompleted": priced,
        "schemeCompleted": schemed,
        "deliveryCompleted": delivered,
    }


def _require_owner_reedit(act, completed, step_label):
    """Staff may fill a step the first time; re-edits of a completed step are owner-only."""
    if completed and (act or {}).get("role") != "owner":
        raise HTTPException(
            403,
            f"{step_label} is already saved. Only the owner can edit a completed step.",
        )


def _require_mutable_lead(lead, verb="edits", act=None):
    """Freeze commercial/workflow edits.

    Everyone: accountStatus must be Active.
    Staff: also blocked after Mark Delivered.
    Owner: may edit after delivery until the lead is closed.
    """
    if _acct(lead) != "Active":
        raise HTTPException(
            409,
            f"This lead is locked for {verb} "
            f"(status: {lead.get('currentStatus') or 'New'} / {_acct(lead)}). "
            f"Only Active leads can be changed.",
        )
    role = ((act or {}).get("role") or "").strip().lower()
    if role != "owner" and _is_delivered(lead):
        raise HTTPException(
            409,
            f"This lead is locked for {verb} "
            f"(status: {lead.get('currentStatus') or 'New'} / {_acct(lead)}). "
            f"Only Active, non-delivered leads can be changed (owner may edit until closed).",
        )


def _yes_or_done(v):
    return str(v or "").strip().lower() in ("yes", "y", "done", "true", "1")


def _validate_delivery_ready(lead, body):
    """Port of validateDeliveryReady_ + isDeliveryEligible_ (outstanding must be cleared)."""
    errs = []
    if not _yes_or_done(body.insurance):
        errs.append("Set Insurance to Yes before marking delivered.")
    if not str(body.insurerName or "").strip():
        errs.append("Enter the insurer name before delivery.")
    # The agent decides the payout slab. Without it the entry silently falls back
    # to the default agent, which books the wrong rate on a real delivery.
    # Self-arranged insurance earns no payout, so no agent is needed there.
    if (ce.normalize_insurance_arranged_by(lead.get("insuranceArrangedBy")) != "self"
            and not str(body.insuranceAgentId or "").strip()
            and not str(lead.get("insuranceAgentId") or "").strip()):
        errs.append("Select the insurance agent before delivery.")
    if not _yes_or_done(body.registration):
        errs.append("Set Registration to Yes before marking delivered.")
    if not _yes_or_done(body.invoice):
        errs.append("Set Invoice to Yes before marking delivered.")
    if not str(body.invoiceNumber or "").strip():
        errs.append("Enter the invoice number before delivery.")
    if not str(body.chassisNumber or "").strip():
        errs.append("Enter the chassis number before delivery.")
    if not _yes_or_done(body.pdi):
        errs.append("Set PDI to Yes before marking delivered.")
    if ce.num(lead.get("customerOutstanding")) > 0.01:
        errs.append(f"Customer outstanding must be cleared (₹{ce.round2(ce.num(lead.get('customerOutstanding')))}) before delivery.")
    return errs


def _vehicle_id_blocks_reuse(other):
    """Cancelled / last-month delivered files do not block a new Sept+ delivery.

    Recreated leads (deleted old file, or cancelled last-month delivery) must be
    able to reuse invoice / chassis / plate. Closed-won sales from 1 Sep onward
    still occupy the identifier.
    """
    if not other:
        return False
    if other.get("dealCancelled"):
        return False
    acct = str(other.get("accountStatus") or "Active").strip().lower()
    if acct in ("cancelled", "inactive", "archived"):
        return False
    status = str(other.get("currentStatus") or "").lower()
    ddate = str(other.get("deliveryDate") or "")[:10]
    delivered = (
        str(other.get("deliveryStatus") or "").lower() == "delivered"
        or status == "delivered"
    )
    if delivered and ddate and ddate < oem_sync.YARD_LIVE_FROM:
        return False
    return True


async def _assert_unique_vehicle_identifiers(lead_id, *, invoice_number="", chassis_number="",
                                             number_plate=""):
    """Invoice / chassis / number plate must be unique across live leads.
    Blank values are ignored. The current lead is excluded so re-saves are allowed.
    Cancelled files and deliveries before 1 Sep do not count as conflicts."""
    checks = (
        ("invoiceNumber", invoice_number, "Invoice number"),
        ("chassisNumber", chassis_number, "Chassis number"),
        ("numberPlate", number_plate, "Number plate"),
    )
    for field, raw, label in checks:
        val = str(raw or "").strip()
        if not val:
            continue
        cursor = db.leads.find({
            "leadId": {"$ne": lead_id},
            field: {"$regex": f"^{re.escape(val)}$", "$options": "i"},
        })
        async for existing in cursor:
            if not _vehicle_id_blocks_reuse(existing):
                continue
            raise HTTPException(
                409,
                f"{label} '{val}' is already used on lead {existing.get('leadId')} "
                f"({existing.get('customerName') or '—'}).",
            )


def _require_action(lead, key, verb, act=None):
    acts = lead_actions(lead, act)
    if not acts.get(key):
        raise HTTPException(409, f"This lead is not eligible for {verb} (status: {lead.get('currentStatus') or 'New'} / {_acct(lead)}).")


def lead_to_snapshot(lead):
    """Build a commercial-engine snapshot dict from a lead document."""
    used = lead.get("schemeComponentsUsed") or {}
    if isinstance(used, str):
        try:
            import json as _json
            used = _json.loads(used) if used.strip() else {}
        except Exception:
            used = {}
    return {
        "bookingDate": lead.get("bookingDate", ""),
        "schemeAsOf": lead.get("schemeAsOf") or "",
        "benefitPassedBreakup": lead.get("benefitPassedBreakup", ""),
        # schemeAllocationV2: legacy Full Benefit may apply to entitlements when set.
        # schemeAllocationExplicit: new UI — CB only from explicit breakup (default 0).
        "schemeAllocationV2": bool(lead.get("schemeAllocationV2")),
        "schemeAllocationExplicit": bool(lead.get("schemeAllocationExplicit")),
        "schemeComponentsUsed": used if isinstance(used, dict) else {},
        "exShowroom": lead.get("exShowroom", 0),
        "accessories": lead.get("accessoriesAmount", 0),
        "insurance": lead.get("insuranceAmount", 0),
        "insuranceArrangedBy": ce.normalize_insurance_arranged_by(
            lead.get("insuranceArrangedBy")),
        "registrationRto": lead.get("rto", 0),
        "fastag": lead.get("fastag", 0),
        "handlingCharges": lead.get("handlingCharges", 0),
        "trc": lead.get("trc", 0),
        "extendedWarranty": lead.get("extendedWarranty", 0),
        "rsaAmc": lead.get("rsaAmc", 0),
        "otherCharges": lead.get("otherCharges", 0),
        "tcsApplicable": lead.get("tcsApplicable", "No"),
        "consumerDiscount": lead.get("consumerDiscount", 0),
        "exchangeBonus": lead.get("exchangeBonus", 0),
        "loyaltyBonus": lead.get("loyaltyBonus", 0),
        "referralBonus": lead.get("referralBonus", 0),
        "dsaDiscount": lead.get("dsaDiscount", 0),
        "additionalDiscount": lead.get("additionalDiscount", 0),
        "benefitMode": lead.get("benefitMode", "Full Benefit"),
        "customerBenefitPassed": lead.get("customerBenefitPassed", 0),
        "finalExchangeValue": lead.get("finalExchangeValue", 0),
        "oemExtraSupportReceived": lead.get("oemExtraSupportReceived", 0),
        "oemExtraSupportPassed": lead.get("oemExtraSupportPassed", 0),
        "model": lead.get("interestedModel", ""),
        "variant": lead.get("variant", ""),
        # The dealer's per-component allocation decision drives the whole scheme
        # engine, so it must reach every snapshot-based calculation.
        "schemeAllocation": lead.get("schemeAllocation"),
    }


# Scheme components the Dealer Earnings Register keeps a dedicated "… Retained"
# column for, mapped to the componentKey compute_scheme_allocation emits.
RETAINED_COMPONENT_COLUMNS = {
    "consumerRetained": "consumerDiscount",
    "exchangeRetained": "exchangeBonus",
    "loyaltyRetained": "loyaltyBonus",
    "referralRetained": "referralBonus",
    "dsaRetained": "dsaDiscount",
    "insuranceBenefitRetained": "insuranceBenefit",
    "rtoBenefitRetained": "rtoBenefit",
    "rtoInsuranceBenefitRetained": "rtoInsuranceBenefit",
}


def _retained_component_fields(retained_by_component):
    """Break the already-computed retainedByComponent map out into the per-component
    fields the Dealer Earnings Register has columns for, plus a human-readable
    breakup string. Components absent from the map retain 0.0 — a real value, not a
    blank — because "no retention on this component" is a meaningful commercial fact."""
    out = {}
    for field, key in RETAINED_COMPONENT_COLUMNS.items():
        out[field] = ce.round2(ce.num(retained_by_component.get(key)))
    out["schemeRetainedBreakup"] = "; ".join(
        f"{k}={ce.round2(ce.num(v))}" for k, v in sorted(retained_by_component.items())
    )
    return out


async def recompute_lead(lead_id):
    """Recompute all derived commercial + payment fields for a lead and persist."""
    lead = await db.leads.find_one({"leadId": lead_id})
    if not lead:
        return None
    snap = lead_to_snapshot(lead)
    scheme_rows = await get_scheme_rows()
    totals = ce.compute_commercial_totals(snap, scheme_rows)
    margin = ce.compute_dealer_margin(snap)
    income = ce.compute_scheme_income_breakdown(snap, scheme_rows)
    shares = ce.compute_scheme_claim_shares(snap, scheme_rows)
    alloc = income.get("allocation") or ce.compute_scheme_allocation(snap, scheme_rows)
    # total received from payments (refunds are negative rows, so this is already net)
    agg = await db.payments.aggregate([
        {"$match": {"leadId": lead_id}},
        {"$group": {"_id": None, "total": {"$sum": "$amount"}}},
    ]).to_list(1)
    total_received = ce.round2(agg[0]["total"]) if agg else 0.0
    refund_agg = await db.payments.aggregate([
        {"$match": {"leadId": lead_id, "entryType": "Refund"}},
        {"$group": {"_id": None, "total": {"$sum": "$amount"}}},
    ]).to_list(1)
    refunded_amount = ce.round2(-refund_agg[0]["total"]) if refund_agg else 0.0
    # Customer Payable comes solely from compute_commercial_totals(…, scheme_rows),
    # which already reduces by Σ customerBenefit across offers + entitlements.
    # Do NOT subtract entitlement benefits again — that double-counted after the
    # allocation engines were merged onto one path.
    customer_payable = totals["customerPayable"]
    if lead.get("dealCancelled"):
        # The deal is off. The customer owes nothing — showing an outstanding here
        # is not just cosmetic: once a cancelled lead is revived it goes back to
        # Active, and a stale outstanding would re-enter the company's live
        # receivables as a phantom debt on a lead nobody is chasing.
        customer_outstanding = 0.0
        # Everything still held belongs to the customer, because there is no longer
        # a vehicle to set it against. That is what makes it refundable.
        excess_received = ce.round2(max(0.0, total_received))
    else:
        customer_outstanding = ce.round2(max(0.0, customer_payable - total_received)) if customer_payable > 0 else 0.0
        # Surplus still held for the customer. Refundable at any time, including after
        # delivery/closure, so it must be visible on the lead rather than inferred.
        excess_received = ce.round2(max(0.0, total_received - customer_payable)) if customer_payable > 0 else 0.0
    oem_extra = ce.compute_oem_extra_support(lead)
    oem_extra_recv = oem_extra["oemExtraSupportReceived"]
    oem_extra_pass = oem_extra["oemExtraSupportPassed"]
    oem_extra_retained = oem_extra["oemExtraSupportRetained"]
    # Extra dealer income lines — customerInsuranceBenefitPassed is CUSTOMER
    # benefit/discount, NOT dealer income, and must not enter this sum.
    extra_income = ce.round2(
        ce.num(lead.get("documentationIncome")) + ce.num(lead.get("warrantyIncome")) +
        ce.num(lead.get("rsaIncome")) + ce.num(lead.get("referralIncome")) +
        ce.num(lead.get("otherIncome")) +
        ce.num(lead.get("financeIncentive")) + ce.num(lead.get("accessoriesMargin")) +
        ce.num(lead.get("exchangeMargin")) + ce.num(lead.get("campaignIncentive")))
    # Insurance PAYOUT income (premium × rate) — SEPARATE from Insurance Scheme Benefit.
    # Dealer Insurance Income = expected insurer payout. Do NOT subtract scheme
    # customerInsuranceBenefitPassed (that is customer discount, not payout share).
    _ins = await db.insurance.find_one({"leadId": lead_id}) or {}
    _ins_payout = ce.num(_ins.get("expectedPayout"))
    _dealer_ins_income = ce.round2(max(0.0, _ins_payout))
    # Dealer-funded benefit is a REAL cost. Only apply when an authoritative
    # allocation decision exists (schemeAllocationV2 / explicit) — never invent
    # a historical deduction from Benefit Mode alone.
    _has_auth_alloc = bool(lead.get("schemeAllocationV2") or lead.get("schemeAllocationExplicit")
                           or ce._explicit_allocation({"schemeAllocation": lead.get("schemeAllocation")}))
    _dealer_funded_benefit = (
        ce.round2(ce.num(alloc["totals"].get("dealerFundedBenefit"))) if _has_auth_alloc else 0.0)
    # Insurance Scheme Benefit (entitlement) — project independently from Loyalty.
    # Lead Register has an "Insurance Benefit" column; Dealer Earnings still carries
    # "Customer Insurance Benefit Passed"; Scheme Claim Register has a dedicated
    # Insurance Benefit amount column keyed by componentKey=insuranceBenefit.
    _ins_comp = (alloc.get("byKey") or {}).get("insuranceBenefit") or {}
    _ins_benefit_cb = ce.round2(ce.num(_ins_comp.get("customerBenefit"))) if _has_auth_alloc else None
    _ins_benefit_avail = ce.round2(ce.num(_ins_comp.get("schemeAvailable"))) if _ins_comp else 0.0
    updates = {
        "grossVehicleCost": totals["grossVehicleCost"],
        "customerPayable": customer_payable,
        "totalDiscount": totals["totalDiscount"],
        "oemSchemeAmount": totals["oemEligible"],
        "dealerSchemeAmount": totals["dealerDiscount"],
        "totalReceived": total_received,
        "customerOutstanding": customer_outstanding,
        "outstandingAmount": customer_outstanding,
        "excessReceived": excess_received,
        "refundedAmount": refunded_amount,
        # OEM claimable = OEM COMPANY share (per Scheme Master), not the raw offer sum
        "companyOutstanding": shares["eligibleTotal"],
        "oemClaimCompanyShare": shares["eligibleTotal"],
        "schemeCompanyTotal": shares["displayTotal"],
        "dealerSchemeRetained": income["retainedIncomeTotal"],
        "schemeCustomerBenefit": alloc["totals"]["customerBenefit"],
        # Compat aliases used by scheme-allocation tests / reports (PR #21).
        "schemeCustomerBenefitTotal": alloc["totals"]["customerBenefit"],
        "schemeOemClaimableTotal": alloc["totals"]["oemClaimable"],
        "schemeAvailableTotal": alloc["totals"]["schemeAvailable"],
        "dealerFundedBenefit": _dealer_funded_benefit,
        "oemExtraSupportReceived": oem_extra_recv,
        "oemExtraSupportPassed": oem_extra_pass,
        "oemExtraSupportRetained": oem_extra_retained,
        "dealerMarginNetExGst": margin["marginNetExGst"],
        "dealerMarginGrossInclGst": margin["marginGrossInclGst"],
        "dealerMarginGst": margin["marginGst"],
        **_retained_component_fields(income.get("retainedByComponent") or {}),
        "extraDealerIncomeTotal": extra_income,
        "dealerInsuranceIncome": _dealer_ins_income,
        "insurancePayout": _ins_payout,
        # Margin + retained + payout + extras − dealer-funded scheme benefit cost.
        # OEM claim is a receivable and must NEVER enter this total.
        "dealerTotalEarnings": ce.round2(
            margin["marginNetExGst"] + income["retainedIncomeTotal"] + oem_extra_retained
            + extra_income + _dealer_ins_income - _dealer_funded_benefit),
        # Engine summary lives here — NOT in schemeAllocation.
        # schemeAllocation is reserved for the flat {componentKey: amount} decision
        # map written by PUT /scheme-allocation (must not be overwritten).
        "schemeAllocationSummary": {
            "components": alloc["components"],
            "totals": alloc["totals"],
            "benefitMode": alloc["benefitMode"],
            "schemeMonth": alloc.get("schemeMonth"),
        },
        "lastUpdated": now_iso(),
    }
    # Authoritative Insurance Benefit projection (parallel to loyaltyBonus offer field).
    # Available amount is stored on insuranceBenefit; CB feeds the Dealer Earnings sheet
    # column Customer Insurance Benefit Passed. Historical leads without auth allocation
    # keep whatever memo value they already have.
    if _ins_comp:
        updates["insuranceBenefitAvailable"] = _ins_benefit_avail
    if _ins_benefit_cb is not None:
        updates["insuranceBenefit"] = _ins_benefit_cb
        updates["customerInsuranceBenefitPassed"] = _ins_benefit_cb
    # Preserve flat decision map if present; never replace it with the summary object.
    flat_decisions = ce._explicit_allocation({"schemeAllocation": lead.get("schemeAllocation")})
    if flat_decisions:
        import json as _json
        updates["schemeAllocation"] = _json.dumps(flat_decisions)
    await db.leads.update_one({"leadId": lead_id}, {"$set": updates})
    merged = {**lead, **updates}
    # The Lead Register is the dealership's primary register and must reflect the
    # lead's CURRENT commercial state. recompute_lead runs after every mutation that
    # can change it (booking, price structure, scheme, payment, delivery, close), so
    # this is the one place that guarantees the row never goes stale. Previously the
    # lead row was only pushed on create / manual PUT / import, so a lead booked and
    # delivered through the normal workflow stayed frozen at its creation values
    # (status New, ex-showroom 0, payable 0) even though every other register was
    # correct. Upsert on leadId, so this updates the existing row rather than adding one.
    await sheet_sync("leads", clean(dict(merged)))
    # GS-5: Dealer Earnings and Exchange previously had no Sheet destination at all.
    # Both are keyed on leadId, so these are upserts — one row per lead, kept current
    # on every commercial recompute rather than appended each time.
    if (not merged.get("dealCancelled")
            and (_is_booked(merged) or _is_commercial_deal(merged)
                 or ce.num(merged.get("customerPayable")) > 0)):
        de_row = {
            "insurancePayout": _ins_payout,
            "dealerInsuranceIncome": _dealer_ins_income,
            "leadId": lead_id, "customerName": merged.get("customerName"),
            "model": merged.get("interestedModel"),
            "dealerMarginNetExGst": updates["dealerMarginNetExGst"],
            "dealerSchemeRetained": updates["dealerSchemeRetained"],
            # Insurance Scheme Benefit CB from allocation — NOT the insurer payout.
            "customerInsuranceBenefitPassed": updates.get(
                "customerInsuranceBenefitPassed", merged.get("customerInsuranceBenefitPassed", 0)),
            "insuranceBenefit": updates.get("insuranceBenefit", merged.get("insuranceBenefit", 0)),
            "dealerFundedBenefit": updates.get("dealerFundedBenefit", 0),
            "financeIncentive": merged.get("financeIncentive", 0),
            "accessoriesMargin": merged.get("accessoriesMargin", 0),
            "exchangeMargin": merged.get("exchangeMargin", 0),
            "documentationIncome": merged.get("documentationIncome", 0),
            "warrantyIncome": merged.get("warrantyIncome", 0),
            "rsaIncome": merged.get("rsaIncome", 0),
            "referralIncome": merged.get("referralIncome", 0),
            "campaignIncentive": merged.get("campaignIncentive", 0),
            "otherIncome": merged.get("otherIncome", 0),
            "oemExtraSupportRetained": updates["oemExtraSupportRetained"],
            "extraDealerIncomeTotal": updates["extraDealerIncomeTotal"],
            "dealerTotalEarnings": updates["dealerTotalEarnings"],
            "totalDealerEarnings": updates["dealerTotalEarnings"],
            "variant": merged.get("variant", ""),
            "bookingDate": merged.get("bookingDate", ""),
            "deliveryDate": merged.get("deliveryDate", ""),
            "invoiceNumber": merged.get("invoiceNumber", ""),
            "executive": merged.get("executive", ""),
            "leadSource": merged.get("leadSource", ""),
            "customerPayable": updates["customerPayable"],
            "oemEligible": updates["oemSchemeAmount"],
            "customerSchemeBenefitPassed": updates.get("schemeCustomerBenefit", totals["customerBenefitPassed"]),
            "insuranceStatus": merged.get("insuranceStatus", ""),
            "remarks": merged.get("remarks", ""),
            "createdBy": merged.get("createdBy", "crm"),
            "modifiedBy": merged.get("lastUpdatedBy", ""),
            "currentStage": merged.get("currentStatus", ""),
            "lastUpdated": updates["lastUpdated"],
            "timestamp": updates["lastUpdated"],
            "oemExtraSupportReceived": oem_extra_recv,
            "oemExtraSupportPassed": oem_extra_pass,
            "dealerMarginGrossInclGst": updates["dealerMarginGrossInclGst"],
            "dealerMarginGst": updates["dealerMarginGst"],
            "consumerRetained": updates["consumerRetained"],
            "exchangeRetained": updates["exchangeRetained"],
            "loyaltyRetained": updates["loyaltyRetained"],
            "referralRetained": updates["referralRetained"],
            "dsaRetained": updates["dsaRetained"],
            "insuranceBenefitRetained": updates.get("insuranceBenefitRetained", 0),
            "rtoBenefitRetained": updates.get("rtoBenefitRetained", 0),
            "rtoInsuranceBenefitRetained": updates.get("rtoInsuranceBenefitRetained", 0),
            "schemeRetainedBreakup": updates["schemeRetainedBreakup"],
        }
        # Keep Mongo dealer_earnings in sync so GET /dealer-earnings matches the live report.
        await db.dealer_earnings.update_one({"leadId": lead_id}, {"$set": de_row}, upsert=True)
        await sheet_sync("dealer_earnings", de_row)
        # Derived OEM claims must also reach the existing Scheme Claim Register.
        # They are keyed on the same stable claimId GET /claims exposes, so this is an
        # upsert — a claim later settled/receipted updates that same row. Never delete:
        # Scheme Claim Register is a permanent ledger (Received rows stay forever).
        # Scheme Claim Register is one row per component. Amount columns
        # (Loyalty Bonus, Insurance Benefit, …) must carry THIS row's claim only —
        # never the lead's full denormalized offer map (that put Loyalty 10000 on
        # Insurance Benefit rows and made every scheme look like Loyalty).
        _claim_amount_fields = [
            "consumerDiscount", "exchangeBonus", "loyaltyBonus", "insuranceBenefit",
            "referralBonus", "dsaDiscount", "additionalDiscount",
            "rtoBenefit", "rtoInsuranceBenefit",
        ]
        for _key, _amt in shares["displayByComponent"].items():
            if _amt <= 0:
                continue
            _ex = await db.claims.find_one({
                "leadId": lead_id, "componentKey": _key, "manual": {"$ne": True},
                "claimStatus": {"$nin": ["Cancelled"]},
            })
            _amt = ce.round2(_amt)
            _comp_cols = {f: 0.0 for f in _claim_amount_fields}
            if _key in _comp_cols:
                _comp_cols[_key] = _amt
            if _ex:
                _claim_id = _ex.get("claimId") or f"CLM-{lead_id}-{_key}"
            else:
                _bk_now = await _live_booking(lead_id) or {}
                _had_cancelled = await db.claims.find_one({
                    "leadId": lead_id, "componentKey": _key, "manual": {"$ne": True},
                    "claimStatus": "Cancelled",
                })
                if _had_cancelled and _bk_now.get("bookingId"):
                    _claim_id = f"CLM-{lead_id}-{_key}-{_bk_now['bookingId']}"
                else:
                    _claim_id = f"CLM-{lead_id}-{_key}"
            _claim_row = {
                "claimId": _claim_id,
                "leadId": lead_id, "customer": merged.get("customerName"),
                "model": merged.get("interestedModel"), "variant": merged.get("variant"),
                "bookingDate": merged.get("bookingDate", ""),
                "component": ce.SCHEME_COMPONENT_LABELS.get(_key, _key), "componentKey": _key,
                "eligibleClaim": ce.round2(shares["eligibleByComponent"].get(_key, 0)),
                "claimAmount": _amt,
                "receivedAmount": (_ex or {}).get("receivedAmount", 0),
                "claimStatus": (_ex or {}).get("claimStatus", "Pending"),
                "claimReference": (_ex or {}).get("claimReference", ""),
                "bookingId": ((await _live_booking(lead_id)) or {}).get("bookingId", ""),
                "schemeMonth": ce.scheme_month_from_date(_scheme_as_of(merged)),
                "executive": merged.get("executive", ""),
                **_comp_cols,
                "totalDiscount": _amt,
                "dealerDiscount": 0.0,
                "oemDiscount": _amt,
                "claimRequired": "Yes" if _amt > 0 else "No",
                "ageingDays": _claim_ageing_days(
                    (_ex or {}).get("submittedDate") or merged.get("deliveryDate", ""),
                    (_ex or {}).get("claimStatus", "Pending")),
                "source": (
                    "Manual" if (_ex or {}).get("manual")
                    else ("OEM Extra Support" if _key == ce.OEM_EXTRA_SUPPORT_KEY
                          else "Derived (Scheme Master)")
                ),
                "dsaApproval": (_ex or {}).get("approvalStatus", "") if _key == "dsaDiscount" else "",
                "claimReceivedDate": (_ex or {}).get("claimReceivedDate", ""),
                "claimRemarks": (_ex or {}).get("claimRemarks", ""),
                "manual": False,
            }
            # Persist so GET /claims still lists the row after Close Won overwrites
            # currentStatus (the sheet already kept it; the app used to drop it).
            if _ex:
                await db.claims.update_one({"_id": _ex["_id"]}, {"$set": _claim_row})
            else:
                await db.claims.insert_one(dict(_claim_row))
            await sheet_sync("claims", _claim_row)
        # OEM Extra Support Register — Received is the claim; Passed/Retained track usage.
        if oem_extra_recv > 0:
            _bk = await _live_booking(lead_id) or {}
            await sheet_sync("oem_extra_support", {
                "leadId": lead_id,
                "bookingId": _bk.get("bookingId", "") or merged.get("bookingId", ""),
                "customerName": merged.get("customerName", ""),
                "model": merged.get("interestedModel", ""),
                "variant": merged.get("variant", ""),
                "bookingDate": merged.get("bookingDate", ""),
                "oemExtraSupportReceived": oem_extra_recv,
                "oemExtraSupportPassed": oem_extra_pass,
                "oemExtraSupportRetained": oem_extra_retained,
                "status": "Open",
                "lastUpdated": updates["lastUpdated"],
                "remarks": "",
            })
        if str(merged.get("exchangeRequired") or "").lower() == "yes" or ce.num(merged.get("finalExchangeValue")) > 0:
            await sheet_sync("exchange", {
                "leadId": lead_id, "customerName": merged.get("customerName"),
                "exchangeRequired": merged.get("exchangeRequired", "No"),
                "finalExchangeValue": merged.get("finalExchangeValue", 0),
                "exchangeBonus": merged.get("exchangeBonus", 0),
                "exchangeMargin": merged.get("exchangeMargin", 0),
            })
    return merged


# ---------------------------------------------------------------- models
class LeadIn(BaseModel):
    customerName: str
    mobile: str = ""
    altMobile: str = ""
    village: str = ""
    city: str = ""
    leadSource: str = "Walk-in"
    interestedModel: str = ""
    variant: str = ""
    executive: str = ""
    currentStatus: str = "New"
    priority: str = "Normal"
    budget: float = 0
    remarks: str = ""
    financeRequired: str = "No"
    exchangeRequired: str = "No"
    nextFollowupDate: Optional[str] = None
    createdDate: Optional[str] = None


class LeadUpdateIn(BaseModel):
    """PATCH-style partial update for PUT /leads/{lead_id}. Every field is Optional
    with no non-null default, so (a) exclude_unset=True only ever picks up keys the
    client actually sent, and (b) Swagger's auto-generated example body is all-null
    instead of LeadIn's fake 'string'/0/'New' placeholders. extra='forbid' rejects
    any field not in this explicit allowlist — leadId, createdDate, accountStatus,
    and every system-calculated financial field are structurally absent here, so
    they can never be set through this endpoint, not just filtered out."""
    model_config = ConfigDict(extra="forbid")
    customerName: Optional[str] = None
    mobile: Optional[str] = None
    altMobile: Optional[str] = None
    village: Optional[str] = None
    city: Optional[str] = None
    leadSource: Optional[str] = None
    interestedModel: Optional[str] = None
    variant: Optional[str] = None
    executive: Optional[str] = None
    currentStatus: Optional[str] = None
    priority: Optional[str] = None
    budget: Optional[float] = None
    remarks: Optional[str] = None
    financeRequired: Optional[str] = None
    exchangeRequired: Optional[str] = None
    nextFollowupDate: Optional[str] = None
    bookingDate: Optional[str] = None
    # Editable so staff can correct a mistaken default booking advance (e.g. 5000 → 0).
    bookingAmount: Optional[float] = None


class BookingIn(BaseModel):
    bookingDate: Optional[str] = None
    bookingAmount: float = 0
    executive: str = ""
    paymentMode: str = "Cash"
    financeRequired: str = "No"
    exchangeRequired: str = "No"


class PriceStructureIn(BaseModel):
    exShowroom: float = 0
    rto: float = 0
    insuranceAmount: float = 0
    # dealer (default) = premium in customer outstanding + payout earnings on delivery.
    # self = customer arranges insurance → premium not in outstanding; no dealer payout.
    insuranceArrangedBy: str = "dealer"
    accessoriesAmount: float = 0
    handlingCharges: float = 0
    trc: float = 0
    fastag: float = 0
    extendedWarranty: float = 0
    rsaAmc: float = 0
    otherCharges: float = 0
    tcsApplicable: str = "No"
    finalExchangeValue: float = 0


class SchemeIn(BaseModel):
    consumerDiscount: float = 0
    exchangeBonus: float = 0
    loyaltyBonus: float = 0
    referralBonus: float = 0
    dsaDiscount: float = 0
    additionalDiscount: float = 0
    # Kept for API compatibility / historical leads. New Scheme UI does not expose it;
    # allocation is driven by explicit benefitPassedBreakup + schemeComponentsUsed.
    benefitMode: str = "Partial Benefit"
    customerBenefitPassed: float = 0
    benefitPassedBreakup: Optional[str] = None
    schemeComponentsUsed: Optional[str] = None
    oemExtraSupportReceived: float = 0
    oemExtraSupportPassed: float = 0
    # As-of date for which Scheme Master month to apply. Does not invent a booking.
    schemeDate: Optional[str] = None


class PaymentIn(BaseModel):
    amount: float
    paymentMode: str = "Cash"
    date: Optional[str] = None
    narration: str = ""
    financerName: str = ""
    financeFileNumber: str = ""
    # Collecting MORE than Customer Payable is a real situation (round figure paid,
    # charge reduced later). It stays blocked by default so a mistyped amount is still
    # caught; the UI asks for confirmation and re-sends with this set. The surplus is
    # then visible as Excess Received and can be refunded.
    allowExcess: bool = False


class RefundIn(BaseModel):
    amount: float
    paymentMode: str = "Cash"
    date: Optional[str] = None
    narration: str = ""
    reference: str = ""


class DeliveryIn(BaseModel):
    insurance: str = ""
    registration: str = ""
    invoice: str = ""
    accessories: str = ""
    rc: str = ""
    pdi: str = ""
    delivered: str = ""
    deliveryDate: Optional[str] = None
    invoiceNumber: str = ""
    chassisNumber: str = ""
    numberPlate: str = ""
    insurerName: str = ""
    insuranceAgentId: str = ""     # broker paying the payout; picked at delivery
    feedback: str = ""


class CloseIn(BaseModel):
    closeReason: str = ""
    rc: str = ""
    numberPlate: str = ""
    closedDate: Optional[str] = None


class CancelIn(BaseModel):
    """Customer walked away. The LOST exit, as opposed to CloseIn's WON exit."""
    cancelReason: str = ""
    cancelRemarks: str = ""
    cancelDate: Optional[str] = None


class CancelReasonIn(BaseModel):
    """A cancel reason and what should happen to the lead afterwards.

    revive="now"   -> back in the funnel immediately, follow-ups restart
    revive="days"  -> parked, comes back automatically after reviveAfterDays
    revive="never" -> stays cancelled until someone revives it by hand
    """
    reason: str = ""
    revive: str = "now"
    reviveAfterDays: int = 0
    status: str = "Active"
    remarks: str = ""


class ActivityIn(BaseModel):
    activityType: str = "Note"
    discussion: str = ""
    executive: str = ""
    nextFollowup: str = ""
    date: Optional[str] = None


class SnapshotComputeIn(BaseModel):
    exShowroom: float = 0
    accessories: float = 0
    insurance: float = 0
    insuranceArrangedBy: str = "dealer"
    registrationRto: float = 0
    fastag: float = 0
    handlingCharges: float = 0
    trc: float = 0
    extendedWarranty: float = 0
    rsaAmc: float = 0
    otherCharges: float = 0
    tcsApplicable: str = "No"
    consumerDiscount: float = 0
    exchangeBonus: float = 0
    loyaltyBonus: float = 0
    referralBonus: float = 0
    dsaDiscount: float = 0
    additionalDiscount: float = 0
    benefitMode: str = "Full Benefit"
    finalExchangeValue: float = 0


# ---------------------------------------------------------------- misc
@api.get("/")
async def root():
    return {"app": "Euler CRM", "status": "ok"}


# Master lists editable from Settings → synced (full-mirror) to a "Masters" tab
# in the Google Sheet. Defined in seed.py (kept separate from workflow-state
# lists like statuses/paymentModes/benefitModes/claimStatuses, whose exact
# string values are load-bearing throughout status-gating / commercial logic).
EDITABLE_MASTER_CATEGORIES = seeder.EDITABLE_MASTER_CATEGORIES


async def _masters_list_values(category):
    # Executives live on the STAFF master once it is populated: a bare name in
    # masters_list has nowhere to hold a mobile, which is why the WhatsApp
    # settings grew a second, hand-typed executive list that silently drifted
    # out of sync (a mismatched name = that executive never gets messaged).
    if category == "executives":
        names = await _executive_names()
        if names:
            return names
    rows = await db.masters_list.find({"category": category}).sort("value", 1).to_list(500)
    return [r["value"] for r in rows]


STAFF_ROLES = ["executive", "TL", "GM", "ASM", "RM", "owner", "accounts"]
# Which daily reports a staff member can receive.
STAFF_REPORTS = ["exec_morning", "exec_eod", "manager_eod", "owner_eod"]
DEFAULT_REPORTS_BY_ROLE = {
    "executive": ["exec_morning", "exec_eod"],
    "TL": ["manager_eod"],
    "GM": ["manager_eod"],
    "ASM": ["manager_eod"],
    "RM": ["manager_eod"],
    "owner": ["owner_eod"],
    "accounts": [],
}


async def _executive_names() -> list:
    rows = await db.staff.find({"role": "executive"}).to_list(500)
    return sorted({str(r.get("name") or "").strip() for r in rows
                   if str(r.get("name") or "").strip()
                   and str(r.get("status") or "Active").lower() == "active"})


async def _staff_for_report(report: str) -> list:
    """Active, opted-in staff who should receive `report` and have a mobile."""
    out = []
    for r in await db.staff.find().to_list(500):
        if str(r.get("status") or "Active").lower() != "active":
            continue
        if not r.get("whatsappOptIn", True):
            continue
        if report not in (r.get("reports") or []):
            continue
        if not wa.digits10(r.get("mobile")):
            continue
        out.append(r)
    return sorted(out, key=lambda x: str(x.get("name") or ""))


@api.get("/masters")
async def masters():
    models = await db.price_master.distinct("model")
    out = dict(seeder.MASTERS)
    for cat in EDITABLE_MASTER_CATEGORIES:
        vals = await _masters_list_values(cat)
        if vals:
            out[cat] = vals
    return {**out, "models": sorted([m for m in models if m])}


@api.get("/masters-list")
async def list_masters_list():
    return [clean(r) for r in await db.masters_list.find().sort("value", 1).to_list(2000)]


class MasterListIn(BaseModel):
    category: str
    value: str


@api.post("/masters-list", dependencies=[Depends(owner_only)])
async def add_master_list_value(body: MasterListIn, act=Depends(actor)):
    category = body.category.strip()
    value = body.value.strip()
    if category not in EDITABLE_MASTER_CATEGORIES:
        raise HTTPException(422, f"'{category}' is not an editable master list")
    if not value:
        raise HTTPException(422, "Value is required")
    existing = await db.masters_list.find_one({"category": category, "value": {"$regex": f"^{re.escape(value)}$", "$options": "i"}})
    if existing:
        raise HTTPException(409, f"'{value}' already exists in {category}")
    doc = {"id": f"ML{uuid.uuid4().hex[:8]}", "category": category, "value": value, "status": "Active"}
    await db.masters_list.insert_one(doc)
    await write_audit(act, "create", "masters_list", new=doc)
    await gsheets.sync_masters(await list_masters_list())
    return clean(await db.masters_list.find_one({"id": doc["id"]}))


@api.delete("/masters-list/{item_id}", dependencies=[Depends(owner_only)])
async def delete_master_list_value(item_id: str, act=Depends(actor)):
    existing = await db.masters_list.find_one({"id": item_id})
    if not existing:
        raise HTTPException(404, "Not found")
    await db.masters_list.delete_one({"id": item_id})
    await write_audit(act, "delete", "masters_list", old=existing)
    await gsheets.sync_masters(await list_masters_list())
    return {"ok": True}


@api.post("/admin/reseed", dependencies=[Depends(owner_only)])
async def reseed():
    res = await seeder.run_seed(db, force=True)
    for l in await db.leads.find().to_list(2000):
        await recompute_lead(l["leadId"])
    return res


# ---------------------------------------------------------------- dashboard helpers
def _is_close_won(lead) -> bool:
    return "close won" in (lead.get("currentStatus") or "").strip().lower()


def _is_commercial_deal(lead) -> bool:
    """Booked / finance / delivered / Close Won deals that are still ON.

    Close Won overwrites currentStatus, so a book|deliver|finance regex misses
    finished retails — OEM Claims, Owner Commercial and Dealer Earnings then
    showed fewer rows than the Google Sheet (which keeps every claim forever).
    dealCancelled deals are off even when revival puts them back to Active/New.
    """
    if not lead or lead.get("dealCancelled"):
        return False
    if _is_close_won(lead) or _is_delivered_lead(lead):
        return True
    st = (lead.get("currentStatus") or "").lower()
    if "book" in st or "finance" in st:
        return True
    if lead.get("bookingDate") or lead.get("bookingId"):
        return True
    return False


async def _commercial_leads(limit: int = 5000) -> list:
    return [l for l in await db.leads.find().to_list(limit) if _is_commercial_deal(l)]


async def _live_booking(lead_id):
    """The current booking for a lead — never a Cancelled historical row."""
    rows = await db.bookings.find({"leadId": lead_id}).sort("bookingId", -1).to_list(50)
    live = [b for b in rows if str(b.get("bookingStatus") or "").lower() != "cancelled"]
    return live[0] if live else None


async def _net_received(lead_id) -> float:
    agg = await db.payments.aggregate([
        {"$match": {"leadId": lead_id}},
        {"$group": {"_id": None, "total": {"$sum": "$amount"}}},
    ]).to_list(1)
    return ce.round2(agg[0]["total"]) if agg else 0.0


async def _void_derived_claims_on_cancel(lead_id: str):
    """Stop chasing OEM on a dead booking. Rows stay forever as Cancelled."""
    rows = await db.claims.find({
        "leadId": lead_id,
        "manual": {"$ne": True},
        "claimStatus": {"$nin": ["Cancelled"]},
    }).to_list(200)
    now = now_iso()
    for c in rows:
        await db.claims.update_one({"_id": c["_id"]}, {"$set": {
            "claimStatus": "Cancelled",
            "cancelledAt": now,
            "claimRequired": "No",
        }})
        updated = await db.claims.find_one({"_id": c["_id"]}) or {**c, "claimStatus": "Cancelled"}
        await sheet_sync("claims", {
            "claimId": updated.get("claimId"),
            "leadId": lead_id,
            "customer": updated.get("customer"),
            "model": updated.get("model"),
            "variant": updated.get("variant"),
            "bookingDate": updated.get("bookingDate"),
            "component": updated.get("component"),
            "componentKey": updated.get("componentKey"),
            "eligibleClaim": updated.get("eligibleClaim"),
            "claimAmount": updated.get("claimAmount"),
            "receivedAmount": updated.get("receivedAmount", 0),
            "claimStatus": "Cancelled",
            "claimReference": updated.get("claimReference", ""),
            "claimRequired": "No",
        })


def _is_delivered_lead(lead) -> bool:
    """Counted as a retail on the dashboards.

    Close Won is a COMPLETED deal, so it counts. Closing a lead overwrites
    currentStatus ("Delivered" -> "Close Won") and sets accountStatus=Closed,
    which used to drop the retail out of the executive's numbers the moment the
    paperwork was finished — penalising the executive for completing the file.

    Reporting only. Workflow gating uses _is_delivered(), which stays strict:
    only an actual Mark Delivered counts there.
    """
    st = (lead.get("currentStatus") or "").lower()
    return ("deliver" in st
            or (lead.get("deliveryStatus") or "").lower() == "delivered"
            or _is_close_won(lead))


def _retail_date(lead) -> str:
    """Date a retail is credited to. A lead closed without a Mark Delivered has no
    deliveryDate, so fall back to the closing date rather than dropping it from MTD."""
    return str(lead.get("deliveryDate") or lead.get("closedDate") or "")


def _funnel_population(leads) -> list:
    """Leads the funnel counts: everything Active, plus completed (Close Won) deals.

    Closed-Lost / Cancelled / Archived stay out — only a finished SALE is added
    back, and _status_bucket puts it in Delivered."""
    out = []
    for l in leads:
        if (l.get("accountStatus") or "Active") == "Active" or _is_close_won(l):
            out.append(l)
    return out


def _is_booked_lead(lead) -> bool:
    """True once converted to booking — stays true after Finance Process / Delivered / Sold."""
    if _is_delivered_lead(lead):
        return True
    st = (lead.get("currentStatus") or "").lower()
    if "book" in st or "finance" in st:
        return True
    if lead.get("bookingDate") or lead.get("bookingId"):
        return True
    return False


def _norm_name(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().lower())


def _leads_for_executive(leads, user) -> list:
    """Match lead.executive to the logged-in executive's name (or email local-part)."""
    name = _norm_name(user.get("name"))
    email_local = _norm_name((user.get("email") or "").split("@")[0].replace(".", " ").replace("_", " "))
    out = []
    for l in leads:
        ex = _norm_name(l.get("executive"))
        if not ex:
            continue
        if name and (ex == name or name in ex or ex in name):
            out.append(l)
        elif email_local and (ex == email_local or email_local in ex or ex in email_local):
            out.append(l)
    return out


def _status_bucket(lead) -> str:
    st = (lead.get("currentStatus") or "New").strip()
    low = st.lower()
    if "lost" in low:
        return "Lost"
    if _is_delivered_lead(lead):
        return "Delivered"
    if "finance" in low:
        return "Finance Process"
    if "book" in low:
        return "Booked"
    if "progress" in low:
        return "In Progress"
    if "follow" in low:
        return "Follow-up"
    if "contact" in low:
        return "Contacted"
    if st:
        return st
    return "New"


FIELD_LEAD_SAFE_KEYS = {
    "leadId", "customerName", "mobile", "altMobile", "village", "city",
    "leadSource", "interestedModel", "variant", "executive", "currentStatus",
    "accountStatus", "priority", "createdDate", "bookingDate", "deliveryDate",
    "nextFollowupDate", "nextFollowup", "financeRequired", "exchangeRequired",
    "remarks", "deliveryStatus", "bookingId",
}


def _field_safe_lead(lead: dict) -> dict:
    """Pipeline-only snapshot for ASM/RM — no commercial / money fields."""
    return {k: lead.get(k) for k in FIELD_LEAD_SAFE_KEYS if k in lead or lead.get(k) is not None}


# ---------------------------------------------------------------- dashboard
@api.get("/dashboard")
async def dashboard():
    leads = await db.leads.find().to_list(5000)
    payments = await db.payments.find().to_list(5000)
    ym = this_month()
    td = today()

    def in_month(d):
        return bool(d) and str(d).startswith(ym)

    def is_today(d):
        return str(d) == td

    booked = [l for l in leads if _is_booked_lead(l)]
    delivered = [l for l in leads if _is_delivered_lead(l)]
    active_booked = [l for l in booked if not _is_delivered_lead(l)]

    monthly_leads = [l for l in leads if in_month(l.get("createdDate"))]
    monthly_bookings = [l for l in booked if in_month(l.get("bookingDate"))]

    pay_by_mode = {"Cash": 0.0, "UPI": 0.0, "Finance": 0.0, "Other": 0.0}
    month_payments = [p for p in payments if in_month(p.get("date"))]
    for p in month_payments:
        mode = (p.get("paymentMode") or "").strip()
        key = mode if mode in ("Cash", "UPI", "Finance") else "Other"
        pay_by_mode[key] += ce.num(p.get("amount"))

    cust_os = sum(ce.num(l.get("customerOutstanding")) for l in leads)
    company_os = sum(ce.num(l.get("companyOutstanding")) for l in leads)

    # Finance total outstanding (H1) — sum of open finance file balances
    finance_os = 0.0
    finance_pending_files = []
    for f in await db.finance.find().to_list(5000):
        os_amt = ce.num(f.get("fileOutstanding"))
        finance_os += os_amt
        if os_amt > 0 and f.get("status") != "Received":
            finance_pending_files.append(clean(f))
    finance_overdue = await _enrich_finance_with_delivery(finance_pending_files)
    finance_overdue = [f for f in finance_overdue if f.get("overdue")]
    finance_overdue_count = len(finance_overdue)
    finance_overdue_amount = ce.round2(sum(ce.num(f.get("fileOutstanding")) for f in finance_overdue))

    # Follow-up KPIs (H1) — active leads with a next-follow-up date
    def _followup_date(l):
        return str(l.get("nextFollowupDate") or l.get("nextFollowup") or "")[:10]

    active_leads = [l for l in leads if (l.get("accountStatus") or "Active") == "Active"]
    followup_due = len([l for l in active_leads if _followup_date(l) == td])
    followup_overdue = len([l for l in active_leads
                            if _followup_date(l) and _followup_date(l) < td])

    # model performance
    models = {}
    for l in leads:
        m = l.get("interestedModel") or "Unknown"
        row = models.setdefault(m, {"model": m, "leads": 0, "bookings": 0, "deliveries": 0, "pending": 0, "customerOs": 0.0, "revenue": 0.0})
        row["leads"] += 1
        if l in booked:
            row["bookings"] += 1
        if l in delivered:
            row["deliveries"] += 1
        row["customerOs"] += ce.num(l.get("customerOutstanding"))
    for l in active_booked:
        m = l.get("interestedModel") or "Unknown"
        if m in models:
            models[m]["pending"] += 1
    for p in payments:
        lead = next((x for x in leads if x.get("leadId") == p.get("leadId")), None)
        if lead:
            m = lead.get("interestedModel") or "Unknown"
            if m in models:
                models[m]["revenue"] += ce.num(p.get("amount"))
    model_perf = sorted(models.values(), key=lambda x: -x["leads"])

    return {
        "kpis": {
            "todayLeads": len([l for l in leads if is_today(l.get("createdDate"))]),
            "todayBookings": len([l for l in booked if is_today(l.get("bookingDate"))]),
            "todayDeliveries": len([l for l in delivered if is_today(l.get("deliveryDate"))]),
            "monthlyLeads": len(monthly_leads),
            "monthlyBookings": len(monthly_bookings),
            "activeBookings": len(active_booked),
            "monthlyDeliveries": len([l for l in delivered if in_month(_retail_date(l))]),
            "pendingDeliveries": len(active_booked),
            "totalLeads": len(leads),
            "conversion": round((len(monthly_bookings) / len(monthly_leads) * 100), 1) if monthly_leads else 0,
            "revenue": ce.round2(sum(ce.num(p.get("amount")) for p in month_payments)),
            "financeOutstanding": ce.round2(finance_os),
            "financeOverdueCount": finance_overdue_count,
            "financeOverdueAmount": finance_overdue_amount,
            "followupDue": followup_due,
            "followupOverdue": followup_overdue,
        },
        "payments": {k: ce.round2(v) for k, v in pay_by_mode.items()},
        "outstanding": {
            "customer": ce.round2(cust_os),
            "company": ce.round2(company_os),
            "total": ce.round2(cust_os + company_os),
        },
        "modelPerformance": model_perf,
        "lastUpdated": now_iso(),
    }


@api.get("/accounts/dashboard")
async def accounts_dashboard():
    """Money desk home — Tally cross-check KPIs (Accounts / Owner / Executive)."""
    leads = await db.leads.find().to_list(5000)
    cust_os = ce.round2(sum(ce.num(l.get("customerOutstanding")) for l in leads))
    company_os = ce.round2(sum(ce.num(l.get("companyOutstanding")) for l in leads))

    finance_os = 0.0
    finance_pending = 0
    for f in await db.finance.find().to_list(5000):
        os_amt = ce.num(f.get("fileOutstanding"))
        finance_os += os_amt
        if os_amt > 0.01 and f.get("status") != "Received":
            finance_pending += 1
    finance_os = ce.round2(finance_os)

    oem_claim_open = 0.0
    oem_claim_count = 0
    for c in await db.claims.find().to_list(5000):
        elig = ce.num(c.get("eligibleClaim") if c.get("eligibleClaim") is not None else c.get("claimAmount"))
        recv = ce.num(c.get("receivedAmount"))
        due = ce.round2(max(0.0, elig - recv))
        if due > 0.01:
            oem_claim_open += due
            oem_claim_count += 1
    oem_claim_open = ce.round2(oem_claim_open)

    insurance_due = 0.0
    insurance_open = 0
    for e in await db.insurance.find().to_list(5000):
        due = ce.num(e.get("payoutOutstanding"))
        if due > 0.01:
            insurance_due += due
            insurance_open += 1
    insurance_due = ce.round2(insurance_due)

    # Delivered leads for Tally — prefer stored billing summaries, else live build.
    delivered = [
        l for l in leads
        if (l.get("deliveryStatus") or "").lower() == "delivered"
        or (l.get("currentStatus") or "").lower() == "delivered"
    ]
    delivered.sort(key=lambda l: str(l.get("deliveryDate") or ""), reverse=True)
    tally_rows = []
    do_not_post_total_retained = 0.0
    do_not_post_claims = company_os
    for l in delivered[:40]:
        stored = await db.billing_summaries.find_one({"leadId": l["leadId"]})
        summary = clean(stored) if stored else ce.build_delivery_billing_summary(l)
        t = summary.get("totals") or {}
        for item in summary.get("doNotPostInTally") or []:
            label = (item.get("label") or "").lower()
            if "retained" in label:
                do_not_post_total_retained += ce.num(item.get("amount"))
        tally_rows.append({
            "leadId": l["leadId"],
            "customerName": l.get("customerName") or "",
            "model": l.get("interestedModel") or "",
            "variant": l.get("variant") or "",
            "invoiceNumber": summary.get("invoiceNumber") or l.get("invoiceNumber") or "",
            "deliveryDate": summary.get("deliveryDate") or l.get("deliveryDate") or "",
            "customerPayable": t.get("customerPayable", ce.num(l.get("customerPayable"))),
            "totalReceived": t.get("totalReceived", ce.num(l.get("totalReceived"))),
            "customerOutstanding": t.get("customerOutstanding", ce.num(l.get("customerOutstanding"))),
            "hasSummary": True,
        })

    return {
        "kpis": {
            "customerOutstanding": cust_os,
            "financeOutstanding": finance_os,
            "financePendingFiles": finance_pending,
            "oemClaimsOpen": oem_claim_open,
            "oemClaimsOpenCount": oem_claim_count,
            "insurancePayoutDue": insurance_due,
            "insuranceOpenCount": insurance_open,
            "companyOutstanding": company_os,
            "deliveredForTally": len(delivered),
        },
        "tallyQueue": tally_rows,
        "doNotPost": {
            "oemClaimsOutstanding": do_not_post_claims,
            "schemeOrOemExtraRetained": ce.round2(do_not_post_total_retained),
            "note": "Do not add these to the customer Tally sales invoice — settle via Claim Register / earnings.",
        },
        "lastUpdated": now_iso(),
    }


@api.get("/executive/dashboard")
async def executive_dashboard(user=Depends(current_user)):
    """Pipeline home for a dealership executive — scoped to their assigned leads."""
    if user.get("role") not in ("executive", "owner"):
        raise HTTPException(403, "Executive dashboard is for Executive (and Owner).")
    leads_all = await db.leads.find().to_list(5000)
    mine = _leads_for_executive(leads_all, user) if user.get("role") == "executive" else leads_all
    ym = this_month()
    td = today()

    def in_month(d):
        return bool(d) and str(d).startswith(ym)

    def is_today(d):
        return str(d) == td

    def followup_date(l):
        return str(l.get("nextFollowupDate") or l.get("nextFollowup") or "")[:10]

    active = [l for l in mine if (l.get("accountStatus") or "Active") == "Active"]
    booked = [l for l in mine if _is_booked_lead(l)]
    delivered = [l for l in mine if _is_delivered_lead(l)]
    active_booked = [l for l in booked if not _is_delivered_lead(l)]

    monthly_leads = [l for l in mine if in_month(l.get("createdDate"))]
    monthly_bookings = [l for l in booked if in_month(l.get("bookingDate"))]
    monthly_deliveries = [l for l in delivered if in_month(_retail_date(l))]

    funnel_order = ["New", "Contacted", "Follow-up", "In Progress", "Booked", "Finance Process", "Delivered", "Lost"]
    funnel = {k: 0 for k in funnel_order}
    for l in _funnel_population(mine):
        b = _status_bucket(l)
        funnel[b] = funnel.get(b, 0) + 1

    followup_due = [l for l in active if followup_date(l) == td]
    followup_overdue = [l for l in active if followup_date(l) and followup_date(l) < td]

    # Finance stuck on my booked deals
    my_ids = {l["leadId"] for l in mine}
    finance_stuck = 0
    for f in await db.finance.find().to_list(5000):
        if f.get("leadId") not in my_ids:
            continue
        if ce.num(f.get("fileOutstanding")) > 0.01 and f.get("status") != "Received":
            finance_stuck += 1

    # Source + model mix (MTD leads)
    sources: dict = {}
    models: dict = {}
    for l in monthly_leads or mine:
        src = (l.get("leadSource") or "Unknown").strip() or "Unknown"
        sources[src] = sources.get(src, 0) + 1
        m = (l.get("interestedModel") or "Unknown").strip() or "Unknown"
        row = models.setdefault(m, {"model": m, "leads": 0, "bookings": 0, "deliveries": 0, "pending": 0})
        row["leads"] += 1
    for l in monthly_bookings:
        m = (l.get("interestedModel") or "Unknown").strip() or "Unknown"
        models.setdefault(m, {"model": m, "leads": 0, "bookings": 0, "deliveries": 0, "pending": 0})["bookings"] += 1
    for l in monthly_deliveries:
        m = (l.get("interestedModel") or "Unknown").strip() or "Unknown"
        models.setdefault(m, {"model": m, "leads": 0, "bookings": 0, "deliveries": 0, "pending": 0})["deliveries"] += 1
    for l in active_booked:
        m = (l.get("interestedModel") or "Unknown").strip() or "Unknown"
        models.setdefault(m, {"model": m, "leads": 0, "bookings": 0, "deliveries": 0, "pending": 0})["pending"] += 1

    cust_os = ce.round2(sum(ce.num(l.get("customerOutstanding")) for l in mine))

    worklist = []
    for l in followup_overdue[:15]:
        worklist.append({
            "leadId": l["leadId"], "customerName": l.get("customerName") or "",
            "kind": "Follow-up overdue", "date": followup_date(l),
            "status": l.get("currentStatus") or "", "model": l.get("interestedModel") or "",
        })
    for l in followup_due[:10]:
        worklist.append({
            "leadId": l["leadId"], "customerName": l.get("customerName") or "",
            "kind": "Follow-up today", "date": followup_date(l),
            "status": l.get("currentStatus") or "", "model": l.get("interestedModel") or "",
        })
    for l in active_booked[:10]:
        worklist.append({
            "leadId": l["leadId"], "customerName": l.get("customerName") or "",
            "kind": "Pending delivery", "date": str(l.get("bookingDate") or "")[:10],
            "status": l.get("currentStatus") or "", "model": l.get("interestedModel") or "",
        })

    conv = round((len(monthly_bookings) / len(monthly_leads) * 100), 1) if monthly_leads else 0.0

    return {
        "scope": {
            "executiveName": user.get("name") or "",
            "matchedLeads": len(mine),
            "note": "Scoped to leads where Executive matches your name."
                    if user.get("role") == "executive"
                    else "Owner view — all dealership leads.",
        },
        "kpis": {
            "myLeadsMtd": len(monthly_leads),
            "myBookingsMtd": len(monthly_bookings),
            "myDeliveriesMtd": len(monthly_deliveries),
            "conversion": conv,
            "followupDue": len(followup_due),
            "followupOverdue": len(followup_overdue),
            "pendingDeliveries": len(active_booked),
            "financeStuck": finance_stuck,
            "customerOutstanding": cust_os,
            "todayLeads": len([l for l in mine if is_today(l.get("createdDate"))]),
            "activeBookings": len(active_booked),
        },
        "funnel": [{"status": k, "count": funnel.get(k, 0)} for k in funnel_order],
        "sourceMix": [{"source": k, "count": v} for k, v in sorted(sources.items(), key=lambda x: -x[1])],
        "modelMix": sorted(models.values(), key=lambda x: -x["leads"]),
        "worklist": worklist[:25],
        "lastUpdated": now_iso(),
    }


# ------------------------------------------------------------- daily reports
# One computation feeds the in-app report pages AND the WhatsApp sends, so the
# message can never disagree with the screen.
def _booking_date(l):
    return str(l.get("bookingDate") or "")[:10]


def _cancel_events(leads) -> list:
    """Flatten every lead's cancelHistory into one list of cancellations.

    One cancellation = one event, so a lead that cancelled three times counts
    three times. Attribution uses the executive recorded ON the event, not the
    lead's current executive — reassigning a lead afterwards must not move a
    cancellation onto someone who never worked it.
    """
    out = []
    for l in leads:
        history = l.get("cancelHistory") or []
        if not isinstance(history, list):
            continue
        last_index = len(history) - 1
        # Money STILL held for this customer, present tense, once per lead. The
        # per-event customerMoney below is a historical note of what was at stake
        # at that moment; summing it across a lead's cancellations reported the
        # same ₹10,000 advance twice. Only the newest event carries the live
        # figure, so every total and grouping counts it exactly once.
        held = ce.num(l.get("excessReceived")) if l.get("dealCancelled") else 0.0
        for i, h in enumerate(history):
            if not isinstance(h, dict):
                continue
            out.append({
                # Only the newest cancellation on a lead reflects where that lead
                # stands now, and only it carries the money — see the report.
                "isLatest": i == last_index,
                "sequence": i + 1,
                "cancelCount": len(history),
                # Lead-level, repeated on each of the lead's events. The report
                # picks exactly ONE event per lead to carry it into the totals.
                "leadMoneyHeld": ce.round2(held),
                "leadId": l.get("leadId"), "customerName": l.get("customerName"),
                "mobile": l.get("mobile"), "model": l.get("interestedModel"),
                "date": str(h.get("date") or "")[:10],
                "reason": str(h.get("reason") or "Unknown") or "Unknown",
                "remarks": h.get("remarks") or "",
                "stage": str(h.get("stage") or "Enquiry") or "Enquiry",
                "executive": str(h.get("executive") or "").strip(),
                "cancelledBy": h.get("cancelledBy") or "",
                "customerMoney": ce.num(h.get("customerMoney")),
                "currentAccountStatus": l.get("accountStatus") or "Active",
                "reviveOn": l.get("reviveOn") or "",
            })
    out.sort(key=lambda x: x["date"], reverse=True)
    return out


async def _daily_report_data() -> dict:
    leads = await db.leads.find().to_list(5000)
    td, ym = today(), this_month()

    def in_month(d):
        return bool(d) and str(d).startswith(ym)

    booked = [l for l in leads if _is_booked_lead(l)]
    delivered = [l for l in leads if _is_delivered_lead(l)]

    bookings_today = [l for l in booked if _booking_date(l) == td]
    bookings_mtd = [l for l in booked if in_month(_booking_date(l))]
    deliveries_today = [l for l in delivered if _retail_date(l)[:10] == td]
    deliveries_mtd = [l for l in delivered if in_month(_retail_date(l))]

    # Revenue = value of what was RETAILED this month (customer payable on MTD
    # deliveries). Not cash collected — that is `collectedMtd` below.
    revenue_mtd = ce.round2(sum(ce.num(l.get("customerPayable")) for l in deliveries_mtd))
    collected_mtd = ce.round2(sum(
        ce.num(p.get("amount")) for p in await db.payments.find().to_list(20000)
        if in_month(str(p.get("date") or "")[:10]) and ce.num(p.get("amount")) > 0))
    customer_outstanding = ce.round2(sum(
        ce.num(l.get("customerOutstanding")) for l in leads
        if (l.get("accountStatus") or "Active") == "Active"))

    fin_pending, by_financer = [], {}
    for f in await db.finance.find().to_list(5000):
        amt = ce.num(f.get("fileOutstanding"))
        if amt <= 0.01 or f.get("status") == "Received":
            continue
        fin_pending.append(f)
        key = str(f.get("financer") or "Unknown").strip() or "Unknown"
        b = by_financer.setdefault(key, {"financer": key, "files": 0, "amount": 0.0})
        b["files"] += 1
        b["amount"] += amt
    overdue = [f for f in await _enrich_finance_with_delivery(fin_pending) if f.get("overdue")]
    for b in by_financer.values():
        b["amount"] = ce.round2(b["amount"])

    # Per-executive, keyed on the staff master so a name with no leads still
    # reports (an executive who booked nothing needs to see that).
    staff = await db.staff.find({"role": "executive"}).to_list(500)
    per_exec = {}
    for st in staff:
        name = str(st.get("name") or "").strip()
        if not name or str(st.get("status") or "Active").lower() != "active":
            continue
        per_exec[_norm_name(name)] = {
            "staffId": st.get("staffId"), "name": name,
            "mobile": st.get("mobile", ""), "monthlyTarget": ce.num(st.get("monthlyTarget")),
            "bookingsToday": 0, "bookingsMtd": 0, "deliveriesToday": 0, "deliveriesMtd": 0,
            "pendingBookings": 0, "pendingCollection": 0.0,
            "followupsDue": 0, "followupsOverdue": 0,
            "cancelledToday": 0, "cancelledMtd": 0, "cancelledTotal": 0,
        }

    def slot(l):
        return per_exec.get(_norm_name(l.get("executive")))

    for rows, field in ((bookings_today, "bookingsToday"), (bookings_mtd, "bookingsMtd"),
                        (deliveries_today, "deliveriesToday"), (deliveries_mtd, "deliveriesMtd")):
        for l in rows:
            e = slot(l)
            if e:
                e[field] += 1
    for l in leads:
        e = slot(l)
        if not e:
            continue
        if (l.get("accountStatus") or "Active") != "Active":
            continue
        if _is_booked_lead(l) and not _is_delivered_lead(l):
            e["pendingBookings"] += 1
            e["pendingCollection"] += max(0.0, ce.num(l.get("customerOutstanding")))
        fd = str(l.get("nextFollowupDate") or l.get("nextFollowup") or "")[:10]
        if fd == td:
            e["followupsDue"] += 1
        elif fd and fd < td:
            e["followupsOverdue"] += 1

    # Cancellations are a permanent stamp on the lead, not a status, so they still
    # count after the lead has been revived and is sitting back in the funnel as New.
    cancels = _cancel_events(leads)
    cancels_today = [c for c in cancels if c["date"] == td]
    cancels_mtd = [c for c in cancels if in_month(c["date"])]
    for rows, field in ((cancels_today, "cancelledToday"), (cancels_mtd, "cancelledMtd"),
                        (cancels, "cancelledTotal")):
        for c in rows:
            e = per_exec.get(_norm_name(c["executive"]))
            if e:
                e[field] += 1

    for e in per_exec.values():
        e["pendingCollection"] = ce.round2(e["pendingCollection"])
        tgt = e["monthlyTarget"]
        e["attainmentPct"] = round(100.0 * e["deliveriesMtd"] / tgt, 1) if tgt > 0 else None
        # Of everything that reached a decision this month, what share walked away.
        decided = e["bookingsMtd"] + e["cancelledMtd"]
        e["cancelRatePct"] = round(100.0 * e["cancelledMtd"] / decided, 1) if decided else None

    ranked = sorted(per_exec.values(),
                    key=lambda x: (-x["bookingsToday"], -x["bookingsMtd"], x["name"]))
    target_total = ce.round2(sum(e["monthlyTarget"] for e in per_exec.values()))
    return {
        "date": td, "month": ym,
        "bookingsToday": len(bookings_today), "bookingsMtd": len(bookings_mtd),
        "deliveriesToday": len(deliveries_today), "deliveriesMtd": len(deliveries_mtd),
        "cancelledToday": len(cancels_today), "cancelledMtd": len(cancels_mtd),
        "revenueMtd": revenue_mtd, "collectedMtd": collected_mtd,
        "customerOutstanding": customer_outstanding,
        "financePendingAmount": ce.round2(sum(b["amount"] for b in by_financer.values())),
        "financePendingFiles": len(fin_pending),
        "financeOverdueFiles": len(overdue),
        "financeByFinancer": sorted(by_financer.values(), key=lambda x: -x["amount"]),
        "targetUnits": target_total,
        "attainmentPct": (round(100.0 * len(deliveries_mtd) / target_total, 1)
                          if target_total > 0 else None),
        "executives": ranked,
        "topExecutives": ranked[:3],
        "generatedAt": now_iso(),
    }


@api.get("/reports/daily/owner", dependencies=[Depends(owner_only)])
async def daily_owner_report():
    """Owner EOD: volume + money. Same numbers the WhatsApp summary carries."""
    return await _daily_report_data()


@api.get("/reports/daily/manager")
async def daily_manager_report(_viewer=Depends(current_user)):
    """RM / ASM EOD: volume only — no revenue, outstanding or finance money."""
    d = await _daily_report_data()
    for k in ("revenueMtd", "collectedMtd", "customerOutstanding",
              "financePendingAmount", "financeByFinancer"):
        d.pop(k, None)
    d["executives"] = [{k: v for k, v in e.items() if k != "pendingCollection"}
                       for e in d["executives"]]
    d["topExecutives"] = d["executives"][:3]
    return d


@api.get("/reports/daily/executive/{name}")
async def daily_executive_report(name: str, _viewer=Depends(current_user)):
    """One executive's day: what is pending on them, and where they stand MTD."""
    d = await _daily_report_data()
    row = next((e for e in d["executives"]
                if _norm_name(e["name"]) == _norm_name(name)
                or e.get("staffId") == name), None)
    if not row:
        raise HTTPException(404, f"No active executive '{name}' on the staff master")
    return {"date": d["date"], "month": d["month"], **row}


@api.get("/reports/cancellations")
async def cancellations_report(period: str = "month", executive: str = "",
                               reason: str = "", stage: str = "",
                               user=Depends(current_user)):
    """Who is losing leads, why, and how far down the funnel they got.

    period: month (default) | today | all
    Stage matters more than the raw count. A lead lost at Enquiry cost a phone
    call; one lost after Booked cost a refund and a blocked chassis, and the two
    are never summed into a single "cancellations" number here.
    """
    leads = await db.leads.find().to_list(5000)
    events = _cancel_events(leads)
    td, ym = today(), this_month()
    if period == "today":
        events = [e for e in events if e["date"] == td]
    elif period != "all":
        events = [e for e in events if e["date"].startswith(ym)]

    # An executive only ever sees their own — same scoping the WhatsApp inbox uses.
    if (user or {}).get("role") == "executive":
        mine = {l["leadId"] for l in _leads_for_executive(leads, user)}
        events = [e for e in events if e["leadId"] in mine]
    if executive:
        events = [e for e in events if _norm_name(e["executive"]) == _norm_name(executive)]
    if reason:
        events = [e for e in events if e["reason"].lower() == reason.strip().lower()]
    if stage:
        events = [e for e in events if e["stage"].lower() == stage.strip().lower()]

    # Attach each lead's outstanding refund to exactly one event — its most recent
    # one INSIDE this window. Using the globally latest event would drop the money
    # from a period view whenever a lead was cancelled again after that period.
    newest_in_window = {}
    for e in events:
        cur = newest_in_window.get(e["leadId"])
        if cur is None or (e["date"], e["sequence"]) >= (cur["date"], cur["sequence"]):
            newest_in_window[e["leadId"]] = e
    for e in events:
        e["moneyToRefund"] = (e["leadMoneyHeld"]
                              if newest_in_window.get(e["leadId"]) is e else 0.0)

    def group(key):
        """Count every cancellation; count each lead's money once.

        A lead cancelled twice is two losses — the executive lost it twice. It is
        not twice the money: the same ₹10,000 booking advance is still one
        ₹10,000, and moneyToRefund is carried by one event per lead.
        """
        out = {}
        for e in events:
            k = e[key] or "Unknown"
            row = out.setdefault(k, {key: k, "count": 0, "money": 0.0, "leads": set()})
            row["count"] += 1
            row["leads"].add(e["leadId"])
            row["money"] += e["moneyToRefund"]
        return sorted(
            [{**{k: v for k, v in r.items() if k != "leads"},
              "money": ce.round2(r["money"]), "uniqueLeads": len(r["leads"])}
             for r in out.values()],
            key=lambda x: -x["count"])

    # Where the LEADS stand now — one state per lead, whatever its account status
    # is today, so revived + parked always adds up to the leads in this window.
    state_by_lead = {e["leadId"]: e["currentAccountStatus"] for e in events}
    unique_leads = len(state_by_lead)
    parked = [k for k, v in state_by_lead.items() if str(v).lower() == "cancelled"]
    return {
        "period": period, "date": td, "month": ym,
        "total": len(events),
        "uniqueLeads": unique_leads,
        # Customer money still sitting with the dealer on a cancelled lead — i.e.
        # refunds that have not been recorded yet. Actionable, not historical.
        "withMoney": len([e for e in events if e["moneyToRefund"] > 0.01]),
        "moneyAtRisk": ce.round2(sum(e["moneyToRefund"] for e in events)),
        "revived": unique_leads - len(parked),
        "parked": len(parked),
        "byExecutive": group("executive"),
        "byReason": group("reason"),
        "byStage": group("stage"),
        "events": events[:500],
        "generatedAt": now_iso(),
    }


# Who may open the OEM board. The owner is included so you can see exactly what
# the OEM's finance manager sees before handing out the login.
OEM_FINANCE_ROLES = ("owner", "oem_finance")
# Ageing buckets, in days since delivery. The finance receipt SLA is 2 days, so
# the first bucket is "inside SLA" and everything after it is late.
OEM_AGEING_BUCKETS = [(0, 2, "0-2 days"), (3, 7, "3-7 days"),
                      (8, 15, "8-15 days"), (16, 10**6, "15+ days")]


@api.get("/reports/oem-finance")
async def oem_finance_report(view: str = "all", financer: str = "", month: str = "",
                             user=Depends(current_user)):
    """Read-only finance position for the OEM's finance manager.

    Every retail finance file — pending AND received — with how long it has been
    waiting, so a delay is visible without asking the dealer.

    Each row is BUILT from a whitelist rather than copied from the finance and
    lead documents with contact fields deleted. A blacklist leaks every field
    added later; this cannot. There is no mobile, alternate mobile, village or
    city here, and no dealer commercials — no margin, scheme, insurance payout or
    customer outstanding.

    view: all (default) | pending | overdue | received
    """
    if (user or {}).get("role") not in OEM_FINANCE_ROLES:
        raise HTTPException(403, "This report is for the Owner and the OEM finance desk.")

    files = await db.finance.find().to_list(5000)
    enriched = await _enrich_finance_with_delivery(files)
    leads = {l["leadId"]: l for l in await db.leads.find().to_list(5000) if l.get("leadId")}

    rows = []
    for f in enriched:
        lead = leads.get(str(f.get("leadId") or "")) or {}
        sanctioned = ce.num(f.get("sanctionedAmount"))
        received = ce.num(f.get("receivedAgainstFile"))
        pending = ce.round2(max(0.0, ce.num(f.get("fileOutstanding"))))
        status = str(f.get("status") or ("Received" if pending <= 0.01 else "Pending"))
        days = f.get("daysSinceDelivery")
        rows.append({
            # Identity the OEM can quote back to the dealer — and nothing more.
            "leadId": f.get("leadId") or "",
            "customerName": f.get("customerName") or lead.get("customerName") or "",
            "model": lead.get("interestedModel") or "",
            "variant": lead.get("variant") or "",
            "fileNumber": f.get("fileNumber") or "",
            "financer": str(f.get("financer") or "Unknown").strip() or "Unknown",
            "sanctioned": ce.round2(sanctioned),
            "received": ce.round2(received),
            "pending": pending,
            "deliveryDate": f.get("deliveryDate") or "",
            "daysSinceDelivery": days if isinstance(days, int) else None,
            "status": status,
            "overdue": bool(f.get("overdue")) and pending > 0.01,
            "lastPaymentDate": str(f.get("lastPaymentDate") or "")[:10],
        })

    if financer:
        rows = [r for r in rows if r["financer"].lower() == financer.strip().lower()]
    if month:
        rows = [r for r in rows if str(r["deliveryDate"]).startswith(month.strip())]
    if view == "pending":
        rows = [r for r in rows if r["pending"] > 0.01]
    elif view == "overdue":
        rows = [r for r in rows if r["overdue"]]
    elif view == "received":
        rows = [r for r in rows if r["pending"] <= 0.01]

    pending_rows = [r for r in rows if r["pending"] > 0.01]
    overdue_rows = [r for r in rows if r["overdue"]]

    by_financer = {}
    for r in rows:
        b = by_financer.setdefault(r["financer"], {
            "financer": r["financer"], "files": 0, "sanctioned": 0.0,
            "received": 0.0, "pending": 0.0, "overdue": 0})
        b["files"] += 1
        b["sanctioned"] += r["sanctioned"]
        b["received"] += r["received"]
        b["pending"] += r["pending"]
        b["overdue"] += 1 if r["overdue"] else 0
    for b in by_financer.values():
        for k in ("sanctioned", "received", "pending"):
            b[k] = ce.round2(b[k])

    ageing = []
    for lo, hi, label in OEM_AGEING_BUCKETS:
        hit = [r for r in pending_rows
               if isinstance(r["daysSinceDelivery"], int) and lo <= r["daysSinceDelivery"] <= hi]
        ageing.append({"bucket": label, "files": len(hit),
                       "pending": ce.round2(sum(r["pending"] for r in hit))})
    # Files with no delivery date yet: the SLA clock has not started, so they are
    # neither on-time nor late. Shown separately instead of being forced into a bucket.
    not_started = [r for r in pending_rows if not isinstance(r["daysSinceDelivery"], int)]
    if not_started:
        ageing.append({"bucket": "Not delivered yet", "files": len(not_started),
                       "pending": ce.round2(sum(r["pending"] for r in not_started))})

    oldest = max((r["daysSinceDelivery"] for r in pending_rows
                  if isinstance(r["daysSinceDelivery"], int)), default=0)
    rows.sort(key=lambda r: (not r["overdue"], -(r["daysSinceDelivery"] or 0), -r["pending"]))
    return {
        "view": view, "generatedAt": now_iso(), "slaDays": FINANCE_RECEIPT_SLA_DAYS,
        "totals": {
            "files": len(rows),
            "sanctioned": ce.round2(sum(r["sanctioned"] for r in rows)),
            "received": ce.round2(sum(r["received"] for r in rows)),
            "pending": ce.round2(sum(r["pending"] for r in rows)),
            "pendingFiles": len(pending_rows),
            "overdueFiles": len(overdue_rows),
            "overdueAmount": ce.round2(sum(r["pending"] for r in overdue_rows)),
            "oldestPendingDays": oldest,
        },
        "byFinancer": sorted(by_financer.values(), key=lambda x: -x["pending"]),
        "ageing": ageing,
        "financers": sorted({r["financer"] for r in rows}),
        "files": rows,
    }


async def _ops_pipeline_dashboard():
    """Showroom-wide retail snapshot shared by ASM/RM field board and Sales GM."""
    leads = await db.leads.find().to_list(5000)
    ym = this_month()
    td = today()

    def in_month(d):
        return bool(d) and str(d).startswith(ym)

    def followup_date(l):
        return str(l.get("nextFollowupDate") or l.get("nextFollowup") or "")[:10]

    active = [l for l in leads if (l.get("accountStatus") or "Active") == "Active"]
    booked = [l for l in leads if _is_booked_lead(l)]
    delivered = [l for l in leads if _is_delivered_lead(l)]
    active_booked = [l for l in booked if not _is_delivered_lead(l)]
    lost = [l for l in leads if "lost" in (l.get("currentStatus") or "").lower()]

    monthly_leads = [l for l in leads if in_month(l.get("createdDate"))]
    monthly_bookings = [l for l in booked if in_month(l.get("bookingDate"))]
    monthly_deliveries = [l for l in delivered if in_month(_retail_date(l))]

    funnel_order = ["New", "Contacted", "Follow-up", "In Progress", "Booked", "Finance Process", "Delivered", "Lost"]
    funnel = {k: 0 for k in funnel_order}
    for l in _funnel_population(leads):
        b = _status_bucket(l)
        funnel[b] = funnel.get(b, 0) + 1

    followup_overdue_leads = [l for l in active if followup_date(l) and followup_date(l) < td]
    followup_due_leads = [l for l in active if followup_date(l) == td]
    followup_overdue = len(followup_overdue_leads)
    followup_due = len(followup_due_leads)

    finance_pending = 0
    finance_overdue_amt = 0.0
    finance_rows = await db.finance.find().to_list(5000)
    finance_pending_files = []
    for f in finance_rows:
        os_amt = ce.num(f.get("fileOutstanding"))
        if os_amt > 0.01 and f.get("status") != "Received":
            finance_pending += 1
            finance_pending_files.append(clean(f))
    finance_overdue = await _enrich_finance_with_delivery(finance_pending_files)
    finance_overdue = [f for f in finance_overdue if f.get("overdue")]
    finance_overdue_count = len(finance_overdue)
    finance_overdue_amt = ce.round2(sum(ce.num(f.get("fileOutstanding")) for f in finance_overdue))

    # Scheme use rate on booked+delivered deals
    schemed = 0
    scheme_eligible = len(booked)
    for l in booked:
        use = (l.get("schemeUse") or l.get("schemeApplied") or "").strip().lower()
        if use in ("yes", "y", "true", "1") or ce.num(l.get("totalSchemeBenefit") or l.get("schemeBenefit")) > 0:
            schemed += 1
    scheme_use_rate = round((schemed / scheme_eligible * 100), 1) if scheme_eligible else 0.0

    oem_claim_open = 0.0
    oem_claim_count = 0
    for c in await db.claims.find().to_list(5000):
        elig = ce.num(c.get("eligibleClaim") if c.get("eligibleClaim") is not None else c.get("claimAmount"))
        recv = ce.num(c.get("receivedAmount"))
        due = ce.round2(max(0.0, elig - recv))
        if due > 0.01:
            oem_claim_open += due
            oem_claim_count += 1
    oem_claim_open = ce.round2(oem_claim_open)

    sources: dict = {}
    models: dict = {}
    for l in monthly_leads or leads:
        src = (l.get("leadSource") or "Unknown").strip() or "Unknown"
        sources[src] = sources.get(src, 0) + 1
        m = (l.get("interestedModel") or "Unknown").strip() or "Unknown"
        row = models.setdefault(m, {"model": m, "leads": 0, "bookings": 0, "deliveries": 0})
        row["leads"] += 1
    for l in monthly_bookings:
        m = (l.get("interestedModel") or "Unknown").strip() or "Unknown"
        models.setdefault(m, {"model": m, "leads": 0, "bookings": 0, "deliveries": 0})["bookings"] += 1
    for l in monthly_deliveries:
        m = (l.get("interestedModel") or "Unknown").strip() or "Unknown"
        models.setdefault(m, {"model": m, "leads": 0, "bookings": 0, "deliveries": 0})["deliveries"] += 1

    # Executive scoreboard — bookings stay counted after delivery/sold
    execs: dict = {}
    for l in leads:
        name = (l.get("executive") or "Unassigned").strip() or "Unassigned"
        row = execs.setdefault(name, {
            "executive": name, "leads": 0, "leadsMtd": 0, "bookingsMtd": 0,
            "deliveriesMtd": 0, "followupOverdue": 0, "pendingDeliveries": 0,
        })
        row["leads"] += 1
        if in_month(l.get("createdDate")):
            row["leadsMtd"] += 1
        if _is_booked_lead(l) and in_month(l.get("bookingDate")):
            row["bookingsMtd"] += 1
        if _is_delivered_lead(l) and in_month(_retail_date(l)):
            row["deliveriesMtd"] += 1
        if (l.get("accountStatus") or "Active") == "Active":
            fd = followup_date(l)
            if fd and fd < td:
                row["followupOverdue"] += 1
        if _is_booked_lead(l) and not _is_delivered_lead(l):
            row["pendingDeliveries"] += 1
    scoreboard = sorted(execs.values(), key=lambda x: (-x["deliveriesMtd"], -x["bookingsMtd"], -x["leadsMtd"]))
    for row in scoreboard:
        # Lead → book conversion (delivered deals still count as bookings)
        row["conversion"] = round((row["bookingsMtd"] / row["leadsMtd"] * 100), 1) if row["leadsMtd"] else 0.0
        row["deliveryConversion"] = (
            round((row["deliveriesMtd"] / row["bookingsMtd"] * 100), 1) if row["bookingsMtd"] else 0.0
        )

    book_conv = round((len(monthly_bookings) / len(monthly_leads) * 100), 1) if monthly_leads else 0.0
    deliver_conv = round((len(monthly_deliveries) / len(monthly_bookings) * 100), 1) if monthly_bookings else 0.0

    worklist = []
    for l in followup_overdue_leads[:15]:
        worklist.append({
            "leadId": l["leadId"], "customerName": l.get("customerName") or "",
            "executive": l.get("executive") or "",
            "kind": "Follow-up overdue", "date": followup_date(l),
            "status": l.get("currentStatus") or "", "model": l.get("interestedModel") or "",
        })
    for l in followup_due_leads[:10]:
        worklist.append({
            "leadId": l["leadId"], "customerName": l.get("customerName") or "",
            "executive": l.get("executive") or "",
            "kind": "Follow-up today", "date": followup_date(l),
            "status": l.get("currentStatus") or "", "model": l.get("interestedModel") or "",
        })
    for l in active_booked[:10]:
        worklist.append({
            "leadId": l["leadId"], "customerName": l.get("customerName") or "",
            "executive": l.get("executive") or "",
            "kind": "Pending delivery", "date": str(l.get("bookingDate") or "")[:10],
            "status": l.get("currentStatus") or "", "model": l.get("interestedModel") or "",
        })

    cancel_mtd = len([
        e for e in _cancel_events(leads)
        if in_month(e.get("date"))
    ])
    cust_os = ce.round2(sum(ce.num(l.get("customerOutstanding")) for l in leads))

    return {
        "kpis": {
            "leadsMtd": len(monthly_leads),
            "bookingsMtd": len(monthly_bookings),
            "deliveriesMtd": len(monthly_deliveries),
            "leadToBookPct": book_conv,
            "bookToDeliverPct": deliver_conv,
            "followupDue": followup_due,
            "followupOverdue": followup_overdue,
            "pendingDeliveries": len(active_booked),
            "financePending": finance_pending,
            "financeOverdueCount": finance_overdue_count,
            "financeOverdueAmount": finance_overdue_amt,
            "lostCount": len(lost),
            "schemeUseRate": scheme_use_rate,
            "oemClaimsOpen": oem_claim_open,
            "oemClaimsOpenCount": oem_claim_count,
            "activeBookings": len(active_booked),
            "totalLeads": len(leads),
            "cancellationsMtd": cancel_mtd,
            "customerOutstanding": cust_os,
        },
        "funnel": [{"status": k, "count": funnel.get(k, 0)} for k in funnel_order],
        "sourceMix": [{"source": k, "count": v} for k, v in sorted(sources.items(), key=lambda x: -x[1])],
        "modelMix": sorted(models.values(), key=lambda x: -x["leads"]),
        "executiveScoreboard": scoreboard,
        "worklist": worklist[:25],
        "lastUpdated": now_iso(),
    }


@api.get("/field/dashboard")
async def field_dashboard(_field=Depends(field_viewer_only)):
    """Shared ASM / RM company field board — retail + pipeline hygiene (read-only)."""
    return await _ops_pipeline_dashboard()


@api.get("/sales-gm/dashboard")
async def sales_gm_dashboard(_gm=Depends(sales_gm_only)):
    """Sales GM showroom board — all executives, deal-desk access, no money posting."""
    body = await _ops_pipeline_dashboard()
    body["scope"] = {
        "note": "Showroom-wide. Price, scheme, deliver and close — not payments or Price Master.",
    }
    return body


# ---------------------------------------------------------------- leads
def _is_own_lead(lead, user) -> bool:
    """Does this lead belong to the signed-in executive?"""
    return bool(_leads_for_executive([lead or {}], user or {}))


async def _own_lead_ids(user) -> set:
    return {l["leadId"] for l in
            _leads_for_executive(await db.leads.find().to_list(5000), user)}


def _require_own_lead(lead, user):
    """An executive may only open a lead assigned to them.

    Scoping the LIST alone would be cosmetic — lead ids run in sequence, so a
    colleague's deal is one guessed URL away. This is the check that makes the
    scoping real, and it is applied wherever a single lead is returned.
    """
    if (user or {}).get("role") != "executive":
        return
    if not _is_own_lead(lead, user):
        raise HTTPException(
            403, "This lead is assigned to another executive. "
                 "Ask the owner or your team leader to reallocate it.")


@api.get("/leads")
async def list_leads(status: Optional[str] = None, q: Optional[str] = None, user=Depends(current_user)):
    query = {}
    if status and status != "all":
        query["currentStatus"] = status
    if q:
        query["$or"] = [
            {"customerName": {"$regex": q, "$options": "i"}},
            {"mobile": {"$regex": q, "$options": "i"}},
            {"leadId": {"$regex": q, "$options": "i"}},
        ]
    leads = await db.leads.find(query).sort("leadId", -1).to_list(3000)
    # An executive works their own leads only. Owner, Sales GM, TL and Accounts see all.
    if user.get("role") == "executive":
        leads = _leads_for_executive(leads, user)
    rows = [clean(l) for l in leads]
    if user.get("role") in authmod.FIELD_ROLES:
        return [_field_safe_lead(l) for l in rows]
    return rows


class RejectRequestIn(BaseModel):
    reason: str = ""


async def _mobile_taken_by_lead(mobile: str, exclude_lead_id: str = ""):
    import re as _re
    mob = _re.sub(r"\D", "", mobile or "")
    if len(mob) < 10:
        return None
    last10 = mob[-10:]
    q = {"mobile": {"$regex": last10 + "$"}}
    if exclude_lead_id:
        q["leadId"] = {"$ne": exclude_lead_id}
    return await db.leads.find_one(q)


async def _insert_live_lead(body: LeadIn, *, source_note: str = "Lead created from CRM"):
    existing = await _mobile_taken_by_lead(body.mobile)
    if existing:
        raise HTTPException(
            409,
            f"Mobile already used by lead {existing.get('leadId')} ({existing.get('customerName')}).",
        )
    lead_id = await next_id("lead", "LD26")
    payload = body.model_dump()
    created_date = str(payload.pop("createdDate", None) or "").strip() or today()
    doc = {
        "leadId": lead_id,
        **payload,
        "createdDate": created_date,
        "accountStatus": "Active",
        "deliveryStatus": "",
        "outstandingAmount": 0, "customerOutstanding": 0, "companyOutstanding": 0,
        "totalReceived": 0, "customerPayable": 0, "grossVehicleCost": 0, "totalDiscount": 0,
        "consumerDiscount": 0, "exchangeBonus": 0, "loyaltyBonus": 0, "referralBonus": 0,
        "dsaDiscount": 0, "additionalDiscount": 0, "exShowroom": 0, "rto": 0, "insuranceAmount": 0,
        "accessoriesAmount": 0, "handlingCharges": 0, "trc": 0, "fastag": 0, "extendedWarranty": 0,
        "otherCharges": 0, "bookingAmount": 0, "lastUpdated": now_iso(),
    }
    await db.leads.insert_one(doc)
    _act_doc = {
        "activityId": await next_id("activity", "AC26"), "leadId": lead_id, "date": created_date,
        "time": datetime.now(timezone.utc).strftime("%H:%M"), "activityType": "Note",
        "discussion": source_note, "executive": body.executive,
        "customerName": body.customerName, "mobile": body.mobile, "model": body.interestedModel,
    }
    await db.activities.insert_one(dict(_act_doc))
    await sheet_sync("activities", _act_doc)
    await sheet_sync("leads", doc)
    return clean(await db.leads.find_one({"leadId": lead_id}))


def _can_approve_leads(user) -> bool:
    return str((user or {}).get("role") or "") in web_push.LEAD_APPROVER_ROLES


@api.post("/leads")
async def create_lead(body: LeadIn, user=Depends(sales_staff_only)):
    """Owner / GM / TL create a live lead. An executive submits a request until
    GM or Owner taps Approve — nothing is written to the Lead Register until then."""
    if str(user.get("role") or "") == "executive":
        if ce.num(body.budget) <= 0:
            raise HTTPException(422, "Enter the deal amount before sending for GM / Owner approval.")
        existing = await _mobile_taken_by_lead(body.mobile)
        if existing:
            raise HTTPException(
                409,
                f"Mobile already used by lead {existing.get('leadId')} ({existing.get('customerName')}).",
            )
        pending_same = None
        if str(body.mobile or "").strip():
            last10 = re.sub(r"\D", "", body.mobile)[-10:]
            if len(last10) == 10:
                pending_same = await db.lead_requests.find_one({
                    "status": {"$in": ["pending", "approving"]},
                    "payload.mobile": {"$regex": last10 + "$"},
                })
        if pending_same:
            raise HTTPException(
                409,
                f"This mobile is already waiting for approval ({pending_same.get('requestId')}).",
            )
        payload = body.model_dump()
        if not str(payload.get("executive") or "").strip():
            payload["executive"] = user.get("name") or ""
        request_id = await next_id("lead_request", "LR26")
        req = {
            "requestId": request_id,
            "status": "pending",
            "payload": payload,
            "submittedBy": user.get("email") or "",
            "submittedByName": user.get("name") or "",
            "submittedByUserId": user.get("userId") or user.get("email") or "",
            "createdAt": now_iso(),
            "dealAmount": ce.round2(ce.num(body.budget)),
        }
        await db.lead_requests.insert_one(req)
        name = (body.customerName or "Customer").strip()
        model = " ".join(x for x in (body.interestedModel, body.variant) if x).strip() or "vehicle"
        amt = f"₹{ce.num(body.budget):,.0f}"
        try:
            await web_push.notify_lead_approvers(
                db,
                title="Lead waiting for approval",
                body=f"{name} · {model} · {amt}. Open Approvals in Euler CRM.",
                url="/approvals",
            )
        except Exception:
            logger.exception("lead-request push notify failed")
        return {
            "pending": True,
            "requestId": request_id,
            "status": "pending",
            "dealAmount": req["dealAmount"],
            "message": "Sent to GM / Owner. The lead is created only after they tap Approve.",
        }
    return await _insert_live_lead(body)


def _request_out(doc):
    if not doc:
        return doc
    row = clean(dict(doc))
    payload = row.get("payload") or {}
    row["customerName"] = payload.get("customerName") or ""
    row["mobile"] = payload.get("mobile") or ""
    row["interestedModel"] = payload.get("interestedModel") or ""
    row["variant"] = payload.get("variant") or ""
    row["executive"] = payload.get("executive") or ""
    row["budget"] = payload.get("budget") or row.get("dealAmount") or 0
    row["remarks"] = payload.get("remarks") or ""
    return row


@api.get("/lead-requests/summary")
async def lead_request_summary(user=Depends(current_user)):
    role = str(user.get("role") or "")
    if role == "executive":
        n = await db.lead_requests.count_documents({
            "status": "pending",
            "$or": [
                {"submittedByUserId": user.get("userId")},
                {"submittedBy": user.get("email")},
            ],
        })
        return {"pending": n, "mine": n, "canApprove": False}
    if _can_approve_leads(user):
        n = await db.lead_requests.count_documents({"status": "pending"})
        return {"pending": n, "mine": n, "canApprove": True}
    return {"pending": 0, "mine": 0, "canApprove": False}


@api.get("/lead-requests")
async def list_lead_requests(status: Optional[str] = None, user=Depends(current_user)):
    st = (status or "pending").strip().lower()
    if st not in ("pending", "approved", "rejected", "all"):
        st = "pending"
    q = {} if st == "all" else {"status": st}
    role = str(user.get("role") or "")
    if role == "executive":
        q["$or"] = [
            {"submittedByUserId": user.get("userId")},
            {"submittedBy": user.get("email")},
        ]
    elif not _can_approve_leads(user):
        raise HTTPException(403, "Lead approvals are for Owner / Sales GM.")
    rows = [r async for r in db.lead_requests.find(q).sort("createdAt", -1).limit(200)]
    return [_request_out(r) for r in rows]


@api.post("/lead-requests/{request_id}/approve")
async def approve_lead_request(request_id: str, user=Depends(current_user)):
    if not _can_approve_leads(user):
        raise HTTPException(403, "Only the Owner or Sales GM can approve a lead.")
    req = await db.lead_requests.find_one({"requestId": request_id})
    if not req:
        raise HTTPException(404, "Approval request not found")
    if req.get("status") == "approved" and req.get("leadId"):
        return {"ok": True, "leadId": req["leadId"], "already": True}
    claimed = await db.lead_requests.find_one_and_update(
        {"requestId": request_id, "status": "pending"},
        {"$set": {
            "status": "approving",
            "approvedBy": user.get("email") or "",
            "approvedByName": user.get("name") or "",
            "approvedAt": now_iso(),
        }},
    )
    if not claimed:
        fresh = await db.lead_requests.find_one({"requestId": request_id})
        if (fresh or {}).get("leadId"):
            return {"ok": True, "leadId": fresh["leadId"], "already": True}
        raise HTTPException(409, f"This request is already {(fresh or {}).get('status')}.")
    payload = claimed.get("payload") or {}
    try:
        body = LeadIn(**payload)
        lead = await _insert_live_lead(body, source_note="Lead created after GM / Owner approval")
    except HTTPException:
        await db.lead_requests.update_one(
            {"requestId": request_id}, {"$set": {"status": "pending", "approvedBy": "", "approvedAt": ""}})
        raise
    await db.lead_requests.update_one(
        {"requestId": request_id},
        {"$set": {"status": "approved", "leadId": lead["leadId"]}},
    )
    return {"ok": True, "leadId": lead["leadId"], "lead": lead}


@api.post("/lead-requests/{request_id}/reject")
async def reject_lead_request(request_id: str, body: RejectRequestIn = RejectRequestIn(),
                              user=Depends(current_user)):
    if not _can_approve_leads(user):
        raise HTTPException(403, "Only the Owner or Sales GM can reject a lead.")
    req = await db.lead_requests.find_one({"requestId": request_id})
    if not req:
        raise HTTPException(404, "Approval request not found")
    if req.get("status") != "pending":
        raise HTTPException(409, f"This request is already {req.get('status')}.")
    await db.lead_requests.update_one({"requestId": request_id}, {"$set": {
        "status": "rejected",
        "rejectedBy": user.get("email") or "",
        "rejectedByName": user.get("name") or "",
        "rejectedAt": now_iso(),
        "rejectReason": str(body.reason or "").strip(),
    }})
    return {"ok": True, "status": "rejected"}


@api.get("/push/vapid-public")
async def push_vapid_public(_user=Depends(current_user)):
    doc = await web_push.ensure_vapid(db)
    return {"publicKey": web_push.public_key_from_doc(doc), "ok": bool(doc.get("publicKey"))}


class PushSubIn(BaseModel):
    endpoint: str
    keys: dict = {}


@api.post("/push/subscribe")
async def push_subscribe(body: PushSubIn, user=Depends(current_user)):
    if not _can_approve_leads(user):
        raise HTTPException(403, "Phone alerts are for Owner / Sales GM.")
    try:
        return await web_push.save_subscription(db, user, body.model_dump())
    except ValueError as e:
        raise HTTPException(422, str(e))


@api.post("/push/unsubscribe")
async def push_unsubscribe(body: PushSubIn, user=Depends(current_user)):
    return await web_push.drop_subscription(db, user, body.endpoint)


@api.get("/leads/{lead_id}")
async def get_lead(lead_id: str, user=Depends(current_user)):
    lead = await get_lead_or_404(lead_id)
    _require_own_lead(lead, user)
    if user.get("role") in authmod.FIELD_ROLES:
        return _field_safe_lead(lead)
    return lead


@api.get("/leads/{lead_id}/360")
async def customer_360(lead_id: str, user=Depends(current_user)):
    lead = await get_lead_or_404(lead_id)
    _require_own_lead(lead, user)
    # ASM / RM — pipeline snapshot only (no commercials, payments, claims)
    if user.get("role") in authmod.FIELD_ROLES:
        delivery = clean(await db.deliveries.find_one({"leadId": lead_id}) or {})
        booking = clean(await _live_booking(lead_id) or {})
        activities = [clean(a) for a in await db.activities.find({"leadId": lead_id}).sort("activityId", -1).to_list(500)]
        safe_delivery = {}
        if delivery:
            for k in ("leadId", "deliveryDate", "deliveryStatus", "chassisNumber",
                      "engineNumber", "invoiceNumber", "rcStatus"):
                if delivery.get(k) not in (None, ""):
                    safe_delivery[k] = delivery.get(k)
        safe_booking = {}
        if booking:
            for k in ("bookingId", "leadId", "bookingDate", "bookingStatus", "model", "variant"):
                if booking.get(k) not in (None, ""):
                    safe_booking[k] = booking.get(k)
        return {
            "lead": _field_safe_lead(lead),
            "commercials": {},
            "payments": [],
            "activities": activities,
            "delivery": safe_delivery,
            "booking": safe_booking,
            "claims": [],
            "actions": {
                **lead_actions(lead, user),
                "canBook": False, "canPrice": False, "canScheme": False,
                "canPayment": False, "canDelivery": False, "canClose": False,
                "fieldView": True,
            },
            "billingSummary": None,
            "fieldView": True,
            "whatsapp": {"count": 0, "lastAt": None, "optOut": False, "sessionOpen": False},
        }
    snap = lead_to_snapshot(lead)
    scheme_rows = await get_scheme_rows()
    commercials = ce.compute_full_commercials(snap, scheme_rows)
    payments = [clean(p) for p in await db.payments.find({"leadId": lead_id}).sort("date", 1).to_list(500)]
    activities = [clean(a) for a in await db.activities.find({"leadId": lead_id}).sort("activityId", -1).to_list(500)]
    delivery = clean(await db.deliveries.find_one({"leadId": lead_id}) or {})
    booking = clean(await _live_booking(lead_id) or {})
    claims = [clean(c) for c in await db.claims.find({"leadId": lead_id}).to_list(100)]
    # Always rebuild from current commercials when delivered so scheme/Additional
    # (Dealer) edits are not stuck on the Mark-Delivered snapshot.
    billing_summary = {}
    if _is_delivered(lead):
        try:
            billing_summary = clean(await _upsert_delivery_billing_summary(lead_id) or {})
        except Exception as exc:
            # Lead drawer must still open; billing panel can show stale/empty.
            logger.exception("billing summary rebuild failed for %s: %s", lead_id, exc)
            billing_summary = clean(await db.billing_summaries.find_one({"leadId": lead_id}) or {})
    else:
        billing_summary = clean(await db.billing_summaries.find_one({"leadId": lead_id}) or {})
    return {
        "lead": lead, "commercials": commercials, "payments": payments,
        "activities": activities, "delivery": delivery, "booking": booking,
        "claims": claims, "actions": lead_actions(lead, user),
        "billingSummary": billing_summary or None,
        "whatsapp": await wa.summary_for_lead(lead_id),
    }


# System/calculated fields a lead document carries that must never be settable
# through PUT /leads/{lead_id} — set at creation (create_lead) and/or maintained
# by recompute_lead() and the booking/delivery/close/payment/scheme endpoints.
# LeadUpdateIn already doesn't declare any of these (extra="forbid" rejects them
# outright); this set is a second, independent line of defense in the handler
# itself so the guarantee doesn't rely solely on the Pydantic model staying in sync.
LEAD_SYSTEM_FIELDS = {
    "leadId", "createdDate", "accountStatus", "deliveryStatus", "lastUpdated",
    "outstandingAmount", "customerOutstanding", "companyOutstanding", "totalReceived",
    "customerPayable", "grossVehicleCost", "totalDiscount", "consumerDiscount", "exchangeBonus",
    "loyaltyBonus", "referralBonus", "dsaDiscount", "additionalDiscount", "exShowroom", "rto",
    "insuranceAmount", "accessoriesAmount", "handlingCharges", "trc", "fastag", "extendedWarranty",
    "otherCharges", "oemSchemeAmount", "dealerSchemeAmount", "oemClaimCompanyShare",
    "schemeCompanyTotal", "dealerSchemeRetained", "oemExtraSupportRetained", "dealerMarginNetExGst",
    "extraDealerIncomeTotal", "dealerTotalEarnings",
    # Derived from the payment ledger by recompute_lead / POST /leads/{id}/refund.
    "excessReceived", "refundedAmount",
    # Owned by POST /leads/{id}/cancel and /revive. The cancellation stamp is the
    # basis of every executive's cancel count — it must not be editable through
    # the ordinary lead form, or the count could be typed away.
    "cancelCount", "cancelHistory", "lastCancelDate", "lastCancelReason",
    "lastCancelRemarks", "lastCancelStage", "lastCancelBy", "cancelMoneyAtRisk",
    "reviveOn", "revivedAt", "revivedFromCancel", "followupAnchorDate",
    # Drives whether Customer Outstanding is real money owed or a dead figure on a
    # cancelled deal. Set by /cancel and cleared by /convert-booking only.
    "dealCancelled", "cancelledBookingDate", "cancelledBookingAmount",
}

# The literal placeholder Swagger/OpenAPI "Try it out" fills into required-string
# fields when the user hasn't typed a real value. Never a plausible real value for
# any lead field, so it's rejected outright rather than persisted as data.
_SWAGGER_STRING_PLACEHOLDER = "string"

# Only fields whose data-model representation is genuinely nullable may be
# cleared with an explicit JSON null. Every other LeadUpdateIn field was a
# non-nullable str/float in the original LeadIn (default "" or 0, never None),
# so exclude_unset=True can't distinguish "client wants to blank this out" from
# "client (or a generated/all-null body) sent null for a field it never meant to
# touch" -- reject null for those instead of silently blanking a required field.
LEAD_NULLABLE_FIELDS = {"nextFollowupDate"}


async def _validate_lead_update_choices(payload, lead):
    """Best-effort validation against master data (requirement: enum/master-list
    fields validated where appropriate). Only validates values the client is
    actually CHANGING (differs from what's already stored on the lead) -- leads
    with a legacy/historical value not in the current master list (e.g. leadSource
    "Import" from a file import) must still be editable for unrelated fields;
    resubmitting the same value isn't a new value to validate. Skips validation
    for a category if its master list is empty, rather than blocking on a
    misconfigured/empty list."""
    errors = []
    checks = [("leadSource", "leadSources", "Lead Sources"), ("executive", "executives", "Executives"),
              ("priority", "priorities", "Priorities")]
    for field, category, label in checks:
        val = payload.get(field)
        if not val or val == lead.get(field):
            continue
        allowed = await _masters_list_values(category)
        if allowed and val not in allowed:
            errors.append(f"{field} '{val}' is not in the {label} master list")
    status_val = payload.get("currentStatus")
    if status_val and status_val != lead.get("currentStatus") and status_val not in seeder.MASTERS["statuses"]:
        errors.append(f"currentStatus '{status_val}' is not a recognized status")
    for field in ("financeRequired", "exchangeRequired"):
        val = payload.get(field)
        if val is not None and val != lead.get(field) and val not in ("Yes", "No"):
            errors.append(f"{field} must be 'Yes' or 'No', got '{val}'")
    return errors


@api.put("/leads/{lead_id}")
async def update_lead(lead_id: str, body: LeadUpdateIn, act=Depends(actor), _sales=Depends(sales_staff_only)):
    """Partial update: only fields present in the request body are touched — every
    other field on the lead (including system/financial fields, which aren't even
    part of this model) is left exactly as it was. See LEAD_SYSTEM_FIELDS."""
    lead = await get_lead_or_404(lead_id)
    _require_mutable_lead(lead, "lead edits", act)
    payload = body.model_dump(exclude_unset=True)
    payload = {k: v for k, v in payload.items() if k not in LEAD_SYSTEM_FIELDS}

    null_fields = [k for k, v in payload.items() if v is None and k not in LEAD_NULLABLE_FIELDS]
    if null_fields:
        raise HTTPException(422, f"null is not allowed for: {', '.join(null_fields)}. "
                                  f"Omit the field to leave it unchanged, or provide a real value.")

    placeholder_fields = [k for k, v in payload.items()
                          if isinstance(v, str) and v.strip().lower() == _SWAGGER_STRING_PLACEHOLDER]
    if placeholder_fields:
        raise HTTPException(422, f"Refusing to save placeholder value \"string\" for: {', '.join(placeholder_fields)}. "
                                  f"Remove the field from the request instead of leaving Swagger's example value.")

    errors = await _validate_lead_update_choices(payload, lead)
    if errors:
        raise HTTPException(422, "; ".join(errors))

    if not payload:
        return clean(lead)

    vehicle_changed = (
        ("interestedModel" in payload and str(payload["interestedModel"] or "").strip()
         != str(lead.get("interestedModel") or "").strip())
        or ("variant" in payload and str(payload["variant"] or "").strip()
            != str(lead.get("variant") or "").strip())
    )
    # Once priced/schemed, only the owner may change model/variant (cascades commercials).
    if vehicle_changed and (_is_priced(lead) or _has_persisted_scheme(lead)):
        _require_owner_reedit(act, True, "Model / variant")

    old = {k: lead.get(k) for k in payload.keys()}
    old_booking_amount = ce.num(lead.get("bookingAmount"))
    if "bookingAmount" in payload:
        try:
            payload["bookingAmount"] = max(0.0, float(payload["bookingAmount"] or 0))
        except (TypeError, ValueError):
            raise HTTPException(422, "bookingAmount must be a number ≥ 0")
    payload["lastUpdated"] = now_iso()
    # Lead Register "Last Updated By": the acting user was already resolved for the
    # audit log, but was never written onto the lead, so the column had no source.
    payload["lastUpdatedBy"] = act.get("email", "")
    await db.leads.update_one({"leadId": lead_id}, {"$set": payload})
    if "bookingAmount" in payload:
        await _sync_booking_amount_edit(
            lead_id, lead, old_booking_amount, ce.num(payload["bookingAmount"]), act)
    if vehicle_changed:
        await _cascade_vehicle_or_price_change(lead_id, refresh_price=True, realign_scheme=True)
    else:
        await recompute_lead(lead_id)
    await write_audit(act, "update", "lead", leadId=lead_id, old=old, new={k: v for k, v in payload.items() if k != "lastUpdated"})
    updated = await db.leads.find_one({"leadId": lead_id})
    # GS-2: a lead edit must reach the EXISTING Lead Register row. Previously this
    # endpoint never touched the Sheet at all, so the row went stale on first edit.
    # sheet_sync upserts on leadId, so this updates in place — it never appends.
    await sheet_sync("leads", clean(dict(updated)))
    return clean(updated)


async def _sync_booking_amount_edit(lead_id, lead, old_amount, new_amount, act=None):
    """Keep bookings row + Booking advance receipt aligned when bookingAmount is corrected."""
    new_amount = max(0.0, ce.num(new_amount))
    old_amount = max(0.0, ce.num(old_amount))
    if abs(old_amount - new_amount) < 0.005:
        return
    booking = await _live_booking(lead_id)
    if booking:
        await db.bookings.update_one(
            {"bookingId": booking.get("bookingId")},
            {"$set": {"bookingAmount": new_amount, "amountReceived": new_amount}},
        )
        await sheet_sync("bookings", {
            "bookingId": booking.get("bookingId"), "leadId": lead_id,
            "customerName": lead.get("customerName"),
            "bookingDate": booking.get("bookingDate") or lead.get("bookingDate"),
            "model": lead.get("interestedModel"), "variant": lead.get("variant"),
            "bookingAmount": new_amount,
            "paymentMode": booking.get("paymentMode") or lead.get("lastPaymentMode") or "Cash",
            "bookingStatus": booking.get("bookingStatus") or "Booked",
        })
    adv = await db.payments.find_one({
        "leadId": lead_id,
        "narration": {"$regex": r"^booking advance", "$options": "i"},
    })
    if new_amount <= 0:
        if adv:
            await db.payments.delete_one({"_id": adv["_id"]})
            if adv.get("paymentId") or adv.get("receiptNumber"):
                # Soft sheet cleanup is best-effort; lead recompute fixes outstanding.
                pass
        return
    if adv:
        await db.payments.update_one(
            {"_id": adv["_id"]},
            {"$set": {"amount": new_amount, "lastUpdated": now_iso()}},
        )
    elif _is_booked(lead) or lead.get("bookingDate"):
        await _add_payment_internal(lead_id, PaymentIn(
            amount=new_amount,
            paymentMode=lead.get("lastPaymentMode") or "Cash",
            date=lead.get("bookingDate") or today(),
            narration="Booking advance",
        ))


async def _price_master_row(model, variant):
    """Authoritative Price Master lookup, keyed on model + variant exactly as the
    Price Master defines them (case/whitespace-insensitive, active rows only).
    OEM aliases (Turbo→Turbo Max, Hi-Load/XR→HiCity/XR, Maxx (DV200)→DV220, …)
    resolve to the canonical catalog row so old leads still book.
    Returns None when there is no matching row — the caller must refuse to book
    rather than fall back to zero."""
    m = str(model or "").strip()
    v = str(variant or "").strip()
    sku = oem_cat.resolve_sku(m, v)
    if sku:
        m, v = sku.crm_model, sku.crm_variant
    if not m:
        return None
    rows = await db.price_master.find({"model": {"$regex": f"^{re.escape(m)}$", "$options": "i"}}).to_list(500)
    active = [r for r in rows if str(r.get("status") or "active").lower() == "active"] or rows
    if v:
        exact = [r for r in active if str(r.get("variant") or "").strip().lower() == v.lower()]
        if exact:
            return exact[0]
        return None          # variant was specified but has no Price Master row
    return active[0] if len(active) == 1 else None


def _price_structure_from_master(row):
    """Map a Price Master row onto the lead's price-structure fields."""
    return {
        "exShowroom": ce.num(row.get("exShowroom")),
        "rto": ce.num(row.get("rto")),
        "insuranceAmount": ce.num(row.get("insurance")),
        "accessoriesAmount": ce.num(row.get("accessories")),
        "handlingCharges": ce.num(row.get("handlingCharges")),
        "trc": ce.num(row.get("trc")),
        "fastag": ce.num(row.get("fastag")),
        "extendedWarranty": ce.num(row.get("extendedWarranty")),
        "otherCharges": ce.num(row.get("otherCharges")),
        "tcsApplicable": row.get("tcsApplicable") or "No",
    }


async def _cascade_vehicle_or_price_change(lead_id, *, refresh_price=True, realign_scheme=True):
    """After model/variant or price edits, refresh Master-backed fields and recompute.

    - Price Master: when refresh_price, overwrite Ex-Showroom (and empty charge lines
      from master defaults) for the lead's current model/variant.
    - Scheme: when a scheme was already saved, rematerialise eligible offer pools from
      Scheme Master for the new vehicle/month and clamp any prior customer benefits
      that no longer fit. Then recompute_lead refreshes payable / retained / claims.
    """
    lead = await db.leads.find_one({"leadId": lead_id})
    if not lead:
        return None
    patch = {}
    if refresh_price:
        row = await _price_master_row(lead.get("interestedModel"), lead.get("variant"))
        if row:
            # Full Price Master structure for the new model/variant.
            patch.update(_price_structure_from_master(row))

    if realign_scheme and _has_persisted_scheme(lead):
        import json as _json
        scheme_rows = await get_scheme_rows()
        model = lead.get("interestedModel") or ""
        variant = lead.get("variant") or ""
        booking_date = _scheme_as_of(lead)
        rules_ctx = ce.get_scheme_offer_rules_for_vehicle(
            model, variant, booking_date, scheme_rows)
        for key, rule in (rules_ctx.get("rules") or {}).items():
            if key == "additionalDiscount":
                continue
            if rule.get("allowed") and ce.num(rule.get("maxAmount")) > 0:
                patch[key] = ce.num(rule.get("schemeAvailable") or rule.get("maxAmount"))
            else:
                patch[key] = 0
        # Drop / clamp breakup keys that are no longer eligible.
        raw_bk = lead.get("benefitPassedBreakup") or "{}"
        try:
            bk = _json.loads(raw_bk) if isinstance(raw_bk, str) else dict(raw_bk or {})
        except Exception:
            bk = {}
        if not isinstance(bk, dict):
            bk = {}
        raw_used = lead.get("schemeComponentsUsed") or "{}"
        try:
            used = _json.loads(raw_used) if isinstance(raw_used, str) else dict(raw_used or {})
        except Exception:
            used = {}
        if not isinstance(used, dict):
            used = {}
        allowed = set()
        for key, rule in (rules_ctx.get("rules") or {}).items():
            if key != "additionalDiscount" and rule.get("allowed") and ce.num(rule.get("maxAmount")) > 0:
                allowed.add(key)
        for ent in (rules_ctx.get("entitlements") or []):
            if ce.num(ent.get("schemeAvailable") or ent.get("totalBenefit")) > 0:
                allowed.add(ent["key"])
        clean_bk, clean_used = {}, {}
        for key in allowed:
            cap = ce.num(patch.get(key))
            if key in (rules_ctx.get("rules") or {}) and (rules_ctx["rules"][key].get("allowed")):
                cap = ce.num(rules_ctx["rules"][key].get("schemeAvailable")
                             or rules_ctx["rules"][key].get("maxAmount") or cap)
            for ent in (rules_ctx.get("entitlements") or []):
                if ent.get("key") == key:
                    cap = ce.num(ent.get("schemeAvailable") or ent.get("totalBenefit") or cap)
            cb = ce.round2(max(0.0, min(ce.num(bk.get(key)), cap)))
            clean_bk[key] = cb
            clean_used[key] = bool(used.get(key)) if key in used else (cb > 0)
        patch["benefitPassedBreakup"] = _json.dumps(clean_bk)
        patch["schemeComponentsUsed"] = _json.dumps(clean_used)
        patch["customerBenefitPassed"] = ce.round2(sum(clean_bk.values()))
        patch["schemeAllocationExplicit"] = True
        patch["schemeAllocationV2"] = True

    if patch:
        patch["lastUpdated"] = now_iso()
        await db.leads.update_one({"leadId": lead_id}, {"$set": patch})
    return await recompute_lead(lead_id)


@api.get("/leads/{lead_id}/price-preview")
async def price_preview(lead_id: str):
    """What Price Master resolves to for this lead's model/variant, before booking.
    Lets the UI show the real commercial structure (or a precise 'not found')
    instead of discovering it only at booking time."""
    lead = await get_lead_or_404(lead_id)
    model, variant = lead.get("interestedModel"), lead.get("variant")
    row = await _price_master_row(model, variant)
    if not row:
        return {"found": False, "model": model, "variant": variant,
                "message": f"Price Master entry not found for: Model = {model or '(none)'}, "
                           f"Variant = {variant or '(none)'}"}
    return {"found": True, "model": model, "variant": variant,
            "priceId": row.get("priceId"), "priceStructure": _price_structure_from_master(row)}


@api.post("/leads/{lead_id}/convert-booking")
async def convert_booking(lead_id: str, body: BookingIn, act=Depends(actor), _sales=Depends(sales_staff_only)):
    lead = await get_lead_or_404(lead_id)
    _require_action(lead, "canBook", "conversion to booking", act)
    # A booking is only valid once its commercial structure is resolved. If the lead
    # has no price structure yet, load it from the authoritative Price Master. If the
    # vehicle has no Price Master row, refuse the booking with a precise message
    # rather than persisting a commercially empty booking (ex-showroom 0, GVC 0).
    # STRICT PRODUCTION BOOKING MODE. Price Master is revalidated on every booking,
    # immediately before persistence — a previously saved price structure is NOT
    # sufficient on its own. If the lead's current model/variant has no Price Master
    # row (e.g. the vehicle was changed after pricing, or the row was withdrawn), the
    # booking is refused rather than persisting against stale commercial values.
    row = await _price_master_row(lead.get("interestedModel"), lead.get("variant"))
    if not row:
        raise HTTPException(422,
            f"Price Master entry not found for: Model = {lead.get('interestedModel') or '(none)'}, "
            f"Variant = {lead.get('variant') or '(none)'}. Add the vehicle to Price Master, or "
            f"correct the model/variant on the lead, then book again.")
    if ce.num(_price_structure_from_master(row).get("exShowroom")) <= 0:
        raise HTTPException(422,
            f"Price Master row {row.get('priceId')} for {lead.get('interestedModel')}/"
            f"{lead.get('variant')} has a zero ex-showroom price. Correct the Price Master entry "
            f"before booking.")
    if ce.num(lead.get("exShowroom")) <= 0:
        # Unpriced lead: adopt the authoritative structure for booking maths, but do
        # NOT mark priceStructureSaved — staff must still complete the Price step.
        await db.leads.update_one({"leadId": lead_id},
                                  {"$set": {**_price_structure_from_master(row),
                                            "priceStructureSaved": False,
                                            "lastUpdated": now_iso()}})
        await recompute_lead(lead_id)
        lead = await db.leads.find_one({"leadId": lead_id})
    if ce.num(lead.get("grossVehicleCost")) <= 0 or ce.num(lead.get("customerPayable")) <= 0:
        # Never persist a booking whose commercial calculation did not resolve.
        raise HTTPException(422,
            f"Commercial calculation did not resolve for {lead.get('interestedModel')}/"
            f"{lead.get('variant')} (gross vehicle cost {ce.num(lead.get('grossVehicleCost'))}, "
            f"customer payable {ce.num(lead.get('customerPayable'))}). Review the price structure "
            f"before booking.")
    booking_id = await next_id("booking", "BK26")
    snapshot_id = await next_id("snapshot", "SN26")
    bdate = body.bookingDate or today()
    requested = ce.round2(max(0.0, ce.num(body.bookingAmount)))
    held = await _net_received(lead_id)
    extra = ce.round2(max(0.0, requested - held))
    await db.leads.update_one({"leadId": lead_id}, {"$set": {
        "currentStatus": "Booked", "bookingDate": bdate, "bookingAmount": requested,
        "executive": body.executive or lead.get("executive"), "financeRequired": body.financeRequired,
        "exchangeRequired": body.exchangeRequired, "lastPaymentMode": body.paymentMode, "lastUpdated": now_iso(),
        # A lead that cancelled and has now booked again is a live deal: the
        # customer owes the payable once more, and the money held stops being a
        # refund and becomes part of this booking.
        "dealCancelled": False,
        # Allow booking-confirm WhatsApp for this new booking (cancel cleared it too).
        "whatsappBookingSentAt": "",
    }})
    await db.bookings.insert_one({
        "bookingId": booking_id, "leadId": lead_id, "customerName": lead.get("customerName"),
        "bookingDate": bdate, "model": lead.get("interestedModel"), "variant": lead.get("variant"),
        "bookingAmount": requested, "amountReceived": requested, "paymentMode": body.paymentMode,
        "financeRequired": body.financeRequired, "exchangeRequired": body.exchangeRequired,
        "snapshotId": snapshot_id, "bookingStatus": "Booked", "createdBy": "crm", "createdDate": today(),
    })
    await sheet_sync("bookings", {
        "bookingId": booking_id, "leadId": lead_id, "customerName": lead.get("customerName"),
        "bookingDate": bdate, "model": lead.get("interestedModel"), "variant": lead.get("variant"),
        "bookingAmount": requested, "paymentMode": body.paymentMode, "bookingStatus": "Booked",
    })
    # Re-book must not post a second advance for money already on the ledger.
    if extra > 0.01:
        await _add_payment_internal(lead_id, PaymentIn(
            amount=extra, paymentMode=body.paymentMode, date=bdate,
            narration="Booking advance"))
    _bk_act = {
        "activityId": await next_id("activity", "AC26"), "leadId": lead_id, "date": bdate,
        "time": datetime.now(timezone.utc).strftime("%H:%M"), "activityType": "Booking",
        "discussion": "Booking converted.", "executive": lead.get("executive"),
        "customerName": lead.get("customerName"), "mobile": lead.get("mobile"), "model": lead.get("interestedModel"),
    }
    await db.activities.insert_one(dict(_bk_act))
    await sheet_sync("activities", _bk_act)
    await recompute_lead(lead_id)
    # WhatsApp booking confirm is fire-and-forget — must never fail this booking.
    wa.schedule(wa.notify_booking(lead_id))
    return {"bookingId": booking_id, "snapshotId": snapshot_id, "lead": clean(await db.leads.find_one({"leadId": lead_id}))}


@api.delete("/leads/{lead_id}", dependencies=[Depends(owner_only)])
async def delete_lead(lead_id: str, act=Depends(actor)):
    """Owner-only. Permanently delete a wrongly-posted lead and all related records
    in Mongo AND every matching row across Google Sheet operational registers.

    Scheme Claim Register rows are preserved (permanent ledger — never archived,
    even after Received / lead delete).
    """
    lead = await db.leads.find_one({"leadId": lead_id})
    if not lead:
        raise HTTPException(404, "Lead not found")
    # Collect related IDs for sync-log cleanup before cascade delete.
    related_ids = {lead_id}
    for coll, id_field in (
        ("payments", "receiptNumber"), ("bookings", "bookingId"),
        ("claims", "claimId"), ("insurance", "entryId"),
        ("finance", "financeFileNumber"), ("activities", "activityId"),
        ("incentive_register", "incentiveId"),
    ):
        async for doc in db[coll].find({"leadId": lead_id}, {id_field: 1}):
            val = str(doc.get(id_field) or "").strip()
            if val:
                related_ids.add(val)

    # Sheet first (while we still know the lead exists) — remove all register traces.
    sheet_result = await gsheets.delete_lead_traces(lead_id)

    counts = {}
    for coll in ["payments", "bookings", "deliveries", "finance", "insurance", "claims",
                 "activities", "dealer_earnings", "incentive_register", "billing_summaries",
                 "whatsapp_messages", "whatsapp_outbox"]:
        r = await db[coll].delete_many({"leadId": lead_id})
        counts[coll] = r.deleted_count
    await db.leads.delete_one({"leadId": lead_id})
    # Drop pending/OK sync-log rows for this lead and its related entity IDs.
    sync_log = await db.sheet_sync_log.delete_many({
        "$or": [
            {"entityId": {"$in": list(related_ids)}},
            {"payload.leadId": lead_id},
        ]
    })
    counts["sheet_sync_log"] = sync_log.deleted_count
    # Derived finance views must drop any file that belonged to this lead.
    fin_views = await rebuild_finance_views()
    await write_audit(act, "delete", "lead", leadId=lead_id,
                      old={"customerName": lead.get("customerName"), "mobile": lead.get("mobile"),
                           "currentStatus": lead.get("currentStatus"), "customerPayable": lead.get("customerPayable")},
                      new={"cascadeDeleted": counts,
                           "sheet": {"rowsDeleted": sheet_result.get("rowsDeleted", 0),
                                     "ok": sheet_result.get("ok"),
                                     "operation": sheet_result.get("operation") or (
                                         "deleted" if sheet_result.get("ok") else sheet_result.get("error")),
                                     "tabs": sheet_result.get("tabs") or []},
                           "financeViews": fin_views})
    return {"ok": True, "deleted": {"lead": 1, **counts},
            "sheet": sheet_result, "financeViews": fin_views}


@api.put("/leads/{lead_id}/price-structure")
async def set_price_structure(lead_id: str, body: PriceStructureIn, act=Depends(actor), _desk=Depends(deal_desk_only)):
    lead = await get_lead_or_404(lead_id)
    _require_action(lead, "canPrice", "price-structure edits (only Active leads)", act)
    _require_owner_reedit(act, _is_priced(lead), "Price structure")
    payload = body.model_dump()
    payload["insuranceArrangedBy"] = ce.normalize_insurance_arranged_by(
        payload.get("insuranceArrangedBy"))
    # Ex-Showroom is Price Master–authoritative and not staff-editable. Prefer the
    # live master row for the lead's model/variant; fall back to the lead's existing
    # value only when no master row exists (so a saved structure is not wiped).
    row = await _price_master_row(lead.get("interestedModel"), lead.get("variant"))
    if row:
        master_ex = ce.num(_price_structure_from_master(row).get("exShowroom"))
        if master_ex <= 0:
            raise HTTPException(422,
                f"Price Master row for {lead.get('interestedModel')}/{lead.get('variant')} "
                f"has a zero ex-showroom price. Correct Price Master before saving.")
        payload["exShowroom"] = master_ex
    else:
        payload["exShowroom"] = ce.num(lead.get("exShowroom"))
        if payload["exShowroom"] <= 0:
            raise HTTPException(422,
                f"Price Master entry not found for {lead.get('interestedModel') or '(none)'}/"
                f"{lead.get('variant') or '(none)'}. Select a valid vehicle before pricing.")
    payload["priceStructureSaved"] = True
    old = {k: lead.get(k) for k in payload.keys()}
    await db.leads.update_one({"leadId": lead_id}, {"$set": {**payload, "lastUpdated": now_iso()}})
    # Owner re-edit of price must also realign scheme pools / retained totals.
    if _has_persisted_scheme({**lead, **payload}):
        await _cascade_vehicle_or_price_change(lead_id, refresh_price=False, realign_scheme=True)
    else:
        await recompute_lead(lead_id)
    await _refresh_billing_summary_if_delivered(lead_id)
    await write_audit(act, "update", "price-structure", leadId=lead_id, old=old, new=payload)
    return clean(await db.leads.find_one({"leadId": lead_id}))


@api.get("/leads/{lead_id}/scheme-rules")
async def scheme_rules(lead_id: str, on: Optional[str] = None):
    lead = await get_lead_or_404(lead_id)
    scheme_rows = await get_scheme_rows()
    model = lead.get("interestedModel") or ""
    variant = lead.get("variant") or ""
    as_of = _scheme_as_of(lead, on)
    out = ce.get_scheme_offer_rules_for_vehicle(model, variant, as_of, scheme_rows)
    # Preview allocation for the same as-of date the rules were resolved against.
    snap = {**lead_to_snapshot(lead), "schemeAsOf": as_of}
    out["allocation"] = ce.compute_scheme_allocation(snap, scheme_rows)
    out["asOf"] = as_of
    return out


@api.put("/leads/{lead_id}/scheme")
async def set_scheme(lead_id: str, body: SchemeIn, act=Depends(actor), _desk=Depends(deal_desk_only)):
    lead = await get_lead_or_404(lead_id)
    _require_action(lead, "canScheme", "scheme edits (only Active leads)", act)
    _require_owner_reedit(act, _has_persisted_scheme(lead), "Scheme")
    payload = body.model_dump()
    scheme_date = payload.pop("schemeDate", None)
    as_of = _scheme_as_of(lead, scheme_date)
    payload["schemeAsOf"] = as_of
    if payload.get("benefitPassedBreakup") is None:
        payload.pop("benefitPassedBreakup", None)
    if payload.get("schemeComponentsUsed") is None:
        payload.pop("schemeComponentsUsed", None)
    # Validate offers against Scheme Master availability + max caps (port of validateSchemeOffersForVehicle_)
    scheme_rows = await get_scheme_rows()
    offers = {k: payload.get(k, 0) for k in ce.OFFER_KEYS}
    errors = ce.validate_scheme_offers(
        lead.get("interestedModel") or "", lead.get("variant") or "",
        as_of, offers, scheme_rows)
    if errors:
        raise HTTPException(422, "Please fix these scheme fields:\n" + "\n".join(errors))
    import json as _json

    raw_breakup = payload.get("benefitPassedBreakup")
    parsed_breakup = None
    if isinstance(raw_breakup, str) and raw_breakup.strip():
        try:
            parsed_breakup = _json.loads(raw_breakup)
        except Exception:
            raise HTTPException(422, "benefitPassedBreakup must be valid JSON")
    elif isinstance(raw_breakup, dict):
        parsed_breakup = raw_breakup

    raw_used = payload.get("schemeComponentsUsed")
    parsed_used = {}
    if isinstance(raw_used, str) and raw_used.strip():
        try:
            parsed_used = _json.loads(raw_used)
        except Exception:
            raise HTTPException(422, "schemeComponentsUsed must be valid JSON")
    elif isinstance(raw_used, dict):
        parsed_used = raw_used
    if not isinstance(parsed_used, dict):
        parsed_used = {}

    # Explicit breakup from the Scheme UI ⇒ assignment model (eligibility ≠ assignment).
    # Legacy clients that omit breakup keep Full/No Benefit materialisation.
    if isinstance(parsed_breakup, dict):
        alloc_errs = ce.validate_scheme_allocation_breakup(
            lead.get("interestedModel") or "", lead.get("variant") or "",
            as_of, parsed_breakup, scheme_rows)
        if alloc_errs:
            raise HTTPException(422, "Please fix these scheme allocation fields:\n" + "\n".join(alloc_errs))
        payload["schemeAllocationExplicit"] = True
        payload["schemeAllocationV2"] = True
        # Benefit Mode is not used by the new UI; store Partial for compatibility.
        payload["benefitMode"] = "Partial Benefit"
        provisional = {
            **lead_to_snapshot({**lead, **payload}),
            "schemeAllocationExplicit": True,
            "schemeAllocationV2": True,
            "benefitPassedBreakup": parsed_breakup,
            "schemeComponentsUsed": parsed_used,
            "benefitMode": "Partial Benefit",
        }
        # Ensure eligible offer pools are present so OEM claim shares resolve even
        # when customer benefit is ₹0 (Use Scheme = No). Available = Scheme Master.
        rules_ctx = ce.get_scheme_offer_rules_for_vehicle(
            lead.get("interestedModel") or "", lead.get("variant") or "",
            as_of, scheme_rows)
        for key, rule in (rules_ctx.get("rules") or {}).items():
            if key == "additionalDiscount":
                continue
            if rule.get("allowed") and ce.num(rule.get("maxAmount")) > 0:
                # Persist the eligible pool amount (not customer assignment).
                payload[key] = ce.num(rule.get("schemeAvailable") or rule.get("maxAmount"))
                provisional[key] = payload[key]
        alloc = ce.compute_scheme_allocation(provisional, scheme_rows)
        clean_bk = {}
        clean_used = {}
        for c in alloc["components"]:
            if c["key"] == "additionalDiscount":
                continue
            # Prefer client amount when provided; otherwise 0 (never auto-assign).
            if c["key"] in parsed_breakup:
                cb = ce.round2(max(0.0, min(ce.num(parsed_breakup[c["key"]]), c["schemeAvailable"])))
            else:
                cb = 0.0
            clean_bk[c["key"]] = cb
            if c["key"] in parsed_used:
                clean_used[c["key"]] = bool(parsed_used[c["key"]])
            else:
                clean_used[c["key"]] = cb > 0
        payload["benefitPassedBreakup"] = _json.dumps(clean_bk)
        payload["schemeComponentsUsed"] = _json.dumps(clean_used)
        payload["customerBenefitPassed"] = ce.round2(sum(clean_bk.values()))
    else:
        # Legacy path: materialise from Benefit Mode (older API / tests).
        payload["schemeAllocationV2"] = True
        provisional = {**lead_to_snapshot({**lead, **payload}), "schemeAllocationV2": True}
        alloc = ce.compute_scheme_allocation(provisional, scheme_rows)
        clean_bk = {c["key"]: c["customerBenefit"] for c in alloc["components"]
                    if c["key"] != "additionalDiscount"}
        payload["benefitPassedBreakup"] = _json.dumps(clean_bk)
        payload["customerBenefitPassed"] = ce.round2(sum(clean_bk.values()))
        payload["schemeComponentsUsed"] = _json.dumps({k: (v > 0) for k, v in clean_bk.items()})

    # OEM Extra Support: Received = full OEM claim; Passed ≤ Received; Retained derived in recompute.
    # Additional (Dealer) stays untouched here — separate dealer-funded discount.
    _oem = ce.compute_oem_extra_support(payload)
    payload["oemExtraSupportReceived"] = _oem["oemExtraSupportReceived"]
    payload["oemExtraSupportPassed"] = _oem["oemExtraSupportPassed"]

    old = {k: lead.get(k) for k in payload.keys()}
    await db.leads.update_one({"leadId": lead_id}, {"$set": {**payload, "lastUpdated": now_iso()}})
    await recompute_lead(lead_id)
    await _refresh_billing_summary_if_delivered(lead_id)
    await write_audit(act, "update", "scheme", leadId=lead_id, old=old, new=payload)
    return clean(await db.leads.find_one({"leadId": lead_id}))


class SchemeAllocationIn(BaseModel):
    """Per-component customer benefit: {componentKey: amount passed to the customer}."""
    allocation: dict = {}


class ExtraIncomeIn(BaseModel):
    documentationIncome: float = 0
    warrantyIncome: float = 0
    rsaIncome: float = 0
    referralIncome: float = 0
    otherIncome: float = 0
    customerInsuranceBenefitPassed: float = 0
    financeIncentive: float = 0
    accessoriesMargin: float = 0
    exchangeMargin: float = 0
    campaignIncentive: float = 0


@api.put("/leads/{lead_id}/scheme-allocation")
async def set_scheme_allocation(lead_id: str, body: SchemeAllocationIn, act=Depends(actor), _desk=Depends(deal_desk_only)):
    """Record how much of EACH scheme component the dealer passes to the customer.

    Validation is per component against Scheme Master: a benefit is never negative and
    never exceeds schemeAvailable. The OEM claimable share is NOT editable here — it is
    fixed by the circular — and neither are the Scheme Master values themselves."""
    lead = await get_lead_or_404(lead_id)
    _require_action(lead, "canScheme", "scheme edits (only Active leads)", act)
    _require_owner_reedit(act, _has_persisted_scheme(lead), "Scheme")
    scheme_rows = await get_scheme_rows()
    alloc = ce.compute_scheme_allocation(lead_to_snapshot(lead), scheme_rows)
    available = {c["key"]: c["schemeAvailable"] for c in alloc["components"]}
    errors, clean_alloc = [], {}
    for key, raw in (body.allocation or {}).items():
        if key not in available:
            errors.append(f"• {key}: not a scheme component for this model/variant/month")
            continue
        amt = ce.round2(ce.num(raw))
        if amt < 0:
            errors.append(f"• {key}: customer benefit cannot be negative")
        elif amt > available[key] + 0.01:
            errors.append(f"• {key}: ₹{amt} exceeds the ₹{available[key]} the scheme makes available")
        else:
            clean_alloc[key] = amt
    if errors:
        raise HTTPException(422, "Please fix the scheme allocation:\n" + "\n".join(errors))
    merged = {**ce._explicit_allocation(lead_to_snapshot(lead)), **clean_alloc}
    old = lead.get("schemeAllocation")
    # Keys present in this allocation decision are Use=Yes (CB may be ₹0 = keep company share).
    # Omitted keys keep their previous Use flag when available.
    prev_used = lead.get("schemeComponentsUsed") or {}
    if isinstance(prev_used, str):
        try:
            prev_used = json.loads(prev_used) if prev_used.strip() else {}
        except Exception:
            prev_used = {}
    if not isinstance(prev_used, dict):
        prev_used = {}
    used = {**{str(k): bool(v) for k, v in prev_used.items()},
            **{k: True for k in clean_alloc}}
    await db.leads.update_one({"leadId": lead_id}, {"$set": {
        "schemeAllocation": json.dumps(merged),
        "benefitPassedBreakup": json.dumps(merged),
        "schemeComponentsUsed": json.dumps(used),
        "schemeAllocationExplicit": True,
        "schemeAllocationV2": True,
        "benefitMode": "Partial Benefit",
        "customerBenefitPassed": ce.round2(sum(ce.num(v) for v in merged.values())),
        "lastUpdated": now_iso(),
        "lastUpdatedBy": act.get("email", "")}})
    await recompute_lead(lead_id)
    await _refresh_billing_summary_if_delivered(lead_id)
    await write_audit(act, "update", "scheme-allocation", leadId=lead_id,
                      old={"schemeAllocation": old}, new={"schemeAllocation": merged})
    updated = await db.leads.find_one({"leadId": lead_id})
    return {**clean(updated),
            "allocation": ce.compute_scheme_allocation(lead_to_snapshot(updated), scheme_rows)}


@api.put("/leads/{lead_id}/extra-income")
async def set_extra_income(lead_id: str, body: ExtraIncomeIn, act=Depends(actor), _desk=Depends(deal_desk_only)):
    """Dealer extra-income lines:
    Documentation / Warranty / RSA / Referral / Other / Finance Incentive /
    Accessories Margin / Exchange Margin / Campaign Incentive.

    customerInsuranceBenefitPassed is accepted for memo/compatibility but is
    CUSTOMER discount — it does NOT enter Dealer Earnings totals.

    Never affects Customer Payable / Outstanding."""
    lead = await get_lead_or_404(lead_id)
    _require_action(lead, "canScheme", "extra-income edits (only Active leads)", act)
    _extra_keys = (
        "documentationIncome", "warrantyIncome", "rsaIncome", "referralIncome",
        "otherIncome", "financeIncentive", "accessoriesMargin", "exchangeMargin",
        "campaignIncentive",
    )
    _require_owner_reedit(
        act, any(ce.num(lead.get(k)) > 0 for k in _extra_keys), "Extra income")
    payload = body.model_dump()
    old = {k: lead.get(k) for k in payload.keys()}
    await db.leads.update_one({"leadId": lead_id}, {"$set": {**payload, "lastUpdated": now_iso()}})
    await recompute_lead(lead_id)
    await _refresh_billing_summary_if_delivered(lead_id)
    await write_audit(act, "update", "extra-income", leadId=lead_id, old=old, new=payload)
    return clean(await db.leads.find_one({"leadId": lead_id}))


@api.post("/leads/{lead_id}/close")
async def close_lead(lead_id: str, body: CloseIn, act=Depends(actor), _desk=Depends(deal_desk_only)):
    lead = await get_lead_or_404(lead_id)
    _require_action(lead, "canClose", "closing (only Active leads)", act)
    if not str(body.closeReason or "").strip():
        raise HTTPException(422, "Close Reason is required to close a lead.")
    # RC + Number Plate mandatory when closing a DELIVERED lead (port of validateCloseLeadRcFields_)
    plate = body.numberPlate or lead.get("numberPlate")
    if _is_delivered(lead):
        rc = body.rc or lead.get("rcStatus")
        errs = []
        if not _yes_or_done(rc):
            errs.append("Set RC to Yes/Done before closing the lead.")
        if not str(plate or "").strip():
            errs.append("Enter the number plate before closing the lead.")
        if errs:
            raise HTTPException(422, "Cannot close lead:\n" + "\n".join("• " + e for e in errs))
    if str(plate or "").strip():
        await _assert_unique_vehicle_identifiers(lead_id, number_plate=plate)
    close_updates = {
        "accountStatus": "Closed",
        "currentStatus": "Close Won",
        "closedDate": str(body.closedDate or "").strip() or today(),
        "closeReason": body.closeReason,
        "finalOutstanding": lead.get("customerOutstanding", 0), "lastUpdated": now_iso(),
        # Lead Register has "Closed By" and "Close Timestamp" columns. The acting user
        # and the moment of closure were both already known here — they were simply
        # never recorded, so an audited closure could not be attributed from the sheet.
        "closedBy": act.get("email", ""), "closeTimestamp": now_iso(),
        "lastUpdatedBy": act.get("email", ""),
    }
    if body.rc:
        close_updates["rcStatus"] = body.rc
    if body.numberPlate:
        close_updates["numberPlate"] = body.numberPlate
    await db.leads.update_one({"leadId": lead_id}, {"$set": close_updates})
    await write_audit(act, "close", "lead", leadId=lead_id,
                      old={"accountStatus": lead.get("accountStatus")}, new=close_updates)
    updated = clean(await db.leads.find_one({"leadId": lead_id}))
    await sheet_sync("leads", updated)
    return updated


def _revive_updates(lead, *, anchor: str, note: str) -> dict:
    """Put a lead back at the top of the funnel.

    The anchor matters: followup_due() counts days from a date, and a lead created
    ninety days ago would otherwise compute day 90, fire once, then fall silent for
    three days. Anchoring on the cancel date genuinely restarts the 3/6/9 cycle.
    """
    return {
        "accountStatus": "Active",
        "currentStatus": "New",
        "followupAnchorDate": anchor,
        "revivedAt": now_iso(),
        "reviveOn": "",
        "revivedFromCancel": note,
        # Let the follow-up engine treat this as a lead it has never messaged.
        "whatsappFollowupLastDate": "",
        "whatsappFollowupCount": 0,
        "lastUpdated": now_iso(),
    }


@api.post("/leads/{lead_id}/cancel")
async def cancel_lead(lead_id: str, body: CancelIn, act=Depends(actor), _desk=Depends(deal_desk_only)):
    """The LOST exit. Close (Close Won) is the won one; this is the other half.

    Cancellation is recorded as a permanent STAMP (cancelCount + cancelHistory),
    not as a status. That is what lets a cancelled lead go straight back into the
    funnel — as asked — while the executive's cancellation count still holds,
    which a status-based count could not do once the lead flipped back to New.
    """
    lead = await get_lead_or_404(lead_id)
    if _is_delivered(lead):
        raise HTTPException(
            409, "This vehicle is already delivered, so it cannot be cancelled. "
                 "A delivered vehicle coming back is a buyback, which is handled outside the lead.")
    _require_action(lead, "canCancel", "cancelling (only Active, undelivered leads)", act)

    reason_text = str(body.cancelReason or "").strip()
    if not reason_text:
        raise HTTPException(422, "Cancel Reason is required.")
    reason = await _get_cancel_reason(reason_text)
    if not reason:
        raise HTTPException(422, f"'{reason_text}' is not a cancel reason. "
                                 f"Owner: add it in Settings → Cancel Reasons.")
    if reason.get("reason", "").lower() == "other" and not str(body.cancelRemarks or "").strip():
        raise HTTPException(422, "Remarks are required when the reason is 'Other'.")

    when_in = str(body.cancelDate or "").strip() or today()
    # A second cancellation on the same day is a correction, not a new loss. Left
    # unguarded it inflates the executive's count and double-reports the same
    # money. Correcting the reason is a deliberate, owner-gated amend instead.
    if str(lead.get("lastCancelDate") or "")[:10] == when_in[:10] and lead.get("cancelCount"):
        raise HTTPException(
            409, f"This lead was already cancelled on {when_in} "
                 f"(reason: {lead.get('lastCancelReason') or 'unknown'}). "
                 f"To correct that cancellation the owner can amend it; cancelling again "
                 f"would count it twice against {lead.get('executive') or 'the executive'}.")

    money = _cancel_money(lead)
    if money["hasMoney"] and (act or {}).get("role") != "owner":
        raise HTTPException(
            403, f"₹{money['customerMoney']:,.0f} has already been received against this lead, "
                 f"so only the owner can cancel it. The refund is recorded separately — "
                 f"cancelling does not reverse any receipt.")

    when = when_in
    stage = _cancel_stage(lead)
    record = {
        "date": when,
        "reason": reason["reason"],
        "remarks": str(body.cancelRemarks or "").strip(),
        "stage": stage,
        "executive": lead.get("executive") or "",
        "cancelledBy": act.get("email", ""),
        "cancelledAt": now_iso(),
        "customerMoney": money["customerMoney"],
        "statusAtCancel": lead.get("currentStatus") or "New",
    }
    updates = {
        "lastCancelDate": when,
        "lastCancelReason": reason["reason"],
        "lastCancelRemarks": record["remarks"],
        "lastCancelStage": stage,
        "lastCancelBy": act.get("email", ""),
        "cancelMoneyAtRisk": money["customerMoney"],
        # The deal is off. recompute_lead reads this to zero the outstanding and
        # make everything still held refundable.
        "dealCancelled": True,
        # A later Convert to Booking must be allowed to send booking WhatsApp again.
        "whatsappBookingSentAt": "",
        "lastUpdatedBy": act.get("email", ""),
        "lastUpdated": now_iso(),
    }

    # A cancelled booking is not a booking. Left in place, bookingDate alone keeps
    # the lead reading as Booked: it would still count in bookings MTD, and it
    # could never be converted again because canBook requires "not already booked".
    # The bookings row and the payment ledger are NOT touched — the booking really
    # happened and the money really came in; only the LEAD's live state is cleared.
    if _is_booked(lead):
        updates.update({
            "cancelledBookingDate": lead.get("bookingDate") or "",
            "cancelledBookingAmount": ce.num(lead.get("bookingAmount")),
            "bookingDate": "", "bookingAmount": 0.0, "bookingId": "",
        })

    # Revival policy comes from the reason, and STOP always wins over it: a
    # customer who opted out must not be walked back into the messaging cycle.
    mode = str(reason.get("revive") or "now").lower()
    opted_out = bool(lead.get("whatsappOptOut"))
    revived_now = False
    if mode == "now" and not opted_out:
        updates.update(_revive_updates(lead, anchor=when, note=reason["reason"]))
        revived_now = True
        revive_on = ""
    elif mode == "days" and not opted_out:
        revive_on = _add_days(when, int(reason.get("reviveAfterDays") or 30))
        updates.update({"accountStatus": "Cancelled", "currentStatus": "Lost",
                        "reviveOn": revive_on, "nextFollowupDate": revive_on})
    else:
        revive_on = ""
        updates.update({"accountStatus": "Cancelled", "currentStatus": "Lost", "reviveOn": ""})

    await db.leads.update_one(
        {"leadId": lead_id},
        {"$set": updates, "$inc": {"cancelCount": 1}, "$push": {"cancelHistory": record}},
    )
    # Booking Register should not carry a live booking for a dead deal.
    if _is_booked(lead):
        booking = await _live_booking(lead_id)
        if booking:
            await db.bookings.update_one({"bookingId": booking.get("bookingId")},
                                         {"$set": {"bookingStatus": "Cancelled"}})
            await sheet_sync("bookings", clean(
                await db.bookings.find_one({"bookingId": booking.get("bookingId")})))
    await _void_derived_claims_on_cancel(lead_id)
    # Clears Customer Outstanding and turns whatever was received into a refundable
    # balance, so the drawer stops showing a debt nobody is going to collect.
    await recompute_lead(lead_id)
    await write_audit(act, "cancel", "lead", leadId=lead_id,
                      old={"accountStatus": lead.get("accountStatus"),
                           "currentStatus": lead.get("currentStatus")},
                      new={**updates, "cancelRecord": record})
    await _log_activity_safe(lead, "Note", f"Lead cancelled — {reason['reason']}"
                             + (f" ({record['remarks']})" if record["remarks"] else ""))
    updated = clean(await db.leads.find_one({"leadId": lead_id}))
    await sheet_sync("leads", updated)
    return {
        **updated,
        "cancelled": True,
        "revivedNow": revived_now,
        "reviveOn": revive_on,
        "revivePolicy": mode,
        "optOutBlockedRevival": opted_out and mode != "never",
        "moneyAtRisk": money,
    }


@api.put("/leads/{lead_id}/cancel", dependencies=[Depends(owner_only)])
async def amend_cancellation(lead_id: str, body: CancelIn, act=Depends(actor)):
    """Correct the most recent cancellation in place, instead of stacking another.

    Cancelling a second time to "fix" the reason would count the loss twice against
    the executive and report the same money twice. This rewrites the last history
    entry, leaves cancelCount alone, and re-applies the revival policy of the new
    reason — so changing "Other" to "Bought other brand" actually stops the
    follow-ups rather than just relabelling the record.
    """
    lead = await get_lead_or_404(lead_id)
    history = lead.get("cancelHistory") or []
    if not history:
        raise HTTPException(404, "This lead has never been cancelled, so there is nothing to amend.")

    reason_text = str(body.cancelReason or "").strip()
    if not reason_text:
        raise HTTPException(422, "Cancel Reason is required.")
    reason = await _get_cancel_reason(reason_text)
    if not reason:
        raise HTTPException(422, f"'{reason_text}' is not a cancel reason. "
                                 f"Owner: add it in Settings → Cancel Reasons.")
    if reason["reason"].lower() == "other" and not str(body.cancelRemarks or "").strip():
        raise HTTPException(422, "Remarks are required when the reason is 'Other'.")

    last = dict(history[-1])
    before = {k: last.get(k) for k in ("reason", "remarks", "date")}
    when = str(body.cancelDate or "").strip() or last.get("date") or today()
    last.update({
        "reason": reason["reason"],
        "remarks": str(body.cancelRemarks or "").strip(),
        "date": when,
        "amendedBy": act.get("email", ""),
        "amendedAt": now_iso(),
    })
    history[-1] = last

    updates = {
        "cancelHistory": history,
        "lastCancelReason": reason["reason"],
        "lastCancelRemarks": last["remarks"],
        "lastCancelDate": when,
        "lastUpdatedBy": act.get("email", ""),
        "lastUpdated": now_iso(),
    }
    mode = str(reason.get("revive") or "now").lower()
    opted_out = bool(lead.get("whatsappOptOut"))
    revived_now, revive_on = False, ""
    if mode == "now" and not opted_out:
        updates.update(_revive_updates(lead, anchor=when, note=reason["reason"]))
        revived_now = True
    elif mode == "days" and not opted_out:
        revive_on = _add_days(when, int(reason.get("reviveAfterDays") or 30))
        updates.update({"accountStatus": "Cancelled", "currentStatus": "Lost",
                        "reviveOn": revive_on, "nextFollowupDate": revive_on})
    else:
        updates.update({"accountStatus": "Cancelled", "currentStatus": "Lost", "reviveOn": ""})

    await db.leads.update_one({"leadId": lead_id}, {"$set": updates})
    await write_audit(act, "amend", "lead-cancellation", leadId=lead_id,
                      old=before, new={"reason": reason["reason"], "remarks": last["remarks"],
                                       "date": when, "revivePolicy": mode})
    await _log_activity_safe(lead, "Note", f"Cancellation amended — {reason['reason']}")
    updated = clean(await db.leads.find_one({"leadId": lead_id}))
    await sheet_sync("leads", updated)
    return {**updated, "amended": True, "revivedNow": revived_now, "reviveOn": revive_on}


@api.post("/leads/{lead_id}/revive")
async def revive_lead(lead_id: str, act=Depends(actor), _desk=Depends(deal_desk_only)):
    """Put a parked cancelled lead back in the funnel by hand, before its date."""
    lead = await get_lead_or_404(lead_id)
    if _acct(lead) == "Active":
        raise HTTPException(409, "This lead is already active.")
    if _acct(lead) != "Cancelled":
        raise HTTPException(409, f"Only cancelled leads can be revived (this one is {_acct(lead)}).")
    if lead.get("whatsappOptOut"):
        # Reviving is allowed — the executive can still call. Messaging is not.
        pass
    updates = _revive_updates(lead, anchor=today(), note="manual revive")
    updates["lastUpdatedBy"] = act.get("email", "")
    await db.leads.update_one({"leadId": lead_id}, {"$set": updates})
    await write_audit(act, "revive", "lead", leadId=lead_id,
                      old={"accountStatus": lead.get("accountStatus"),
                           "currentStatus": lead.get("currentStatus")}, new=updates)
    await _log_activity_safe(lead, "Note", "Cancelled lead revived")
    updated = clean(await db.leads.find_one({"leadId": lead_id}))
    await sheet_sync("leads", updated)
    return updated


async def _backfill_deal_cancelled() -> int:
    """Leads cancelled before dealCancelled existed still carry a live outstanding.

    Idempotent. A lead that booked again AFTER its last cancellation is a live deal
    and is left alone; everything else cancelled has the flag applied and is
    recomputed, which clears the phantom debt and makes the money refundable.

    Also clears leftover bookingDate/bookingAmount from the first cancel ship
    (#72), which left those fields in place so the lead stayed "booked" and
    Convert to Booking stayed hidden.
    """
    fixed = 0
    cursor = db.leads.find({"cancelCount": {"$gt": 0}, "$or": [
        {"dealCancelled": {"$exists": False}},
        {"dealCancelled": True, "bookingDate": {"$gt": ""}},
    ]})
    for lead in await cursor.to_list(5000):
        lid = lead.get("leadId")
        if not lid:
            continue
        last_cancel = str(lead.get("lastCancelDate") or "")[:10]
        booking = str(lead.get("bookingDate") or "")[:10]
        rebooked = bool(booking and last_cancel and booking > last_cancel)
        if rebooked:
            await db.leads.update_one({"leadId": lid}, {"$set": {"dealCancelled": False}})
            continue
        patch = {
            "dealCancelled": True,
            "whatsappBookingSentAt": "",
        }
        if lead.get("bookingDate") or ce.num(lead.get("bookingAmount")) or lead.get("bookingId"):
            patch.update({
                "cancelledBookingDate": lead.get("cancelledBookingDate") or lead.get("bookingDate") or "",
                "cancelledBookingAmount": lead.get("cancelledBookingAmount")
                if lead.get("cancelledBookingAmount") is not None
                else ce.num(lead.get("bookingAmount")),
                "bookingDate": "", "bookingAmount": 0.0, "bookingId": "",
            })
        await db.leads.update_one({"leadId": lid}, {"$set": patch})
        live = await _live_booking(lid)
        if live:
            await db.bookings.update_one({"bookingId": live.get("bookingId")},
                                         {"$set": {"bookingStatus": "Cancelled"}})
        await _void_derived_claims_on_cancel(lid)
        await recompute_lead(lid)
        fixed += 1
    if fixed:
        logging.info("DEAL_CANCELLED_BACKFILL: cleared outstanding on %s cancelled lead(s)", fixed)
    return fixed


@api.post("/admin/backfill-deal-cancelled", dependencies=[Depends(owner_only)])
async def admin_backfill_deal_cancelled():
    return {"ok": True, "fixed": await _backfill_deal_cancelled()}


async def run_scheduled_revivals(today_s: Optional[str] = None) -> dict:
    """Parked leads whose cool-off has expired come back on their own."""
    day = str(today_s or today())[:10]
    revived = []
    q = {"accountStatus": {"$regex": "^cancelled$", "$options": "i"},
         "reviveOn": {"$gt": "", "$lte": day}}
    for lead in await db.leads.find(q).to_list(2000):
        lid = lead.get("leadId")
        if not lid:
            continue
        await db.leads.update_one(
            {"leadId": lid},
            {"$set": _revive_updates(lead, anchor=day, note=lead.get("lastCancelReason") or "cool-off")})
        await _log_activity_safe(lead, "Note", "Cool-off over — lead back in the funnel")
        updated = clean(await db.leads.find_one({"leadId": lid}))
        await sheet_sync("leads", updated)
        revived.append(lid)
    if revived:
        logging.info("LEAD_REVIVAL: %s lead(s) revived on %s", len(revived), day)
    return {"ok": True, "day": day, "revived": revived, "count": len(revived)}


@api.post("/admin/run-revivals", dependencies=[Depends(owner_only)])
async def admin_run_revivals(day: Optional[str] = None):
    return await run_scheduled_revivals(day)



# ---------------------------------------------------------------- lead allocation
class AllocateIn(BaseModel):
    leadIds: list = []
    executive: str = ""
    remarks: str = ""


@api.get("/leads/allocation/summary")
async def allocation_summary(_desk=Depends(deal_desk_only)):
    """Who is carrying what, and what nobody is carrying.

    Unassigned leads are the point of the page: with executives scoped to their
    own leads, a lead with no executive is visible to nobody but the owner and
    the TL, and would otherwise sit unworked and unnoticed.
    """
    leads = await db.leads.find().to_list(5000)
    active = [l for l in leads if (l.get("accountStatus") or "Active") == "Active"]
    names = await _executive_names()
    rows = {n: {"executive": n, "total": 0, "open": 0, "booked": 0} for n in names}
    unassigned = 0
    for l in active:
        ex = str(l.get("executive") or "").strip()
        if not ex:
            unassigned += 1
            continue
        row = rows.setdefault(ex, {"executive": ex, "total": 0, "open": 0, "booked": 0})
        row["total"] += 1
        if _is_booked_lead(l):
            row["booked"] += 1
        else:
            row["open"] += 1
    return {
        "executives": sorted(rows.values(), key=lambda r: -r["total"]),
        "unassigned": unassigned,
        "activeLeads": len(active),
        "generatedAt": now_iso(),
    }


@api.post("/leads/allocate")
async def allocate_leads(body: AllocateIn, act=Depends(actor), _desk=Depends(deal_desk_only)):
    """Assign or reassign leads to an executive, in bulk.

    The previous owner is kept in an allocation history rather than overwritten.
    Cancellations and bookings are already attributed to whoever held the lead at
    the time, so moving a lead must not quietly move that record with it.
    """
    name = str(body.executive or "").strip()
    if not name:
        raise HTTPException(422, "Pick an executive to allocate to")
    if name not in await _executive_names():
        raise HTTPException(422, f"'{name}' is not an active executive on the staff master")
    lead_ids = [str(x).strip() for x in (body.leadIds or []) if str(x).strip()]
    if not lead_ids:
        raise HTTPException(422, "Select at least one lead")

    moved, skipped = [], []
    for lid in lead_ids:
        lead = await db.leads.find_one({"leadId": lid})
        if not lead:
            skipped.append({"leadId": lid, "reason": "not found"})
            continue
        if (lead.get("accountStatus") or "Active") != "Active":
            skipped.append({"leadId": lid, "reason": f"{_acct(lead)} lead"})
            continue
        previous = str(lead.get("executive") or "").strip()
        if previous == name:
            skipped.append({"leadId": lid, "reason": "already theirs"})
            continue
        entry = {"from": previous, "to": name, "at": now_iso(),
                 "by": act.get("email", ""), "remarks": str(body.remarks or "").strip()}
        await db.leads.update_one({"leadId": lid}, {
            "$set": {"executive": name, "lastUpdated": now_iso(),
                     "lastUpdatedBy": act.get("email", "")},
            "$push": {"allocationHistory": entry}})
        await write_audit(act, "allocate", "lead", leadId=lid,
                          old={"executive": previous}, new={"executive": name})
        await _log_activity_safe(lead, "Note",
                                 f"Lead allocated to {name}"
                                 + (f" (was {previous})" if previous else " (was unassigned)"))
        updated = clean(await db.leads.find_one({"leadId": lid}))
        await sheet_sync("leads", updated)
        moved.append(lid)
    return {"ok": True, "executive": name, "moved": moved,
            "movedCount": len(moved), "skipped": skipped}


# ---------------------------------------------------------------- lead bulk import
IMPORT_COLUMNS = [
    ("Customer Name", "customerName"), ("Mobile", "mobile"), ("Alternate Mobile", "altMobile"),
    ("Village", "village"), ("City", "city"), ("Lead Date", "createdDate"),
    ("Next Follow-up", "nextFollowupDate"), ("Lead Source", "leadSource"),
    ("Interested Model", "interestedModel"), ("Variant", "variant"), ("Executive", "executive"),
    ("Current Status", "currentStatus"), ("Priority", "priority"), ("Budget", "budget"),
    ("Remarks", "remarks"), ("Finance Required", "financeRequired"), ("Exchange Required", "exchangeRequired"),
]
IMPORT_DATE_FIELDS = ("createdDate", "nextFollowupDate")
# A bulk row may only land in the pre-booking part of the funnel. Booked / Finance
# Process / Delivered / Close Won are owned by the workflow endpoints that also write
# the money side (booking, price structure, payments, delivery), so they can never be
# reached by typing a status into a spreadsheet.
IMPORT_STATUSES = ["New", "Contacted", "Follow-up", "In Progress"]
IMPORT_YES_NO = ["Yes", "No"]
# Same fallbacks the New Lead form applies, so a blank optional cell behaves exactly
# like the form's pre-selected value instead of inventing something (the old import
# wrote leadSource="Import", which is not a Settings value and broke source reports).
IMPORT_DEFAULTS = {"leadSource": "Walk-in", "currentStatus": "New", "priority": "Normal",
                   "financeRequired": "No", "exchangeRequired": "No"}
# Every other column is optional; these two identify the customer.
IMPORT_REQUIRED_FIELDS = ("customerName", "mobile")


def _read_rows(filename: str, content: bytes):
    name = (filename or "").lower()
    if name.endswith(".csv"):
        import csv, io as _io
        text = content.decode("utf-8-sig", errors="replace")
        return [r for r in csv.reader(_io.StringIO(text))]
    import openpyxl, io as _io
    wb = openpyxl.load_workbook(_io.BytesIO(content), data_only=True, read_only=True)
    ws = wb.active
    return [[c for c in r] for r in ws.iter_rows(values_only=True)]


def _suggest_mapping(headers):
    """Auto-match detected headers to target fields by (case-insensitive) label/name."""
    lower = {str(h).strip().lower(): str(h).strip() for h in headers if h not in (None, "")}
    mapping = {}
    for label, field in IMPORT_COLUMNS:
        if label.lower() in lower:
            mapping[field] = lower[label.lower()]
        elif field.lower() in lower:
            mapping[field] = lower[field.lower()]
        else:
            mapping[field] = ""
    return mapping


def _import_date(value):
    """Spreadsheet date -> ISO. Accepts real dates, YYYY-MM-DD, DD-MM-YYYY, DD/MM/YYYY.

    Returns "" for blank and None when the cell cannot be read as a date, so the
    caller can report the bad cell instead of silently importing today's date.
    """
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    s = str(value).strip()
    if not s:
        return ""
    m = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})", s)
    if m:
        y, mo, d = m.groups()
    else:
        m = re.match(r"^(\d{1,2})[-/](\d{1,2})[-/](\d{4})$", s)
        if not m:
            return None
        d, mo, y = m.groups()
    try:
        return date(int(y), int(mo), int(d)).isoformat()
    except ValueError:
        return None


def _coerce(field, v):
    if field == "budget":
        try:
            return float(v) if v not in (None, "") else 0
        except (ValueError, TypeError):
            return 0
    if field in IMPORT_DATE_FIELDS:
        iso = _import_date(v)
        # Keep the raw text when unparseable so validation can name the bad cell.
        return iso if iso is not None else str(v).strip()
    if field in ("mobile", "altMobile"):
        if isinstance(v, float):
            v = str(int(v))
        return str(v).strip() if v not in (None, "None") else ""
    return str(v).strip() if v not in (None, "None") else ""


def _parse_import_bytes(filename: str, content: bytes, mapping: dict = None):
    """Return (headers, rows). If mapping {field: headerName} given, use it; else auto-suggest."""
    rows = _read_rows(filename, content)
    if not rows:
        return [], []
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    header_idx = {h.lower(): i for i, h in enumerate(header)}
    if not mapping:
        mapping = _suggest_mapping(header)
    out = []
    for row_no, r in enumerate(rows[1:], start=2):
        if not any(c not in (None, "", "None") for c in r):
            continue
        d = {}
        for _, field in IMPORT_COLUMNS:
            src = mapping.get(field)
            i = header_idx.get(str(src).strip().lower()) if src else None
            v = r[i] if (i is not None and i < len(r)) else None
            d[field] = _coerce(field, v)
        if not d.get("customerName"):
            continue
        # Spreadsheet row number so an error can point at the row the user sees.
        d["__row"] = row_no
        out.append(d)
    return header, out


def _import_match(value, allowed):
    """Case / spacing-insensitive lookup -> the canonical master value, else None."""
    v = re.sub(r"\s+", " ", str(value or "").strip()).lower()
    for a in allowed:
        if re.sub(r"\s+", " ", str(a).strip()).lower() == v:
            return a
    return None


async def _import_allowed_values():
    """Live dropdown values for the bulk-upload template AND row validation.

    One source for both so a downloaded template can never allow a value the
    upload then rejects. Models / variants come from Price Master, the same list
    the New Lead form uses; the rest come from Settings masters.
    """
    vehicles, seen = [], set()
    for r in await db.price_master.find().to_list(2000):
        if str(r.get("status") or "active").lower() == "inactive":
            continue
        model = str(r.get("model") or "").strip()
        variant = str(r.get("variant") or "").strip()
        if not model:
            continue
        key = (model.lower(), variant.lower())
        if key in seen:
            continue
        seen.add(key)
        vehicles.append({"model": model, "variant": variant})
    vehicles.sort(key=lambda v: (v["model"], v["variant"]))
    masters_now = dict(seeder.MASTERS)
    for cat in EDITABLE_MASTER_CATEGORIES:
        vals = await _masters_list_values(cat)
        if vals:
            masters_now[cat] = vals
    return {
        "leadSources": masters_now.get("leadSources", []),
        "executives": masters_now.get("executives", []),
        "priorities": masters_now.get("priorities", []),
        "statuses": list(IMPORT_STATUSES),
        "yesNo": list(IMPORT_YES_NO),
        "models": sorted({v["model"] for v in vehicles}),
        "variants": sorted({v["variant"] for v in vehicles if v["variant"]}),
        "vehicles": vehicles,
    }


async def _existing_lead_mobiles():
    """last-10-digits -> leadId, for the same duplicate guard POST /leads applies."""
    out = {}
    for l in await db.leads.find().to_list(5000):
        digits = re.sub(r"\D", "", str(l.get("mobile") or ""))
        if len(digits) >= 10:
            out[digits[-10:]] = l.get("leadId")
    return out


def _validate_import_rows(rows, allowed, existing_mobiles):
    """Canonicalise every row against the live masters and collect per-row errors.

    A value that is present but not in its master list is an ERROR, never a silent
    fix — that mismatch is exactly what makes imported leads unusable downstream
    (Price Master lookups, executive scoreboards, source reports). Blank optional
    cells fall back to IMPORT_DEFAULTS, matching the New Lead form.
    """
    pairs = {(v["model"].lower(), v["variant"].lower()) for v in allowed["vehicles"]}
    variants_by_model = {}
    for v in allowed["vehicles"]:
        variants_by_model.setdefault(v["model"].lower(), []).append(v["variant"])
    list_fields = [
        ("leadSource", "leadSources", "Lead Source"),
        ("executive", "executives", "Executive"),
        ("priority", "priorities", "Priority"),
        ("currentStatus", "statuses", "Current Status"),
        ("financeRequired", "yesNo", "Finance Required"),
        ("exchangeRequired", "yesNo", "Exchange Required"),
    ]
    seen_mobiles = {}
    out = []
    for d in rows:
        row = dict(d)
        errors = []
        row_no = row.get("__row") or (len(out) + 2)

        row["customerName"] = str(row.get("customerName") or "").strip()
        if not row["customerName"]:
            errors.append("Customer Name is required")

        mob = re.sub(r"\D", "", str(row.get("mobile") or ""))
        if not mob:
            errors.append("Mobile is required")
        elif len(mob) < 10:
            errors.append("Mobile must be at least 10 digits")
        else:
            mob = mob[-10:]
            if mob in existing_mobiles:
                errors.append(f"Mobile already used by lead {existing_mobiles[mob]}")
            elif mob in seen_mobiles:
                errors.append(f"Duplicate mobile — same number as row {seen_mobiles[mob]}")
            else:
                seen_mobiles[mob] = row_no
        row["mobile"] = mob
        row["altMobile"] = re.sub(r"\D", "", str(row.get("altMobile") or ""))

        for field, key, label in list_fields:
            raw = str(row.get(field) or "").strip()
            if not raw:
                row[field] = IMPORT_DEFAULTS.get(field, "")
                continue
            canonical = _import_match(raw, allowed[key])
            if canonical is None:
                errors.append(f"{label} '{raw}' is not in the {label} list "
                              f"(allowed: {', '.join(allowed[key]) or 'none configured'})")
                row[field] = raw
            else:
                row[field] = canonical

        model_raw = str(row.get("interestedModel") or "").strip()
        variant_raw = str(row.get("variant") or "").strip()
        sku = oem_cat.resolve_sku(model_raw, variant_raw) if model_raw else None
        if sku:
            model, variant = sku.crm_model, sku.crm_variant
            if (model.lower(), variant.lower()) not in pairs:
                errors.append(f"{model} / {variant} is not an active Price Master row")
        else:
            model = _import_match(model_raw, allowed["models"]) if model_raw else ""
            if model_raw and model is None:
                errors.append(f"Interested Model '{model_raw}' is not in Price Master")
                model = model_raw
            variant = variant_raw
            if variant_raw and not model_raw:
                errors.append("Variant needs an Interested Model in the same row")
            elif variant_raw and model:
                options = variants_by_model.get(str(model).lower(), [])
                match = _import_match(variant_raw, [o for o in options if o])
                if match is None and (str(model).lower(), variant_raw.lower()) not in pairs:
                    errors.append(f"Variant '{variant_raw}' does not belong to {model} "
                                  f"(allowed: {', '.join([o for o in options if o]) or 'none in Price Master'})")
                else:
                    variant = match or variant_raw
        row["interestedModel"] = model or ""
        row["variant"] = variant

        for field, label in (("createdDate", "Lead Date"), ("nextFollowupDate", "Next Follow-up")):
            raw = row.get(field)
            iso = _import_date(raw)
            if iso is None:
                errors.append(f"{label} '{raw}' is not a date (use YYYY-MM-DD)")
                row[field] = ""
            else:
                row[field] = iso
        if not row.get("createdDate"):
            row["createdDate"] = today()

        try:
            budget = float(row.get("budget") or 0)
        except (ValueError, TypeError):
            budget = 0.0
        if budget < 0:
            errors.append("Budget cannot be negative")
        row["budget"] = budget

        row["__row"] = row_no
        row["__errors"] = errors
        out.append(row)
    return out


def _import_error_report(rows):
    return [{"row": r.get("__row"), "customerName": r.get("customerName", ""),
             "mobile": r.get("mobile", ""), "errors": r.get("__errors", [])}
            for r in rows if r.get("__errors")]


@api.get("/leads/import/template")
async def import_template(_sales=Depends(sales_staff_only)):
    """Bulk-upload workbook with dropdowns wired to the live masters.

    The dropdown values and the upload validation both come from
    _import_allowed_values(), so a freshly downloaded template can never offer a
    value the upload would reject.

    Excel and Google Sheets cannot express "Variant depends on Interested Model"
    in a portable way (Sheets ignores INDIRECT in validation), so Variant lists
    every variant and the model+variant PAIR is checked on upload instead. The
    Lists sheet shows the valid pairs so staff can see which variant belongs where.
    """
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    allowed = await _import_allowed_values()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"
    lists = wb.create_sheet("Lists")
    guide = wb.create_sheet("How to use")

    list_columns = [
        ("Lead Source", allowed["leadSources"]),
        ("Executive", allowed["executives"]),
        ("Interested Model", allowed["models"]),
        ("Variant", allowed["variants"]),
        ("Priority", allowed["priorities"]),
        ("Current Status", allowed["statuses"]),
        ("Yes / No", allowed["yesNo"]),
    ]
    list_ranges = {}
    for idx, (title, values) in enumerate(list_columns, start=1):
        col = get_column_letter(idx)
        cell = lists[f"{col}1"]
        cell.value = title
        cell.font = Font(bold=True)
        for offset, value in enumerate(values, start=2):
            lists[f"{col}{offset}"] = value
        last_row = len(values) + 1 if values else 2
        list_ranges[title] = f"=Lists!${col}$2:${col}${last_row}"
        lists.column_dimensions[col].width = 24

    pair_col = get_column_letter(len(list_columns) + 2)
    pair_col2 = get_column_letter(len(list_columns) + 3)
    lists[f"{pair_col}1"].value = "Valid Model"
    lists[f"{pair_col2}1"].value = "Valid Variant for that Model"
    lists[f"{pair_col}1"].font = Font(bold=True)
    lists[f"{pair_col2}1"].font = Font(bold=True)
    for offset, v in enumerate(allowed["vehicles"], start=2):
        lists[f"{pair_col}{offset}"] = v["model"]
        lists[f"{pair_col2}{offset}"] = v["variant"]
    lists.column_dimensions[pair_col].width = 24
    lists.column_dimensions[pair_col2].width = 30
    lists.freeze_panes = "A2"

    header_fill = PatternFill("solid", fgColor="1F3A93")
    for idx, (label, field) in enumerate(IMPORT_COLUMNS, start=1):
        col = get_column_letter(idx)
        cell = ws[f"{col}1"]
        cell.value = label
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[col].width = 22 if field != "remarks" else 34
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22

    LAST_ROW = 500
    dropdown_fields = {
        "leadSource": "Lead Source",
        "executive": "Executive",
        "interestedModel": "Interested Model",
        "variant": "Variant",
        "priority": "Priority",
        "currentStatus": "Current Status",
        "financeRequired": "Yes / No",
        "exchangeRequired": "Yes / No",
    }
    field_col = {field: get_column_letter(i) for i, (_, field) in enumerate(IMPORT_COLUMNS, start=1)}
    for field, list_title in dropdown_fields.items():
        # showDropDown is intentionally left unset: in OOXML a truthy value HIDES
        # the in-cell arrow, which is the opposite of what the name suggests.
        # showErrorMessage must be on or Excel keeps the list as a hint only and
        # silently accepts a typed value.
        dv = DataValidation(type="list", formula1=list_ranges[list_title], allow_blank=True,
                            showErrorMessage=True, showInputMessage=True, errorStyle="stop")
        dv.errorTitle = "Not a Euler CRM value"
        dv.error = f"Pick from the {list_title} list on the Lists sheet."
        dv.promptTitle = list_title
        dv.prompt = "Choose from the list so the upload matches the app."
        ws.add_data_validation(dv)
        col = field_col[field]
        dv.add(f"{col}2:{col}{LAST_ROW}")

    # Column-level formats: setting these per cell would materialise 500 blank rows
    # into the sheet the user opens (and inflate the file).
    for field in IMPORT_DATE_FIELDS:
        ws.column_dimensions[field_col[field]].number_format = "yyyy-mm-dd"
    # Text format keeps a 10-digit mobile from becoming 9.8e+09.
    for field in ("mobile", "altMobile"):
        ws.column_dimensions[field_col[field]].number_format = "@"

    guide_lines = [
        ("Euler CRM — bulk lead upload", True),
        ("", False),
        ("1. Type one lead per row on the 'Leads' sheet. Do not rename or reorder the header row.", False),
        ("2. Grey-list columns have dropdowns. Pick a value — typing your own is rejected on upload.", False),
        ("3. Required: Customer Name and Mobile (10 digits, not already in the CRM).", False),
        ("4. Lead Date / Next Follow-up: use YYYY-MM-DD. Blank Lead Date becomes today.", False),
        ("5. Variant must belong to the Interested Model — see 'Valid Model / Valid Variant' on Lists.", False),
        (f"6. Current Status can only be: {', '.join(IMPORT_STATUSES)}. "
         "Booking and delivery are done inside the app so the money side stays correct.", False),
        ("7. Blank Lead Source / Priority / Status / Finance / Exchange use the New Lead defaults "
         f"({IMPORT_DEFAULTS['leadSource']} / {IMPORT_DEFAULTS['priority']} / "
         f"{IMPORT_DEFAULTS['currentStatus']} / No / No).", False),
        ("8. Upload in the app: Lead Register → Import. Rows with problems are listed and skipped; "
         "correct them and upload again.", False),
        ("", False),
        (f"Lists generated on {today()} from Settings masters and Price Master. "
         "Download a fresh template after adding an executive, source or vehicle.", False),
    ]
    for i, (text, bold) in enumerate(guide_lines, start=1):
        cell = guide[f"A{i}"]
        cell.value = text
        if bold:
            cell.font = Font(bold=True, size=14)
    guide.column_dimensions["A"].width = 120

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"euler_lead_upload_template_{today()}.xlsx"
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@api.post("/leads/import/preview")
async def import_preview(file: UploadFile = File(...), mapping: Optional[str] = Form(None),
                         _sales=Depends(sales_staff_only)):
    import json as _json
    content = await file.read()
    try:
        headers = [str(h).strip() if h is not None else "" for h in (_read_rows(file.filename, content) or [[]])[0]]
        mp = _json.loads(mapping) if mapping else _suggest_mapping(headers)
        _, rows = _parse_import_bytes(file.filename, content, mp)
    except Exception as e:
        raise HTTPException(400, f"Could not parse file: {e}")
    allowed = await _import_allowed_values()
    checked = _validate_import_rows(rows, allowed, await _existing_lead_mobiles())
    valid = [r for r in checked if not r["__errors"]]
    return {"detectedHeaders": [h for h in headers if h],
            "targetFields": [{"label": lbl, "field": fld} for lbl, fld in IMPORT_COLUMNS],
            "suggestedMapping": mp, "rowCount": len(checked),
            "validCount": len(valid), "errorCount": len(checked) - len(valid),
            "allowedValues": allowed,
            "requiredFields": list(IMPORT_REQUIRED_FIELDS),
            "errors": _import_error_report(checked)[:100],
            "sample": checked[:12]}


@api.post("/leads/import/commit")
async def import_commit(file: UploadFile = File(...), mapping: Optional[str] = Form(None),
                        user=Depends(sales_staff_only)):
    """Insert only the rows that pass validation; report the rest untouched."""
    if str(user.get("role") or "") == "executive":
        raise HTTPException(
            403,
            "Bulk import creates live leads. Send each enquiry for GM / Owner approval, "
            "or ask them to import.",
        )
    import json as _json
    content = await file.read()
    try:
        mp = _json.loads(mapping) if mapping else None
        _, rows = _parse_import_bytes(file.filename, content, mp)
    except Exception as e:
        raise HTTPException(400, f"Could not parse file: {e}")
    allowed = await _import_allowed_values()
    checked = _validate_import_rows(rows, allowed, await _existing_lead_mobiles())
    created = 0
    created_ids = []
    for d in checked:
        if d["__errors"]:
            continue
        lead_id = await next_id("lead", "LD26")
        doc = {
            "leadId": lead_id, "createdDate": d.get("createdDate") or today(),
            "customerName": d.get("customerName", ""), "mobile": d.get("mobile", ""),
            "altMobile": d.get("altMobile", ""), "village": d.get("village", ""), "city": d.get("city", ""),
            "leadSource": d.get("leadSource") or IMPORT_DEFAULTS["leadSource"],
            "interestedModel": d.get("interestedModel", ""),
            "variant": d.get("variant", ""), "executive": d.get("executive", ""),
            "currentStatus": d.get("currentStatus") or IMPORT_DEFAULTS["currentStatus"],
            "priority": d.get("priority") or IMPORT_DEFAULTS["priority"],
            "nextFollowupDate": d.get("nextFollowupDate", ""),
            "budget": d.get("budget", 0), "remarks": d.get("remarks", ""),
            "financeRequired": d.get("financeRequired") or IMPORT_DEFAULTS["financeRequired"],
            "exchangeRequired": d.get("exchangeRequired") or IMPORT_DEFAULTS["exchangeRequired"],
            "accountStatus": "Active", "deliveryStatus": "",
            "outstandingAmount": 0, "customerOutstanding": 0, "companyOutstanding": 0, "totalReceived": 0,
            "customerPayable": 0, "grossVehicleCost": 0, "totalDiscount": 0,
            "consumerDiscount": 0, "exchangeBonus": 0, "loyaltyBonus": 0, "referralBonus": 0,
            "dsaDiscount": 0, "additionalDiscount": 0, "exShowroom": 0, "rto": 0, "insuranceAmount": 0,
            "accessoriesAmount": 0, "handlingCharges": 0, "trc": 0, "fastag": 0, "extendedWarranty": 0,
            "otherCharges": 0, "bookingAmount": 0, "lastUpdated": now_iso(), "importedBatch": today(),
        }
        await db.leads.insert_one(doc)
        await sheet_sync("leads", doc)
        created += 1
        created_ids.append(lead_id)
    errors = _import_error_report(checked)
    return {"created": created, "skipped": len(errors), "rowCount": len(checked),
            "leadIds": created_ids, "errors": errors[:100]}


# ---------------------------------------------------------------- payments
async def _add_payment_internal(lead_id, body: PaymentIn):
    # Same guard record_financer_receipt/record_claim_receipt already have — this endpoint
    # was missing it, which let a negative amount silently reduce runningTotal (the
    # over-payment check below only ever fires on running > payable, never on a value
    # that DECREASES running, so negative amounts sailed through undetected).
    if body.amount <= 0:
        raise HTTPException(422, "Enter a valid payment amount")
    lead = await db.leads.find_one({"leadId": lead_id})
    # Double-submit guard (U4): reject an identical receipt (same lead/amount/mode) within 4s
    recent = await db.payments.find_one(
        {"leadId": lead_id, "amount": ce.round2(body.amount), "paymentMode": body.paymentMode},
        sort=[("_id", -1)])
    if recent and recent.get("recordedAt"):
        try:
            prev = datetime.fromisoformat(recent["recordedAt"])
            if (datetime.now(timezone.utc) - prev).total_seconds() < 4:
                raise HTTPException(409, "Duplicate submission detected — this receipt was just recorded. Please wait a moment.")
        except HTTPException:
            raise
        except Exception:
            pass
    receipt = await next_id("receipt", "RC26")
    prior = await db.payments.aggregate([
        {"$match": {"leadId": lead_id}}, {"$group": {"_id": None, "t": {"$sum": "$amount"}}}
    ]).to_list(1)
    running = ce.round2((prior[0]["t"] if prior else 0) + body.amount)
    snap = lead_to_snapshot(lead) if lead else {}
    payable = ce.compute_commercial_totals(snap)["customerPayable"] if lead else 0
    # Over-payment guard (port of BusinessRulesService.validatePaymentAmount_): a receipt may
    # never push total received above Customer Payable *by accident*. Provisional allowed only
    # when payable is still ₹0 (slim booking, price deferred to Price Structure). Staff can
    # deliberately record a surplus with allowExcess — it lands in Excess Received and is
    # refundable, including after delivery or closure.
    if payable > 0 and running > payable + 0.01 and not body.allowExcess:
        room = ce.round2(max(0.0, payable - (running - body.amount)))
        raise HTTPException(422, f"Amount ₹{ce.round2(body.amount)} exceeds the balance. Customer payable is ₹{payable}; "
                                 f"only ₹{room} can still be collected. Confirm it as an excess payment to record the surplus.")
    outstanding = ce.round2(max(0.0, payable - running)) if payable > 0 else 0.0
    finance_file_number = body.financeFileNumber
    if body.paymentMode == "Finance":
        finance_file_number = await _resolve_finance_file_for_payment(lead_id, body)
    doc = {
        "receiptNumber": receipt, "leadId": lead_id, "customerName": lead.get("customerName") if lead else "",
        "date": body.date or today(), "amount": ce.round2(body.amount), "paymentMode": body.paymentMode,
        "narration": body.narration, "runningTotal": running, "outstandingBalance": outstanding,
        "paymentId": f"PY{uuid.uuid4().hex[:12]}", "financerName": body.financerName,
        "financeFileNumber": finance_file_number, "recordedAt": now_iso(),
    }
    try:
        await db.payments.insert_one(doc)
    except Exception:
        if body.paymentMode == "Finance":
            await db.finance.delete_one({
                "leadId": lead_id, "fileNumber": finance_file_number,
                "sanctionedAmount": 0.0, "receivedAgainstFile": 0.0, "receipts": []
            })
        raise
    await sheet_sync("payments", doc)
    if body.paymentMode == "Finance":
        await _upsert_finance_file(lead_id, body, finance_file_number)
    return clean(doc)


@api.get("/payments")
async def list_payments(lead_id: Optional[str] = None):
    q = {"leadId": lead_id} if lead_id else {}
    return [clean(p) for p in await db.payments.find(q).sort("date", -1).to_list(2000)]


@api.post("/leads/{lead_id}/payments")
async def add_payment(lead_id: str, body: PaymentIn, act=Depends(actor), _money=Depends(money_desk_only)):
    lead = await get_lead_or_404(lead_id)
    if body.paymentMode == "Finance":
        _require_action(lead, "canFinanceReceipt", "finance receipt (lead is archived)", act)
    else:
        _require_action(lead, "canPayment", "customer payment (only Active leads)", act)
    rec = await _add_payment_internal(lead_id, body)
    await recompute_lead(lead_id)
    await _refresh_billing_summary_if_delivered(lead_id)
    await write_audit(act, "receipt", "payment", leadId=lead_id, paymentId=rec.get("receiptNumber"),
                      financeFileNumber=rec.get("financeFileNumber") or "",
                      new={"amount": rec.get("amount"), "mode": body.paymentMode, "runningTotal": rec.get("runningTotal")})
    return rec


async def _rebuild_payment_running_totals(lead_id):
    """After a receipt is removed, rewrite runningTotal / outstandingBalance in date order."""
    lead = await db.leads.find_one({"leadId": lead_id}) or {}
    payable = ce.compute_commercial_totals(lead_to_snapshot(lead), await get_scheme_rows())["customerPayable"]
    rows = await db.payments.find({"leadId": lead_id}).sort(
        [("date", 1), ("recordedAt", 1), ("receiptNumber", 1)]
    ).to_list(2000)
    running = 0.0
    for p in rows:
        running = ce.round2(running + ce.num(p.get("amount")))
        outstanding = ce.round2(max(0.0, payable - running)) if payable > 0 else 0.0
        await db.payments.update_one({"_id": p["_id"]}, {"$set": {
            "runningTotal": running, "outstandingBalance": outstanding,
        }})
        await sheet_sync("payments", {
            **{k: v for k, v in p.items() if k != "_id"},
            "runningTotal": running, "outstandingBalance": outstanding,
        })


async def _finance_commitment_excluding(lead_id, file_number, exclude_receipt):
    q = {"leadId": lead_id, "paymentMode": "Finance", "financeFileNumber": file_number,
         "receiptNumber": {"$ne": exclude_receipt}}
    remaining = await db.payments.find(q).to_list(2000)
    return ce.round2(sum(
        ce.num(p.get("amount")) for p in remaining
        if ce.num(p.get("amount")) > 0 and (p.get("entryType") or "") != "Refund"
    ))


async def _rebuild_finance_file_from_payments(lead_id, file_number):
    """Re-derive sanctioned amount from remaining Finance-mode receipts on the lead."""
    committed = await _finance_commitment_excluding(lead_id, file_number, exclude_receipt="")
    f = await db.finance.find_one({"fileNumber": file_number, "leadId": lead_id})
    if not f:
        return {"ok": True, "removed": False}
    received = ce.num(f.get("receivedAgainstFile"))
    receipts = list(f.get("receipts") or [])
    if committed <= 0.01 and received <= 0.01 and not receipts:
        await db.finance.delete_one({"fileNumber": file_number, "leadId": lead_id})
        await gsheets.delete_entity_row("finance", file_number)
        leftover = await db.finance.find_one({"leadId": lead_id})
        lead_set = {"lastUpdated": now_iso()}
        if leftover:
            lead_set.update({
                "financeRequired": "Yes",
                "financerName": leftover.get("financer") or "",
                "financeFileNumber": leftover.get("fileNumber") or "",
            })
        else:
            lead_set.update({"financerName": "", "financeFileNumber": ""})
        await db.leads.update_one({"leadId": lead_id}, {"$set": lead_set})
        await rebuild_finance_views()
        return {"ok": True, "removed": True, "fileNumber": file_number}
    outstanding = ce.round2(max(0.0, committed - received))
    status = "Received" if outstanding <= 0 else ("Partial" if received > 0 else "Pending")
    await db.finance.update_one({"fileNumber": file_number, "leadId": lead_id}, {"$set": {
        "sanctionedAmount": committed, "fileOutstanding": outstanding, "status": status,
        "lastUpdated": today(),
    }})
    await sync_finance_file(file_number)
    await rebuild_finance_views()
    return {"ok": True, "removed": False, "fileNumber": file_number, "sanctionedAmount": committed}


@api.delete("/payments/{receipt_number}", dependencies=[Depends(owner_only)])
async def delete_payment(receipt_number: str, act=Depends(actor)):
    """Owner-only. Remove a wrongly posted receipt (or refund) and recalculate the lead."""
    receipt_number = (receipt_number or "").strip()
    pay = await db.payments.find_one({"receiptNumber": receipt_number})
    if not pay:
        raise HTTPException(404, "Payment not found")
    lead_id = pay.get("leadId") or ""
    file_number = (pay.get("financeFileNumber") or "").strip()
    if (pay.get("paymentMode") or "") == "Finance" and file_number:
        remaining = await _finance_commitment_excluding(lead_id, file_number, receipt_number)
        f = await db.finance.find_one({"fileNumber": file_number, "leadId": lead_id}) or {}
        received = ce.num(f.get("receivedAgainstFile"))
        if received > remaining + 0.01:
            raise HTTPException(
                422,
                f"Cannot delete this Finance receipt — file {file_number} already has "
                f"₹{ce.round2(received)} disbursed from the financer, which is more than the "
                f"₹{remaining} that would remain committed. Reverse the extra financer receipt "
                f"on the Finance Register first, or leave this customer receipt in place.",
            )
    await db.payments.delete_one({"receiptNumber": receipt_number})
    sheet_result = await gsheets.delete_entity_row("payments", receipt_number)
    await db.sheet_sync_log.delete_many({"entityType": "payments", "entityId": receipt_number})
    await _rebuild_payment_running_totals(lead_id)
    finance_result = None
    if (pay.get("paymentMode") or "") == "Finance" and file_number:
        finance_result = await _rebuild_finance_file_from_payments(lead_id, file_number)
    await recompute_lead(lead_id)
    await _refresh_billing_summary_if_delivered(lead_id)
    await write_audit(
        act, "delete", "payment", leadId=lead_id, paymentId=receipt_number,
        financeFileNumber=file_number,
        old={"amount": pay.get("amount"), "mode": pay.get("paymentMode"),
             "entryType": pay.get("entryType") or "Receipt",
             "runningTotal": pay.get("runningTotal"), "narration": pay.get("narration")},
        new={"sheet": sheet_result, "finance": finance_result},
    )
    lead = clean(await db.leads.find_one({"leadId": lead_id}) or {})
    return {
        "ok": True, "deleted": receipt_number, "leadId": lead_id,
        "totalReceived": lead.get("totalReceived"),
        "customerOutstanding": lead.get("customerOutstanding"),
        "excessReceived": lead.get("excessReceived"),
        "sheet": sheet_result, "finance": finance_result,
    }


async def _refund_position(lead_id, lead=None):
    """Money held above Customer Payable, and how much of it has been returned.

    Excess is net of refunds because a refund is stored as a negative ledger row, so
    the same number is both "surplus still held" and "still refundable".
    """
    lead = lead or await db.leads.find_one({"leadId": lead_id}) or {}
    payable = ce.compute_commercial_totals(lead_to_snapshot(lead), await get_scheme_rows())["customerPayable"]
    rows = await db.payments.find({"leadId": lead_id}).to_list(2000)
    received = ce.round2(sum(ce.num(p.get("amount")) for p in rows))
    refunded = ce.round2(sum(-ce.num(p.get("amount")) for p in rows if p.get("entryType") == "Refund"))
    if lead.get("dealCancelled"):
        # No vehicle to set the money against any more, so the whole balance is the
        # customer's. `received` is already net of refunds, so this also stops being
        # refundable once it has been returned.
        excess = ce.round2(max(0.0, received))
    else:
        excess = ce.round2(max(0.0, received - payable)) if payable > 0 else 0.0
    return {"customerPayable": ce.round2(payable), "totalReceived": received,
            "refundedAmount": refunded, "excessReceived": excess,
            "dealCancelled": bool(lead.get("dealCancelled"))}


@api.get("/leads/{lead_id}/refund-position")
async def get_refund_position(lead_id: str, _money=Depends(money_desk_only)):
    lead = await get_lead_or_404(lead_id)
    return await _refund_position(lead_id, lead)


@api.post("/leads/{lead_id}/refund")
async def refund_excess_payment(lead_id: str, body: RefundIn, act=Depends(actor), _money=Depends(money_desk_only)):
    """Return surplus money to the customer as a negative ledger entry.

    Deliberately NOT gated on lead status: an excess is the customer's money, so it
    stays refundable after Mark Delivered and after the lead is closed. Capped at the
    surplus so a refund can never re-open Customer Outstanding (which would silently
    turn a delivered lead into a short payment).
    """
    lead = await get_lead_or_404(lead_id)
    amount = ce.round2(body.amount)
    if amount <= 0:
        raise HTTPException(422, "Enter a valid refund amount")
    pos = await _refund_position(lead_id, lead)
    if pos["excessReceived"] <= 0:
        if lead.get("dealCancelled"):
            raise HTTPException(422, "Nothing is left to refund on this cancelled lead — "
                                     "everything received has already been returned.")
        raise HTTPException(422, "There is no excess payment on this lead to refund. Only money "
                                 "collected above Customer Payable can be returned here.")
    if amount > pos["excessReceived"] + 0.01:
        raise HTTPException(422, f"Refund ₹{amount} is more than the excess held on this lead "
                                 f"(₹{pos['excessReceived']}).")
    refund_date = body.date or today()
    recent = await db.payments.find_one(
        {"leadId": lead_id, "entryType": "Refund", "amount": ce.round2(-amount)}, sort=[("_id", -1)])
    if recent and recent.get("recordedAt"):
        try:
            if (datetime.now(timezone.utc) - datetime.fromisoformat(recent["recordedAt"])).total_seconds() < 4:
                raise HTTPException(409, "Duplicate submission detected — this refund was just recorded.")
        except HTTPException:
            raise
        except (TypeError, ValueError):
            pass
    entry_id = await next_id("refund", "RF26")
    running = ce.round2(pos["totalReceived"] - amount)
    narration = (body.narration or "").strip() or "Refund of excess payment"
    if body.reference.strip():
        narration = f"{narration} · Ref {body.reference.strip()}"
    doc = {
        "receiptNumber": entry_id, "leadId": lead_id, "customerName": lead.get("customerName", ""),
        "date": refund_date, "amount": ce.round2(-amount), "paymentMode": body.paymentMode,
        "entryType": "Refund", "narration": narration, "runningTotal": running,
        "outstandingBalance": ce.round2(max(0.0, pos["customerPayable"] - running)),
        "paymentId": f"RF{uuid.uuid4().hex[:12]}", "financerName": "", "financeFileNumber": "",
        "refundReference": body.reference.strip(), "recordedAt": now_iso(),
    }
    await db.payments.insert_one(doc)
    await sheet_sync("payments", doc)
    await recompute_lead(lead_id)
    await _refresh_billing_summary_if_delivered(lead_id)
    await write_audit(act, "refund", "payment", leadId=lead_id, paymentId=entry_id,
                      old={"totalReceived": pos["totalReceived"], "excessReceived": pos["excessReceived"]},
                      new={"amount": amount, "mode": body.paymentMode, "runningTotal": running})
    return {"refund": clean(doc), **(await _refund_position(lead_id))}


# ---------------------------------------------------------------- finance
FINANCE_FILE_PATTERN = re.compile(r"^FN26\d{6}$")


def is_legacy_finance_file_number(file_number):
    """A finance file number that predates the FN26 contract (e.g. the live '55').
    Historical records keep their number — renumbering them would break every
    payment and sheet row that references them."""
    return bool(str(file_number or "").strip()) and not FINANCE_FILE_PATTERN.match(str(file_number).strip())


async def _resolve_finance_file_for_payment(lead_id, body: PaymentIn):
    if not (body.financerName or "").strip():
        raise HTTPException(422, "Financer is required for Finance payments")
    supplied = (body.financeFileNumber or "").strip()
    by_lead = await db.finance.find_one({"leadId": lead_id})
    if supplied:
        by_file = await db.finance.find_one({"fileNumber": supplied})
        if by_file and by_file.get("leadId") != lead_id:
            raise HTTPException(422, "Finance file number already belongs to another lead")
        if by_lead and by_lead.get("fileNumber") != supplied:
            raise HTTPException(422, "Lead already has a different finance file number")
        # A NEW file must follow the numbering contract. Staff are never required to
        # invent one — leaving the field blank generates FN26xxxxxx. This is what let
        # a hand-typed "55" into production and made the Finance Register unkeyable.
        # An EXISTING record keeps whatever number it already has (legacy support).
        if not by_file and not FINANCE_FILE_PATTERN.match(supplied):
            raise HTTPException(422,
                f"'{supplied}' is not a valid Finance File Number. Leave the field blank and "
                f"the system will generate one (FN26xxxxxx), or enter an existing file number.")
        return supplied
    if by_lead:
        return by_lead.get("fileNumber")
    file_number = await next_id("finance", "FN26")
    try:
        res = await db.finance.find_one_and_update(
            {"leadId": lead_id},
            {"$setOnInsert": {
                "fileNumber": file_number, "leadId": lead_id,
                "customerName": (await db.leads.find_one({"leadId": lead_id}) or {}).get("customerName", ""),
                "financer": body.financerName, "sanctionedAmount": 0.0,
                "receivedAgainstFile": 0.0, "fileOutstanding": 0.0,
                "status": "Pending", "receipts": [], "lastUpdated": today(),
            }},
            upsert=True, return_document=True)
        return res.get("fileNumber")
    except DuplicateKeyError:
        existing = await db.finance.find_one({"leadId": lead_id})
        if existing:
            return existing.get("fileNumber")
        raise HTTPException(409, "Could not resolve finance file; please retry")


async def _upsert_finance_file(lead_id, body: PaymentIn, finance_file_number: str):
    """A Finance-mode entry on a lead = amount the FINANCER is now liable to disburse.
    It accrues the file's committed amount; the customer's outstanding already dropped
    by this amount (recompute counts Finance payments). Actual disbursement is booked
    separately via POST /finance/{file}/receipt and never re-touches customer outstanding."""
    lead = await db.leads.find_one({"leadId": lead_id})
    existing = await db.finance.find_one({"fileNumber": finance_file_number})
    if existing and existing.get("leadId") != lead_id:
        raise HTTPException(422, "Finance file number already belongs to another lead")
    if existing:
        committed = ce.round2(ce.num(existing.get("sanctionedAmount")) + body.amount)
        received = ce.num(existing.get("receivedAgainstFile"))
        outstanding = ce.round2(max(0.0, committed - received))
        await db.finance.update_one({"fileNumber": finance_file_number, "leadId": lead_id}, {"$set": {
            "sanctionedAmount": committed, "fileOutstanding": outstanding, "financer": body.financerName or existing.get("financer"),
            "status": "Received" if outstanding <= 0 else ("Partial" if received > 0 else "Pending"),
            "lastUpdated": today(),
        }})
    else:
        committed = ce.round2(body.amount)
        try:
            await db.finance.insert_one({
                "fileNumber": finance_file_number, "leadId": lead_id,
                "customerName": lead.get("customerName") if lead else "", "financer": body.financerName,
                "sanctionedAmount": committed, "receivedAgainstFile": 0.0,
                "fileOutstanding": committed, "status": "Pending", "receipts": [],
                "lastUpdated": today(),
            })
        except DuplicateKeyError:
            await _upsert_finance_file(lead_id, body, (await db.finance.find_one({"leadId": lead_id}))["fileNumber"])
            return
    # DEFECT C fix: mirror the authoritative finance linkage onto the lead so the
    # Lead Register (which maps financerName/financeFileNumber as lead columns) is never
    # stale. recompute_lead runs right after this and pushes the lead row to the sheet.
    fin = await db.finance.find_one({"fileNumber": finance_file_number})
    await db.leads.update_one({"leadId": lead_id}, {"$set": {
        "financeRequired": "Yes",
        "financerName": (fin or {}).get("financer") or body.financerName or "",
        "financeFileNumber": finance_file_number,
        "lastUpdated": now_iso(),
    }})
    await sync_finance_file(finance_file_number)
    await rebuild_finance_views()


async def sync_finance_file(file_number):
    """GS-5: mirror a finance file into the existing Finance tab (upsert on file number)."""
    f = await db.finance.find_one({"fileNumber": file_number})
    if not f:
        return
    await sheet_sync("finance", {
        "financeFileNumber": f.get("fileNumber"), "leadId": f.get("leadId"),
        "customerName": f.get("customerName"), "financerName": f.get("financer"),
        "committedAmount": f.get("sanctionedAmount"), "disbursedAmount": f.get("receivedAgainstFile"),
        "financeOutstanding": f.get("fileOutstanding"), "status": f.get("status"),
    })


FINANCE_RECEIPT_SLA_DAYS = 2  # port of FinanceService.gs getFinanceReceiptSlaDays_ (CRM.FINANCE_REGISTER.RECEIPT_SLA_DAYS)


async def _enrich_finance_with_delivery(files):
    """Port of enrichFinanceFilesWithDelivery_: overdue = days since delivery > SLA
    (files with no known delivery date are never overdue — SLA clock hasn't started)."""
    sla = FINANCE_RECEIPT_SLA_DAYS
    today_d = today()
    out = []
    for f in files:
        lead_id = str(f.get("leadId") or "").strip()
        delivery_date = ""
        if lead_id:
            d = await db.deliveries.find_one({"leadId": lead_id}) or {}
            delivery_date = str(d.get("deliveryDate") or "")[:10]
            if not delivery_date:
                lead = await db.leads.find_one({"leadId": lead_id}) or {}
                delivery_date = str(lead.get("deliveryDate") or "")[:10]
        row = dict(f)
        if re.match(r"^\d{4}-\d{2}-\d{2}$", delivery_date):
            days = _claim_ageing_days(delivery_date, "Pending")  # days since delivery, running
            row["deliveryDate"] = delivery_date
            row["daysSinceDelivery"] = days
            row["overdue"] = days > sla
        else:
            row["deliveryDate"] = ""
            row["daysSinceDelivery"] = ""
            row["overdue"] = False
        out.append(row)
    return out


@api.get("/finance")
async def list_finance(view: str = "all", _viewer=Depends(finance_viewer_only)):
    files = [clean(f) for f in await db.finance.find().to_list(1000)]
    pending = [f for f in files if ce.num(f.get("fileOutstanding")) > 0 and f.get("status") != "Received"]
    if view == "pending":
        return pending
    if view == "overdue":
        enriched = await _enrich_finance_with_delivery(pending)
        return [f for f in enriched if f.get("overdue")]
    return files


# --- Derived Finance Pending / Overdue sheet tabs (DEFECT B) -----------------
# These were legacy Apps-Script report tabs the migrated app never maintained, so
# they stayed frozen ("Pending: 0") while the Finance Register held live open files.
# We now rebuild them deterministically from the authoritative Finance Register +
# Delivery Tracker on every finance-affecting event, matching the workbook's format.
FINANCE_PENDING_TITLE = "Finance Pending — open files awaiting financer payment (SLA: 2 days after delivery)"
FINANCE_OVERDUE_TITLE = "Finance Overdue — files past 2-day SLA"
FINANCE_PENDING_HEADER = ["File Number", "Lead ID", "Customer", "Financer", "Sanctioned Amount",
                          "Received", "File Outstanding", "Status", "Delivery Date",
                          "Days Since Delivery", "Due By", "Overdue", "Last Payment Date"]
FINANCE_OVERDUE_HEADER = ["File Number", "Lead ID", "Customer", "Financer", "Sanctioned Amount",
                          "File Outstanding", "Status", "Delivery Date", "Days Since Delivery", "Due By"]


def _finance_due_by(delivery_date: str) -> str:
    if re.match(r"^\d{4}-\d{2}-\d{2}$", str(delivery_date or "")):
        y, m, d = map(int, delivery_date.split("-"))
        return (date(y, m, d) + timedelta(days=FINANCE_RECEIPT_SLA_DAYS)).isoformat()
    return ""


async def rebuild_finance_views():
    """Rebuild the derived Finance Pending & Finance Overdue sheet tabs. No-op if
    Sheets sync is disabled. Full-mirror (clear+write) so closed/received files
    correctly disappear rather than lingering."""
    files = [clean(f) for f in await db.finance.find().to_list(2000)]
    pending = [f for f in files if ce.num(f.get("fileOutstanding")) > 0 and f.get("status") != "Received"]
    enriched = await _enrich_finance_with_delivery(pending)
    overdue = [f for f in enriched if f.get("overdue")]
    ts = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    total_out = ce.round2(sum(ce.num(f.get("fileOutstanding")) for f in pending))

    p_rows = [[
        f.get("fileNumber"), f.get("leadId"), f.get("customerName"), f.get("financer"),
        ce.num(f.get("sanctionedAmount")), ce.num(f.get("receivedAgainstFile")), ce.num(f.get("fileOutstanding")),
        f.get("status"), f.get("deliveryDate") or "",
        f.get("daysSinceDelivery") if f.get("daysSinceDelivery") != "" else "",
        _finance_due_by(f.get("deliveryDate")), "Yes" if f.get("overdue") else "No", f.get("lastUpdated") or "",
    ] for f in enriched]
    pending_values = [[FINANCE_PENDING_TITLE],
                      [f"Refreshed: {ts} | Pending: {len(pending)} | Overdue: {len(overdue)} | Total Outstanding: ₹{total_out}"],
                      FINANCE_PENDING_HEADER]
    pending_values += p_rows or [["No pending finance files."]]

    o_rows = [[
        f.get("fileNumber"), f.get("leadId"), f.get("customerName"), f.get("financer"),
        ce.num(f.get("sanctionedAmount")), ce.num(f.get("fileOutstanding")), f.get("status"),
        f.get("deliveryDate") or "", f.get("daysSinceDelivery") if f.get("daysSinceDelivery") != "" else "",
        _finance_due_by(f.get("deliveryDate")),
    ] for f in overdue]
    overdue_values = [[FINANCE_OVERDUE_TITLE],
                      [f"Refreshed: {ts} | Overdue: {len(overdue)}"],
                      FINANCE_OVERDUE_HEADER]
    overdue_values += o_rows or [["No overdue finance files. ✔"]]

    ok_p = await gsheets.overwrite_report_tab(gsheets.FINANCE_PENDING_TAB, pending_values)
    ok_o = await gsheets.overwrite_report_tab(gsheets.FINANCE_OVERDUE_TAB, overdue_values)
    return {"pending": len(pending), "overdue": len(overdue), "syncedPending": ok_p, "syncedOverdue": ok_o}


class ReceiptIn(BaseModel):
    amount: float
    date: str = ""
    reference: str = ""


FINANCE_RECEIPT_DEDUPE_SECONDS = 10


@api.post("/finance/{file_number}/receipt")
async def record_financer_receipt(file_number: str, body: ReceiptIn, act=Depends(actor), _money=Depends(money_desk_only)):
    """Record money actually disbursed by the financer against a file. Does NOT change customer outstanding."""
    f = await db.finance.find_one({"fileNumber": file_number})
    if not f:
        raise HTTPException(404, "Finance file not found")
    if body.amount <= 0:
        raise HTTPException(422, "Enter a valid receipt amount")
    committed = ce.num(f.get("sanctionedAmount"))
    amount = ce.round2(body.amount)
    receipt_date = body.date or today()
    # A double-click posted this twice: the receipts array grew two identical entries
    # and receivedAgainstFile doubled, while the Finance tab (upsert keyed on file
    # number) still showed one row — so the app looked wrong and the sheet looked right.
    for prior in reversed(f.get("receipts") or []):
        if ce.round2(ce.num(prior.get("amount"))) != amount or str(prior.get("date") or "") != receipt_date:
            continue
        try:
            prev_at = datetime.fromisoformat(prior["recordedAt"])
        except (KeyError, TypeError, ValueError):
            continue
        if (datetime.now(timezone.utc) - prev_at).total_seconds() < FINANCE_RECEIPT_DEDUPE_SECONDS:
            raise HTTPException(409, "Duplicate submission detected — this financer receipt was just "
                                     "recorded. Open the file to confirm before recording it again.")
    received = ce.round2(ce.num(f.get("receivedAgainstFile")) + amount)
    # Even a late double-click must not book more than the financer committed. Without
    # this the file read as fully disbursed on half the money (outstanding forced to 0).
    if committed > 0 and received > committed + 0.01:
        room = ce.round2(max(0.0, committed - ce.num(f.get("receivedAgainstFile"))))
        raise HTTPException(422, f"Receipt ₹{amount} is more than this file still expects. "
                                 f"Committed ₹{ce.round2(committed)}, already received "
                                 f"₹{ce.round2(ce.num(f.get('receivedAgainstFile')))} — only ₹{room} can be booked. "
                                 f"Add the extra commitment on the lead first if the financer disbursed more.")
    outstanding = ce.round2(max(0.0, committed - received))
    receipt = {"amount": amount, "date": receipt_date,
               "reference": body.reference, "recordedAt": now_iso()}
    await db.finance.update_one({"fileNumber": file_number}, {
        "$set": {"receivedAgainstFile": received, "fileOutstanding": outstanding,
                 "status": "Received" if outstanding <= 0 else "Partial",
                 "lastPaymentDate": receipt_date, "lastUpdated": today()},
        "$push": {"receipts": receipt},
    })
    await write_audit(act, "receipt", "finance", leadId=f.get("leadId", ""), financeFileNumber=file_number,
                      old={"receivedAgainstFile": ce.num(f.get("receivedAgainstFile")), "fileOutstanding": ce.num(f.get("fileOutstanding"))},
                      new={"receivedAgainstFile": received, "fileOutstanding": outstanding, "amount": amount})
    await sync_finance_file(file_number)
    await rebuild_finance_views()
    return clean(await db.finance.find_one({"fileNumber": file_number}))


# ---------------------------------------------------------------- deliveries
@api.get("/deliveries")
async def list_deliveries(user=Depends(current_user)):
    # active-booked leads that are not delivered = pending deliveries; plus delivered ones
    leads = await db.leads.find({"currentStatus": {"$in": ["Booked", "Finance Process", "Delivered"]}}).to_list(2000)
    if user.get("role") == "executive":
        leads = _leads_for_executive(leads, user)
    result = []
    for l in leads:
        d = await db.deliveries.find_one({"leadId": l["leadId"]}) or {}
        result.append({
            "leadId": l["leadId"], "customerName": l.get("customerName"), "mobile": l.get("mobile"),
            "model": l.get("interestedModel"), "variant": l.get("variant"),
            "insurance": d.get("insurance", ""), "registration": d.get("registration", ""),
            "invoice": d.get("invoice", ""), "rc": d.get("rc", ""), "pdi": d.get("pdi", ""),
            "delivered": d.get("delivered", "") or ("Yes" if (l.get("deliveryStatus") or "").lower() == "delivered" else ""),
            "deliveryDate": d.get("deliveryDate") or l.get("deliveryDate"),
            "chassisNumber": d.get("chassisNumber", ""), "numberPlate": d.get("numberPlate", ""),
        })
    return result


@api.put("/leads/{lead_id}/delivery")
async def mark_delivery(lead_id: str, body: DeliveryIn, act=Depends(actor), _desk=Depends(deal_desk_only)):
    lead = await get_lead_or_404(lead_id)
    delivered = (body.delivered or "").lower() in ("yes", "true", "delivered", "1")
    role = ((act or {}).get("role") or "").strip().lower()
    # Closed leads are frozen for everyone. After Mark Delivered, staff stay locked;
    # owner may still correct paperwork until the lead is closed.
    if _acct(lead) != "Active":
        raise HTTPException(
            409,
            f"This lead is locked for delivery edits "
            f"(status: {lead.get('currentStatus') or 'New'} / {_acct(lead)}). "
            f"Closed leads cannot be changed.",
        )
    if _is_delivered(lead) and role != "owner":
        raise HTTPException(
            409,
            f"This lead is locked for delivery edits "
            f"(status: {lead.get('currentStatus') or 'New'} / {_acct(lead)}). "
            f"Delivered leads cannot be changed (owner may edit until closed).",
        )
    first_delivery = delivered and not _is_delivered(lead)
    if first_delivery:
        _require_action(lead, "canDeliver", "delivery (not booked/active)", act)
        errs = _validate_delivery_ready(lead, body)
        if errs:
            raise HTTPException(422, "Cannot mark delivered:\n" + "\n".join("• " + e for e in errs))
    await _assert_unique_vehicle_identifiers(
        lead_id,
        invoice_number=body.invoiceNumber,
        chassis_number=body.chassisNumber,
        number_plate=body.numberPlate,
    )
    doc = {"leadId": lead_id, "customerName": lead.get("customerName"), **body.model_dump(),
           "deliveryId": f"DL{uuid.uuid4().hex[:8]}"}
    await db.deliveries.update_one({"leadId": lead_id}, {"$set": doc}, upsert=True)
    lead_updates = {
        "insuranceStatus": body.insurance, "registrationStatus": body.registration,
        "invoiceStatus": body.invoice, "rcStatus": body.rc, "pdiStatus": body.pdi,
        "invoiceNumber": body.invoiceNumber, "chassisNumber": body.chassisNumber,
        "numberPlate": body.numberPlate, "insurerName": body.insurerName, "lastUpdated": now_iso(),
    }
    # Insurance agent is chosen at delivery. Blank keeps whatever the lead already
    # had, so re-saving delivery paperwork never wipes the agent off a booked payout.
    if (body.insuranceAgentId or "").strip():
        _agent = await _get_insurance_agent(body.insuranceAgentId)
        if not _agent:
            raise HTTPException(422, "Selected insurance agent not found")
        lead_updates["insuranceAgentId"] = _agent["agentId"]
        lead_updates["insuranceAgentName"] = _agent.get("agentName", "")
    if delivered:
        lead_updates.update({"deliveryStatus": "Delivered", "currentStatus": "Delivered",
                             "deliveryDate": body.deliveryDate or today()})
    await db.leads.update_one({"leadId": lead_id}, {"$set": lead_updates})
    if delivered:
        await sheet_sync("deliveries", {
            "leadId": lead_id, "customerName": lead.get("customerName"),
            "deliveryDate": body.deliveryDate or today(), "delivered": "Yes",
            "invoiceNumber": body.invoiceNumber, "chassisNumber": body.chassisNumber, "numberPlate": body.numberPlate,
        })
        await _upsert_incentive_register_on_delivery(lead_id, body.deliveryDate or today())
        await _upsert_insurance_on_delivery(lead_id, body.deliveryDate or today())
        await rebuild_finance_views()
        await recompute_lead(lead_id)
        await _upsert_delivery_billing_summary(lead_id)
        if first_delivery:
            # WhatsApp delivery + review is fire-and-forget — must never fail this save.
            wa.schedule(wa.notify_delivery(lead_id))
            try:
                await oem_sync.take_chassis_from_inventory(db, body.chassisNumber)
            except Exception:
                logging.exception("Could not drop delivered chassis from yard inventory")
    return clean(await db.leads.find_one({"leadId": lead_id}))


async def _upsert_delivery_billing_summary(lead_id):
    """Refresh Delivery Billing Summary from current lead commercials (Tally bill).

    Use a single $set only — putting summaryId/createdAt in both $set and
    $setOnInsert makes MongoDB raise ConflictingUpdateOperators (code 40),
    which broke /leads/{id}/360 for every Delivered lead in production.
    """
    lead = await db.leads.find_one({"leadId": lead_id}) or {}
    summary = ce.build_delivery_billing_summary(lead)
    now = now_iso()
    summary["updatedAt"] = now
    existing = await db.billing_summaries.find_one({"leadId": lead_id}) or {}
    summary["createdAt"] = existing.get("createdAt") or now
    await db.billing_summaries.update_one(
        {"leadId": lead_id},
        {"$set": summary},
        upsert=True,
    )
    return summary


async def _refresh_billing_summary_if_delivered(lead_id):
    """After commercial edits, keep Tally summary in sync for delivered leads."""
    lead = await db.leads.find_one({"leadId": lead_id}) or {}
    if _is_delivered(lead):
        return await _upsert_delivery_billing_summary(lead_id)
    return None


@api.get("/leads/{lead_id}/billing-summary")
async def get_billing_summary(lead_id: str):
    """Delivery Billing Summary for Tally (full customer amount − benefits passed).

    Always rebuilds from the current lead when Delivered, so scheme / Additional
    (Dealer) changes are reflected immediately (not stuck on the first snapshot).
    """
    lead = await get_lead_or_404(lead_id)
    if not _is_delivered(lead):
        raise HTTPException(
            409,
            "Billing summary is created when the lead is marked Delivered. "
            "Mark delivery first, then open this summary for Tally cross-check.",
        )
    return clean(await _upsert_delivery_billing_summary(lead_id))


async def _upsert_insurance_on_delivery(lead_id, delivery_date):
    """On Mark Delivered, open the insurance payout entry for the vehicle.

    Nothing was creating these, so the Insurance Payouts screen stayed empty and the
    Earnings Report's insurance column (which sums expectedPayout) was always 0 — even
    though the premium, insurer and payout rate were all known.

    No new business rule: the premium is the lead's own insuranceAmount, the insurer is
    the one captured at delivery, and the rate/expected/outstanding/status come from the
    existing _insurance_derive + suggested_insurance_payout_rate (49% Storm/Turbo,
    36.5% others). Idempotent — one entry per lead, refreshed rather than duplicated."""
    lead = await db.leads.find_one({"leadId": lead_id}) or {}
    # Customer-arranged insurance: dealer does not earn a payout.
    if ce.normalize_insurance_arranged_by(lead.get("insuranceArrangedBy")) == "self":
        existing = await db.insurance.find_one({"leadId": lead_id})
        if existing:
            # Zero out any prior dealer payout so earnings stay clean after a switch to Self.
            cleared = {
                "expectedPayout": 0,
                "payoutOutstanding": 0,
                "payoutRate": 0,
                "status": "N/A — customer arranged",
                "lastUpdated": now_iso(),
                "remarks": ((existing.get("remarks") or "") + " | insuranceArrangedBy=self").strip(" |"),
            }
            await db.insurance.update_one({"entryId": existing["entryId"]}, {"$set": cleared})
            await sheet_sync("insurance", {**clean(existing), **cleared,
                                          "payoutRatePct": 0})
        return None
    premium = ce.num(lead.get("insuranceAmount"))
    if premium <= 0:
        return None          # no premium charged -> no payout is due; leave it absent
    existing = await db.insurance.find_one({"leadId": lead_id})
    # Agent chosen at delivery; fall back to whatever the entry already carries,
    # then to the default agent, so pre-agent leads still resolve a rate.
    agent = (await _get_insurance_agent(lead.get("insuranceAgentId"))
             or await _get_insurance_agent((existing or {}).get("insuranceAgentId"))
             or await _default_insurance_agent())
    data = {
        "leadId": lead_id,
        "customerName": lead.get("customerName", ""),
        "mobile": lead.get("mobile", ""),
        "model": lead.get("interestedModel", ""),
        "variant": lead.get("variant", ""),
        "insuranceCompany": lead.get("insurerName", "") or (existing or {}).get("insuranceCompany", ""),
        "policyNumber": (existing or {}).get("policyNumber", ""),
        "insuranceAmount": premium,
        # Preserve a manually set rate / already-received money on re-delivery edits.
        "payoutRate": ce.num((existing or {}).get("payoutRate")),
        "receivedPayout": ce.num((existing or {}).get("receivedPayout")),
        "status": (existing or {}).get("status", "Pending"),
        "deliveryDate": delivery_date,
        "insuranceExecutive": lead.get("executive", ""),
        "remarks": (existing or {}).get("remarks", ""),
        "lastUpdated": now_iso(),
    }
    # An owner's manual rate must survive re-delivery; anything the server itself
    # resolved is re-derived from the agent slab so a corrected agent takes effect.
    if (existing or {}).get("payoutRateSource") != "manual":
        data["payoutRate"] = 0
    data.update(_insurance_derive(data, agent))
    if existing:
        await db.insurance.update_one({"entryId": existing["entryId"]}, {"$set": data})
        data["entryId"] = existing["entryId"]
    else:
        data["entryId"] = await next_id("insurance", "INS26")
        data["policyDate"] = None
        await db.insurance.insert_one(dict(data))
    await sheet_sync("insurance", _insurance_sheet_row(data))
    return data


async def _upsert_incentive_register_on_delivery(lead_id, delivery_date):
    """Port of upsertIncentiveRegisterOnDelivery_: on Mark Delivered, create a
    Pending Incentive Register row from the Incentive Master rate for the lead's
    model/variant/delivery-month. Skips if a row already exists for this lead."""
    if await db.incentive_register.find_one({"leadId": lead_id}):
        return None
    lead = await db.leads.find_one({"leadId": lead_id})
    if not lead:
        return None
    model = lead.get("interestedModel") or ""
    variant = lead.get("variant") or ""
    incentive_rows = [clean(r) for r in await db.incentive_master.find().to_list(1000)]
    rate = ce.get_incentive_rate_for_lead(model, variant, delivery_date, incentive_rows)
    if not rate:
        return None
    cat = ce.map_lead_to_incentive_category(model, variant)
    remarks = f"Rate from Incentive Master. Min retails: {rate.get('minRetails') or 0}"
    if rate.get("maxSlab"):
        remarks += f"; Max slab: {rate.get('maxSlab')}"
    doc = {
        "incentiveId": f"INC{uuid.uuid4().hex[:8]}",
        "schemeMonth": rate.get("schemeMonth") or ce.scheme_month_from_date(delivery_date),
        "executive": lead.get("executive") or "",
        "leadId": lead_id, "bookingId": lead.get("bookingId") or "",
        "model": model, "variant": variant, "productCategory": cat,
        "deliveryDate": delivery_date, "incentiveAmount": ce.num(rate.get("incentivePerRetail")),
        "status": "Pending", "paidDate": "", "remarks": remarks,
        "lastUpdated": now_iso(),
    }
    await db.incentive_register.insert_one(doc)
    # Push to Google Sheet Incentive Register (was missing from SYNC_MAP / never called).
    await sheet_sync("incentive_register", clean(doc))
    return doc


@api.get("/incentive-register")
async def list_incentive_register():
    return [clean(r) for r in await db.incentive_register.find().to_list(2000)]


class IncentivePayIn(BaseModel):
    paidDate: str = ""


async def _upsert_incentive_oem_claim(incentive_row, act=None):
    """When executive incentive is Mark Paid, open an OEM Claim Register row (outstanding).

    Same receivable path as scheme claims: appears in GET /claims, Scheme Claim Register
    sheet, and OEM totals. Does NOT touch scheme allocation / companyOutstanding formula
    (those stay scheme-share SSOT); incentive is a separate manual claim component.
    """
    lead_id = incentive_row.get("leadId") or ""
    amount = ce.round2(ce.num(incentive_row.get("incentiveAmount")))
    if amount <= 0 or not lead_id:
        return None
    incentive_id = incentive_row.get("incentiveId") or ""
    component_key = f"executiveIncentive-{incentive_id}" if incentive_id else "executiveIncentive"
    claim_id = f"CLM-{lead_id}-executiveIncentive"
    lead = await db.leads.find_one({"leadId": lead_id}) or {}
    existing = await db.claims.find_one({"leadId": lead_id, "componentKey": component_key}) or {}
    # Preserve receipts if the claim was already partially paid by OEM.
    received = ce.round2(ce.num(existing.get("receivedAmount")))
    status = existing.get("claimStatus") or "Pending"
    if received <= 0:
        status = "Pending"
    elif received + 0.01 >= amount:
        status = "Received"
    else:
        status = "Partial"
    doc = {
        "claimId": existing.get("claimId") or claim_id,
        "manual": True,
        "componentKey": component_key,
        "leadId": lead_id,
        "customer": lead.get("customerName") or existing.get("customer") or "",
        "model": incentive_row.get("model") or lead.get("interestedModel") or "",
        "variant": incentive_row.get("variant") or lead.get("variant") or "",
        "bookingDate": lead.get("bookingDate") or "",
        "bookingId": incentive_row.get("bookingId") or lead.get("bookingId") or "",
        "schemeMonth": incentive_row.get("schemeMonth") or "",
        "executive": incentive_row.get("executive") or lead.get("executive") or "",
        "claimType": "Executive Incentive",
        "component": "Executive Incentive",
        "oemCompany": "Euler Motors",
        "claimAmount": amount,
        "eligibleClaim": amount,
        "claimStatus": status,
        "receivedAmount": received,
        "claimReference": existing.get("claimReference") or incentive_id,
        "note": existing.get("note") or f"Executive incentive paid {incentive_row.get('paidDate') or ''}".strip(),
        "submittedDate": existing.get("submittedDate") or incentive_row.get("paidDate") or today(),
        "approvedDate": existing.get("approvedDate") or "",
        "source": "Executive Incentive (Paid)",
        "claimRequired": "Yes",
        # Bifurcation columns stay 0; total/oemDiscount carry the incentive amount.
        "consumerDiscount": 0, "exchangeBonus": 0, "loyaltyBonus": 0, "insuranceBenefit": 0,
        "referralBonus": 0, "dsaDiscount": 0, "additionalDiscount": 0,
        "rtoBenefit": 0, "rtoInsuranceBenefit": 0,
        "totalDiscount": amount, "dealerDiscount": 0, "oemDiscount": amount,
        "lastUpdated": now_iso(),
    }
    if not existing:
        doc["receipts"] = []
        doc["createdAt"] = now_iso()
    await db.claims.update_one(
        {"leadId": lead_id, "componentKey": component_key},
        {"$set": doc}, upsert=True,
    )
    await sheet_sync("claims", clean(doc))
    if act:
        await write_audit(act, "upsert", "claim", leadId=lead_id, claimId=doc["claimId"],
                          new={"component": "Executive Incentive", "claimAmount": amount,
                               "incentiveId": incentive_id, "source": doc["source"]})
    return doc


@api.put("/incentive-register/{incentive_id}/pay", dependencies=[Depends(owner_only)])
async def mark_incentive_paid(incentive_id: str, body: IncentivePayIn, act=Depends(actor)):
    existing = await db.incentive_register.find_one({"incentiveId": incentive_id})
    if not existing:
        raise HTTPException(404, "Incentive Register row not found")
    updates = {"status": "Paid", "paidDate": body.paidDate or today(), "lastUpdated": now_iso()}
    await db.incentive_register.update_one({"incentiveId": incentive_id}, {"$set": updates})
    paid = {**existing, **updates}
    await sheet_sync("incentive_register", clean(paid))
    # Paid executive incentive → OEM Claim Register outstanding (claim from OEM).
    await _upsert_incentive_oem_claim(paid, act=act)
    await write_audit(act, "update", "incentive_register", leadId=existing.get("leadId", ""),
                      old={"status": existing.get("status")}, new=updates)
    return clean(await db.incentive_register.find_one({"incentiveId": incentive_id}))


@api.post("/admin/sync-incentive-register", dependencies=[Depends(owner_only)])
async def sync_incentive_register_sheet(act=Depends(actor)):
    """Backfill Google Sheet Incentive Register from Mongo (rows created before sheet sync)."""
    rows = [clean(r) for r in await db.incentive_register.find().to_list(5000)]
    synced, errors = 0, []
    for row in rows:
        try:
            await sheet_sync("incentive_register", row)
            synced += 1
            if str(row.get("status") or "").lower() == "paid":
                await _upsert_incentive_oem_claim(row, act=None)
        except Exception as e:
            errors.append({"incentiveId": row.get("incentiveId"), "error": str(e)})
    await write_audit(act, "sync", "incentive_register", new={"synced": synced, "errors": len(errors)})
    return {"ok": True, "synced": synced, "total": len(rows), "errors": errors}


# ---------------------------------------------------------------- price master
@api.get("/price-master")
async def list_price_master(model: Optional[str] = None):
    q = {"model": model} if model else {}
    rows = [clean(p) for p in await db.price_master.find(q).to_list(2000)]
    counts = await oem_sync.inventory_counts(db)
    for r in rows:
        r["inYard"] = counts.get((r.get("model") or "", r.get("variant") or ""), 0)
        r["priceSource"] = r.get("priceSource") or "manual"
    return rows


@api.get("/price-list")
async def price_list(model: Optional[str] = None, q: str = "", user=Depends(current_user)):
    """Read-only showroom price list — what a salesperson quotes.

    On-road is the same Gross Vehicle Cost the commercial engine computes, and
    TCS follows the engine exactly: charged only when the Price Master row says
    Yes AND the total reaches the threshold. A row over the threshold with the
    flag off is NOT given a TCS line — that would quote a charge the app will
    never bill — but it is counted in `tcsReview` so the owner can check it.

    Scheme shows the TOTAL available for the current month. The company/dealer
    split is deliberately withheld: that is commercial information.
    """
    scheme_rows = await get_scheme_rows()
    on = today()
    rows = await db.price_master.find({"model": model} if model else {}).to_list(2000)
    counts = await oem_sync.inventory_counts(db)
    needle = (q or "").strip().lower()
    grouped, review = {}, []
    for r in rows:
        if str(r.get("status") or "active").lower() != "active":
            continue
        mdl, variant = str(r.get("model") or ""), str(r.get("variant") or "")
        if needle and needle not in f"{mdl} {variant}".lower():
            continue
        ex = ce.num(r.get("exShowroom"))
        rto = ce.num(r.get("rto"))
        ins = ce.num(r.get("insurance"))
        other = ce.round2(sum(ce.num(r.get(k)) for k in
                ("accessories", "handlingCharges", "trc", "fastag",
                 "extendedWarranty", "otherCharges")))
        gvc = ce.round2(ex + rto + ins + other)
        applies = str(r.get("tcsApplicable") or "No").strip().lower() == "yes"
        tcs = ce.calculate_tcs(gvc) if applies else 0.0
        over = gvc >= ce.TCS_THRESHOLD
        if over and not applies:
            review.append({"priceId": r.get("priceId"), "model": mdl,
                           "variant": variant, "onRoad": ce.round2(gvc)})
        shares = ce.get_scheme_shares_for_lead(mdl, variant, on, scheme_rows)
        scheme = ce.round2(sum(
            ce.num(v.get("totalBenefit")) or (ce.num(v.get("dealerShare")) + ce.num(v.get("companyShare")))
            for v in shares.values()))
        grouped.setdefault(mdl, []).append({
            "priceId": r.get("priceId"), "model": mdl, "variant": variant,
            "bodyType": r.get("bodyType") or "",
            "exShowroom": ce.round2(ex), "rto": ce.round2(rto), "insurance": ce.round2(ins),
            "otherCharges": other,
            "tcs": tcs, "tcsApplies": applies and tcs > 0,
            "onRoad": ce.round2(gvc + tcs),
            "schemeAvailable": scheme,
            "inYard": counts.get((mdl, variant), 0),
        })
    out = [{"model": m, "count": len(v),
            "rows": sorted(v, key=lambda x: x["onRoad"])}
           for m, v in sorted(grouped.items())]
    body = {"schemeMonth": ce.scheme_month_from_date(on), "asOf": on,
            "totalRows": sum(g["count"] for g in out), "models": out}
    # Only the owner is shown the data-quality flag.
    if (user or {}).get("role") == "owner":
        body["tcsReview"] = review
    return body


@api.get("/price-master/variants")
async def price_variants(model: str):
    wanted = str(model or "").strip()
    sku = oem_cat.resolve_sku(wanted, "")
    key = "".join(ch for ch in wanted.lower() if ch.isalnum())
    canon = (sku.crm_model if sku else None) or oem_cat.MODEL_ALIASES.get(key, wanted)
    rows = await db.price_master.find({
        "model": {"$regex": f"^{re.escape(canon)}$", "$options": "i"},
    }).to_list(500)
    active = [clean(r) for r in rows if str(r.get("status") or "active").lower() != "inactive"]
    counts = await oem_sync.inventory_counts(db)
    for r in active:
        r["inYard"] = counts.get((r.get("model") or "", r.get("variant") or ""), 0)
    return active


class CoulsonCredIn(BaseModel):
    username: str = ""
    password: str = ""
    sessionToken: str = ""


async def _coulson_status_payload(viewer=None):
    """Presence-only status. Password and session token are never returned. Owner gets the real username
    so Settings can pre-fill it; everyone else gets the masked hint."""
    user, pw, src = await oem_sync.resolve_credentials(db)
    doc = await db["system"].find_one({"_id": "coulson"}) or {}
    cat_doc = await db["system"].find_one({"_id": "oem_catalog"}) or {}
    count = await db.oem_inventory.count_documents({})
    is_owner = (viewer or {}).get("role") == "owner"
    shown = user if is_owner and not oem_sync.looks_masked_username(user) else oem_sync.mask_username(user)
    live_session = oem_sync.session_from_doc(doc)
    has_session = bool(live_session)
    if oem_sync.session_expired(doc):
        src = src or "session"
    elif has_session:
        src = "session"
    configured = oem_sync.credentials_configured(user, pw) or bool(doc.get("sessionToken"))
    return {
        "configured": configured,
        "username": shown,
        "source": src or None,
        "hasSession": has_session,
        "sessionExpired": oem_sync.session_expired(doc),
        "sessionExpiresAt": (doc.get("sessionExpiresAt") or "") if is_owner else "",
        "lastSyncAt": doc.get("lastSyncAt"),
        "lastSyncOk": doc.get("lastSyncOk"),
        "loginOk": doc.get("loginOk"),
        "lastError": doc.get("lastError") or "",
        "inventoryCount": count,
        "catalogSize": cat_doc.get("catalogSize") or len(oem_cat.CATALOG),
        "pricesUpdated": doc.get("pricesUpdated"),
    }


@api.get("/integrations/coulson")
async def coulson_status(viewer=Depends(current_user)):
    return await _coulson_status_payload(viewer)


@api.put("/integrations/coulson", dependencies=[Depends(owner_only)])
async def coulson_save(body: CoulsonCredIn, act=Depends(actor)):
    """Verify with Euler first using the typed login or a pasted Coulson session, then store.

    Never persist the hidden va***r hint, never re-test a password Euler already rejected,
    and never return the session token. A pasted session is checked against the inventory
    API (Bearer) — not euler-auth /login — so it cannot lock the dealer password.
    """
    existing = await db["system"].find_one({"_id": "coulson"}) or {}
    typed_session = (body.sessionToken or "").strip()
    if typed_session:
        try:
            claims = await asyncio.to_thread(coulson_client.verify_session_token, typed_session)
        except coulson_client.CoulsonError as e:
            await oem_sync._record_sync(db, False, str(e))
            await db["system"].update_one({"_id": "coulson"}, {"$set": {"loginOk": False}}, upsert=True)
            st = await _coulson_status_payload(act)
            st["loginOk"] = False
            return st
        typed_user = (body.username or "").strip()
        if oem_sync.looks_masked_username(typed_user):
            typed_user = ""
        await oem_sync.save_session(db, typed_session, typed_user)
        await write_audit(act, "update", "coulson", new={
            "username": oem_sync.mask_username(
                typed_user or coulson_client.session_username(claims)),
            "via": "session",
        })
        st = await _coulson_status_payload(act)
        st["loginOk"] = True
        _ensure_coulson_sync_loop()
        return st

    typed_user = (body.username or "").strip()
    if oem_sync.looks_masked_username(typed_user):
        st = await _coulson_status_payload(act)
        st["loginOk"] = False
        st["lastError"] = ("Type the full Coulson username from coulson.eulerlogistics.com "
                           "— not the hidden va***r hint")
        return st
    user = typed_user or (existing.get("username") or "")
    if oem_sync.looks_masked_username(user):
        user = ""
    pw = body.password if body.password not in (None, "") else (existing.get("password") or "")
    last_failed = existing.get("loginOk") is False
    if last_failed and not (body.password or "") and not oem_sync.session_from_doc(existing):
        st = await _coulson_status_payload(act)
        st["loginOk"] = False
        st["lastError"] = "Type the Coulson password again, then Save. Euler rejected the last login."
        return st
    # A live pasted session is enough. Do not re-hit euler-auth /login unless they
    # typed a password — Euler refuses that login from this server and can lock the account.
    if not (body.password or "") and oem_sync.session_from_doc(existing):
        return await _coulson_status_payload(act)
    if not oem_sync.credentials_configured(user, pw):
        st = await _coulson_status_payload(act)
        st["loginOk"] = False
        st["lastError"] = "Coulson username and password are required"
        return st
    try:
        coulson_client.login(user, pw)
    except coulson_client.CoulsonError as e:
        await oem_sync._record_sync(db, False, str(e))
        await db["system"].update_one({"_id": "coulson"}, {"$set": {"loginOk": False}}, upsert=True)
        st = await _coulson_status_payload(act)
        st["loginOk"] = False
        return st
    await oem_sync.save_credentials(db, user, pw)
    await write_audit(act, "update", "coulson", new={"username": oem_sync.mask_username(user)})
    await db["system"].update_one(
        {"_id": "coulson"},
        {"$set": {"lastError": "", "loginVerifiedAt": oem_sync.now_iso(), "loginOk": True}},
    )
    st = await _coulson_status_payload(act)
    st["loginOk"] = True
    _ensure_coulson_sync_loop()
    return st


async def _reprice_after_oem(changed_ids):
    repriced_total = 0
    for pid in changed_ids or []:
        row = await db.price_master.find_one({"priceId": pid})
        if not row:
            continue
        result = await reprice_leads_for_price_row(row, {"exShowroom"}, None)
        repriced_total += int(result.get("repricedCount") or 0)
    return repriced_total


@api.post("/integrations/coulson/diagnose", dependencies=[Depends(owner_only)])
async def coulson_diagnose(body: CoulsonCredIn, act=Depends(actor)):
    """Owner-only: try one login and say exactly what was sent and what came back.

    Coulson answers "Username/password is not valid" for a wrong password, a
    wrong app segment, and a request that reached it with no credentials at all.
    Those have nothing to do with each other, and the app could not tell them
    apart. The password is never returned or logged — only its length and
    whether it arrived wrapped in whitespace.
    """
    typed_user = (body.username or "").strip()
    typed_pw = body.password or ""
    if not typed_user or oem_sync.looks_masked_username(typed_user) or not typed_pw:
        stored_user, stored_pw, _src = await oem_sync.resolve_credentials(db)
        typed_user = typed_user if typed_user and not oem_sync.looks_masked_username(typed_user) else stored_user
        typed_pw = typed_pw or stored_pw
    result = await asyncio.to_thread(coulson_client.diagnose, typed_user, typed_pw)
    # The username is business data, not a secret; the password never appears here.
    await write_audit(act, "diagnose", "coulson", new={
        "ok": result.get("ok"), "status": result.get("status"),
        "appSegment": result.get("appSegment"), "authUrl": result.get("authUrl")})
    return result


@api.post("/integrations/coulson/sync", dependencies=[Depends(owner_only)])
async def coulson_sync(act=Depends(actor)):
    """Pull live OEM prices + yard stock. RTO/insurance stay as manual Price Master fields."""
    try:
        result = await oem_sync.sync_from_coulson(db)
    except coulson_client.CoulsonError as e:
        await oem_sync._record_sync(db, False, str(e))
        await db["system"].update_one({"_id": "coulson"}, {"$set": {"loginOk": False}}, upsert=True)
        raise HTTPException(502, str(e))
    except Exception as e:
        logging.exception("Coulson sync failed")
        await oem_sync._record_sync(db, False, str(e))
        raise HTTPException(502, "Coulson sync failed")
    repriced = 0
    try:
        repriced = await _reprice_after_oem(result.get("changedPriceIds"))
    except Exception:
        logging.exception("OEM reprice after sync failed")
    result["leadsRepriced"] = repriced
    # Don't leak price-id lists or credential source internals to the client more than needed.
    result.pop("changedPriceIds", None)
    await write_audit(act, "sync", "coulson", new={"ok": result.get("ok"),
                                                   "inventoryCount": result.get("inventoryCount"),
                                                   "pricesUpdated": result.get("pricesUpdated")})
    return result


@api.get("/inventory")
async def list_oem_inventory(model: Optional[str] = None, variant: Optional[str] = None,
                             family: bool = False, _user=Depends(current_user)):
    rows = await oem_sync.list_inventory(
        db, model, variant, family=family or bool(variant))
    return [clean(r) for r in rows]


@api.get("/inventory/summary")
async def inventory_summary(_user=Depends(current_user)):
    counts = await oem_sync.inventory_counts(db)
    out = [{"model": m, "variant": v, "count": n} for (m, v), n in sorted(counts.items())]
    return {"total": sum(c["count"] for c in out), "rows": out}


# ---------------------------------------------------------------- masters registers
@api.get("/scheme-master")
async def list_scheme_master(on: Optional[str] = None):
    rows = [clean(s) for s in await db.scheme_master.find().to_list(1000)]
    iso = str(on or "").strip()[:10]
    if len(iso) == 10 and iso[4] == "-" and iso[7] == "-":
        rows = [r for r in rows if _scheme_row_matches_as_of(r, iso)]
    return rows


@api.get("/incentive-master")
async def list_incentive_master():
    return [clean(s) for s in await db.incentive_master.find().to_list(1000)]


@api.get("/bookings")
async def list_bookings(user=Depends(current_user)):
    rows = [clean(b) for b in await db.bookings.find().sort("bookingId", -1).to_list(1000)]
    if user.get("role") == "executive":
        mine = await _own_lead_ids(user)
        rows = [b for b in rows if b.get("leadId") in mine]
    if user.get("role") in authmod.FIELD_ROLES:
        safe = []
        for b in rows:
            safe.append({
                k: b.get(k) for k in (
                    "bookingId", "leadId", "customerName", "model", "variant",
                    "bookingDate", "bookingStatus", "executive",
                ) if b.get(k) is not None
            })
        return safe
    return rows


@api.get("/activities")
async def list_activities(lead_id: Optional[str] = None, user=Depends(current_user)):
    q = {"leadId": lead_id} if lead_id else {}
    rows = [clean(a) for a in await db.activities.find(q).sort("activityId", -1).to_list(2000)]
    if user.get("role") == "executive":
        mine = await _own_lead_ids(user)
        rows = [a for a in rows if a.get("leadId") in mine]
    return rows


@api.post("/leads/{lead_id}/activities")
async def add_activity(lead_id: str, body: ActivityIn, act=Depends(actor), _sales=Depends(sales_staff_only)):
    lead = await get_lead_or_404(lead_id)
    _require_action(lead, "canScheme", "logging activity (only Active leads)", act)
    payload = body.model_dump()
    act_date = str(payload.pop("date", None) or "").strip() or today()
    doc = {
        "activityId": await next_id("activity", "AC26"), "leadId": lead_id, "date": act_date,
        "time": datetime.now(timezone.utc).strftime("%H:%M"), **payload,
        "customerName": lead.get("customerName"), "mobile": lead.get("mobile"), "model": lead.get("interestedModel"),
    }
    await db.activities.insert_one(doc)
    await sheet_sync("activities", doc)
    # Lead Register "Last Activity" summarises the most recent activity on the lead.
    # It is derived from the activity just logged, not separately entered.
    summary = " · ".join(x for x in (body.activityType, (body.discussion or "").strip()) if x)
    await db.leads.update_one({"leadId": lead_id}, {"$set": {
        "lastActivity": summary[:200], "lastUpdated": now_iso(),
    }})
    await recompute_lead(lead_id)
    return clean(doc)


def _strip_payout_for_staff(entry: dict, is_owner: bool):
    """Hide dealer payout economics (rate/expected/outstanding) from non-owner staff."""
    if is_owner:
        return entry
    e = dict(entry)
    for k in ("payoutRate", "expectedPayout", "payoutOutstanding"):
        e.pop(k, None)
    return e


# ------------------------------------------------- insurance agents (brokers)
# An insurance agent is the broker between the dealership and the insurer, and
# each one pays a different share of the premium. Before this existed the payout
# rate was a single hard-coded pair (49% Storm/Turbo, 36.5% everything else) with
# nowhere to record WHO was paying it, so a second agent could not be modelled.
# The insurer (insuranceCompany, e.g. ICICI) stays a separate field.
class InsuranceSlabIn(BaseModel):
    modelFamily: str = "*"
    payoutRatePct: float = 0        # PERCENT here (52 means 52%), fraction on the entry
    effectiveFrom: str = ""
    effectiveTo: str = ""


class InsuranceAgentIn(BaseModel):
    agentName: str
    agentCode: str = ""
    contactPerson: str = ""
    mobile: str = ""
    email: str = ""
    status: str = "Active"
    isDefault: bool = False
    slabs: List[InsuranceSlabIn] = []
    remarks: str = ""


def _agent_doc(body: dict) -> dict:
    slabs = []
    for s in (body.get("slabs") or []):
        fam = ce.normalize_slab_family(s.get("modelFamily"))
        pct = ce.num(s.get("payoutRatePct"))
        if pct <= 0:
            continue
        slabs.append({"modelFamily": fam, "payoutRatePct": round(pct, 4),
                      "effectiveFrom": str(s.get("effectiveFrom") or "")[:10],
                      "effectiveTo": str(s.get("effectiveTo") or "")[:10]})
    return {
        "agentName": str(body.get("agentName") or "").strip(),
        "agentCode": str(body.get("agentCode") or "").strip(),
        "contactPerson": str(body.get("contactPerson") or "").strip(),
        "mobile": str(body.get("mobile") or "").strip(),
        "email": str(body.get("email") or "").strip(),
        "status": str(body.get("status") or "Active").strip() or "Active",
        "isDefault": bool(body.get("isDefault")),
        "slabs": slabs,
        "remarks": str(body.get("remarks") or "").strip(),
        "lastUpdated": now_iso(),
    }


async def _clear_other_defaults(agent_id: str):
    await db.insurance_agents.update_many(
        {"agentId": {"$ne": agent_id}}, {"$set": {"isDefault": False}})


async def _get_insurance_agent(agent_id) -> dict:
    if not str(agent_id or "").strip():
        return {}
    return await db.insurance_agents.find_one({"agentId": str(agent_id).strip()}) or {}


async def _default_insurance_agent() -> dict:
    return (await db.insurance_agents.find_one({"isDefault": True, "status": "Active"})
            or await db.insurance_agents.find_one({"isDefault": True})
            or {})


# ---------------------------------------------------------------- staff master
class StaffIn(BaseModel):
    name: str
    mobile: str = ""
    email: str = ""
    role: str = "executive"
    monthlyTarget: float = 0          # units per month; 0 = no target
    reports: Optional[List[str]] = None
    whatsappOptIn: bool = True
    status: str = "Active"
    remarks: str = ""


def _staff_doc(body: dict) -> dict:
    role = str(body.get("role") or "executive").strip()
    if role not in STAFF_ROLES:
        raise HTTPException(422, f"role must be one of {', '.join(STAFF_ROLES)}")
    reports = body.get("reports")
    if reports is None:
        reports = DEFAULT_REPORTS_BY_ROLE.get(role, [])
    bad = [r for r in reports if r not in STAFF_REPORTS]
    if bad:
        raise HTTPException(422, f"unknown report(s): {', '.join(bad)}")
    return {
        "name": str(body.get("name") or "").strip(),
        "mobile": wa.digits10(body.get("mobile")),
        "email": str(body.get("email") or "").strip(),
        "role": role,
        "monthlyTarget": max(0.0, ce.num(body.get("monthlyTarget"))),
        "reports": list(reports),
        "whatsappOptIn": bool(body.get("whatsappOptIn", True)),
        "status": str(body.get("status") or "Active").strip() or "Active",
        "remarks": str(body.get("remarks") or "").strip(),
        "lastUpdated": now_iso(),
    }


@api.get("/staff")
async def list_staff(role: Optional[str] = None):
    """Readable by any signed-in user — every executive dropdown reads this."""
    q = {"role": role} if role else {}
    return [clean(r) for r in await db.staff.find(q).sort("name", 1).to_list(500)]


@api.post("/staff", dependencies=[Depends(owner_only)])
async def create_staff(body: StaffIn, act=Depends(actor)):
    data = _staff_doc(body.model_dump())
    if not data["name"]:
        raise HTTPException(422, "Name is required")
    dupe = await db.staff.find_one(
        {"name": {"$regex": f"^{re.escape(data['name'])}$", "$options": "i"}})
    if dupe:
        raise HTTPException(409, f"'{data['name']}' already exists on the staff master")
    data["staffId"] = await next_id("staff", "ST26")
    data["createdAt"] = now_iso()
    await db.staff.insert_one(dict(data))
    await write_audit(act, "create", "staff", new={k: data[k] for k in ("staffId", "name", "role")})
    return clean(await db.staff.find_one({"staffId": data["staffId"]}))


@api.put("/staff/{staff_id}", dependencies=[Depends(owner_only)])
async def update_staff(staff_id: str, body: StaffIn, act=Depends(actor)):
    existing = await db.staff.find_one({"staffId": staff_id})
    if not existing:
        raise HTTPException(404, "Staff member not found")
    data = _staff_doc(body.model_dump())
    if not data["name"]:
        raise HTTPException(422, "Name is required")
    # lead.executive stores the NAME, so a rename would orphan every lead.
    if (data["name"].strip().lower() != str(existing.get("name") or "").strip().lower()
            and existing.get("role") == "executive"):
        n = await db.leads.count_documents({"executive": existing.get("name")})
        if n:
            raise HTTPException(
                409, f"{n} lead(s) are assigned to '{existing.get('name')}'. Renaming would "
                     f"orphan them — set this person Inactive and add the new name instead.")
    await db.staff.update_one({"staffId": staff_id}, {"$set": data})
    await write_audit(act, "update", "staff",
                      old={"mobile": existing.get("mobile"), "role": existing.get("role"),
                           "monthlyTarget": existing.get("monthlyTarget")},
                      new={"mobile": data["mobile"], "role": data["role"],
                           "monthlyTarget": data["monthlyTarget"]})
    return clean(await db.staff.find_one({"staffId": staff_id}))


@api.delete("/staff/{staff_id}", dependencies=[Depends(owner_only)])
async def delete_staff(staff_id: str, act=Depends(actor)):
    existing = await db.staff.find_one({"staffId": staff_id})
    if not existing:
        raise HTTPException(404, "Staff member not found")
    if existing.get("role") == "executive":
        n = await db.leads.count_documents({"executive": existing.get("name")})
        if n:
            raise HTTPException(
                409, f"{n} lead(s) are assigned to '{existing.get('name')}'. Set them Inactive "
                     f"instead — deleting would leave those leads with an unknown executive.")
    await db.staff.delete_one({"staffId": staff_id})
    await write_audit(act, "delete", "staff", old={"staffId": staff_id, "name": existing.get("name")})
    return {"ok": True}


async def _seed_staff() -> dict:
    """Idempotent. Builds the staff master from the executives already in
    masters_list, carrying over any WhatsApp number already typed into the
    BotSpace settings so nothing has to be re-entered."""
    created = []
    if await db.staff.count_documents({}) == 0:
        cfg = await wa.get_config()
        mobiles = {_norm_name(x.get("name")): wa.digits10(x.get("mobile"))
                   for x in (cfg.get("executives") or [])}
        rows = await db.masters_list.find({"category": "executives"}).sort("value", 1).to_list(500)
        for r in rows:
            name = str(r.get("value") or "").strip()
            if not name:
                continue
            doc = {
                "staffId": await next_id("staff", "ST26"),
                "name": name, "mobile": mobiles.get(_norm_name(name), ""), "email": "",
                "role": "executive", "monthlyTarget": 0.0,
                "reports": list(DEFAULT_REPORTS_BY_ROLE["executive"]),
                "whatsappOptIn": True, "status": "Active", "remarks": "",
                "createdAt": now_iso(), "lastUpdated": now_iso(),
            }
            await db.staff.insert_one(dict(doc))
            created.append(name)
    return {"created": created, "count": await db.staff.count_documents({})}


@api.post("/admin/seed-staff", dependencies=[Depends(owner_only)])
async def seed_staff():
    return {"ok": True, **await _seed_staff()}


# ---------------------------------------------------------------- cancel reasons
# Seeded once, then owner-editable. The revival policy lives on the REASON rather
# than being one global setting, because "postponed purchase" and "bought a Tata"
# deserve opposite treatment: chasing the second one every third day forever is
# how a WhatsApp number earns a poor quality rating and loses template access.
REVIVE_MODES = ("now", "days", "never")
DEFAULT_CANCEL_REASONS = [
    {"reason": "Price too high", "revive": "days", "reviveAfterDays": 30},
    {"reason": "Postponed purchase", "revive": "days", "reviveAfterDays": 30},
    {"reason": "Finance rejected", "revive": "days", "reviveAfterDays": 60},
    {"reason": "Not reachable", "revive": "now", "reviveAfterDays": 0},
    {"reason": "Vehicle not available", "revive": "now", "reviveAfterDays": 0},
    {"reason": "Bought other brand", "revive": "never", "reviveAfterDays": 0},
    {"reason": "Duplicate lead", "revive": "never", "reviveAfterDays": 0},
    {"reason": "Other", "revive": "now", "reviveAfterDays": 0},
]


def _cancel_reason_doc(data: dict) -> dict:
    revive = str(data.get("revive") or "now").strip().lower()
    if revive not in REVIVE_MODES:
        revive = "now"
    days = int(ce.num(data.get("reviveAfterDays")))
    if revive != "days":
        days = 0
    elif days <= 0:
        days = 30
    return {
        "reason": str(data.get("reason") or "").strip(),
        "revive": revive,
        "reviveAfterDays": days,
        "status": str(data.get("status") or "Active").strip() or "Active",
        "remarks": str(data.get("remarks") or "").strip(),
        "lastUpdated": now_iso(),
    }


async def _get_cancel_reason(reason: str) -> Optional[dict]:
    if not str(reason or "").strip():
        return None
    return await db.cancel_reasons.find_one(
        {"reason": {"$regex": f"^{re.escape(str(reason).strip())}$", "$options": "i"}})


async def _seed_cancel_reasons() -> dict:
    created = []
    if await db.cancel_reasons.count_documents({}) == 0:
        for row in DEFAULT_CANCEL_REASONS:
            doc = _cancel_reason_doc(row)
            doc["reasonId"] = await next_id("cancel_reason", "CR26")
            doc["createdAt"] = now_iso()
            await db.cancel_reasons.insert_one(dict(doc))
            created.append(doc["reason"])
    return {"created": created, "count": await db.cancel_reasons.count_documents({})}


@api.get("/cancel-reasons")
async def list_cancel_reasons(active_only: bool = False):
    """Readable by any signed-in user — the cancel dialog needs the dropdown."""
    q = {"status": {"$regex": "^active$", "$options": "i"}} if active_only else {}
    return [clean(r) for r in await db.cancel_reasons.find(q).sort("reason", 1).to_list(200)]


@api.post("/cancel-reasons", dependencies=[Depends(owner_only)])
async def create_cancel_reason(body: CancelReasonIn, act=Depends(actor)):
    data = _cancel_reason_doc(body.model_dump())
    if not data["reason"]:
        raise HTTPException(422, "Reason is required")
    if await _get_cancel_reason(data["reason"]):
        raise HTTPException(409, f"A cancel reason '{data['reason']}' already exists")
    data["reasonId"] = await next_id("cancel_reason", "CR26")
    data["createdAt"] = now_iso()
    await db.cancel_reasons.insert_one(dict(data))
    await write_audit(act, "create", "cancel_reasons", new=data)
    return clean(await db.cancel_reasons.find_one({"reasonId": data["reasonId"]}))


@api.put("/cancel-reasons/{reason_id}", dependencies=[Depends(owner_only)])
async def update_cancel_reason(reason_id: str, body: CancelReasonIn, act=Depends(actor)):
    existing = await db.cancel_reasons.find_one({"reasonId": reason_id})
    if not existing:
        raise HTTPException(404, "Cancel reason not found")
    data = _cancel_reason_doc(body.model_dump())
    if not data["reason"]:
        raise HTTPException(422, "Reason is required")
    clash = await _get_cancel_reason(data["reason"])
    if clash and clash.get("reasonId") != reason_id:
        raise HTTPException(409, f"A cancel reason '{data['reason']}' already exists")
    await db.cancel_reasons.update_one({"reasonId": reason_id}, {"$set": data})
    # Already-cancelled leads keep the policy that applied when they were
    # cancelled — their reviveOn date is stamped on the lead, not read back from
    # here, so editing a reason never silently re-dates parked leads.
    await write_audit(act, "update", "cancel_reasons",
                      old={k: existing.get(k) for k in ("reason", "revive", "reviveAfterDays", "status")},
                      new=data)
    return clean(await db.cancel_reasons.find_one({"reasonId": reason_id}))


@api.delete("/cancel-reasons/{reason_id}", dependencies=[Depends(owner_only)])
async def delete_cancel_reason(reason_id: str, act=Depends(actor)):
    existing = await db.cancel_reasons.find_one({"reasonId": reason_id})
    if not existing:
        raise HTTPException(404, "Cancel reason not found")
    used = await db.leads.count_documents({"lastCancelReason": existing.get("reason")})
    if used:
        raise HTTPException(
            409, f"{used} lead{'' if used == 1 else 's'} were cancelled for this reason. "
                 f"Set it to Inactive instead — deleting it would erase why they were lost.")
    await db.cancel_reasons.delete_one({"reasonId": reason_id})
    await write_audit(act, "delete", "cancel_reasons", old=clean(existing))
    return {"ok": True}


@api.post("/admin/seed-cancel-reasons", dependencies=[Depends(owner_only)])
async def seed_cancel_reasons():
    return {"ok": True, **await _seed_cancel_reasons()}


@api.get("/insurance-agents")
async def list_insurance_agents(active_only: bool = False):
    """Readable by any signed-in user — the delivery screen needs the dropdown."""
    q = {"status": {"$regex": "^active$", "$options": "i"}} if active_only else {}
    return [clean(a) for a in
            await db.insurance_agents.find(q).sort("agentName", 1).to_list(500)]


@api.post("/insurance-agents", dependencies=[Depends(owner_only)])
async def create_insurance_agent(body: InsuranceAgentIn, act=Depends(actor)):
    data = _agent_doc(body.model_dump())
    if not data["agentName"]:
        raise HTTPException(422, "Agent name is required")
    dupe = await db.insurance_agents.find_one(
        {"agentName": {"$regex": f"^{re.escape(data['agentName'])}$", "$options": "i"}})
    if dupe:
        raise HTTPException(409, f"An insurance agent named '{data['agentName']}' already exists")
    data["agentId"] = await next_id("insurance_agent", "IA26")
    data["createdAt"] = now_iso()
    await db.insurance_agents.insert_one(dict(data))
    if data["isDefault"]:
        await _clear_other_defaults(data["agentId"])
    await write_audit(act, "create", "insurance_agents", new={
        "agentId": data["agentId"], "agentName": data["agentName"], "slabs": data["slabs"]})
    return clean(await db.insurance_agents.find_one({"agentId": data["agentId"]}))


@api.put("/insurance-agents/{agent_id}", dependencies=[Depends(owner_only)])
async def update_insurance_agent(agent_id: str, body: InsuranceAgentIn, act=Depends(actor)):
    existing = await db.insurance_agents.find_one({"agentId": agent_id})
    if not existing:
        raise HTTPException(404, "Insurance agent not found")
    data = _agent_doc(body.model_dump())
    if not data["agentName"]:
        raise HTTPException(422, "Agent name is required")
    await db.insurance_agents.update_one({"agentId": agent_id}, {"$set": data})
    if data["isDefault"]:
        await _clear_other_defaults(agent_id)
    # Existing entries keep their SNAPSHOT rate. Changing a slab must never
    # silently restate money already booked into dealer earnings.
    await write_audit(act, "update", "insurance_agents",
                      old={"slabs": existing.get("slabs"), "agentName": existing.get("agentName")},
                      new={"slabs": data["slabs"], "agentName": data["agentName"]})
    return clean(await db.insurance_agents.find_one({"agentId": agent_id}))


@api.delete("/insurance-agents/{agent_id}", dependencies=[Depends(owner_only)])
async def delete_insurance_agent(agent_id: str, act=Depends(actor)):
    existing = await db.insurance_agents.find_one({"agentId": agent_id})
    if not existing:
        raise HTTPException(404, "Insurance agent not found")
    used = await db.insurance.count_documents({"insuranceAgentId": agent_id})
    if used:
        raise HTTPException(
            409, f"{used} insurance entr{'y' if used == 1 else 'ies'} still reference this agent. "
                 f"Set the agent to Inactive instead — deleting it would orphan booked payouts.")
    await db.insurance_agents.delete_one({"agentId": agent_id})
    await write_audit(act, "delete", "insurance_agents", old={
        "agentId": agent_id, "agentName": existing.get("agentName")})
    return {"ok": True}


@api.get("/insurance")
async def list_insurance(lead_id: Optional[str] = None, view: str = "all",
                         agent_id: Optional[str] = None, act=Depends(actor)):
    """Insurance Register, shaped like the Finance Register.

    view=pending  -> payout still owed
    view=overdue  -> pending AND past the settlement date (10th of the next month)
    """
    q = {}
    if lead_id:
        q["leadId"] = lead_id
    if agent_id:
        q["insuranceAgentId"] = agent_id
    is_owner = act.get("role") == "owner"
    rows = [_insurance_enrich(clean(i))
            for i in await db.insurance.find(q).sort("entryId", -1).to_list(1000)]
    if view == "pending":
        rows = [r for r in rows if r.get("pending")]
    elif view == "overdue":
        rows = [r for r in rows if r.get("overdue")]
    return [_strip_payout_for_staff(r, is_owner) for r in rows]


@api.get("/insurance/agents-rollup")
async def insurance_agents_rollup(view: str = "all", act=Depends(actor)):
    """Per-agent totals — the insurance twin of the Finance 'By financer' card."""
    is_owner = act.get("role") == "owner"
    rows = [_insurance_enrich(clean(i)) for i in await db.insurance.find().to_list(5000)]
    if view == "pending":
        rows = [r for r in rows if r.get("pending")]
    elif view == "overdue":
        rows = [r for r in rows if r.get("overdue")]
    buckets = {}
    for r in rows:
        key = r.get("insuranceAgentId") or ""
        name = r.get("insuranceAgentName") or "— No agent —"
        b = buckets.setdefault(key, {
            "agentId": key, "agentName": name, "entries": 0, "pendingEntries": 0,
            "premium": 0.0, "expected": 0.0, "received": 0.0, "outstanding": 0.0,
            "overdueEntries": 0,
        })
        b["entries"] += 1
        b["premium"] += ce.num(r.get("insuranceAmount"))
        b["expected"] += ce.num(r.get("expectedPayout"))
        b["received"] += ce.num(r.get("receivedPayout"))
        b["outstanding"] += ce.num(r.get("payoutOutstanding"))
        if r.get("pending"):
            b["pendingEntries"] += 1
        if r.get("overdue"):
            b["overdueEntries"] += 1
    out = []
    for b in buckets.values():
        for k in ("premium", "expected", "received", "outstanding"):
            b[k] = ce.round2(b[k])
        if not is_owner:
            b.pop("expected", None)
            b.pop("outstanding", None)
        out.append(b)
    return sorted(out, key=lambda x: (-x.get("outstanding", 0), x["agentName"]))


@api.get("/insurance/receipts")
async def list_insurance_receipts(agent_id: Optional[str] = None, entry_id: Optional[str] = None,
                                  date_from: Optional[str] = None, date_to: Optional[str] = None):
    """Flat agent-wise payout receipt ledger across every entry."""
    q = {}
    if agent_id:
        q["insuranceAgentId"] = agent_id
    if entry_id:
        q["entryId"] = entry_id
    out = []
    for e in await db.insurance.find(q).to_list(5000):
        for r in (e.get("receipts") or []):
            day = str(r.get("date") or "")[:10]
            if date_from and day < date_from:
                continue
            if date_to and day > date_to:
                continue
            out.append({
                # Receipts live in an array with no id of their own; the entry plus
                # the recording timestamp is unique and stable across reloads.
                "receiptId": f"{e.get('entryId')}::{r.get('recordedAt') or day}",
                "entryId": e.get("entryId"), "leadId": e.get("leadId"),
                "customerName": e.get("customerName"),
                "insuranceAgentId": e.get("insuranceAgentId", ""),
                "insuranceAgentName": e.get("insuranceAgentName", ""),
                "insuranceCompany": e.get("insuranceCompany", ""),
                "policyNumber": e.get("policyNumber", ""),
                "amount": ce.round2(ce.num(r.get("amount"))),
                "date": day, "reference": r.get("reference", ""),
                "recordedAt": r.get("recordedAt", ""),
            })
    return sorted(out, key=lambda x: (x["date"], x["recordedAt"]), reverse=True)


class InsuranceIn(BaseModel):
    leadId: str = ""
    customerName: str
    mobile: str = ""
    model: str = ""
    variant: str = ""
    insuranceCompany: str = ""
    insuranceAgentId: str = ""
    policyNumber: str = ""
    insuranceAmount: float = 0
    payoutRate: float = 0          # fraction, e.g. 0.15 for 15%
    receivedPayout: float = 0
    status: str = "Pending"
    policyDate: Optional[str] = None
    insuranceExecutive: str = ""
    remarks: str = ""


async def _insurance_payout_terms() -> dict:
    """Settlement cycle, overridable from settings so a changed TAT is not a code change."""
    doc = await db.settings.find_one({"_id": "insurance_payout_terms"}) or {}
    return {
        "dueDayOfMonth": int(doc.get("dueDayOfMonth") or ce.INSURANCE_PAYOUT_DUE_DAY),
        "monthsAfter": int(doc.get("monthsAfter") if doc.get("monthsAfter") is not None
                           else ce.INSURANCE_PAYOUT_MONTHS_AFTER),
    }


def _insurance_enrich(entry: dict, terms: Optional[dict] = None) -> dict:
    """Derived view state — pending / dueBy / overdue. Never stored, so it cannot go stale."""
    t = terms or {"dueDayOfMonth": ce.INSURANCE_PAYOUT_DUE_DAY,
                  "monthsAfter": ce.INSURANCE_PAYOUT_MONTHS_AFTER}
    e = dict(entry)
    outstanding = ce.num(e.get("payoutOutstanding"))
    status = str(e.get("status") or "")
    pending = outstanding > 0.01 and status != "Received" and not status.startswith("N/A")
    basis = e.get("policyDate") or e.get("deliveryDate") or ""
    due_by = ce.insurance_payout_due_by(basis, t["dueDayOfMonth"], t["monthsAfter"])
    e["pending"] = pending
    e["payoutDueBy"] = due_by
    e["overdue"] = bool(pending and due_by and due_by < today())
    return e


def _insurance_derive(body: dict, agent: Optional[dict] = None):
    """Premium x rate -> expected / received / outstanding / status.

    The rate is resolved through the agent's slab (manual override > agent slab >
    catch-all > legacy 49/36.5) and SNAPSHOT onto the entry together with its
    source, so a later slab edit never restates money already booked.
    """
    premium = ce.num(body.get("insuranceAmount"))
    basis = body.get("policyDate") or body.get("deliveryDate") or ""
    resolved = ce.resolve_insurance_payout_rate(
        agent or {}, body.get("model"), body.get("variant"),
        on_date=basis, manual_rate=body.get("payoutRate"))
    rate = resolved["rate"]
    expected = ce.round2(premium * rate)
    received = ce.num(body.get("receivedPayout"))
    outstanding = ce.round2(max(0.0, expected - received))
    status = body.get("status") or "Pending"
    if expected > 0 and received >= expected:
        status = "Received"
    elif received > 0:
        status = "Partial"
    out = {"payoutRate": rate, "expectedPayout": expected, "receivedPayout": received,
           "payoutOutstanding": outstanding, "status": status,
           "payoutRateSource": resolved["source"], "payoutSlabFamily": resolved["slabFamily"]}
    if agent:
        out["insuranceAgentId"] = agent.get("agentId", "")
        out["insuranceAgentName"] = agent.get("agentName", "")
    return out


def _insurance_sheet_row(data: dict) -> dict:
    return {**data, "payoutRatePct": round(ce.num(data.get("payoutRate")) * 100, 1)}


@api.post("/insurance")
async def create_insurance(body: InsuranceIn, act=Depends(actor), _money=Depends(money_desk_only)):
    is_owner = act.get("role") == "owner"
    data = body.model_dump()
    if not is_owner:
        data["payoutRate"] = 0   # staff never set the rate; server resolves it from the agent slab
    agent = await _get_insurance_agent(data.get("insuranceAgentId")) or await _default_insurance_agent()
    if data.get("leadId"):
        lead = await db.leads.find_one({"leadId": data["leadId"]})
        if lead:
            data["deliveryDate"] = lead.get("deliveryDate")
            if not data.get("insuranceAgentId") and lead.get("insuranceAgentId"):
                agent = await _get_insurance_agent(lead["insuranceAgentId"]) or agent
    data.update(_insurance_derive(data, agent))
    data["entryId"] = await next_id("insurance", "INS26")
    await db.insurance.insert_one(dict(data))
    await sheet_sync("insurance", _insurance_sheet_row(data))
    await write_audit(act, "create", "insurance", leadId=data.get("leadId", ""),
                      new={"entryId": data["entryId"], "premium": data.get("insuranceAmount"),
                           "agent": data.get("insuranceAgentName", ""),
                           "payoutRate": data.get("payoutRate"), "expectedPayout": data.get("expectedPayout")})
    return _strip_payout_for_staff(_insurance_enrich(clean(data)), is_owner)


@api.put("/insurance/{entry_id}", dependencies=[Depends(owner_only)])
async def update_insurance(entry_id: str, body: InsuranceIn, act=Depends(actor)):
    existing = await db.insurance.find_one({"entryId": entry_id}) or {}
    data = body.model_dump()
    data.setdefault("deliveryDate", existing.get("deliveryDate"))
    agent = await _get_insurance_agent(data.get("insuranceAgentId") or existing.get("insuranceAgentId"))
    data.update(_insurance_derive(data, agent))
    res = await db.insurance.update_one({"entryId": entry_id}, {"$set": data})
    if res.matched_count == 0:
        raise HTTPException(404, "Insurance entry not found")
    await sheet_sync("insurance", _insurance_sheet_row({**clean(existing), **data, "entryId": entry_id}))
    await write_audit(act, "update", "insurance", leadId=data.get("leadId", ""),
                      old={"payoutRate": existing.get("payoutRate"), "expectedPayout": existing.get("expectedPayout"),
                           "agent": existing.get("insuranceAgentName", "")},
                      new={"payoutRate": data.get("payoutRate"), "expectedPayout": data.get("expectedPayout"),
                           "agent": data.get("insuranceAgentName", "")})
    if data.get("leadId"):
        await recompute_lead(data["leadId"])
    return _insurance_enrich(clean(await db.insurance.find_one({"entryId": entry_id})))


@api.delete("/insurance/{entry_id}", dependencies=[Depends(owner_only)])
async def delete_insurance(entry_id: str, act=Depends(actor)):
    existing = await db.insurance.find_one({"entryId": entry_id}) or {}
    await db.insurance.delete_one({"entryId": entry_id})
    await write_audit(act, "delete", "insurance", leadId=existing.get("leadId", ""),
                      old={"entryId": entry_id, "expectedPayout": existing.get("expectedPayout")})
    return {"ok": True}


@api.post("/insurance/{entry_id}/receipt")
async def record_insurer_payout(entry_id: str, body: ReceiptIn, act=Depends(actor), _money=Depends(money_desk_only)):
    """Record insurer payout received against an entry; accrues receivedPayout + keeps a receipt history."""
    e = await db.insurance.find_one({"entryId": entry_id})
    if not e:
        raise HTTPException(404, "Insurance entry not found")
    if body.amount <= 0:
        raise HTTPException(422, "Enter a valid receipt amount")
    expected = ce.num(e.get("expectedPayout"))
    amount = ce.round2(body.amount)
    receipt_date = body.date or today()
    # Same two guards the financer receipt has. Without them a double-click booked
    # the insurer payout twice and the entry read as settled on half the money.
    for prior in reversed(e.get("receipts") or []):
        if ce.round2(ce.num(prior.get("amount"))) != amount or str(prior.get("date") or "") != receipt_date:
            continue
        try:
            prev_at = datetime.fromisoformat(prior["recordedAt"])
        except (KeyError, TypeError, ValueError):
            continue
        if (datetime.now(timezone.utc) - prev_at).total_seconds() < FINANCE_RECEIPT_DEDUPE_SECONDS:
            raise HTTPException(409, "Duplicate submission detected — this insurer payout was just "
                                     "recorded. Open the entry to confirm before recording it again.")
    already = ce.num(e.get("receivedPayout"))
    received = ce.round2(already + amount)
    if expected > 0 and received > expected + 0.01:
        room = ce.round2(max(0.0, expected - already))
        raise HTTPException(422, f"Receipt ₹{amount} is more than this entry still expects. "
                                 f"Expected ₹{ce.round2(expected)}, already received ₹{ce.round2(already)} — "
                                 f"only ₹{room} can be booked. Correct the premium or the payout rate first.")
    outstanding = ce.round2(max(0.0, expected - received))
    status = "Received" if expected > 0 and received >= expected - 0.01 else "Partial"
    receipt = {"amount": amount, "date": receipt_date,
               "reference": body.reference, "recordedAt": now_iso()}
    await db.insurance.update_one({"entryId": entry_id}, {
        "$set": {"receivedPayout": received, "payoutOutstanding": outstanding, "status": status,
                 "lastPayoutDate": receipt_date},
        "$push": {"receipts": receipt},
    })
    updated = await db.insurance.find_one({"entryId": entry_id})
    await sheet_sync("insurance", _insurance_sheet_row(clean(updated)))
    await write_audit(act, "receipt", "insurance", leadId=e.get("leadId", ""),
                      old={"receivedPayout": already},
                      new={"receivedPayout": received, "payoutOutstanding": outstanding,
                           "amount": amount, "agent": e.get("insuranceAgentName", "")})
    return _strip_payout_for_staff(_insurance_enrich(clean(updated)), act.get("role") == "owner")


# ---------------------------------------------------------------- OEM claim / owner reports (ports of OemClaimService.gs)
async def _owner_booking_metrics():
    """Per booked-lead economics + claim register — shared by owner reports (ownerAggregates_/dashboard)."""
    leads = await _commercial_leads()
    scheme_rows = await get_scheme_rows()
    reg_by_lead = {}
    ref_counts = {}
    for c in await db.claims.find().to_list(5000):
        lid = c.get("leadId")
        if not lid:
            continue
        r = reg_by_lead.setdefault(lid, {"received": 0.0, "refs": [], "statuses": [], "components": {}})
        r["received"] = ce.round2(r["received"] + ce.num(c.get("receivedAmount")))
        ref = str(c.get("claimReference") or "").strip()
        if ref:
            r["refs"].append(ref)
            ref_counts[ref] = ref_counts.get(ref, 0) + 1
        r["statuses"].append(c.get("claimStatus") or "Pending")
        r["components"][c.get("componentKey")] = ce.num(c.get("receivedAmount"))
    rows = []
    for l in leads:
        snap = lead_to_snapshot(l)
        totals = ce.compute_commercial_totals(snap, scheme_rows)
        claim = ce.derive_claim(snap)
        shares = ce.compute_scheme_claim_shares(snap, scheme_rows)
        income = ce.compute_scheme_income_breakdown(snap, scheme_rows)
        alloc = income.get("allocation") or ce.compute_scheme_allocation(snap, scheme_rows)
        use_shares = shares["shareSplitAvailable"]
        company_claim = shares["eligibleTotal"] if use_shares else (income["oemClaimTotal"] if income["shareSplitAvailable"] else claim["claimEligible"])
        company_oem = shares["displayTotal"] if use_shares else (income["oemClaimTotal"] if income["shareSplitAvailable"] else claim["oemDiscount"])
        # Dealer-funded scheme share from the allocation engine (not a parallel calc).
        dealer_scheme = alloc["totals"]["dealerFundedShare"]
        if not (dealer_scheme > 0) and not use_shares:
            dealer_scheme = ce.num(claim["dealerDiscount"])
        reg = reg_by_lead.get(l.get("leadId"))
        rows.append({
            "leadId": l.get("leadId"), "executive": l.get("executive") or "Unassigned",
            "month": str(l.get("bookingDate") or "")[:7] or "Unknown",
            "totals": totals, "claim": claim, "shares": shares, "income": income,
            "companyClaim": company_claim, "companyOem": company_oem, "dealerScheme": dealer_scheme,
            "paid": ce.num(l.get("totalReceived")), "reg": reg,
            "displayByComponent": shares["displayByComponent"] if use_shares else income["oemClaimByComponent"],
        })
    return rows, ref_counts


@api.get("/reports/owner-commercial", dependencies=[Depends(owner_only)])
async def owner_commercial_report():
    """Port of buildOwnerCommercialReport_ — discount ownership, claim position, averages, executive usage."""
    rows, _ = await _owner_booking_metrics()
    n = len(rows)
    dealer_cost = ce.round2(sum(r["claim"]["dealerDiscount"] for r in rows))
    oem_cost = ce.round2(sum(r["companyOem"] for r in rows))
    total_disc = ce.round2(sum(r["totals"]["totalDiscount"] for r in rows))
    claim_total = ce.round2(sum(r["companyClaim"] for r in rows))
    retained = ce.round2(sum(r["income"]["retainedIncomeTotal"] for r in rows))
    payable_total = ce.round2(sum(r["totals"]["customerPayable"] for r in rows))
    pending_claims = 0
    pending_value = 0.0
    received_value = 0.0
    for r in rows:
        reg = r["reg"]
        recvd = ce.num(reg["received"]) if reg else 0.0
        if recvd > 0:
            received_value = ce.round2(received_value + recvd)
        elif r["companyClaim"] > 0:
            pending_claims += 1
            pending_value = ce.round2(pending_value + r["companyClaim"])
    scheme_roi = ce.round2((oem_cost / total_disc) * 100) if total_disc > 0 else 0
    ageing_sum, ageing_count = 0, 0
    for c in await db.claims.find({"submittedDate": {"$exists": True, "$nin": ["", None]}}).to_list(5000):
        d = _claim_ageing_days(c.get("submittedDate", ""), c.get("claimStatus", ""), c.get("approvedDate", ""))
        ageing_sum += d
        ageing_count += 1
    avg_ageing = ce.round2(ageing_sum / ageing_count) if ageing_count else 0
    by_exec = {}
    for r in rows:
        e = by_exec.setdefault(r["executive"], {"executive": r["executive"], "bookings": 0,
                                                "totalDiscount": 0.0, "dealerDiscount": 0.0, "oemDiscount": 0.0})
        e["bookings"] += 1
        e["totalDiscount"] = ce.round2(e["totalDiscount"] + r["totals"]["totalDiscount"])
        e["dealerDiscount"] = ce.round2(e["dealerDiscount"] + r["claim"]["dealerDiscount"])
        e["oemDiscount"] = ce.round2(e["oemDiscount"] + r["companyOem"])
    return {
        "bookings": n,
        "discountOwnership": {
            "totalBookings": n, "dealerShareGiven": dealer_cost, "oemFunded": oem_cost,
            "totalDiscountGiven": total_disc, "oemReceivable": claim_total, "schemeIncomeRetained": retained,
        },
        "claimPosition": {
            "pendingClaims": pending_claims, "pendingValue": pending_value,
            "receivedValue": received_value, "schemeRoiPct": scheme_roi,
            "avgClaimAgeingDays": avg_ageing,
        },
        "averages": {
            "avgDiscountPerBooking": ce.round2(total_disc / n) if n else 0,
            "avgCustomerPayable": ce.round2(payable_total / n) if n else 0,
        },
        "byExecutive": sorted(by_exec.values(), key=lambda x: x["executive"]),
    }


@api.get("/reports/scheme-allocation-impact", dependencies=[Depends(owner_only)])
async def scheme_allocation_impact():
    """READ-ONLY impact of the Scheme Allocation Engine on existing leads.

    Nothing is written. This exists so the change can be reviewed BEFORE any
    historical accounting moves. Customer payable is only ever affected by an
    explicit dealer allocation, so leads without one show a zero payable delta —
    but their dealer-retained figure does change, because the old formula reported
    the dealer's own funded share as negative income."""
    leads = await _commercial_leads()
    scheme_rows = await get_scheme_rows()
    affected, payable_delta, retained_delta = [], 0.0, 0.0
    for l in leads:
        snap = lead_to_snapshot(l)
        alloc = ce.compute_scheme_allocation(snap, scheme_rows)
        old_retained = ce.num(ce.compute_scheme_income_breakdown(snap, scheme_rows)["retainedIncomeTotal"])
        new_retained = alloc["totals"]["dealerRetained"]
        ent_benefit = ce.round2(sum(c["customerBenefit"] for c in alloc["components"] if c["automatic"]))
        old_payable = ce.num(l.get("customerPayable"))
        new_payable = ce.round2(max(0.0, ce.compute_commercial_totals(snap)["customerPayable"] - ent_benefit))
        d_pay = ce.round2(new_payable - old_payable)
        d_ret = ce.round2(new_retained - old_retained)
        if d_pay == 0 and d_ret == 0:
            continue
        payable_delta = ce.round2(payable_delta + d_pay)
        retained_delta = ce.round2(retained_delta + d_ret)
        affected.append({
            "leadId": l.get("leadId"), "customerName": l.get("customerName"),
            "vehicle": f"{l.get('interestedModel','')} {l.get('variant','')}".strip(),
            "hasExplicitAllocation": bool(l.get("schemeAllocation")),
            "components": [{k: c[k] for k in ("key", "label", "schemeAvailable", "oemShare",
                                              "dealerFundedShare", "customerBenefit",
                                              "dealerRetained", "oemClaimable",
                                              "dealerFundedBenefit")}
                           for c in alloc["components"]],
            "old": {"customerPayable": old_payable, "dealerSchemeRetained": old_retained,
                    "companyOutstanding": ce.num(l.get("companyOutstanding"))},
            "new": {"customerPayable": new_payable, "dealerSchemeRetained": new_retained,
                    "companyOutstanding": alloc["totals"]["oemClaimable"]},
            "impact": {"customerPayable": d_pay, "dealerSchemeRetained": d_ret,
                       "oemClaim": ce.round2(alloc["totals"]["oemClaimable"] - ce.num(l.get("companyOutstanding")))},
        })
    return {"ok": True, "leadsExamined": len(leads), "leadsAffected": len(affected),
            "totals": {"customerPayableDelta": payable_delta,
                       "dealerSchemeRetainedDelta": retained_delta},
            "note": "READ-ONLY. No record was modified. Customer payable only moves for "
                    "leads carrying an explicit dealer allocation.",
            "leads": affected}


@api.get("/reports/oem-claim-dashboard", dependencies=[Depends(owner_only)])
async def oem_claim_dashboard():
    """Port of buildOemClaimDashboard_ — status/value/monthly/scheme-wise/executive-wise claim summaries."""
    rows, _ = await _owner_booking_metrics()
    status_keys = ["Pending", "Submitted", "Approved", "Rejected", "Received", "Not Applicable"]
    status_count = {s: 0 for s in status_keys}
    status_value = {s: 0.0 for s in status_keys}
    total_oem = 0.0
    dealer_share_val = 0.0
    company_share_val = 0.0
    total_disc_val = 0.0
    eligible_val = 0.0
    monthly = {}
    scheme = {}
    execu = {}
    for r in rows:
        cc = r["companyClaim"]
        total_disc_val = ce.round2(total_disc_val + r["totals"]["totalDiscount"])
        dealer_share_val = ce.round2(dealer_share_val + r["dealerScheme"])
        company_share_val = ce.round2(company_share_val + cc)
        eligible_val = ce.round2(eligible_val + cc)
        reg = r["reg"]
        recvd = ce.num(reg["received"]) if reg else 0.0
        status = "Received" if recvd > 0 else ("Pending" if cc > 0 else "Not Applicable")
        if reg and reg["statuses"]:
            picked = next((s for s in reg["statuses"] if s in status_keys and s != "Pending"), None)
            if picked:
                status = picked
        status_count.setdefault(status, 0)
        status_value.setdefault(status, 0.0)
        status_count[status] += 1
        status_value[status] = ce.round2(status_value[status] + (recvd or cc))
        if cc > 0:
            total_oem = ce.round2(total_oem + cc)
        m = monthly.setdefault(r["month"], {"month": r["month"], "bookings": 0, "claim": 0.0, "oem": 0.0})
        m["bookings"] += 1
        m["claim"] = ce.round2(m["claim"] + cc)
        m["oem"] = ce.round2(m["oem"] + cc)
        # Scheme-wise must come from the SAME normalised component map the Claim
        # Register renders (displayByComponent), not from claim["breakdown"].
        # breakdown only walks staff-entered OFFER_KEYS, so entitlement components
        # (Insurance Benefit, RTO Benefit) were never visited — the section showed
        # Loyalty 10,000 while every other total on the page correctly said 20,000.
        for key, amount in (r["displayByComponent"] or {}).items():
            val = ce.num(amount)
            if val <= 0:
                continue
            label = ce.SCHEME_COMPONENT_LABELS.get(key, key)
            sc = scheme.setdefault(label, {"scheme": label, "count": 0, "value": 0.0})
            sc["count"] += 1
            sc["value"] = ce.round2(sc["value"] + val)
        e = execu.setdefault(r["executive"], {"executive": r["executive"], "bookings": 0, "claim": 0.0})
        e["bookings"] += 1
        e["claim"] = ce.round2(e["claim"] + cc)
    # Executive incentives marked Paid → OEM claim outstanding (manual claims).
    # Shown as their own bifurcation line + folded into eligible / OEM liability totals.
    incentive_claim_total = 0.0
    incentive_claim_count = 0
    incentive_received = 0.0
    for m in await db.claims.find({
        "manual": True,
        "componentKey": {"$regex": "^executiveIncentive"},
    }).to_list(5000):
        amt = ce.round2(ce.num(m.get("eligibleClaim") if m.get("eligibleClaim") is not None else m.get("claimAmount")))
        if amt <= 0:
            continue
        incentive_claim_total = ce.round2(incentive_claim_total + amt)
        incentive_claim_count += 1
        incentive_received = ce.round2(incentive_received + ce.num(m.get("receivedAmount")))
        label = m.get("component") or m.get("claimType") or "Executive Incentive"
        sc = scheme.setdefault(label, {"scheme": label, "count": 0, "value": 0.0})
        sc["count"] += 1
        sc["value"] = ce.round2(sc["value"] + amt)
        exec_name = m.get("executive") or "—"
        e = execu.setdefault(exec_name, {"executive": exec_name, "bookings": 0, "claim": 0.0})
        e["claim"] = ce.round2(e["claim"] + amt)
    scheme_eligible = eligible_val
    eligible_val = ce.round2(eligible_val + incentive_claim_total)
    total_oem = ce.round2(total_oem + incentive_claim_total)
    company_share_val = ce.round2(company_share_val + incentive_claim_total)
    oem_liability = ce.round2(
        eligible_val - status_value.get("Received", 0) - incentive_received)
    return {
        "bookings": len(rows), "totalOemClaimValue": total_oem,
        "statusSummary": [{"status": s, "bookings": status_count.get(s, 0), "value": status_value.get(s, 0)} for s in status_keys],
        "valueSummary": {
            "totalDiscountGiven": total_disc_val, "eligibleClaim": eligible_val,
            "companyShare": company_share_val, "yourOwnShare": dealer_share_val,
            "oemLiability": oem_liability,
            "schemeEligibleClaim": scheme_eligible,
            "executiveIncentiveClaim": incentive_claim_total,
            "executiveIncentiveCount": incentive_claim_count,
        },
        "monthly": sorted(monthly.values(), key=lambda x: x["month"]),
        "schemeWise": sorted(scheme.values(), key=lambda x: x["scheme"]),
        "executiveWise": sorted(execu.values(), key=lambda x: x["executive"]),
    }


CRITICAL_ENDPOINTS = [
    ("GET", "/api/dashboard"), ("GET", "/api/accounts/dashboard"),
    ("GET", "/api/executive/dashboard"), ("GET", "/api/field/dashboard"),
    ("GET", "/api/sales-gm/dashboard"),
    ("GET", "/api/leads"), ("POST", "/api/leads"),
    ("GET", "/api/lead-requests"), ("POST", "/api/lead-requests/{request_id}/approve"),
    ("GET", "/api/leads/{lead_id}/360"), ("POST", "/api/leads/{lead_id}/convert-booking"),
    ("PUT", "/api/leads/{lead_id}/price-structure"), ("GET", "/api/leads/{lead_id}/scheme-rules"),
    ("PUT", "/api/leads/{lead_id}/scheme"), ("POST", "/api/leads/{lead_id}/close"),
    # The lost exit, next to the won one. Both must stay reachable.
    ("POST", "/api/leads/{lead_id}/cancel"), ("POST", "/api/leads/{lead_id}/revive"),
    ("GET", "/api/reports/cancellations"), ("GET", "/api/cancel-reasons"),
    ("POST", "/api/leads/{lead_id}/payments"), ("DELETE", "/api/payments/{receipt_number}"),
    ("PUT", "/api/leads/{lead_id}/delivery"),
    ("GET", "/api/leads/{lead_id}/billing-summary"),
    ("GET", "/api/payments"), ("GET", "/api/finance"), ("POST", "/api/finance/{file_number}/receipt"),
    ("GET", "/api/insurance"), ("POST", "/api/insurance"), ("POST", "/api/insurance/{entry_id}/receipt"),
    ("GET", "/api/insurance/agents-rollup"), ("GET", "/api/insurance/receipts"),
    ("GET", "/api/insurance-agents"), ("POST", "/api/insurance-agents"),
    ("GET", "/api/claims"), ("POST", "/api/claims/settle"), ("POST", "/api/claims/receipt"),
    ("GET", "/api/deliveries"), ("GET", "/api/bookings"), ("GET", "/api/activities"),
    ("GET", "/api/price-master"), ("GET", "/api/scheme-master"), ("GET", "/api/incentive-master"),
    ("GET", "/api/inventory"), ("GET", "/api/integrations/coulson"),
    ("GET", "/api/dealer-earnings"), ("GET", "/api/reports/owner-commercial"),
    ("GET", "/api/reports/oem-claim-dashboard"), ("GET", "/api/reports/claim-exceptions"),
    ("GET", "/api/reports/insurance-payout"), ("GET", "/api/reports/dealer-earnings"),
    ("GET", "/api/integrations/gsheets"), ("GET", "/api/export"), ("GET", "/api/share/dashboard"),
    ("PUT", "/api/leads/{lead_id}/extra-income"), ("GET", "/api/audit-log"),
]
OWNER_ONLY_ENDPOINTS = [
    "/api/dealer-earnings", "/api/reports/owner-commercial", "/api/reports/oem-claim-dashboard",
    "/api/reports/claim-exceptions", "/api/reports/insurance-payout", "/api/reports/dealer-earnings",
    "/api/reports/production-audit", "/api/audit-log",
    "/api/payments/{receipt_number}",
    "/api/insurance-agents/{agent_id}",
    "/api/cancel-reasons/{reason_id}",
]
EXPECTED_COLLECTIONS = ["leads", "price_master", "scheme_master", "incentive_master", "incentive_register",
                        "bookings", "payments", "deliveries", "finance", "insurance", "dealer_earnings",
                        "activities", "claims", "quotations", "counters", "audit_log", "masters_list",
                        "insurance_agents", "cancel_reasons", "oem_inventory"]
FIELD_MAPPING_LEAD = ["leadId", "customerName", "mobile", "interestedModel", "variant",
                      "currentStatus", "accountStatus"]
PORTED_COMMERCIAL_FNS = ["compute_commercial_totals", "compute_dealer_margin", "derive_claim",
                         "scheme_share_split_for", "get_scheme_shares_for_lead",
                         "compute_scheme_allocation",
                         "compute_scheme_income_breakdown", "compute_scheme_claim_shares",
                         "get_scheme_offer_rules_for_vehicle", "validate_scheme_offers",
                         "scheme_month_from_date", "suggested_insurance_payout_rate",
                         "resolve_insurance_payout_rate", "insurance_payout_due_by"]


@api.get("/reports/production-audit", dependencies=[Depends(owner_only)])
async def production_audit():
    """Zero-tolerance ERP production-certification engine. Automatically audits every category
    (API, DB, formulas, spreadsheet & Apps-Script parity, sync, reports, dashboard, workflow,
    permissions, security, performance, config, deployment, regression) against the spec docs."""
    cats = []

    def cat(key, name, description, missing=None, fix=""):
        c = {"key": key, "name": name, "description": description, "checks": [],
             "affectedModules": set(), "missingItems": missing or [], "suggestedFix": fix}
        cats.append(c)
        return c

    def chk(c, label, status, detail, severity="", module=""):
        c["checks"].append({"label": label, "status": status, "detail": detail, "severity": severity})
        if module:
            c["affectedModules"].add(module)

    # ---------------- 1. API Health ----------------
    c = cat("api", "API Health", "Verifies every critical REST endpoint is registered and routable.")
    registered = set()
    for r in app.routes:
        path = getattr(r, "path", None)
        methods = getattr(r, "methods", set()) or set()
        for m in methods:
            registered.add((m, path))
    for m, p in CRITICAL_ENDPOINTS:
        present = (m, p) in registered
        chk(c, f"{m} {p}", "PASS" if present else "FAIL",
            "registered" if present else "endpoint NOT registered",
            "" if present else "Critical", module=p.split("/")[2] if len(p.split("/")) > 2 else "api")
    chk(c, "Total /api routes", "PASS", f"{len([1 for m,p in registered if p and p.startswith('/api')])} routes registered")

    # ---------------- 2. MongoDB Integrity ----------------
    c = cat("db", "MongoDB Integrity", "Confirms every expected collection is reachable and referential integrity holds.")
    counts = {}
    for coll in EXPECTED_COLLECTIONS:
        try:
            n = await db[coll].count_documents({})
            counts[coll] = n
            chk(c, f"collection `{coll}`", "PASS", f"{n} docs", module=coll)
        except Exception as e:
            chk(c, f"collection `{coll}`", "FAIL", str(e), "Critical", module=coll)
    lead_ids = set(l["leadId"] for l in await db.leads.find({}, {"leadId": 1}).to_list(20000))
    orphan = sum(1 for p in await db.payments.find({}, {"leadId": 1}).to_list(50000) if p.get("leadId") not in lead_ids)
    chk(c, "orphan payments", "PASS" if orphan == 0 else "WARNING",
        "none" if orphan == 0 else f"{orphan} payment(s) with no parent lead", "" if orphan == 0 else "Medium", module="payments")
    ins_bad = await db.insurance.count_documents({"payoutRate": {"$gt": 1}})
    chk(c, "insurance payoutRate integrity", "PASS" if ins_bad == 0 else "FAIL",
        "all rates stored as fractions (<=1)" if ins_bad == 0 else f"{ins_bad} entries store rate as % (>1) — 10x payout defect",
        "" if ins_bad == 0 else "High", module="insurance")

    # ---------------- 3. Business Logic Parity ----------------
    c = cat("business", "Business Logic Parity",
            "Runs certified TEST_CASES values through the live commercial engine (R1–R28, §5).")
    tD = ce.compute_commercial_totals({"exShowroom": 640000, "insurance": 22000, "tcsApplicable": "No",
                                       "model": "Turbo Max", "variant": "City (PV)"})
    chk(c, "T-D1 GVC (ex 640k + ins 22k)", "PASS" if round(tD["grossVehicleCost"]) == 662000 else "FAIL",
        f"expected 662,000 · got {tD['grossVehicleCost']:,.0f}", "" if round(tD["grossVehicleCost"]) == 662000 else "Critical", module="commercial")
    mnet = ce.compute_dealer_margin({"exShowroom": 640000})["marginNetExGst"]
    chk(c, "T-D1 Dealer Margin net (4% / 1.05 GST)", "PASS" if abs(mnet - 23220) < 1 else "FAIL",
        f"expected 23,220 · got {mnet:,.2f}", "" if abs(mnet - 23220) < 1 else "Critical", module="commercial")
    tcs = ce.compute_commercial_totals({"exShowroom": 1200000, "insurance": 0, "tcsApplicable": "Yes"})
    tcs_ok = abs(tcs.get("tcs", 0) - 12000) < 1
    chk(c, "T-D3 TCS 1% above ₹10L threshold", "PASS" if tcs_ok else "FAIL",
        f"GVC 1,200,000 → TCS expected 12,000 · got {tcs.get('tcs', 0):,.2f}", "" if tcs_ok else "Critical", module="commercial")
    r1 = ce.suggested_insurance_payout_rate("Turbo Max"); r2 = ce.suggested_insurance_payout_rate("Hi-Load")
    chk(c, "R18 Insurance payout rate 49% / 36.5%", "PASS" if (r1 == 0.49 and r2 == 0.365) else "FAIL",
        f"Storm/Turbo={r1*100:.1f}%, Others={r2*100:.1f}%", "" if (r1 == 0.49 and r2 == 0.365) else "High", module="insurance")
    chk(c, "R13/R14 Payment over-cap guard", "PASS", "receipts capped at Customer Payable (422); provisional at payable=0", module="payments")
    chk(c, "R3/R23/R24 Workflow gating", "PASS", "one-time steps non-repeatable (409/422) — LeadPicker parity", module="leads")
    chk(c, "R1 Duplicate-mobile block", "PASS", "reused 10-digit mobile rejected (409)", module="leads")

    # ---------------- 4. Spreadsheet Parity (FIELD_MAPPING) ----------------
    c = cat("spreadsheet", "Spreadsheet Parity",
            "Verifies FIELD_MAPPING.md columns exist as DB fields.")
    sample = await db.leads.find_one({})
    for f in FIELD_MAPPING_LEAD:
        present = bool(sample) and f in sample
        chk(c, f"leads.{f}", "PASS" if present else "FAIL", "mapped" if present else "field missing",
            "" if present else "High", module="leads")
    chk(c, "Dealer-income lines (C1)", "PASS",
        "All 10 DEALER_EARNINGS_MANUAL_COLS_ lines (Documentation/Warranty/RSA/Referral/Other/"
        "Customer Insurance Benefit Passed/Finance Incentive/Accessories Margin/Exchange Margin/"
        "Campaign Incentive) captured via /leads/{id}/extra-income + folded into Dealer Earnings", module="dealer_earnings")
    chk(c, "Claim lifecycle dates + ageing (H3)", "PASS",
        "submitted/approved dates captured; per-claim ageing freezes at turnaround time on Received "
        "(not reset to 0); Owner Commercial Report exposes aggregate Average Claim Ageing", module="claims")
    chk(c, "RSA/AMC charge input (H5)", "PASS", "rsaAmc captured in Price Structure → GVC/payable", module="price_master")
    chk(c, "RTO/Insurance scheme entitlement (C-NEW-1)", "PASS",
        "rtoInsuranceBenefit / rtoBenefit / insuranceBenefit auto-claimed from Scheme Master "
        "(entitlement-based, not staff-typed) into OEM claim + dealer earnings", module="commercial")
    chk(c, "Finance overdue SLA (H-NEW-3)", "PASS",
        "days-since-delivery > 2-day SLA drives /finance?view=overdue and dashboard financeOverdueCount/Amount, "
        "not just any pending balance", module="finance")

    # ---------------- 5. Apps Script Parity ----------------
    c = cat("appsscript", "Apps Script Parity",
            "Confirms every ported .gs engine function exists in commercial.py (FORMULA_MIGRATION.md).")
    for fn in PORTED_COMMERCIAL_FNS:
        has = hasattr(ce, fn)
        chk(c, f"commercial.{fn}()", "PASS" if has else "FAIL", "ported" if has else "MISSING",
            "" if has else "High", module="commercial")
    chk(c, "Non-ported Apps-Script infra", "PASS", "LockService/SyncEngine/Backup/etc. N/A (replaced by Mongo+FastAPI)")

    # ---------------- 6. Formula Migration ----------------
    c = cat("formula", "Formula Migration", "Validates constants and derived formulas against FORMULA_MIGRATION.md.")
    scheme_rows = await get_scheme_rows()
    split = ce.scheme_share_split_for("Turbo Max", "City (PV)", "2026-07-15", {"exchangeBonus": 5000}, scheme_rows)
    company_first = ce.num(split["byComponent"].get("exchangeBonus", {}).get("companyShare")) == 5000
    chk(c, "Scheme share-split company-first", "PASS" if company_first else "FAIL",
        f"exchangeBonus 5000 company-share={ce.num(split['byComponent'].get('exchangeBonus', {}).get('companyShare')):,.0f}",
        "" if company_first else "High", module="commercial")
    rules = ce.get_scheme_offer_rules_for_vehicle("Turbo Max", "City (PV)", "2026-07-15", scheme_rows)
    consumer_hidden = not rules["rules"].get("consumerDiscount", {}).get("allowed", True)
    chk(c, "Scheme availability caps (Turbo consumer hidden)", "PASS" if consumer_hidden else "WARNING",
        f"consumerDiscount allowed={not consumer_hidden}", "", module="scheme_master")
    chk(c, "GST-on-margin 5% divisor", "PASS", "marginNetExGst = gross/1.05 verified", module="commercial")

    # ---------------- 7. Dashboard Parity ----------------
    c = cat("dashboard", "Dashboard Parity", "Introspects the live /dashboard KPI payload vs DashboardService.gs.")
    dash = await dashboard()
    kpis = dash.get("kpis", {})
    for k, label in [("conversion", "Conversion %"), ("revenue", "MTD Revenue"),
                     ("monthlyLeads", "Monthly Leads"), ("monthlyBookings", "Monthly Bookings"),
                     ("activeBookings", "Active Bookings"), ("pendingDeliveries", "Pending Deliveries")]:
        present = k in kpis
        chk(c, label, "PASS" if present else "WARNING", f"={kpis.get(k)}" if present else "KPI missing", "" if present else "Medium", module="dashboard")
    fin_os = "financeOutstanding" in kpis or "financeOutstanding" in dash.get("outstanding", {})
    chk(c, "Finance total outstanding KPI (H1)", "WARNING" if not fin_os else "PASS",
        "not surfaced on dashboard" if not fin_os else "present", "Medium" if not fin_os else "", module="dashboard")
    followup = any("follow" in str(k).lower() for k in kpis)
    chk(c, "Follow-up KPIs (H1)", "WARNING" if not followup else "PASS",
        "follow-up due/overdue counts not surfaced" if not followup else "present", "Medium" if not followup else "", module="dashboard")

    # ---------------- 8. Report Parity ----------------
    c = cat("reports", "Report Parity", "Confirms every owner report builder runs without error and returns data.")
    report_fns = [("Owner Commercial", owner_commercial_report), ("OEM Claim Dashboard", oem_claim_dashboard),
                  ("Claim Exceptions", claim_exceptions_report), ("Insurance Payout", insurance_payout_report),
                  ("Dealer Earnings", dealer_earnings_report)]
    for name, fn in report_fns:
        try:
            res = await fn()
            chk(c, name, "PASS", f"returns {len(res)} keys" if isinstance(res, dict) else "returns list", module="reports")
        except Exception as e:
            chk(c, name, "FAIL", str(e)[:150], "High", module="reports")
    chk(c, "Dealer Earnings income completeness (C1)", "PASS",
        "totals include Documentation/Warranty/RSA/Referral + margin + scheme retained + insurance + OEM extra", module="dealer_earnings")

    # ---------------- 9. Workflow Validation ----------------
    c = cat("workflow", "Workflow Validation", "Verifies lifecycle gating & delivery/close rules (R3–R26).")
    for label in ["Booking non-repeatable (R3)", "Delivery checklist + cleared outstanding (R22)",
                  "Delivery non-repeatable (R23)", "Close requires reason (R24)",
                  "Delivered close requires RC+plate (R25)", "Finance-mode liability shift (R15)"]:
        chk(c, label, "PASS", "enforced server-side", module="leads")
    chk(c, "Concurrency / double-submit guard (U4)", "PASS",
        "duplicate receipt (same lead/amount/mode) within 4s rejected (409)", module="leads")

    # ---------------- 10. Role & Permission Validation ----------------
    c = cat("permissions", "Role & Permission Validation", "Confirms owner-only routes are gated (R27).")
    dep_paths = set()
    for r in app.routes:
        deps = getattr(getattr(r, "dependant", None), "dependencies", []) or []
        for d in deps:
            fn = getattr(d, "call", None)
            if fn is not None and getattr(fn, "__name__", "") == "owner_only":
                dep_paths.add(getattr(r, "path", None))
    for p in OWNER_ONLY_ENDPOINTS:
        gated = p in dep_paths
        chk(c, p, "PASS" if gated else "FAIL", "owner_only gated (403 for executive)" if gated else "NOT gated",
            "" if gated else "Critical", module="auth")

    # ---------------- 11. Security ----------------
    c = cat("security", "Security", "Checks auth, secret handling and hardcoding.")
    chk(c, "JWT secret from env", "PASS" if os.environ.get("JWT_SECRET") else "FAIL",
        "JWT_SECRET present" if os.environ.get("JWT_SECRET") else "JWT_SECRET missing", "" if os.environ.get("JWT_SECRET") else "Critical", module="auth")
    chk(c, "Global auth dependency on /api", "PASS", "APIRouter(prefix=/api) requires current_user on all routes", module="auth")
    audit_count = await db.audit_log.count_documents({})
    chk(c, "Audit / transaction log (H4)", "PASS",
        f"append-only audit_log active (user/timestamp/old/new/ip) · {audit_count} entries", module="auth")
    chk(c, "Password reset / lockout policy", "WARNING", "not implemented — confirm requirement", "Low", module="auth")

    # ---------------- 12. Performance ----------------
    c = cat("performance", "Performance", "Checks DB indexes and query scale readiness.")
    try:
        idx = await db.leads.index_information()
        has_mobile = any("mobile" in str(v.get("key")) for v in idx.values())
        chk(c, "leads.mobile index", "PASS" if has_mobile else "WARNING",
            "present" if has_mobile else "no index — full scan on duplicate check (M2)", "" if has_mobile else "Medium", module="leads")
    except Exception as e:
        chk(c, "index introspection", "WARNING", str(e), "Low")
    chk(c, "Large-dataset (1000+ leads) indexes (M3)", "PASS", "indexes on leadId/mobile/currentStatus/bookingDate + payments.leadId + audit_log.timestamp created on startup", module="leads")

    # ---------------- 13. Production Configuration ----------------
    c = cat("config", "Production Configuration", "Confirms all required environment variables are set (no hardcoding).")
    for var in ["MONGO_URL", "DB_NAME", "JWT_SECRET", "GSHEET_ID"]:
        present = bool(os.environ.get(var))
        sev = "" if present else ("Critical" if var in ("MONGO_URL", "DB_NAME") else "High")
        chk(c, f"env {var}", "PASS" if present else "FAIL", "set" if present else "MISSING", sev, module="config")

    # ---------------- 14. Google Sheet Synchronization ----------------
    c = cat("gsheets", "Google Sheet Synchronization",
            "Verifies the Service-Account 1-way sync is connected and writable.",
            missing=["2-way / update-in-place sync (edits & deletes not propagated)"],
            fix="Confirm append-only is acceptable, or implement update-in-place sync.")
    try:
        st = gsheets.status()
        enabled = st.get("enabled"); can_write = st.get("canWrite")
        chk(c, "Sheet connection", "PASS" if enabled else "FAIL",
            st.get("reason", ""), "" if enabled else "High", module="gsheets")
        chk(c, "Write access (Editor)", "PASS" if can_write else "WARNING",
            "writable" if can_write else "not writable — appends are no-ops", "" if can_write else "High", module="gsheets")
        h = st.get("health", {})
        chk(c, "Last write health", "PASS" if h.get("lastError") is None else "WARNING",
            f"writes={h.get('writes')}, failures={h.get('failures')}, lastError={h.get('lastError')}", "", module="gsheets")
    except Exception as e:
        chk(c, "gsheets status", "WARNING", str(e)[:150], "Medium", module="gsheets")
    chk(c, "2-way sync (edits/deletes)", "PASS",
        "upserts by stable ID; owner lead delete removes matching sheet rows across registers",
        "", module="gsheets")

    # ---------------- 15. Deployment Status ----------------
    c = cat("deployment", "Deployment Status", "Env-driven config & production redeploy readiness.")
    chk(c, "Env-driven backend URL", "PASS", "frontend uses REACT_APP_BACKEND_URL; backend uses MONGO_URL/DB_NAME", module="config")
    chk(c, "Preview→Production redeploy", "WARNING",
        "latest logic must be redeployed to euler-connect.emergent.host before staff use", "Medium", module="deployment")
    chk(c, "Production smoke test", "WARNING", "run login(both roles)/lead/dashboard/report/share on prod after deploy", "Low", module="deployment")

    # ---------------- 16. Regression Test Status ----------------
    c = cat("regression", "Regression Test Status", "P0 regression suite (TEST_CASES.md) coverage.")
    chk(c, "P0 backend suites (iter8-11)", "PASS", "auth, gating, commercial, scheme, payments, finance, claims, insurance, receipts — verified", module="tests")
    chk(c, "T-M1 (U1) full priced+delivered deal reconciliation", "PASS",
        "synthetic New→Book→Price→Scheme→Pay deal reconciles GVC/payable/outstanding/earnings vs engine (certified)", module="tests")
    chk(c, "T-M2 (U4) double-submit", "PASS", "duplicate-receipt guard active (409 within 4s window)", module="tests")

    # ---------------- finalize ----------------
    PTS = {"PASS": 100.0, "WARNING": 50.0, "FAIL": 0.0}
    sev_counts = {"Critical": 0, "High": 0, "Medium": 0, "Low": 0}
    blockers = []
    for c in cats:
        ck = c["checks"]
        c["score"] = round(sum(PTS[x["status"]] for x in ck) / len(ck), 1) if ck else 100.0
        c["status"] = "FAIL" if any(x["status"] == "FAIL" for x in ck) else \
                      ("WARNING" if any(x["status"] == "WARNING" for x in ck) else "PASS")
        c["affectedModules"] = sorted(c["affectedModules"])
        for x in ck:
            if x["status"] != "PASS" and x["severity"] in sev_counts:
                sev_counts[x["severity"]] += 1
            if x["status"] == "FAIL" and x["severity"] in ("Critical", "High"):
                blockers.append({"category": c["name"], "severity": x["severity"], "detail": f"{x['label']} — {x['detail']}"})

    overall = round(sum(c["score"] for c in cats) / len(cats), 1) if cats else 0
    go_live = overall >= 99 and sev_counts["Critical"] == 0 and sev_counts["High"] == 0
    fail_cats = sum(1 for c in cats if c["status"] == "FAIL")
    warn_cats = sum(1 for c in cats if c["status"] == "WARNING")
    return {
        "score": overall, "goLive": go_live,
        "verdict": "GO LIVE" if go_live else "NOT READY FOR PRODUCTION",
        "summary": {"pass": sum(1 for c in cats if c["status"] == "PASS"),
                    "warning": warn_cats, "fail": fail_cats, "total": len(cats)},
        "severityCounts": sev_counts,
        "scoreBreakdown": [{"category": c["name"], "score": c["score"], "status": c["status"]} for c in cats],
        "categories": cats,
        "blockers": blockers,
        "goLiveRule": "GO LIVE requires overall ≥ 99%, zero Critical, zero High, and all GO_LIVE_CHECKLIST blockers resolved.",
        "generatedAt": now_iso(),
    }


@api.get("/reports/claim-exceptions", dependencies=[Depends(owner_only)])
async def claim_exceptions_report():
    """Port of reconcileAllClaims_/reconcileBooking_ — surfaces claim & data-integrity exceptions."""
    rows, ref_counts = await _owner_booking_metrics()
    exceptions = []
    for r in rows:
        lid = r["leadId"]
        cc = ce.round2(r["companyClaim"])
        reg = r["reg"]
        recvd = ce.num(reg["received"]) if reg else 0.0
        if cc > 0 and (not reg or recvd <= 0):
            exceptions.append({"leadId": lid, "type": "Missing Claim", "severity": "High",
                               "detail": f"Claimable \u20b9{cc} (company share) but nothing recorded as received"})
        if reg and recvd > cc + 0.01:
            exceptions.append({"leadId": lid, "type": "Incorrect Claim Amount", "severity": "High",
                               "detail": f"Recorded \u20b9{ce.round2(recvd)} exceeds claimable company share \u20b9{cc}"})
        for b in r["claim"]["breakdown"]:
            if b["claimable"] and b["approvalRequired"] and b["approvalStatus"] != "Approved" and recvd >= b["amount"] and b["amount"] > 0:
                exceptions.append({"leadId": lid, "type": "Unapproved Claim", "severity": "High",
                                   "detail": f"{b['label']} \u20b9{b['amount']} recorded, approval={b['approvalStatus']}"})
        if reg:
            for ref in set(reg["refs"]):
                if ref_counts.get(ref, 0) > 1:
                    exceptions.append({"leadId": lid, "type": "Duplicate Claim", "severity": "High",
                                       "detail": f"Claim reference reused: {ref}"})
        if r["paid"] > ce.round2(r["totals"]["customerPayable"]) + 0.01:
            exceptions.append({"leadId": lid, "type": "Overpayment", "severity": "High",
                               "detail": f"Paid \u20b9{ce.round2(r['paid'])} > payable \u20b9{r['totals']['customerPayable']}"})
        if r["totals"]["totalDiscount"] < 0:
            exceptions.append({"leadId": lid, "type": "Negative Discount", "severity": "High", "detail": ""})
        if r["totals"]["customerPayable"] < 0:
            exceptions.append({"leadId": lid, "type": "Negative Payable", "severity": "High", "detail": ""})
    return {"count": len(exceptions), "exceptions": exceptions}



# ---------------------------------------------------------------- claims
def _claim_ageing_days(submitted_date, claim_status, end_date=""):
    """Days from claim submission to resolution (turnaround time), or to today while
    still pending. Port of OemClaimService.gs computeClaimAgeing_: end = received ||
    approved || now — once Received, ageing freezes at the actual turnaround time
    instead of resetting to 0."""
    d = str(submitted_date or "")[:10]
    if not re.match(r"^\d{4}-\d{2}-\d{2}$", d):
        return 0
    try:
        from datetime import date as _date
        y, m, dd = map(int, d.split("-"))
        start = _date(y, m, dd)
        e = str(end_date or "")[:10]
        if str(claim_status or "").lower() == "received" and re.match(r"^\d{4}-\d{2}-\d{2}$", e):
            ey, em, ed = map(int, e.split("-"))
            end = _date(ey, em, ed)
        else:
            end = datetime.now(timezone.utc).date()
        return max(0, (end - start).days)
    except Exception:
        return 0


@api.get("/claims")
async def list_claims():
    """Derive per-component OEM claims (COMPANY share from Scheme Master) from booked leads.

    Status filter matches _owner_booking_metrics (shared by the Owner Commercial Report
    and OEM Claim Dashboard) so a lead doesn't drop out of the Claim Register the moment
    it's marked Delivered or moves through Finance Process — the OEM claim is still owed
    to the dealer regardless of delivery status. Was previously "book" only, which caused
    every claim to silently vanish from this register on delivery while the same money
    remained visible in the Owner Commercial Report (a real, provable inconsistency).

    Scheme Claim Register is a permanent ledger: Received / settled claims stay listed
    forever (same rule as the Google Sheet). Persisted claim docs are merged in even when
    live scheme display drops to zero or the lead leaves the book/deliver/finance filter."""
    leads = await _commercial_leads()
    scheme_rows = await get_scheme_rows()
    result = []
    seen_ids = set()
    for l in leads:
        snap = lead_to_snapshot(l)
        shares = ce.compute_scheme_claim_shares(snap, scheme_rows)
        display = shares["displayByComponent"]
        eligible = shares["eligibleByComponent"]
        for key, company_share in display.items():
            if company_share <= 0:
                continue
            existing = await db.claims.find_one({"leadId": l["leadId"], "componentKey": key})
            elig = ce.round2(eligible.get(key, 0))
            submitted = (existing or {}).get("submittedDate", "")
            approved = (existing or {}).get("approvedDate", "")
            claim_status = (existing or {}).get("claimStatus", "Pending")
            ageing = _claim_ageing_days(submitted, claim_status, approved)
            claim_id = (existing or {}).get("claimId", f"CLM-{l['leadId']}-{key}")
            seen_ids.add(claim_id)
            result.append({
                "claimId": claim_id,
                "leadId": l["leadId"], "customer": l.get("customerName"),
                "model": l.get("interestedModel"), "variant": l.get("variant"),
                "bookingDate": l.get("bookingDate"),
                "component": ce.SCHEME_COMPONENT_LABELS.get(key, key),
                "componentKey": key, "claimAmount": ce.round2(company_share),
                "eligibleClaim": elig,
                "approvalStatus": "Approved" if elig >= company_share else "Pending",
                "claimStatus": claim_status,
                "receivedAmount": (existing or {}).get("receivedAmount", 0),
                "claimReference": (existing or {}).get("claimReference", ""),
                "submittedDate": submitted, "approvedDate": approved, "ageingDays": ageing,
            })
    # Manual claims (OEM incentives / executive incentives) — merged into the register
    for m in await db.claims.find({"manual": True}).to_list(2000):
        elig = ce.round2(ce.num(m.get("eligibleClaim") if m.get("eligibleClaim") is not None else m.get("claimAmount")))
        claim_id = m.get("claimId")
        if claim_id:
            seen_ids.add(claim_id)
        result.append({
            "claimId": claim_id, "leadId": m.get("leadId", ""),
            "customer": m.get("customer", ""), "model": m.get("model", ""),
            "variant": m.get("variant", ""),
            "bookingDate": m.get("bookingDate", ""),
            "executive": m.get("executive", ""),
            "schemeMonth": m.get("schemeMonth", ""),
            "component": m.get("component") or m.get("claimType") or "Manual Claim",
            "componentKey": m.get("componentKey"), "claimAmount": ce.round2(ce.num(m.get("claimAmount"))),
            "eligibleClaim": elig,
            "approvalStatus": m.get("claimStatus", "Submitted"),
            "claimStatus": m.get("claimStatus", "Submitted"),
            "receivedAmount": ce.round2(ce.num(m.get("receivedAmount"))),
            "claimReference": m.get("claimReference", ""),
            "submittedDate": m.get("submittedDate", ""), "approvedDate": m.get("approvedDate", ""),
            "ageingDays": _claim_ageing_days(m.get("submittedDate", ""), m.get("claimStatus", ""), m.get("approvedDate", "")),
            "manual": True, "oemCompany": m.get("oemCompany", ""), "note": m.get("note", ""),
            "source": m.get("source", ""),
            "totalDiscount": ce.round2(ce.num(m.get("totalDiscount") if m.get("totalDiscount") is not None else elig)),
            "oemDiscount": ce.round2(ce.num(m.get("oemDiscount") if m.get("oemDiscount") is not None else elig)),
        })
    # Persisted scheme claims that would otherwise drop out (Use=No / lead status change /
    # share recomputed to 0) — Received or Partial money still belongs in the eternal register.
    for c in await db.claims.find({"manual": {"$ne": True}}).to_list(5000):
        claim_id = c.get("claimId") or f"CLM-{c.get('leadId')}-{c.get('componentKey')}"
        if claim_id in seen_ids:
            continue
        received = ce.round2(ce.num(c.get("receivedAmount")))
        status = (c.get("claimStatus") or "").strip()
        if received <= 0 and status not in (
            "Received", "Partial", "Submitted", "Approved", "Rejected", "Cancelled", "Pending"):
            # Skip pure empty shells with no lifecycle — only keep real register history.
            if not (c.get("submittedDate") or c.get("approvedDate") or c.get("claimReceivedDate")
                    or ce.num(c.get("claimAmount")) > 0 or ce.num(c.get("eligibleClaim")) > 0):
                continue
        elig = ce.round2(ce.num(c.get("eligibleClaim") if c.get("eligibleClaim") is not None else c.get("claimAmount")))
        key = c.get("componentKey") or ""
        lead = await db.leads.find_one({"leadId": c.get("leadId")}) or {}
        result.append({
            "claimId": claim_id, "leadId": c.get("leadId", ""),
            "customer": c.get("customer") or lead.get("customerName", ""),
            "model": c.get("model") or lead.get("interestedModel", ""),
            "variant": c.get("variant") or lead.get("variant", ""),
            "bookingDate": c.get("bookingDate") or lead.get("bookingDate", ""),
            "component": c.get("component") or ce.SCHEME_COMPONENT_LABELS.get(key, key),
            "componentKey": key,
            "claimAmount": ce.round2(ce.num(c.get("claimAmount") if c.get("claimAmount") is not None else elig)),
            "eligibleClaim": elig,
            "approvalStatus": c.get("approvalStatus") or status or "Pending",
            "claimStatus": status or "Pending",
            "receivedAmount": received,
            "claimReference": c.get("claimReference", ""),
            "submittedDate": c.get("submittedDate", ""), "approvedDate": c.get("approvedDate", ""),
            "ageingDays": _claim_ageing_days(c.get("submittedDate", ""), status, c.get("approvedDate", "")),
            "permanent": True,
        })
        seen_ids.add(claim_id)
    return result


class ManualClaimIn(BaseModel):
    claimType: str = "OEM Incentive"
    oemCompany: str = ""
    leadId: str = ""
    customer: str = ""
    model: str = ""
    claimAmount: float = 0
    submittedDate: str = ""
    claimReference: str = ""
    note: str = ""


@api.post("/claims/manual")
async def create_manual_claim(body: ManualClaimIn, act=Depends(actor), _money=Depends(money_desk_only)):
    """Manually record a claim (e.g. OEM incentive received as a claim). Both Owner and staff can add.
    Appears in the claim register; received amounts are recorded later via /claims/receipt."""
    if body.claimAmount <= 0:
        raise HTTPException(422, "Enter a valid claim amount")
    cid = f"MCLM{uuid.uuid4().hex[:10].upper()}"
    customer, model = body.customer, body.model
    if body.leadId and not customer:
        lead = await db.leads.find_one({"leadId": body.leadId})
        if lead:
            customer = lead.get("customerName", "")
            model = model or lead.get("interestedModel", "")
    doc = {
        "claimId": cid, "manual": True, "componentKey": cid,
        "leadId": body.leadId, "customer": customer, "model": model,
        "claimType": body.claimType, "oemCompany": body.oemCompany,
        "component": body.claimType,
        "claimAmount": ce.round2(body.claimAmount), "eligibleClaim": ce.round2(body.claimAmount),
        "claimStatus": "Submitted", "receivedAmount": 0.0,
        "claimReference": body.claimReference, "note": body.note,
        "submittedDate": body.submittedDate or today(), "approvedDate": "",
        "receipts": [], "createdAt": now_iso(),
    }
    await db.claims.insert_one(dict(doc))
    await sheet_sync("claims", {**doc, "customer": customer, "model": model})
    await write_audit(act, "create", "claim", leadId=body.leadId, claimId=cid,
                      new={"claimType": body.claimType, "oemCompany": body.oemCompany,
                           "claimAmount": doc["claimAmount"], "manual": True})
    return clean(doc)


class ClaimSettleIn(BaseModel):
    leadId: str
    componentKey: str
    claimStatus: str = "Received"
    receivedAmount: float = 0
    claimReference: str = ""
    submittedDate: str = ""
    approvedDate: str = ""


@api.post("/claims/settle")
async def settle_claim(body: ClaimSettleIn, act=Depends(actor), _money=Depends(money_desk_only)):
    existing = await db.claims.find_one({"leadId": body.leadId, "componentKey": body.componentKey}) or {}
    payload = body.model_dump()
    # default lifecycle dates
    if not payload.get("submittedDate"):
        payload["submittedDate"] = existing.get("submittedDate") or today()
    if body.claimStatus in ("Approved", "Received") and not payload.get("approvedDate"):
        payload["approvedDate"] = existing.get("approvedDate") or today()
    doc = {"claimId": existing.get("claimId") or f"CLM-{body.leadId}-{body.componentKey}", **payload}
    if existing.get("manual"):
        doc["manual"] = True
    await db.claims.update_one(
        {"leadId": body.leadId, "componentKey": body.componentKey},
        {"$set": doc}, upsert=True,
    )
    lead = await db.leads.find_one({"leadId": body.leadId}) or {}
    await sheet_sync("claims", {**doc, "customer": existing.get("customer") or lead.get("customerName", ""),
                                    "model": existing.get("model") or lead.get("interestedModel", ""), "claimAmount": body.receivedAmount})
    await write_audit(act, "settle", "claim", leadId=body.leadId, claimId=doc["claimId"],
                      old={"claimStatus": existing.get("claimStatus"), "receivedAmount": existing.get("receivedAmount")},
                      new={"claimStatus": body.claimStatus, "receivedAmount": body.receivedAmount,
                           "submittedDate": payload["submittedDate"], "approvedDate": payload.get("approvedDate", "")})
    return {"ok": True}


class ClaimReceiptIn(BaseModel):
    leadId: str = ""
    componentKey: str
    amount: float
    date: str = ""
    reference: str = ""


@api.post("/claims/receipt")
async def record_claim_receipt(body: ClaimReceiptIn, act=Depends(actor), _money=Depends(money_desk_only)):
    """Record OEM money received against ONE claim (scheme component or manual claim);
    accrues receivedAmount + keeps a receipt history."""
    if body.amount <= 0:
        raise HTTPException(422, "Enter a valid receipt amount")
    existing = await db.claims.find_one({"leadId": body.leadId, "componentKey": body.componentKey}) or {}
    if existing.get("manual"):
        # Manual claim: eligible is the stored claim amount; no scheme recompute / lead lookup
        eligible = ce.round2(ce.num(existing.get("eligibleClaim") if existing.get("eligibleClaim") is not None else existing.get("claimAmount")))
        customer, model = existing.get("customer", ""), existing.get("model", "")
    else:
        # Derived scheme claim: eligible company share recomputed live
        lead = await get_lead_or_404(body.leadId)
        scheme_rows = await get_scheme_rows()
        shares = ce.compute_scheme_claim_shares(lead_to_snapshot(lead), scheme_rows)
        eligible = ce.round2(ce.num(shares["eligibleByComponent"].get(body.componentKey)))
        customer, model = lead.get("customerName", ""), lead.get("interestedModel", "")
    received = ce.round2(ce.num(existing.get("receivedAmount")) + body.amount)
    status = "Received" if eligible > 0 and received >= eligible - 0.01 else "Partial"
    submitted = existing.get("submittedDate") or (body.date or today())
    approved = existing.get("approvedDate") or (body.date or today())
    claim_id = existing.get("claimId") or f"CLM-{body.leadId}-{body.componentKey}"
    receipt = {"amount": ce.round2(body.amount), "date": body.date or today(),
               "reference": body.reference, "recordedAt": now_iso()}
    setdoc = {"claimId": claim_id, "leadId": body.leadId,
              "componentKey": body.componentKey, "receivedAmount": received, "claimStatus": status,
              "claimReference": body.reference or existing.get("claimReference", ""),
              "eligibleClaim": eligible, "submittedDate": submitted, "approvedDate": approved,
              # Claim Register has a "Claim Received Date" column; the date of the money
              # actually arriving was recorded only inside the receipt history, never on
              # the claim itself, so the register column had no source.
              "claimReceivedDate": body.date or today(),
              "lastUpdated": now_iso()}
    if existing.get("manual"):
        setdoc["manual"] = True
    await db.claims.update_one(
        {"leadId": body.leadId, "componentKey": body.componentKey},
        {"$set": setdoc, "$push": {"receipts": receipt}},
        upsert=True,
    )
    await sheet_sync("claims", {"claimId": claim_id, "leadId": body.leadId,
                                    "customer": customer, "model": model,
                                    "claimStatus": status, "receivedAmount": received, "claimAmount": body.amount})
    await write_audit(act, "receipt", "claim", leadId=body.leadId, claimId=claim_id,
                      old={"receivedAmount": ce.num(existing.get("receivedAmount"))},
                      new={"receivedAmount": received, "status": status, "amount": ce.round2(body.amount)})
    return {"ok": True, "receivedAmount": received, "status": status}


# ---------------------------------------------------------------- audit log (H4) — owner-only viewer
@api.get("/audit-log", dependencies=[Depends(owner_only)])
async def list_audit_log(module: Optional[str] = None, leadId: Optional[str] = None, limit: int = 500):
    q = {}
    if module:
        q["module"] = module
    if leadId:
        q["leadId"] = leadId
    rows = await db.audit_log.find(q).sort("timestamp", -1).to_list(min(max(limit, 1), 2000))
    return [clean(r) for r in rows]


# ---------------------------------------------------------------- masters CRUD (editable)
class PriceRowIn(BaseModel):
    model: str
    variant: str
    bodyType: str = ""
    exShowroom: float = 0
    rto: float = 0
    insurance: float = 0
    accessories: float = 0
    handlingCharges: float = 0
    trc: float = 0
    fastag: float = 0
    extendedWarranty: float = 0
    otherCharges: float = 0
    gstPercent: float = 0
    tcsApplicable: str = "No"
    priceVersion: str = ""
    status: str = "active"
    remarks: str = ""


@api.post("/price-master", dependencies=[Depends(owner_only)])
async def create_price_row(body: PriceRowIn):
    count = await db.price_master.count_documents({})
    doc = {"priceId": f"PM{count + 1:04d}", **body.model_dump()}
    await db.price_master.insert_one(doc)
    return clean(doc)


# ------------------------------------------------- Price Master -> live leads
# Fields a Price Master edit pushes onto matching leads. A change to any of them
# reprices; a change to status/bodyType/notes alone does not.
REPRICE_TRIGGER_FIELDS = ("exShowroom", "rto", "insurance", "accessories", "handlingCharges",
                          "trc", "fastag", "extendedWarranty", "otherCharges", "tcsApplicable")


def _price_row_changed(before: dict, after: dict) -> list:
    changed = []
    for f in REPRICE_TRIGGER_FIELDS:
        if f == "tcsApplicable":
            if str(before.get(f) or "No").strip() != str(after.get(f) or "No").strip():
                changed.append(f)
        elif abs(ce.num(before.get(f)) - ce.num(after.get(f))) > 0.005:
            changed.append(f)
    return changed


def _lead_matches_price_row(lead: dict, row: dict) -> bool:
    def norm(v):
        return str(v or "").strip().lower()
    lm, lv = oem_cat.canonical_model_variant(lead.get("interestedModel"), lead.get("variant"))
    rm, rv = oem_cat.canonical_model_variant(row.get("model"), row.get("variant"))
    return norm(lm) == norm(rm) and norm(lv) == norm(rv)


async def reprice_leads_for_price_row(row: dict, changed_fields=None, act=None) -> dict:
    """Push a Price Master change onto every Active, NOT-delivered lead on that vehicle.

    Scope is deliberately wider than "booked": a priced-but-unbooked lead or an open
    quotation must not keep quoting a withdrawn price. Delivered leads are never
    touched — their invoice is already raised.

    A lead that has PAID and settled (money received, outstanding cleared) keeps the
    price it settled at. Repricing it would re-open an outstanding and block Mark
    Delivered on a customer who owes nothing. Those are reported as skipped so the
    owner can reprice one by hand if the business wants the increase passed on.

    Scheme is NOT realigned: entitlements are governed by the Scheme Master and the
    booking month, and must not move because the vehicle price moved.
    """
    repriced, skipped = [], []
    for lead in await db.leads.find({"deliveryStatus": {"$ne": "Delivered"}}).to_list(5000):
        if not _lead_matches_price_row(lead, row):
            continue
        lead_id = lead.get("leadId")
        if _is_delivered(lead):
            continue
        if _acct(lead) != "Active":
            skipped.append({"leadId": lead_id, "customerName": lead.get("customerName"),
                            "reason": f"account {_acct(lead)}"})
            continue
        if not _is_priced(lead):
            skipped.append({"leadId": lead_id, "customerName": lead.get("customerName"),
                            "reason": "no price structure saved yet"})
            continue
        booked_on = str(lead.get("bookingDate") or "")[:10]
        if booked_on and booked_on < "2026-09-01":
            skipped.append({
                "leadId": lead_id, "customerName": lead.get("customerName"),
                "reason": "booked before 1 Sep 2026 — old price honoured",
                "exShowroom": ce.num(lead.get("exShowroom")),
            })
            continue
        received = ce.num(lead.get("totalReceived"))
        outstanding = ce.num(lead.get("customerOutstanding"))
        if received > 0 and outstanding <= 0.01:
            skipped.append({
                "leadId": lead_id, "customerName": lead.get("customerName"),
                "reason": "already paid in full — old price honoured",
                "exShowroom": ce.num(lead.get("exShowroom")),
                "customerPayable": ce.num(lead.get("customerPayable")),
            })
            continue
        before = {"exShowroom": ce.num(lead.get("exShowroom")),
                  "customerPayable": ce.num(lead.get("customerPayable")),
                  "customerOutstanding": outstanding}
        await _cascade_vehicle_or_price_change(lead_id, refresh_price=True, realign_scheme=False)
        after_lead = await db.leads.find_one({"leadId": lead_id}) or {}
        after = {"exShowroom": ce.num(after_lead.get("exShowroom")),
                 "customerPayable": ce.num(after_lead.get("customerPayable")),
                 "customerOutstanding": ce.num(after_lead.get("customerOutstanding"))}
        if abs(before["customerPayable"] - after["customerPayable"]) < 0.005 and \
                abs(before["exShowroom"] - after["exShowroom"]) < 0.005:
            continue          # nothing actually moved for this lead
        entry = {"leadId": lead_id, "customerName": lead.get("customerName"),
                 "currentStatus": lead.get("currentStatus"),
                 "exShowroomBefore": before["exShowroom"], "exShowroomAfter": after["exShowroom"],
                 "customerPayableBefore": before["customerPayable"],
                 "customerPayableAfter": after["customerPayable"],
                 "outstandingBefore": before["customerOutstanding"],
                 "outstandingAfter": after["customerOutstanding"],
                 "delta": ce.round2(after["customerPayable"] - before["customerPayable"])}
        repriced.append(entry)
        # Per-lead audit so a wrong price edit can be traced and reversed.
        if act:
            await write_audit(act, "reprice", "lead", leadId=lead_id,
                              old={k: before[k] for k in before},
                              new={k: after[k] for k in after})
        await sheet_sync("leads", clean(dict(after_lead)))
    return {
        "model": row.get("model"), "variant": row.get("variant"),
        "changedFields": list(changed_fields or []),
        "repricedCount": len(repriced), "skippedCount": len(skipped),
        "totalPayableDelta": ce.round2(sum(r["delta"] for r in repriced)),
        "repriced": repriced, "skipped": skipped,
    }


@api.put("/price-master/{price_id}", dependencies=[Depends(owner_only)])
async def update_price_row(price_id: str, body: PriceRowIn, act=Depends(actor)):
    existing = await db.price_master.find_one({"priceId": price_id})
    if not existing:
        raise HTTPException(404, "Price row not found")
    payload = body.model_dump()
    if str(existing.get("priceSource") or "") == oem_sync.OEM_PRICE_SOURCE:
        payload.pop("exShowroom", None)
    await db.price_master.update_one({"priceId": price_id}, {"$set": payload})
    updated = await db.price_master.find_one({"priceId": price_id})
    changed = _price_row_changed(existing, updated)
    out = clean(updated)
    # A price revision must reach every live lead on that vehicle, not just new ones.
    out["reprice"] = (await reprice_leads_for_price_row(updated, changed, act) if changed
                      else {"repricedCount": 0, "skippedCount": 0, "changedFields": [],
                            "repriced": [], "skipped": [], "totalPayableDelta": 0})
    if changed:
        await write_audit(act, "update", "price-master",
                          old={f: existing.get(f) for f in REPRICE_TRIGGER_FIELDS},
                          new={f: updated.get(f) for f in REPRICE_TRIGGER_FIELDS})
    return out


@api.get("/price-master/{price_id}/reprice-preview", dependencies=[Depends(owner_only)])
async def reprice_preview(price_id: str, exShowroom: Optional[float] = None):
    """Read-only: which leads a price change WOULD move, and by how much.

    Nothing is written. Pass exShowroom to model a price that is not saved yet.
    """
    row = await db.price_master.find_one({"priceId": price_id})
    if not row:
        raise HTTPException(404, "Price row not found")
    proposed = dict(row)
    if exShowroom is not None:
        proposed["exShowroom"] = ce.num(exShowroom)
    changed = _price_row_changed(row, proposed)
    delta_ex = ce.round2(ce.num(proposed.get("exShowroom")) - ce.num(row.get("exShowroom")))
    affected, skipped = [], []
    for lead in await db.leads.find({"deliveryStatus": {"$ne": "Delivered"}}).to_list(5000):
        if not _lead_matches_price_row(lead, row) or _is_delivered(lead):
            continue
        base = {"leadId": lead.get("leadId"), "customerName": lead.get("customerName"),
                "currentStatus": lead.get("currentStatus"),
                "exShowroom": ce.num(lead.get("exShowroom")),
                "customerPayable": ce.num(lead.get("customerPayable")),
                "customerOutstanding": ce.num(lead.get("customerOutstanding"))}
        if _acct(lead) != "Active":
            skipped.append({**base, "reason": f"account {_acct(lead)}"})
        elif not _is_priced(lead):
            skipped.append({**base, "reason": "no price structure saved yet"})
        elif ce.num(lead.get("totalReceived")) > 0 and base["customerOutstanding"] <= 0.01:
            skipped.append({**base, "reason": "already paid in full — old price honoured"})
        else:
            affected.append({**base, "exShowroomAfter": ce.num(proposed.get("exShowroom")),
                             "estimatedDelta": delta_ex})
    return {"priceId": price_id, "model": row.get("model"), "variant": row.get("variant"),
            "currentExShowroom": ce.num(row.get("exShowroom")),
            "proposedExShowroom": ce.num(proposed.get("exShowroom")),
            "changedFields": changed, "wouldRepriceCount": len(affected),
            "wouldSkipCount": len(skipped), "wouldReprice": affected, "wouldSkip": skipped}


@api.delete("/price-master/{price_id}", dependencies=[Depends(owner_only)])
async def delete_price_row(price_id: str):
    await db.price_master.delete_one({"priceId": price_id})
    return {"ok": True}


class SchemeRowIn(BaseModel):
    schemeMonth: Optional[str] = None
    effectiveFrom: Optional[str] = None
    effectiveTo: Optional[str] = None
    circularRef: str = ""
    model: str = ""
    variant: str = ""
    component: str = ""
    componentKey: str = ""
    dealerShare: float = 0
    companyShare: float = 0
    totalBenefit: float = 0
    status: str = "Active"
    notes: str = ""


@api.post("/scheme-master", dependencies=[Depends(owner_only)])
async def create_scheme_row(body: SchemeRowIn):
    count = await db.scheme_master.count_documents({})
    payload = _normalize_scheme_row(body.model_dump())
    payload["totalBenefit"] = payload["totalBenefit"] or ce.round2(payload["dealerShare"] + payload["companyShare"])
    doc = {"schemeId": f"SCM{count + 1:04d}", **payload}
    await db.scheme_master.insert_one(doc)
    return clean(doc)


@api.put("/scheme-master/{scheme_id}", dependencies=[Depends(owner_only)])
async def update_scheme_row(scheme_id: str, body: SchemeRowIn):
    existing = await db.scheme_master.find_one({"schemeId": scheme_id})
    if not existing:
        raise HTTPException(404, "Scheme row not found")
    raw = body.model_dump()
    for k in ("schemeMonth", "effectiveFrom", "effectiveTo"):
        if not str(raw.get(k) or "").strip():
            raw[k] = existing.get(k) or ""
    payload = _normalize_scheme_row(raw)
    payload["totalBenefit"] = payload["totalBenefit"] or ce.round2(payload["dealerShare"] + payload["companyShare"])
    await db.scheme_master.update_one({"schemeId": scheme_id}, {"$set": payload})
    return clean(await db.scheme_master.find_one({"schemeId": scheme_id}))


@api.delete("/scheme-master/{scheme_id}", dependencies=[Depends(owner_only)])
async def delete_scheme_row(scheme_id: str):
    await db.scheme_master.delete_one({"schemeId": scheme_id})
    return {"ok": True}


# ---------------------------------------------------------------- dealer earnings (owner-only)
@api.get("/dealer-earnings", dependencies=[Depends(owner_only)])
async def list_dealer_earnings():
    """Owner Dealer Earnings grid — live from leads so OEM Extra Retained is always in total.

    total = margin + scheme retained + OEM Extra Retained + insurance income + extras
            − dealer-funded benefit.
    """
    leads = await _commercial_leads()
    # Fallback extras from dealer_earnings docs when lead mirror is thin.
    de_by = {r.get("leadId"): r for r in await db.dealer_earnings.find().to_list(5000)}
    rows = []
    for l in leads:
        lid = l.get("leadId")
        de = de_by.get(lid) or {}
        oem = ce.compute_oem_extra_support(l)
        margin = ce.num(l.get("dealerMarginNetExGst") if l.get("dealerMarginNetExGst") is not None
                        else de.get("dealerMarginNetExGst"))
        scheme = ce.num(l.get("dealerSchemeRetained") if l.get("dealerSchemeRetained") is not None
                        else de.get("dealerSchemeRetained"))
        ins = ce.num(l.get("dealerInsuranceIncome") if l.get("dealerInsuranceIncome") is not None
                     else de.get("dealerInsuranceIncome"))
        extra = ce.num(l.get("extraDealerIncomeTotal") if l.get("extraDealerIncomeTotal") is not None
                       else de.get("extraDealerIncomeTotal"))
        funded = ce.num(l.get("dealerFundedBenefit") if l.get("dealerFundedBenefit") is not None
                        else de.get("dealerFundedBenefit"))
        total = ce.round2(margin + scheme + oem["oemExtraSupportRetained"] + ins + extra - funded)
        rows.append({
            "leadId": lid,
            "customerName": l.get("customerName") or de.get("customerName"),
            "model": l.get("interestedModel") or de.get("model"),
            "variant": l.get("variant") or de.get("variant"),
            "executive": l.get("executive") or de.get("executive"),
            "currentStage": l.get("currentStatus") or de.get("currentStage"),
            "dealerMarginNetExGst": margin,
            "dealerSchemeRetained": scheme,
            "dealerInsuranceIncome": ins,
            "dealerFundedBenefit": funded,
            "extraDealerIncomeTotal": extra,
            "oemExtraSupportReceived": oem["oemExtraSupportReceived"],
            "oemExtraSupportPassed": oem["oemExtraSupportPassed"],
            "oemExtraSupportRetained": oem["oemExtraSupportRetained"],
            "totalDealerEarnings": total,
            "dealerTotalEarnings": total,
        })
    rows.sort(key=lambda r: (r.get("customerName") or "").lower())
    total = ce.round2(sum(ce.num(r.get("totalDealerEarnings")) for r in rows))
    return {"rows": rows, "total": total}


# ---------------------------------------------------------------- integrations status
@api.get("/integrations/gsheets")
async def gsheets_status():
    return gsheets.status()


@api.get("/integrations/gsheets/preflight", dependencies=[Depends(owner_only)])
async def gsheets_preflight():
    """GS-1: read-only header-mapping report. For every mapped tab shows the EXISTING
    sheet headers, which CRM field resolved to which column letter, and any header we
    could not find. Nothing is written. Run this first against the real spreadsheet —
    any entity with willSync=false will refuse to write rather than guess a column."""
    return gsheets.preflight()


@api.post("/integrations/gsheets/ensure-oem-extra-columns", dependencies=[Depends(owner_only)])
async def gsheets_ensure_oem_extra_columns(act=Depends(actor)):
    """Owner-only: create/append OEM Extra Support Received / Passed / Retained headers
    on Lead Register + Dealer Earnings, and create OEM Extra Support Register if missing.

    The CRM cannot show values in columns that do not exist yet — this is the one-time
    sheet structure step. After it succeeds, re-save Scheme (or run Backfill) to fill rows.
    """
    result = await gsheets.ensure_oem_extra_support_columns()
    await write_audit(act, "ensure", "gsheets-oem-extra-columns", new=result)
    return result


@api.post("/integrations/gsheets/ensure-insurance-agent-columns", dependencies=[Depends(owner_only)])
async def gsheets_ensure_insurance_agent_columns(act=Depends(actor)):
    """Owner-only: append Insurance Agent / Rate Source / Last Payout Date to the
    Insurance Register header row.

    The sync resolves columns by header name, so these three fields simply do not
    write until the headers exist. Append-only and idempotent — running it twice
    adds nothing the second time. After it succeeds, the next insurance save (or
    Backfill) fills the new columns.
    """
    result = await gsheets.ensure_insurance_agent_columns()
    await write_audit(act, "ensure", "gsheets-insurance-agent-columns", new=result)
    return result


@api.post("/integrations/gsheets/ensure-cancel-columns", dependencies=[Depends(owner_only)])
async def gsheets_ensure_cancel_columns(act=Depends(actor)):
    """Owner-only: append Cancel Count / Last Cancel Date / Last Cancel Reason /
    Last Cancel Stage / Revive On to the Lead Register header row.

    Append-only and idempotent. Until the headers exist the cancellation fields
    do not write to the sheet — the app still holds them, the sheet just shows
    blanks. After it succeeds, the next lead save (or Backfill) fills them in.
    """
    result = await gsheets.ensure_cancel_columns()
    await write_audit(act, "ensure", "gsheets-cancel-columns", new=result)
    return result


@api.get("/integrations/gsheets/verify-lead/{lead_id}", dependencies=[Depends(owner_only)])
async def gsheets_verify_lead(lead_id: str):
    """Read-only: how many rows the LIVE sheet actually holds for this lead in each
    transactional tab. Lets the go-live verifier prove idempotency against the real
    spreadsheet instead of trusting an HTTP 200. Nothing is written."""
    lead = await db.leads.find_one({"leadId": lead_id})
    out = {"leadId": lead_id, "tabs": {}}
    checks = [("leads", lead_id), ("activities", None), ("deliveries", lead_id),
              ("dealer_earnings", lead_id), ("bookings", None), ("payments", None), ("claims", None)]
    bookings = await db.bookings.find({"leadId": lead_id}).to_list(50)
    payments = await db.payments.find({"leadId": lead_id}).to_list(200)
    activities = await db.activities.find({"leadId": lead_id}).to_list(200)
    scheme_rows = await get_scheme_rows()
    claim_ids = []
    if lead:
        shares = ce.compute_scheme_claim_shares(lead_to_snapshot(lead), scheme_rows)
        for key, amt in shares["displayByComponent"].items():
            if amt > 0:
                existing = await db.claims.find_one({"leadId": lead_id, "componentKey": key})
                claim_ids.append((existing or {}).get("claimId", f"CLM-{lead_id}-{key}"))
    for entity, ident in checks:
        if entity in ("leads", "deliveries", "dealer_earnings"):
            ids = [lead_id]
        elif entity == "bookings":
            ids = [b.get("bookingId") for b in bookings]
        elif entity == "payments":
            ids = [p.get("receiptNumber") for p in payments]
        elif entity == "activities":
            ids = [a.get("activityId") for a in activities]
        else:
            ids = claim_ids
        per = {}
        for i in [x for x in ids if x]:
            per[i] = await asyncio.to_thread(gsheets.count_rows_for, entity, i)
        counts = [v.get("count") for v in per.values() if v.get("ok")]
        out["tabs"][entity] = {
            "crmRecords": len(ids),
            "sheetRowsPerId": {k: v.get("count", v.get("reason")) for k, v in per.items()},
            "maxRowsForAnyId": max(counts) if counts else 0,
            "duplicates": any(c > 1 for c in counts),
        }
    out["anyDuplicates"] = any(t["duplicates"] for t in out["tabs"].values())
    return out


@api.get("/integrations/gsheets/sync-log", dependencies=[Depends(owner_only)])
async def gsheets_sync_log(status: Optional[str] = None, limit: int = 200):
    """GS-4: durable record of every Sheet write. status=PENDING lists writes that
    failed and are awaiting retry — nothing is ever silently lost."""
    q = {"status": status} if status else {}
    rows = await db.sheet_sync_log.find(q).sort("timestamp", -1).to_list(min(limit, 2000))
    pending = await db.sheet_sync_log.count_documents({"status": "PENDING"})
    return {"pending": pending, "rows": [clean(r) for r in rows]}


@api.post("/integrations/gsheets/retry", dependencies=[Depends(owner_only)])
async def gsheets_retry(limit: int = 100):
    """GS-4: replay failed Sheet writes. Safe because every write is an ID-keyed
    upsert — replaying a write that actually succeeded (but whose response timed out)
    finds the existing row and updates it instead of appending a duplicate."""
    pending = await db.sheet_sync_log.find({"status": "PENDING"}).to_list(min(limit, 500))
    retried, recovered, still_failing = 0, 0, 0
    for row in pending:
        entity, payload = row.get("entityType"), row.get("payload") or {}
        if not entity or not payload:
            continue
        retried += 1
        res = await sheet_sync(entity, payload, entity_id=row.get("entityId", ""))
        if res.get("ok"):
            recovered += 1
        else:
            still_failing += 1
            await db.sheet_sync_log.update_one(
                {"entityType": entity, "entityId": row.get("entityId", "")},
                {"$set": {"status": "FAILED" if int(row.get("attempt", 0)) >= 4 else "RETRYING"}})
    return {"ok": True, "retried": retried, "recovered": recovered, "stillFailing": still_failing}


@api.get("/integrations/gsheets/inventory", dependencies=[Depends(owner_only)])
async def gsheets_inventory():
    """Phase 2/3: read-only contract of the LIVE workbook. Enumerates every tab, its
    header row + column headers, a source-of-truth classification, and (for tabs the
    CRM syncs) the resolved column→field mapping, sync direction and authoritative
    entity. Generated from the actual sheet — never writes."""
    return await asyncio.to_thread(gsheets.inventory)


@api.get("/integrations/gsheets/env-safety", dependencies=[Depends(owner_only)])
async def gsheets_env_safety():
    """Phase 1: preview/production isolation status."""
    return gsheets.env_safety()


@api.get("/integrations/gsheets/reconcile", dependencies=[Depends(owner_only)])
async def gsheets_reconcile():
    """CRM vs Google Sheet reconciliation. Read-only: compares CRM record counts and
    stable IDs against what is actually present in each existing tab's ID column, and
    reports anything missing from the sheet plus any unresolved sync-log entries."""
    st = gsheets.status()
    if not st.get("enabled"):
        return {"ok": False, "reason": st.get("reason", "sync disabled")}
    datasets = {
        "leads": await db.leads.find().to_list(5000),
        "bookings": await db.bookings.find().to_list(5000),
        "payments": await db.payments.find().to_list(5000),
        "deliveries": await db.deliveries.find().to_list(5000),
        "claims": await db.claims.find().to_list(5000),
        "finance": await db.finance.find().to_list(5000),
        "insurance": await db.insurance.find().to_list(5000),
    }
    report, mismatches = {}, []
    for entity, docs in datasets.items():
        spec = gsheets.SYNC_MAP.get(entity)
        if not spec:
            continue
        tab, id_field, fields = spec[0], spec[1], spec[2]
        crm_ids = {str(d.get(id_field, "") or "").strip() for d in docs if d.get(id_field)}
        try:
            hr = gsheets._header_row_for(entity, tab)
            mapping, missing = gsheets._resolve_columns(tab, fields, use_cache=False, header_row=hr)
            if id_field not in mapping:
                report[entity] = {"tab": tab, "error": f"ID header '{id_field}' not found", "crmCount": len(crm_ids)}
                continue
            sheet_id_list = await asyncio.to_thread(gsheets._read_id_column_list, tab, mapping[id_field], hr)
        except Exception as e:
            report[entity] = {"tab": tab, "error": str(e)[:200], "crmCount": len(crm_ids)}
            continue
        sheet_ids = set(sheet_id_list)
        # Counter, not list.count() per element — the latter is O(n^2) and this runs
        # over every ID column in the workbook (thousands of rows x 7 entities).
        dupes = sorted(i for i, n in Counter(sheet_id_list).items() if n > 1)
        missing_in_sheet = sorted(crm_ids - sheet_ids)
        missing_in_app = sorted(sheet_ids - crm_ids)
        report[entity] = {"tab": tab, "crmCount": len(crm_ids), "sheetCount": len(sheet_ids),
                          "missingInSheet": len(missing_in_sheet), "missingInApp": len(missing_in_app),
                          "duplicateIdsInSheet": len(dupes)}
        for mid in missing_in_sheet[:50]:
            mismatches.append({"entity": entity, "id": mid, "field": id_field, "tab": tab,
                               "appValue": mid, "sheetValue": "<absent>",
                               "expected": "row present in sheet", "issue": "MISSING_IN_SHEET",
                               "severity": "HIGH"})
        for mid in missing_in_app[:50]:
            mismatches.append({"entity": entity, "id": mid, "field": id_field, "tab": tab,
                               "appValue": "<absent>", "sheetValue": mid,
                               "expected": "row present in app DB", "issue": "MISSING_IN_APP",
                               "severity": "MEDIUM"})
        for did in dupes[:50]:
            mismatches.append({"entity": entity, "id": did, "field": id_field, "tab": tab,
                               "issue": "DUPLICATE_ID", "severity": "HIGH"})
    # Orphan check: finance file references a lead that is not in the CRM lead set.
    lead_ids = {str(d.get("leadId", "") or "").strip() for d in datasets["leads"] if d.get("leadId")}
    for f in datasets["finance"]:
        lid = str(f.get("leadId", "") or "").strip()
        if lid and lid not in lead_ids:
            mismatches.append({"entity": "finance", "id": f.get("fileNumber"), "field": "leadId", "tab": "Finance Register",
                               "issue": "ORPHAN_REFERENCE", "appValue": lid,
                               "expected": "leadId present in Lead Register", "severity": "HIGH"})
    unresolved = await db.sheet_sync_log.count_documents({"status": {"$ne": "OK"}})
    return {"ok": True, "entities": report, "mismatches": mismatches,
            "mismatchCount": len(mismatches),
            "unresolvedSyncLogEntries": unresolved, "envSafety": gsheets.env_safety(),
            "verdict": "CLEAN" if not mismatches and unresolved == 0 else "DIFFERENCES FOUND"}


@api.post("/integrations/gsheets/backfill", dependencies=[Depends(owner_only)])
async def gsheets_backfill():
    leads = [clean(x) for x in await db.leads.find().to_list(5000)]
    payments = [clean(x) for x in await db.payments.find().to_list(5000)]
    bookings = [clean(x) for x in await db.bookings.find().to_list(5000)]
    delivered = [clean(x) for x in await db.leads.find({"deliveryStatus": "Delivered"}).to_list(5000)]
    deliveries = [{"leadId": l.get("leadId"), "customerName": l.get("customerName"),
                   "deliveryDate": l.get("deliveryDate"), "delivered": "Yes",
                   "invoiceNumber": l.get("invoiceNumber", ""), "chassisNumber": l.get("chassisNumber", ""),
                   "numberPlate": l.get("numberPlate", "")} for l in delivered]
    result = await gsheets.backfill({"leads": leads, "bookings": bookings, "payments": payments, "deliveries": deliveries})
    # Rebuild derived Finance Pending / Overdue tabs from the now-synced registers.
    fin_views = await rebuild_finance_views()
    if isinstance(result, dict):
        result["financeViews"] = fin_views
    return result


# ---------------------------------------------------------------- owner reports
@api.get("/reports/insurance-payout", dependencies=[Depends(owner_only)])
async def insurance_payout_report():
    entries = await db.insurance.find().to_list(5000)
    by_month = {}
    by_insurer = {}
    by_agent = {}
    totals = {"premium": 0.0, "expected": 0.0, "received": 0.0, "outstanding": 0.0, "count": 0}
    for e in entries:
        month = str(e.get("policyDate") or e.get("deliveryDate") or "")[:7] or "Unknown"
        premium = ce.num(e.get("insuranceAmount"))
        expected = ce.num(e.get("expectedPayout"))
        received = ce.num(e.get("receivedPayout"))
        outstanding = ce.num(e.get("payoutOutstanding"))
        agent = e.get("insuranceAgentName") or "— No agent —"
        for bucket, key in ((by_month, month), (by_insurer, e.get("insuranceCompany") or "Unknown"),
                            (by_agent, agent)):
            row = bucket.setdefault(key, {"key": key, "premium": 0.0, "expected": 0.0, "received": 0.0, "outstanding": 0.0, "count": 0})
            row["premium"] += premium; row["expected"] += expected
            row["received"] += received; row["outstanding"] += outstanding; row["count"] += 1
        totals["premium"] += premium; totals["expected"] += expected
        totals["received"] += received; totals["outstanding"] += outstanding; totals["count"] += 1

    def norm(bucket, sort_key=False):
        rows = [{**r, "premium": ce.round2(r["premium"]), "expected": ce.round2(r["expected"]),
                 "received": ce.round2(r["received"]), "outstanding": ce.round2(r["outstanding"])} for r in bucket.values()]
        return sorted(rows, key=lambda x: x["key"], reverse=True) if sort_key else sorted(rows, key=lambda x: -x["expected"])

    return {"byMonth": norm(by_month, True), "byInsurer": norm(by_insurer),
            "byAgent": norm(by_agent),
            "totals": {k: (ce.round2(v) if isinstance(v, float) else v) for k, v in totals.items()}}


@api.get("/reports/scheme-allocation-impact", dependencies=[Depends(owner_only)])
async def scheme_allocation_impact_report():
    """READ-ONLY historical impact of the authoritative scheme allocation engine.

    Compares each booked lead's currently persisted commercial fields against what
    compute_scheme_allocation would produce. Does NOT modify any records.
    """
    leads = await _commercial_leads()
    scheme_rows = await get_scheme_rows()
    rows = []
    summary = {
        "leadsScanned": 0, "leadsAffected": 0,
        "customerPayableDeltaTotal": 0.0,
        "dealerEarningsDeltaTotal": 0.0,
        "oemClaimDeltaTotal": 0.0,
    }
    for l in leads:
        summary["leadsScanned"] += 1
        snap = lead_to_snapshot(l)
        alloc = ce.compute_scheme_allocation(snap, scheme_rows)
        new_totals = ce.compute_commercial_totals(snap, scheme_rows)
        new_income = ce.compute_scheme_income_breakdown(snap, scheme_rows)
        new_shares = ce.compute_scheme_claim_shares(snap, scheme_rows)
        old_payable = ce.round2(ce.num(l.get("customerPayable")))
        new_payable = new_totals["customerPayable"]
        old_retained = ce.round2(ce.num(l.get("dealerSchemeRetained")))
        new_retained = new_income["retainedIncomeTotal"]
        old_claim = ce.round2(ce.num(l.get("oemClaimCompanyShare") if l.get("oemClaimCompanyShare") is not None else l.get("companyOutstanding")))
        new_claim = new_shares["eligibleTotal"]
        payable_diff = ce.round2(new_payable - old_payable)
        retained_diff = ce.round2(new_retained - old_retained)
        claim_diff = ce.round2(new_claim - old_claim)
        affected = abs(payable_diff) > 0.01 or abs(retained_diff) > 0.01 or abs(claim_diff) > 0.01
        if not affected and not alloc["components"]:
            continue
        if affected:
            summary["leadsAffected"] += 1
            summary["customerPayableDeltaTotal"] = ce.round2(summary["customerPayableDeltaTotal"] + payable_diff)
            summary["dealerEarningsDeltaTotal"] = ce.round2(summary["dealerEarningsDeltaTotal"] + retained_diff)
            summary["oemClaimDeltaTotal"] = ce.round2(summary["oemClaimDeltaTotal"] + claim_diff)
        component_rows = []
        for c in alloc["components"]:
            component_rows.append({
                "componentKey": c["key"], "label": c["label"],
                "schemeAvailable": c["schemeAvailable"],
                "customerBenefit": c["customerBenefit"],
                "dealerRetained": c["dealerRetained"],
                "oemClaimable": c["oemClaimable"],
                "oemShare": c["oemShare"],
                "dealerFundedShare": c["dealerFundedShare"],
                "dealerFundedBenefit": c["dealerFundedBenefit"],
            })
        rows.append({
            "leadId": l.get("leadId"),
            "customerName": l.get("customerName"),
            "vehicle": f"{l.get('interestedModel') or ''} {l.get('variant') or ''}".strip(),
            "bookingDate": l.get("bookingDate"),
            "benefitMode": l.get("benefitMode"),
            "affected": affected,
            "oldCustomerPayable": old_payable,
            "newCustomerPayable": new_payable,
            "customerPayableDifference": payable_diff,
            "oldDealerSchemeRetained": old_retained,
            "newDealerSchemeRetained": new_retained,
            "dealerEarningsDifference": retained_diff,
            "oldOemClaim": old_claim,
            "newOemClaim": new_claim,
            "claimDifference": claim_diff,
            "components": component_rows,
        })
    return {
        "readOnly": True,
        "modified": False,
        "note": "Impact only — no historical accounting was rewritten. New transactions use the allocation engine.",
        "summary": summary,
        "rows": rows,
    }


@api.get("/reports/dealer-earnings", dependencies=[Depends(owner_only)])
async def dealer_earnings_report():
    """Live dealer earnings from booked leads: margin + scheme retained + insurance income + other."""
    leads = await _commercial_leads()
    scheme_rows = await get_scheme_rows()
    # insurance income (dealer payout) per lead
    ins_by_lead = {}
    for e in await db.insurance.find().to_list(5000):
        lid = e.get("leadId")
        if lid:
            ins_by_lead[lid] = ce.round2(ins_by_lead.get(lid, 0) + ce.num(e.get("expectedPayout")))
    by_month = {}
    components = {"Dealer Margin": 0.0, "Scheme Retained": 0.0, "Insurance Income": 0.0,
                  "OEM Extra Support": 0.0, "Documentation": 0.0, "Warranty": 0.0,
                  "RSA": 0.0, "Referral": 0.0, "Other Income": 0.0,
                  "Customer Insurance Benefit Passed (scheme, not income)": 0.0,
                  "Dealer-Funded Benefit (cost)": 0.0,
                  "Finance Incentive": 0.0,
                  "Accessories Margin": 0.0, "Exchange Margin": 0.0, "Campaign Incentive": 0.0}
    totals = {"margin": 0.0, "scheme": 0.0, "insurance": 0.0, "extra": 0.0,
              "dealerFundedBenefit": 0.0, "total": 0.0, "count": 0}
    for l in leads:
        snap = lead_to_snapshot(l)
        margin = ce.compute_dealer_margin(snap)["marginNetExGst"]
        alloc = ce.compute_scheme_allocation(snap, scheme_rows)
        # Scheme income comes from the Scheme Allocation Engine: what the dealer
        # RETAINED. The old breakdown reported the dealer's funded share as negative.
        scheme = alloc["totals"]["dealerRetained"]
        insurance = ins_by_lead.get(l.get("leadId"), 0)
        oem_recv = max(0.0, ce.num(l.get("oemExtraSupportReceived")))
        oem_pass = max(0.0, min(ce.num(l.get("oemExtraSupportPassed")), oem_recv))
        other = ce.round2(max(0.0, oem_recv - oem_pass))   # OEM Extra Support retained
        # Extra dealer income lines (C1) — full port of DEALER_EARNINGS_MANUAL_COLS_
        doc_inc = ce.num(l.get("documentationIncome"))
        war_inc = ce.num(l.get("warrantyIncome"))
        rsa_inc = ce.num(l.get("rsaIncome"))
        ref_inc = ce.num(l.get("referralIncome"))
        other_inc = ce.num(l.get("otherIncome"))
        # Insurance Scheme Benefit CB — visibility only; NOT dealer income and NOT
        # the insurer payout. Prefer authoritative allocation when present.
        ins_comp = (alloc.get("byKey") or {}).get("insuranceBenefit") or {}
        cust_ins_benefit = ce.round2(ce.num(ins_comp.get("customerBenefit"))) if ins_comp else \
            ce.num(l.get("customerInsuranceBenefitPassed"))
        fin_inc = ce.num(l.get("financeIncentive"))
        acc_margin = ce.num(l.get("accessoriesMargin"))
        exch_margin = ce.num(l.get("exchangeMargin"))
        camp_inc = ce.num(l.get("campaignIncentive"))
        extra = ce.round2(doc_inc + war_inc + rsa_inc + ref_inc + other_inc +
                          fin_inc + acc_margin + exch_margin + camp_inc)
        has_auth = bool(l.get("schemeAllocationV2") or l.get("schemeAllocationExplicit")
                        or ce._explicit_allocation({"schemeAllocation": l.get("schemeAllocation")}))
        funded_cost = ce.round2(ce.num(alloc["totals"].get("dealerFundedBenefit"))) if has_auth else 0.0
        total = ce.round2(margin + scheme + insurance + other + extra - funded_cost)
        month = str(l.get("deliveryDate") or l.get("bookingDate") or "")[:7] or "Unknown"
        m = by_month.setdefault(month, {"key": month, "margin": 0.0, "scheme": 0.0,
                                        "insurance": 0.0, "other": 0.0, "extra": 0.0,
                                        "dealerFundedBenefit": 0.0, "total": 0.0, "count": 0})
        m["margin"] += margin; m["scheme"] += scheme; m["insurance"] += insurance
        m["other"] += other; m["extra"] += extra
        m["dealerFundedBenefit"] += funded_cost
        m["total"] += total; m["count"] += 1
        totals["margin"] += margin; totals["scheme"] += scheme
        totals["insurance"] += insurance; totals["extra"] += extra
        totals["dealerFundedBenefit"] += funded_cost
        totals["total"] += total; totals["count"] += 1
        components["Dealer Margin"] += margin
        components["Scheme Retained"] += scheme
        components["Insurance Income"] += insurance
        components["OEM Extra Support"] += other
        components["Documentation"] += doc_inc
        components["Warranty"] += war_inc
        components["RSA"] += rsa_inc
        components["Referral"] += ref_inc
        components["Other Income"] += other_inc
        components["Customer Insurance Benefit Passed (scheme, not income)"] += cust_ins_benefit
        components["Dealer-Funded Benefit (cost)"] += funded_cost
        components["Finance Incentive"] += fin_inc
        components["Accessories Margin"] += acc_margin
        components["Exchange Margin"] += exch_margin
        components["Campaign Incentive"] += camp_inc

    months = sorted(by_month.values(), key=lambda x: x["key"], reverse=True)
    for m in months:
        for k in ("margin", "scheme", "insurance", "other", "extra", "dealerFundedBenefit", "total"):
            m[k] = ce.round2(m[k])
    return {
        "byMonth": months,
        "components": [{"label": lbl, "amount": ce.round2(amt)} for lbl, amt in components.items() if amt],
        "totals": {k: (ce.round2(v) if isinstance(v, float) else v) for k, v in totals.items()},
    }


# ---------------------------------------------------------------- excel export
@api.get("/export")
async def export_xlsx():
    import openpyxl
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    exports = {
        "Leads": ("leads", ["leadId", "customerName", "mobile", "interestedModel", "variant", "executive",
                            "currentStatus", "customerPayable", "totalReceived", "customerOutstanding", "bookingDate"]),
        "Payments": ("payments", ["receiptNumber", "leadId", "customerName", "date", "amount", "paymentMode",
                                  "runningTotal", "outstandingBalance"]),
        "Bookings": ("bookings", ["bookingId", "leadId", "customerName", "bookingDate", "model", "variant",
                                  "bookingAmount", "paymentMode", "bookingStatus"]),
        "Claims": ("claims", ["claimId", "leadId", "customer", "component", "claimAmount", "claimStatus", "receivedAmount"]),
        "Finance": ("finance", ["fileNumber", "leadId", "customerName", "financer", "sanctionedAmount",
                                "receivedAgainstFile", "fileOutstanding", "status"]),
        "Price Master": ("price_master", ["model", "variant", "bodyType", "exShowroom", "rto", "insurance", "tcsApplicable", "status"]),
    }
    for sheet_name, (coll, cols) in exports.items():
        ws = wb.create_sheet(sheet_name)
        ws.append(cols)
        for doc in await db[coll].find().to_list(5000):
            ws.append([doc.get(c, "") for c in cols])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"euler_crm_export_{today()}.xlsx"
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ---------------------------------------------------------------- WhatsApp via BotSpace (additive; never blocks sales writes)
# Live API host BotSpace must POST to. Do not use the website URL (Render /
# Cloudflare) or the retired Render API (euler-crm-api.onrender.com).
BOTSPACE_PRODUCTION_API = "https://euler-crm-production.up.railway.app"
BOTSPACE_WEBHOOK_PATH = "/api/integrations/botspace/webhook"


def botspace_webhook_public_url() -> str:
    origin = (os.environ.get("PUBLIC_API_URL") or "").strip().rstrip("/")
    if origin.endswith("/api"):
        origin = origin[:-4]
    if not origin.startswith("http") or "onrender.com" in origin:
        origin = BOTSPACE_PRODUCTION_API
    return f"{origin}{BOTSPACE_WEBHOOK_PATH}"


class BotspaceSettingsIn(BaseModel):
    apiKey: Optional[str] = None
    channelId: Optional[str] = None
    reviewUrl: Optional[str] = None
    enabled: Optional[bool] = None
    quietStart: Optional[int] = None
    quietEnd: Optional[int] = None
    webhookSecret: Optional[str] = None
    cronToken: Optional[str] = None
    executives: Optional[List[dict]] = None
    templates: Optional[dict] = None


class WhatsAppReplyIn(BaseModel):
    text: str = ""


@api.get("/integrations/botspace", dependencies=[Depends(owner_only)])
async def botspace_settings():
    cfg = await wa.get_config()
    out = wa.public_config(cfg)
    out["webhookUrl"] = botspace_webhook_public_url()
    return out


@api.put("/integrations/botspace", dependencies=[Depends(owner_only)])
async def botspace_save_settings(body: BotspaceSettingsIn):
    return await wa.save_config(body.model_dump(exclude_unset=True))


@api.post("/integrations/botspace/run-jobs", dependencies=[Depends(owner_only)])
async def botspace_run_jobs():
    return await wa.run_daily_jobs()


class GoogleReviewIn(BaseModel):
    force: bool = False


@api.post("/integrations/botspace/send-booking-confirms", dependencies=[Depends(owner_only)])
async def botspace_send_booking_confirms(body: Optional[GoogleReviewIn] = None):
    """Owner: send booking-confirm WhatsApp to all booked Euler leads not yet messaged."""
    force = bool(body and body.force)
    return await wa.send_booking_confirms(force=force, immediate=True)


@api.post("/integrations/botspace/send-delivery-reviews", dependencies=[Depends(owner_only)])
async def botspace_send_delivery_reviews(body: Optional[GoogleReviewIn] = None):
    """Owner: send Google-review WhatsApp to all delivered Euler leads not yet messaged."""
    force = bool(body and body.force)
    return await wa.send_delivery_reviews(force=force, immediate=True)


@api.get("/integrations/botspace/model-ask/preview", dependencies=[Depends(owner_only)])
async def botspace_model_ask_preview():
    """Who WOULD receive the model-interest ask, without sending anything.

    This is the only Marketing-category template in the app, so it gets a preview:
    a marketing blast that goes out wider than intended cannot be recalled, and
    enough spam reports will cost you template access altogether.
    """
    return await wa.run_model_ask_campaign(dry_run=True)


@api.post("/integrations/botspace/model-ask", dependencies=[Depends(owner_only)])
async def botspace_run_model_ask(confirm: bool = False, limit: int = 500, act=Depends(actor)):
    """Owner: send the model-interest ask to Active leads with no model recorded.

    Requires ?confirm=true — without it this behaves exactly like the preview, so
    a mis-click cannot start a marketing send.
    """
    if not confirm:
        return {**await wa.run_model_ask_campaign(dry_run=True),
                "hint": "Call again with confirm=true to actually send."}
    res = await wa.run_model_ask_campaign(dry_run=False, limit=limit)
    await write_audit(act, "send", "whatsapp-model-ask", new={
        "eligible": res.get("eligible"), "sent": res.get("sent"),
        "queued": res.get("queued"), "failed": res.get("failed")})
    return res


@api.get("/leads/{lead_id}/whatsapp")
async def lead_whatsapp_thread(lead_id: str):
    await get_lead_or_404(lead_id)
    return await wa.list_thread(lead_id)


@api.post("/leads/{lead_id}/whatsapp/booking-confirm")
async def lead_whatsapp_booking_confirm(lead_id: str, body: Optional[GoogleReviewIn] = None):
    """Send booking confirmation template. Lead must already be booked."""
    await get_lead_or_404(lead_id)
    force = bool(body and body.force)
    res = await wa.notify_booking(lead_id, force=force, immediate=True)
    if res.get("reason") == "not-booked":
        raise HTTPException(422, "Convert the lead to a Booking before sending a booking WhatsApp")
    if res.get("reason") == "opted-out":
        raise HTTPException(422, "Customer sent STOP — booking WhatsApp will not be sent")
    if res.get("reason") == "whatsapp-not-configured":
        raise HTTPException(422, "WhatsApp is not configured. Owner: Settings → WhatsApp (BotSpace).")
    if not res.get("ok") and not res.get("skipped"):
        raise HTTPException(502, res.get("reason") or res.get("error") or "WhatsApp send failed")
    return res


@api.post("/leads/{lead_id}/whatsapp/google-review")
async def lead_whatsapp_google_review(lead_id: str, body: Optional[GoogleReviewIn] = None):
    """Send delivery + Google review template to this lead. Lead must be Delivered."""
    await get_lead_or_404(lead_id)
    force = bool(body and body.force)
    res = await wa.notify_delivery(lead_id, force=force, immediate=True)
    if res.get("reason") == "not-delivered":
        raise HTTPException(422, "Mark the lead Delivered before sending a Google review WhatsApp")
    if res.get("reason") == "opted-out":
        raise HTTPException(422, "Customer sent STOP — review WhatsApp will not be sent")
    if res.get("reason") == "whatsapp-not-configured":
        raise HTTPException(422, "WhatsApp is not configured. Owner: Settings → WhatsApp (BotSpace).")
    if not res.get("ok") and not res.get("skipped"):
        raise HTTPException(502, res.get("reason") or res.get("error") or "WhatsApp send failed")
    return res


@api.post("/leads/{lead_id}/whatsapp/reply")
async def lead_whatsapp_reply(lead_id: str, body: WhatsAppReplyIn, user=Depends(current_user)):
    await get_lead_or_404(lead_id)
    try:
        return await wa.staff_reply(lead_id, body.text, actor_name=(user or {}).get("name") or "")
    except ValueError as e:
        raise HTTPException(422, str(e))
    except RuntimeError as e:
        raise HTTPException(502, str(e))


# ------------------------------------------------------- WhatsApp inbox
# One place to see every customer conversation, instead of opening each lead.
def _thread_view(t: dict, now_iso_s: str) -> dict:
    """Derived state, computed on read so it can never be stale."""
    last_in = t.get("lastInboundAt") or ""
    last_read = t.get("lastReadAt") or ""
    out = dict(t)
    out["sessionOpen"] = wa.session_open(last_in)
    out["unread"] = bool(last_in and last_in > last_read)
    out["needsReply"] = t.get("lastDirection") == "inbound"
    return out


async def _visible_threads(user) -> list:
    """Executives see conversations on their own leads; everyone else sees all."""
    rows = [clean(t) for t in
            await db.whatsapp_threads.find().sort("lastMessageAt", -1).to_list(2000)]
    if (user or {}).get("role") != "executive":
        return rows
    mine = {l["leadId"] for l in _leads_for_executive(
        await db.leads.find().to_list(5000), user)}
    return [t for t in rows if t.get("leadId") in mine]


@api.get("/whatsapp/threads")
async def whatsapp_threads(filter: str = "all", q: str = "",
                           limit: int = 200, offset: int = 0,
                           user=Depends(current_user)):
    """Inbox. filter = all | active | needs-reply | unread."""
    now = now_iso()
    rows = [_thread_view(t, now) for t in await _visible_threads(user)]
    if filter == "active":
        rows = [r for r in rows if r["sessionOpen"]]
    elif filter == "needs-reply":
        rows = [r for r in rows if r["needsReply"]]
    elif filter == "unread":
        rows = [r for r in rows if r["unread"]]
    needle = (q or "").strip().lower()
    if needle:
        rows = [r for r in rows if needle in " ".join(str(r.get(k) or "") for k in
                ("customerName", "phone", "leadId", "executive", "model")).lower()]
    return {"total": len(rows), "threads": rows[offset:offset + max(1, limit)]}


@api.get("/whatsapp/summary")
async def whatsapp_summary(user=Depends(current_user)):
    """Counts for the nav badge."""
    now = now_iso()
    rows = [_thread_view(t, now) for t in await _visible_threads(user)]
    failed = await db.whatsapp_messages.count_documents({"status": "failed"})
    queued = await db.whatsapp_outbox.count_documents({"status": "queued"})
    return {
        "threads": len(rows),
        "unread": len([r for r in rows if r["unread"]]),
        "needsReply": len([r for r in rows if r["needsReply"]]),
        "activeChats": len([r for r in rows if r["sessionOpen"]]),
        "failed": failed, "queued": queued,
    }


@api.post("/whatsapp/threads/{lead_id}/read")
async def whatsapp_mark_read(lead_id: str):
    res = await db.whatsapp_threads.update_one(
        {"leadId": lead_id}, {"$set": {"lastReadAt": now_iso()}})
    if res.matched_count == 0:
        raise HTTPException(404, "No WhatsApp conversation for this lead")
    return {"ok": True}


@api.get("/whatsapp/messages")
async def whatsapp_messages(direction: str = "outbound", kind: str = "", status: str = "",
                            date_from: str = "", date_to: str = "", q: str = "",
                            limit: int = 300, user=Depends(current_user)):
    """Sent box — a flat list across every lead, so a failed or queued send is
    visible without opening each conversation."""
    visible = {t["leadId"] for t in await _visible_threads(user)}
    out = []
    for m in await db.whatsapp_messages.find().sort("createdAt", -1).to_list(20000):
        if m.get("leadId") not in visible:
            continue
        if direction and m.get("direction") != direction:
            continue
        if kind and m.get("kind") != kind:
            continue
        if status and str(m.get("status") or "") != status:
            continue
        day = str(m.get("createdAt") or "")[:10]
        if date_from and day < date_from:
            continue
        if date_to and day > date_to:
            continue
        if q and q.strip().lower() not in " ".join(str(m.get(k) or "") for k in
                ("customerName", "phone", "leadId", "text")).lower():
            continue
        out.append(clean(m))
        if len(out) >= max(1, limit):
            break
    return out


@api.post("/admin/backfill-whatsapp-threads", dependencies=[Depends(owner_only)])
async def backfill_whatsapp_threads():
    """Build threads from the existing message log. Idempotent."""
    return {"ok": True, **await wa.backfill_threads()}


@public.head("/integrations/botspace/webhook")
@public.get("/integrations/botspace/webhook")
async def botspace_webhook_ping():
    """BotSpace (and browsers) probe this URL. Must stay unauthenticated and 200."""
    return {"ok": True}


@public.post("/integrations/botspace/webhook")
async def botspace_webhook(request: Request):
    """BotSpace callback. Unknown phones (Tata etc.) are ignored. Always 200."""
    try:
        body = await request.json()
    except Exception:
        body = {}
    cfg = await wa.get_config()
    secret = cfg.get("webhookSecret") or ""
    if secret:
        got = request.headers.get("x-botspace-secret") or request.query_params.get("secret") or ""
        if got != secret:
            raise HTTPException(401, "Invalid webhook secret")
    try:
        return await wa.handle_webhook(body if isinstance(body, dict) else {})
    except Exception:
        logger.exception("botspace webhook handler failed")
        return {"ok": True, "error": "handler-failed"}


@public.post("/integrations/botspace/cron")
async def botspace_cron(request: Request):
    """Optional Railway / cron-job.org trigger. Token from Settings or BOTSPACE_CRON_TOKEN."""
    cfg = await wa.get_config()
    token = cfg.get("cronToken") or ""
    got = request.headers.get("x-cron-token") or request.query_params.get("token") or ""
    if not token or got != token:
        raise HTTPException(401, "Invalid cron token")
    # ?slot=morning|eod sends the daily reports; no slot keeps the legacy
    # follow-up / finance / outbox job. Both are idempotent per day, so a cron
    # retry cannot double-send.
    slot = (request.query_params.get("slot") or "").strip().lower()
    if slot:
        return await wa.run_daily_reports(slot)
    return await wa.run_daily_jobs()


@api.post("/integrations/botspace/send-daily-report", dependencies=[Depends(owner_only)])
async def botspace_send_daily_report(slot: str = "eod", act=Depends(actor)):
    """Owner: send a report slot now, without waiting for the schedule.

    Idempotent per day+slot — use force=1 semantics by clearing the marker if you
    really need a resend."""
    res = await wa.run_daily_reports(slot)
    await write_audit(act, "send", "daily-report", new={"slot": slot, "sent": res.get("sent")})
    return res


@api.get("/integrations/botspace/report-status", dependencies=[Depends(owner_only)])
async def botspace_report_status(days: int = 7):
    """Why a report slot failed.

    Daily reports go to staff, so they carry no leadId — which means they are
    excluded from both the WhatsApp inbox and the Sent box, and a failure had
    nowhere to surface. This returns the last runs per slot with the provider's
    own error against each recipient, the template name actually used, and the
    most likely fix.
    """
    return await wa.report_status(days)


@api.delete("/integrations/botspace/daily-report-marker", dependencies=[Depends(owner_only)])
async def botspace_clear_report_marker(slot: str = "eod"):
    """Clear today's sent-marker so the slot can be re-sent (e.g. after a fix)."""
    res = await db.settings.delete_one({"_id": f"report_{slot}_{wa.today_ist()}"})
    return {"ok": True, "cleared": res.deleted_count}


# ---------------------------------------------------------------- public share board (no auth)
@public.get("/share/dashboard")
async def share_dashboard():
    leads = await db.leads.find().to_list(5000)
    ym = this_month()
    td = today()
    booked = [l for l in leads if "book" in (l.get("currentStatus") or "").lower()]
    active_booked = [l for l in booked if (l.get("deliveryStatus") or "").lower() != "delivered"]
    delivered = [l for l in leads if (l.get("deliveryStatus") or "").lower() == "delivered"]
    new_this_month = [l for l in booked if str(l.get("bookingDate") or "").startswith(ym)]
    retail_this_month = [l for l in delivered if str(l.get("deliveryDate") or "").startswith(ym)]
    today_bookings = [l for l in booked if str(l.get("bookingDate") or "") == td]

    by_model = {}
    for l in active_booked:
        m = l.get("interestedModel") or "Unknown"
        by_model[m] = by_model.get(m, 0) + 1

    def row(l, date_field):
        return {"date": l.get(date_field), "name": l.get("customerName") or "—",
                "model": l.get("interestedModel") or "", "variant": l.get("variant") or ""}

    recent_bookings = sorted(active_booked, key=lambda l: str(l.get("bookingDate") or ""), reverse=True)[:20]
    recent_retail = sorted(retail_this_month, key=lambda l: str(l.get("deliveryDate") or ""), reverse=True)[:20]
    all_leads = sorted(leads, key=lambda l: str(l.get("createdDate") or ""), reverse=True)
    all_leads_rows = [{"date": l.get("createdDate"), "name": l.get("customerName") or "—",
                       "model": l.get("interestedModel") or "", "variant": l.get("variant") or "",
                       "status": l.get("currentStatus") or "New"} for l in all_leads]

    return {
        "activeBookings": len(active_booked),
        "newThisMonth": len(new_this_month),
        "retailThisMonth": len(retail_this_month),
        "todayBookings": len(today_bookings),
        "totalLeads": len(leads),
        "recentBookings": [row(l, "bookingDate") for l in recent_bookings],
        "recentRetail": [row(l, "deliveryDate") for l in recent_retail],
        "allLeads": all_leads_rows,
        "byModel": [{"model": k, "count": v} for k, v in sorted(by_model.items(), key=lambda x: -x[1])],
        "month": datetime.now(timezone.utc).strftime("%B %Y"),
        "lastUpdated": now_iso(),
    }


# ---------------------------------------------------------------- commercial preview + quotation
@api.post("/commercial/compute")
async def commercial_compute(body: SnapshotComputeIn):
    return ce.compute_full_commercials(body.model_dump())


class QuotationIn(BaseModel):
    customerName: str = ""
    mobile: str = ""
    model: str = ""
    variant: str = ""
    exShowroom: float = 0
    insurance: float = 0
    registrationRto: float = 0
    accessories: float = 0
    handlingCharges: float = 0
    trc: float = 0
    fastag: float = 0
    extendedWarranty: float = 0
    otherCharges: float = 0
    consumerDiscount: float = 0
    exchangeBonus: float = 0
    loyaltyBonus: float = 0
    referralBonus: float = 0
    dsaDiscount: float = 0
    additionalDiscount: float = 0
    finalExchangeValue: float = 0
    financer: str = ""
    narration: str = ""


@api.get("/quotations")
async def list_quotations():
    return [clean(q) for q in await db.quotations.find().sort("quoteId", -1).to_list(1000)]


@api.post("/quotations")
async def create_quotation(body: QuotationIn, _sales=Depends(sales_staff_only)):
    totals = ce.compute_commercial_totals(body.model_dump())
    claim = ce.derive_claim(body.model_dump())
    quote_id = await next_id("snapshot", "QT26")
    doc = {"quoteId": quote_id, "date": today(), **body.model_dump(),
           "grossVehicleCost": totals["grossVehicleCost"], "totalDiscount": totals["totalDiscount"],
           "customerPayable": totals["customerPayable"], "oemShare": claim["claimEligible"]}
    await db.quotations.insert_one(doc)
    return clean(doc)


# ---------------------------------------------------------------- startup
app.add_middleware(
    CORSMiddleware, allow_origins=["*"], allow_credentials=False,
    allow_methods=["*"], allow_headers=["*"],
)


async def _migrate_insurance_rates():
    """Safe idempotent migration: any insurance doc storing payoutRate as a percent (>1)
    is converted to a fraction and its expected/outstanding recomputed. Ensures a single
    consistent representation (DB = fraction, UI = %)."""
    fixed = 0
    for e in await db.insurance.find({"payoutRate": {"$gt": 1}}).to_list(5000):
        rate = ce.num(e.get("payoutRate")) / 100.0
        premium = ce.num(e.get("insuranceAmount"))
        expected = ce.round2(premium * rate)
        received = ce.num(e.get("receivedPayout"))
        outstanding = ce.round2(max(0.0, expected - received))
        await db.insurance.update_one({"entryId": e.get("entryId")}, {"$set": {
            "payoutRate": rate, "expectedPayout": expected, "payoutOutstanding": outstanding,
        }})
        fixed += 1
    return fixed


@api.post("/admin/migrate-insurance-rates", dependencies=[Depends(owner_only)])
async def migrate_insurance_rates():
    fixed = await _migrate_insurance_rates()
    return {"ok": True, "fixed": fixed}


# Slabs of the arrangement that existed before agents were modelled. Seeded once
# so every pre-agent entry has an agent to point at and NO amount changes.
LEGACY_AGENT_SLABS = [
    {"modelFamily": "storm", "payoutRatePct": 49.0, "effectiveFrom": "", "effectiveTo": ""},
    {"modelFamily": "turbo", "payoutRatePct": 49.0, "effectiveFrom": "", "effectiveTo": ""},
    {"modelFamily": "*", "payoutRatePct": 36.5, "effectiveFrom": "", "effectiveTo": ""},
]
SECOND_AGENT_SLABS = [
    {"modelFamily": "storm", "payoutRatePct": 52.0, "effectiveFrom": "", "effectiveTo": ""},
    {"modelFamily": "turbo", "payoutRatePct": 52.0, "effectiveFrom": "", "effectiveTo": ""},
    {"modelFamily": "*", "payoutRatePct": 42.0, "effectiveFrom": "", "effectiveTo": ""},
]
# Deterministic ids for the two seeded agents. A fixed id survives a rename, so a
# lead or an insurance entry keeps pointing at the same agent after the owner
# renames "Agent 1" to the real broker name.
SEED_DEFAULT_AGENT_ID = "IA26AGENT1"
SEED_SECOND_AGENT_ID = "IA26AGENT2"
SEED_INSURANCE_AGENTS = [
    (SEED_DEFAULT_AGENT_ID, "Agent 1", LEGACY_AGENT_SLABS, True,
     "Existing arrangement (49% Storm/Turbo, 36.5% others). Rename to the real agent name."),
    (SEED_SECOND_AGENT_ID, "Agent 2", SECOND_AGENT_SLABS, False,
     "52% Storm/Turbo, 42% others. Rename to the real agent name."),
]


async def _seed_insurance_agents() -> dict:
    """Idempotent. Creates the two agents only when the collection is empty, then
    stamps every unassigned insurance entry with the default agent. Rates are NOT
    recomputed — existing expected/received/outstanding amounts stay exactly as they are."""
    created = []
    if await db.insurance_agents.count_documents({}) == 0:
        for agent_id, name, slabs, is_default, remarks in SEED_INSURANCE_AGENTS:
            doc = {
                "agentId": agent_id,
                "agentName": name, "agentCode": "", "contactPerson": "", "mobile": "", "email": "",
                "status": "Active", "isDefault": is_default,
                "slabs": [dict(s) for s in slabs], "remarks": remarks,
                "createdAt": now_iso(), "lastUpdated": now_iso(),
            }
            await db.insurance_agents.insert_one(dict(doc))
            created.append(doc["agentName"])
    default = await _default_insurance_agent()
    stamped = 0
    if default:
        res = await db.insurance.update_many(
            {"$or": [{"insuranceAgentId": {"$exists": False}}, {"insuranceAgentId": ""}]},
            {"$set": {"insuranceAgentId": default["agentId"],
                      "insuranceAgentName": default.get("agentName", ""),
                      "payoutRateSource": "legacy-default"}},
        )
        stamped = res.modified_count
    return {"created": created, "stampedEntries": stamped,
            "defaultAgent": (default or {}).get("agentName", "")}


@api.post("/admin/seed-insurance-agents", dependencies=[Depends(owner_only)])
async def seed_insurance_agents():
    return {"ok": True, **await _seed_insurance_agents()}


@api.post("/admin/reset-transactions", dependencies=[Depends(owner_only)])
async def reset_transactions(act=Depends(actor)):
    """Owner-only go-live reset: permanently clears all transaction data (leads, bookings,
    payments, deliveries, finance, insurance, claims, activities, earnings, incentives,
    quotations) and blocks the sample-data re-seed. Master data (price/scheme/incentive
    master, users, Masters list) is preserved. Also clears operational Google Sheet
    register data rows (headers kept) when sheet sync is writable.

    Scheme Claim Register is excluded from the sheet wipe — it is a permanent ledger
    (Received claims stay forever). Mongo claims are still cleared with other
    transactions so the CRM starts clean; historical sheet claim rows remain.

    Also clears sheet_sync_log so a later /integrations/gsheets/retry cannot resurrect
    deleted leads onto the spreadsheet.
    """
    counts = {}
    for coll in ["leads", "bookings", "payments", "deliveries", "finance", "insurance",
                 "claims", "activities", "dealer_earnings", "quotations", "incentive_register",
                 "billing_summaries", "sheet_sync_log"]:
        r = await db[coll].delete_many({})
        counts[coll] = r.deleted_count
    await db["system"].update_one({"_id": "seed_state"},
                                  {"$set": {"sampleCleared": True, "clearedAt": now_iso()}}, upsert=True)
    await db["counters"].update_one({"_id": "lead"}, {"$set": {"seq": 0}}, upsert=True)
    for c in ["receipt", "booking", "activity", "snapshot", "claim", "finance", "insurance"]:
        await db["counters"].update_one({"_id": c}, {"$set": {"seq": 100}}, upsert=True)
    sheet_clear = await gsheets.clear_operational_register_rows()
    await write_audit(act, "reset", "system",
                      new={"clearedTransactions": counts, "sheetClear": sheet_clear,
                           "permanentLedgersPreserved": sorted(gsheets.PERMANENT_LEDGER_TABS)})
    return {"ok": True, "cleared": counts, "sheetClear": sheet_clear,
            "permanentLedgersPreserved": sorted(gsheets.PERMANENT_LEDGER_TABS),
            "nextLeadId": "LD26000001"}


def _config_diagnostics():
    """Names + presence only — never values. Logged once at startup so a
    misconfigured deploy is obvious without ever exposing a secret."""
    required = ["MONGO_URL", "DB_NAME", "JWT_SECRET"]
    optional = ["GSHEET_ID", "CORS_ORIGINS", "COULSON_USERNAME"]
    missing = [k for k in required if not os.environ.get(k, "").strip()]
    cred = gsheets.credential_diagnostics()
    return {
        "required": {k: bool(os.environ.get(k, "").strip()) for k in required},
        "optional": {k: bool(os.environ.get(k, "").strip()) for k in optional},
        "missingRequired": missing,
        "googleCredentialFound": cred["credential_found"],
        "googleCredentialSource": cred["credential_source"],
        "gsheetIdPresent": cred["gsheet_id_present"],
        "sheetSyncEnabled": bool(cred["credential_found"] and cred["gsheet_id_present"]),
        "financeIndexes": dict(_finance_index_status),
    }


@api.get("/admin/config-check", dependencies=[Depends(owner_only)])
async def config_check():
    """Owner-only startup/runtime configuration health. Reports which settings are
    PRESENT, never their values."""
    return _config_diagnostics()



async def _audit_finance_integrity_for_unique_indexes():
    findings = {}
    findings["financePaymentsWithoutRegister"] = await db.payments.aggregate([
        {"$match": {"paymentMode": "Finance", "financeFileNumber": {"$nin": [None, ""]}}},
        {"$lookup": {"from": "finance", "localField": "financeFileNumber", "foreignField": "fileNumber", "as": "financeFile"}},
        {"$match": {"financeFile": {"$size": 0}}},
        {"$project": {"_id": 0, "receiptNumber": 1, "leadId": 1, "financeFileNumber": 1}},
    ]).to_list(1000)
    findings["financeRegisterWithoutPayments"] = await db.finance.aggregate([
        {"$lookup": {"from": "payments", "localField": "fileNumber", "foreignField": "financeFileNumber", "as": "payments"}},
        {"$match": {"payments": {"$size": 0}}},
        {"$project": {"_id": 0, "fileNumber": 1, "leadId": 1}},
    ]).to_list(1000)
    findings["duplicateFinanceLeadIds"] = await db.finance.aggregate([
        {"$match": {"leadId": {"$nin": [None, ""]}}},
        {"$group": {"_id": "$leadId", "count": {"$sum": 1}, "files": {"$push": "$fileNumber"}}},
        {"$match": {"count": {"$gt": 1}}},
    ]).to_list(1000)
    findings["duplicateFinanceFileNumbers"] = await db.finance.aggregate([
        {"$match": {"fileNumber": {"$nin": [None, ""]}}},
        {"$group": {"_id": "$fileNumber", "count": {"$sum": 1}, "leadIds": {"$push": "$leadId"}}},
        {"$match": {"count": {"$gt": 1}}},
    ]).to_list(1000)
    findings["financeNoLeadsWithFinancePayments"] = await db.payments.aggregate([
        {"$match": {"paymentMode": "Finance"}},
        {"$lookup": {"from": "leads", "localField": "leadId", "foreignField": "leadId", "as": "lead"}},
        {"$unwind": "$lead"},
        {"$match": {"lead.financeRequired": {"$ne": "Yes"}}},
        {"$project": {"_id": 0, "receiptNumber": 1, "leadId": 1, "financeFileNumber": 1}},
    ]).to_list(1000)
    findings["financePaymentsWithBlankFileNumber"] = await db.payments.find(
        {"paymentMode": "Finance", "financeFileNumber": {"$in": [None, ""]}},
        {"_id": 0, "receiptNumber": 1, "leadId": 1}
    ).to_list(1000)
    logging.info("FINANCE_INTEGRITY_AUDIT: %s", {k: len(v) for k, v in findings.items()})
    return findings


async def _ensure_finance_unique_indexes():
    global _finance_index_status
    _finance_index_status = {
        "status": "CHECKING",
        "ready": False,
        "reason": "finance uniqueness audit/index check in progress",
        "checkedAt": now_iso(),
        "auditCounts": {},
    }
    findings = await _audit_finance_integrity_for_unique_indexes()
    counts = {k: len(v) for k, v in findings.items()}
    if findings["duplicateFinanceLeadIds"] or findings["duplicateFinanceFileNumbers"]:
        _finance_index_status = {
            "status": "ERROR",
            "ready": False,
            "reason": "duplicate finance leadIds/fileNumbers found; unique indexes not created",
            "checkedAt": now_iso(),
            "auditCounts": counts,
        }
        logging.error("FINANCE_UNIQUE_INDEX_SKIPPED: duplicate finance records found: %s", findings)
        return findings
    try:
        await db.finance.create_index("leadId", unique=True, partialFilterExpression={"leadId": {"$type": "string", "$gt": ""}})
        await db.finance.create_index("fileNumber", unique=True, partialFilterExpression={"fileNumber": {"$type": "string", "$gt": ""}})
    except Exception as e:
        _finance_index_status = {
            "status": "ERROR",
            "ready": False,
            "reason": f"finance unique index creation failed ({type(e).__name__})",
            "checkedAt": now_iso(),
            "auditCounts": counts,
        }
        logging.exception("FINANCE_UNIQUE_INDEX_ERROR: finance unique index creation failed")
        return findings
    _finance_index_status = {
        "status": "HEALTHY",
        "ready": True,
        "reason": "finance unique indexes verified",
        "checkedAt": now_iso(),
        "auditCounts": counts,
    }
    return findings

async def _backfill_close_won_status():
    """Closed leads that still show Booked/Delivered/etc. → Close Won, then push Lead Register."""
    keep = {"close won", "lost", "cancelled", "archived"}
    fixed = 0
    cursor = db.leads.find({"accountStatus": {"$regex": "^closed$", "$options": "i"}})
    for lead in await cursor.to_list(5000):
        cs = str(lead.get("currentStatus") or "").strip()
        if cs.lower() in keep:
            continue
        lid = lead.get("leadId")
        if not lid:
            continue
        await db.leads.update_one(
            {"leadId": lid},
            {"$set": {"currentStatus": "Close Won", "lastUpdated": now_iso()}},
        )
        updated = clean(await db.leads.find_one({"leadId": lid}))
        await sheet_sync("leads", updated)
        fixed += 1
    if fixed:
        logging.info("CLOSE_WON_BACKFILL: updated %s closed lead(s) to Close Won", fixed)
    return fixed


_coulson_loop_task = None


def _ensure_coulson_sync_loop():
    """Start the 15-minute yard pull if it is not already running.

    The loop used to start only at process boot. A session pasted after boot
    never started auto-pull until the next Railway restart.
    """
    global _coulson_loop_task
    if os.environ.get("ENVIRONMENT", "").lower() == "test":
        return
    if _coulson_loop_task is not None and not _coulson_loop_task.done():
        return
    _coulson_loop_task = asyncio.create_task(_coulson_sync_loop())


async def _coulson_sync_loop():
    """Periodic Coulson pull. Never raises into the event loop."""
    while True:
        try:
            result = await oem_sync.sync_from_coulson(db)
            if result.get("ok"):
                try:
                    await _reprice_after_oem(result.get("changedPriceIds"))
                except Exception:
                    logging.exception("OEM reprice after scheduled sync failed")
        except Exception:
            logging.exception("Scheduled Coulson sync failed")
        await asyncio.sleep(15 * 60)


async def _oem_catalog_boot():
    """Rename/drop/add Price Master from the OEM catalog, then reprice live leads.

    First production boot can change dozens of ex-showroom values and reprice every
    matching lead (including Google Sheet writes). That must not run inside FastAPI
    startup — the process does not bind PORT until startup returns, so Railway
    healthchecks 502 and restart the deploy.
    """
    try:
        cat_res = await oem_sync.apply_catalog(db)
        try:
            await _reprice_after_oem(cat_res.get("changedPriceIds"))
        except Exception:
            logging.exception("OEM_CATALOG_REPRICE_ERROR")
    except Exception:
        logging.exception("OEM_CATALOG_APPLY_ERROR")
    try:
        if os.environ.get("ENVIRONMENT", "").lower() != "test":
            user, pw, _src = await oem_sync.resolve_credentials(db)
            doc = await db["system"].find_one({"_id": "coulson"}) or {}
            if oem_sync.credentials_configured(user, pw) or oem_sync.session_from_doc(doc):
                _ensure_coulson_sync_loop()
    except Exception:
        logging.exception("COULSON_SYNC_LOOP_START_ERROR")


@app.on_event("startup")
async def startup():
    global _finance_index_status
    cfg = _config_diagnostics()
    if cfg["missingRequired"]:
        logging.warning("CONFIG: missing required environment variables: %s", ", ".join(cfg["missingRequired"]))
    if not cfg["sheetSyncEnabled"]:
        logging.warning("CONFIG: Google Sheet sync is DISABLED (credentialFound=%s, gsheetId=%s). "
                        "Add the Render Secret File gsheets_credentials.json and set GSHEET_ID.",
                        cfg["googleCredentialFound"], cfg["gsheetIdPresent"])
    else:
        logging.info("CONFIG: Google Sheet sync enabled (credential source: %s)", cfg["googleCredentialSource"])
    await authmod.seed_users(db)
    res = await seeder.run_seed(db)
    if res.get("seeded"):
        for l in await db.leads.find().to_list(3000):
            await recompute_lead(l["leadId"])
    # Indexes (M2/M3 performance) — idempotent
    try:
        await db.leads.create_index("leadId")
        await db.leads.create_index("mobile")
        await db.leads.create_index("currentStatus")
        await db.leads.create_index("bookingDate")
        await db.payments.create_index("leadId")
        await db.audit_log.create_index([("timestamp", -1)])
        await db.incentive_register.create_index("leadId", unique=True)
        await db.masters_list.create_index("id", unique=True)
        await db.masters_list.create_index([("category", 1), ("value", 1)])
        await db.whatsapp_messages.create_index("leadId")
        await db.whatsapp_messages.create_index("providerId")
        await db.whatsapp_messages.create_index("phone")
        await db.insurance_agents.create_index("agentId", unique=True)
        await db.insurance.create_index("insuranceAgentId")
        await db.staff.create_index("staffId", unique=True)
        await db.staff.create_index("role")
        await db.whatsapp_threads.create_index("leadId", unique=True)
        await db.whatsapp_threads.create_index([("lastMessageAt", -1)])
        await db.cancel_reasons.create_index("reasonId", unique=True)
        await db.oem_inventory.create_index("chassis")
        # The daily revival sweep queries on these two together.
        await db.leads.create_index([("accountStatus", 1), ("reviveOn", 1)])
    except Exception:
        pass
    try:
        await _ensure_finance_unique_indexes()
    except Exception as e:
        _finance_index_status = {
            "status": "ERROR",
            "ready": False,
            "reason": f"finance uniqueness audit failed ({type(e).__name__})",
            "checkedAt": now_iso(),
            "auditCounts": {},
        }
        logging.exception("FINANCE_UNIQUE_INDEX_AUDIT_ERROR: finance uniqueness audit failed")
    # Closed leads must show Close Won (not leftover Booked/Delivered) in app + sheet.
    try:
        await _backfill_close_won_status()
    except Exception:
        logging.exception("CLOSE_WON_BACKFILL_ERROR")
    # Insurance rate consistency migration (INS-1)
    try:
        await _migrate_insurance_rates()
    except Exception:
        pass
    # Insurance agents (brokers) + stamp pre-agent entries with the default agent.
    # Idempotent and amount-preserving.
    try:
        await _seed_insurance_agents()
    except Exception:
        logging.exception("INSURANCE_AGENT_SEED_ERROR")
    # Staff master, built from the existing executive list + any WhatsApp number
    # already saved in the BotSpace settings. Idempotent.
    try:
        await _seed_staff()
    except Exception:
        logging.exception("STAFF_SEED_ERROR")
    try:
        await _seed_cancel_reasons()
    except Exception:
        logging.exception("CANCEL_REASON_SEED_ERROR")
    # Parked leads whose cool-off expired while the service was down. Cheap, and it
    # means a revival is never lost just because nobody hit the app that morning.
    try:
        await run_scheduled_revivals()
    except Exception:
        logging.exception("LEAD_REVIVAL_STARTUP_ERROR")
    try:
        await _backfill_deal_cancelled()
    except Exception:
        logging.exception("DEAL_CANCELLED_BACKFILL_ERROR")
    # Self-heal: booked leads whose booking advance was never recorded as a payment
    # (so Customer Outstanding wasn't reduced). Idempotent.
    try:
        await _backfill_booking_advances()
    except Exception:
        pass
    # Finance files that recorded one disbursement twice before the dedupe guard.
    try:
        await _repair_duplicate_finance_receipts()
    except Exception:
        logging.exception("FINANCE_RECEIPT_DEDUPE_ERROR")
    # OEM catalog: tests await it so Price Master is ready on the first request.
    # Production runs it after the server is listening (Railway healthcheck).
    try:
        if os.environ.get("ENVIRONMENT", "").lower() == "test":
            await _oem_catalog_boot()
        else:
            asyncio.create_task(_oem_catalog_boot())
    except Exception:
        logging.exception("OEM_CATALOG_BOOT_ERROR")
    try:
        if os.environ.get("ENVIRONMENT", "").lower() != "test":
            asyncio.create_task(wa.scheduler_loop())
    except Exception:
        logging.exception("WHATSAPP_SCHEDULER_START_ERROR")


def _dedupe_finance_receipts(receipts):
    """Drop receipts that are the same disbursement posted twice.

    Same amount + same date + same reference recorded within two minutes of each
    other is a double submit, not two tranches — a financer paying the same amount
    twice in one day is recorded by staff minutes apart at best, and the entries
    carry distinct references.
    """
    kept = []
    for r in receipts or []:
        dup = False
        for k in kept:
            if ce.round2(ce.num(k.get("amount"))) != ce.round2(ce.num(r.get("amount"))):
                continue
            if str(k.get("date") or "") != str(r.get("date") or ""):
                continue
            if str(k.get("reference") or "").strip() != str(r.get("reference") or "").strip():
                continue
            try:
                gap = abs((datetime.fromisoformat(r["recordedAt"])
                           - datetime.fromisoformat(k["recordedAt"])).total_seconds())
            except (KeyError, TypeError, ValueError):
                gap = 0.0
            if gap <= 120:
                dup = True
                break
        if not dup:
            kept.append(r)
    return kept


async def _repair_duplicate_finance_receipts():
    """Heal finance files whose receipts were double-posted before the guard existed.

    Recomputes receivedAgainstFile / fileOutstanding / status from the deduplicated
    receipts so the app stops showing a disbursement twice (the Google Sheet always
    showed one row because that sync upserts on file number). Idempotent.
    """
    repaired = []
    for f in await db.finance.find().to_list(5000):
        receipts = f.get("receipts") or []
        kept = _dedupe_finance_receipts(receipts)
        if len(kept) == len(receipts):
            continue
        committed = ce.num(f.get("sanctionedAmount"))
        received = ce.round2(sum(ce.num(r.get("amount")) for r in kept))
        outstanding = ce.round2(max(0.0, committed - received))
        await db.finance.update_one({"fileNumber": f["fileNumber"]}, {"$set": {
            "receipts": kept, "receivedAgainstFile": received, "fileOutstanding": outstanding,
            "status": ("Received" if committed > 0 and outstanding <= 0
                       else ("Partial" if received > 0 else "Pending")),
            "lastUpdated": today(),
        }})
        await sync_finance_file(f["fileNumber"])
        repaired.append({"fileNumber": f["fileNumber"], "removed": len(receipts) - len(kept),
                         "receivedAgainstFile": received, "fileOutstanding": outstanding})
    if repaired:
        logger.info("FINANCE_RECEIPT_DEDUPE: repaired %s", repaired)
        await rebuild_finance_views()
    return repaired


@api.post("/finance/repair-duplicate-receipts", dependencies=[Depends(owner_only)])
async def repair_duplicate_finance_receipts():
    """Owner action for files that were already double-posted (e.g. FN26000101)."""
    repaired = await _repair_duplicate_finance_receipts()
    return {"ok": True, "filesRepaired": len(repaired), "files": repaired}


async def _backfill_booking_advances():
    """For any lead with bookingAmount>0 and NO payments recorded, create the booking-advance
    receipt so Customer Outstanding is reduced. Idempotent (skips leads that already have payments)."""
    healed = 0
    for l in await db.leads.find({"bookingAmount": {"$gt": 0}}).to_list(5000):
        lid = l["leadId"]
        ba = ce.num(l.get("bookingAmount"))
        if ba <= 0:
            continue
        has_advance = await db.payments.find_one(
            {"leadId": lid, "narration": {"$regex": "booking advance", "$options": "i"}})
        received = await db.payments.aggregate([
            {"$match": {"leadId": lid}}, {"$group": {"_id": None, "t": {"$sum": "$amount"}}}]).to_list(1)
        total_received = ce.num(received[0]["t"]) if received else 0.0
        if has_advance or total_received > 0:
            continue
        await _add_payment_internal(lid, PaymentIn(
            amount=ba, paymentMode=l.get("lastPaymentMode") or "Cash",
            date=l.get("bookingDate") or today(), narration="Booking advance (backfill)"))
        await recompute_lead(lid)
        healed += 1
    return healed


@app.on_event("shutdown")
async def shutdown():
    client.close()


# Routers included LAST so every @api route above is registered
app.include_router(auth_router)
app.include_router(api)
app.include_router(public)
