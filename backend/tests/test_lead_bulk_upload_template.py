"""Bulk lead upload — template dropdowns and row validation.

The template's dropdown values and the upload's validation must come from the same
live masters, so a downloaded template can never offer a value the upload rejects.
Anything typed that is NOT in a master list is reported and skipped, never silently
imported (that mismatch is what breaks Price Master lookups and executive reports).
"""
import io
import os
import sys

import pytest
import pytest_asyncio

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ.setdefault("MONGO_URL", "mongodb://localhost:27017")
os.environ.setdefault("DB_NAME", "lead_bulk_upload")
os.environ.setdefault("JWT_SECRET", "bulk-upload-secret")

import motor.motor_asyncio  # noqa: E402
from mongomock_motor import AsyncMongoMockClient  # noqa: E402

motor.motor_asyncio.AsyncIOMotorClient = AsyncMongoMockClient

import httpx  # noqa: E402
import openpyxl  # noqa: E402
import server  # noqa: E402

HEADERS = [label for label, _ in server.IMPORT_COLUMNS]


def _csv(rows):
    body = ",".join(HEADERS) + "\n"
    for r in rows:
        body += ",".join("" if v is None else str(v) for v in r) + "\n"
    return body.encode()


def _row(name, mobile, **over):
    """A row in IMPORT_COLUMNS order with valid defaults."""
    values = {
        "Customer Name": name, "Mobile": mobile, "Alternate Mobile": "",
        "Village": "Bassi", "City": "Jaipur", "Lead Date": "2026-08-10",
        "Next Follow-up": "", "Lead Source": "Walk-in",
        "Interested Model": "Turbo Max", "Variant": "Maxx (PV)", "Executive": "Amit",
        "Current Status": "New", "Priority": "Normal", "Budget": 750000,
        "Remarks": "", "Finance Required": "No", "Exchange Required": "No",
    }
    values.update(over)
    return [values[label] for label in HEADERS]


@pytest_asyncio.fixture
async def client(monkeypatch):
    """Run against a dedicated mongomock database.

    Every test module shares one `server` import (and therefore one DB_NAME), so
    wiping leads / price_master here would corrupt other modules' fixtures. Auth
    keeps using the original database because the auth router captured it at
    import time; seed_users is idempotent and never deletes.
    """
    isolated = server.client["lead_bulk_upload_isolated"]
    for name in ("leads", "price_master", "masters_list", "counters", "activities"):
        await isolated[name].delete_many({})
    await isolated.price_master.insert_many([
        {"priceId": "PM1", "model": "Turbo Max", "variant": "Maxx (PV)", "exShowroom": 770000, "status": "Active"},
        {"priceId": "PM2", "model": "Turbo Max", "variant": "Maxx (DV2000)", "exShowroom": 780000, "status": "Active"},
        {"priceId": "PM3", "model": "Storm EV", "variant": "Storm (PV)", "exShowroom": 640000, "status": "Active"},
    ])
    monkeypatch.setattr(server, "db", isolated)
    await server.authmod.seed_users(server.client[os.environ["DB_NAME"]])
    transport = httpx.ASGITransport(app=server.app)
    async with httpx.AsyncClient(transport=transport, base_url="http://test") as c:
        r = await c.post("/api/auth/login", json={"email": "owner@euler.com", "password": "euler@123"})
        c.headers["Authorization"] = f"Bearer {r.json()['token']}"
        yield c


# ------------------------------------------------------------------ template
@pytest.mark.asyncio
async def test_template_has_dropdowns_from_live_masters(client):
    r = await client.get("/api/leads/import/template")
    assert r.status_code == 200, r.text
    assert "spreadsheetml" in r.headers["content-type"]

    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    assert wb.sheetnames == ["Leads", "Lists", "How to use"]

    leads, lists = wb["Leads"], wb["Lists"]
    assert [c.value for c in leads[1]] == HEADERS

    # Every list column the user must not free-type carries a validation range.
    validated = set()
    for dv in leads.data_validations.dataValidation:
        assert dv.type == "list"
        assert dv.formula1.startswith("=Lists!")
        validated.update(str(sqref) for sqref in dv.sqref.ranges)
    assert len(validated) == 8, validated

    # Dropdown values come from Price Master / Settings, not a hardcoded list.
    by_title = {}
    for col in lists.iter_cols(min_row=1, max_row=lists.max_row):
        title = col[0].value
        if title:
            by_title[title] = [c.value for c in col[1:] if c.value]
    assert by_title["Interested Model"] == ["Storm EV", "Turbo Max"]
    assert "Maxx (DV2000)" in by_title["Variant"]
    assert "Amit" in by_title["Executive"]
    assert by_title["Current Status"] == server.IMPORT_STATUSES
    assert by_title["Yes / No"] == ["Yes", "No"]
    # Valid model+variant pairs are listed because validation cannot be dependent.
    assert ("Turbo Max", "Maxx (PV)") in list(zip(by_title["Valid Model"],
                                                  by_title["Valid Variant for that Model"]))


@pytest.mark.asyncio
async def test_template_needs_sales_role(client):
    r = await client.post("/api/auth/login", json={"email": "asm@euler.com", "password": "euler@123"})
    asm = {"Authorization": f"Bearer {r.json()['token']}"}
    assert (await client.get("/api/leads/import/template", headers=asm)).status_code == 403


# ------------------------------------------------------------------ validation
async def _preview(client, rows):
    r = await client.post("/api/leads/import/preview",
                          files={"file": ("leads.csv", _csv(rows), "text/csv")})
    assert r.status_code == 200, r.text
    return r.json()


@pytest.mark.asyncio
async def test_clean_template_rows_import(client):
    rows = [_row("Ramesh Kumar", "9800000001"), _row("Sita Devi", "9800000002", **{
        "Interested Model": "Storm EV", "Variant": "Storm (PV)", "Lead Source": "Referral",
        "Current Status": "Follow-up", "Next Follow-up": "2026-08-20"})]
    body = await _preview(client, rows)
    assert (body["rowCount"], body["validCount"], body["errorCount"]) == (2, 2, 0)

    r = await client.post("/api/leads/import/commit",
                          files={"file": ("leads.csv", _csv(rows), "text/csv")})
    assert r.status_code == 200, r.text
    assert r.json()["created"] == 2 and r.json()["skipped"] == 0

    lead = await server.db.leads.find_one({"customerName": "Sita Devi"})
    assert lead["leadSource"] == "Referral"
    assert lead["interestedModel"] == "Storm EV" and lead["variant"] == "Storm (PV)"
    assert lead["currentStatus"] == "Follow-up"
    assert lead["nextFollowupDate"] == "2026-08-20"
    assert lead["createdDate"] == "2026-08-10"
    assert lead["accountStatus"] == "Active"


@pytest.mark.asyncio
async def test_values_outside_masters_are_skipped_not_imported(client):
    rows = [
        _row("Bad Source", "9800000011", **{"Lead Source": "Instagram DM"}),
        _row("Bad Exec", "9800000012", **{"Executive": "Someone Else"}),
        _row("Bad Model", "9800000013", **{"Interested Model": "Turbo Maxx", "Variant": ""}),
        _row("Wrong Variant", "9800000014", **{"Interested Model": "Turbo Max", "Variant": "Storm (PV)"}),
        _row("Booked Status", "9800000015", **{"Current Status": "Delivered"}),
        _row("Good Row", "9800000016"),
    ]
    body = await _preview(client, rows)
    assert body["validCount"] == 1 and body["errorCount"] == 5
    problems = {e["customerName"]: " ".join(e["errors"]) for e in body["errors"]}
    assert "Lead Source" in problems["Bad Source"]
    assert "Executive" in problems["Bad Exec"]
    assert "Price Master" in problems["Bad Model"]
    assert "does not belong to Turbo Max" in problems["Wrong Variant"]
    assert "Current Status" in problems["Booked Status"]

    r = await client.post("/api/leads/import/commit",
                          files={"file": ("leads.csv", _csv(rows), "text/csv")})
    assert r.json()["created"] == 1 and r.json()["skipped"] == 5
    assert await server.db.leads.count_documents({}) == 1
    assert (await server.db.leads.find_one({}))["customerName"] == "Good Row"


@pytest.mark.asyncio
async def test_case_and_spacing_differences_are_accepted(client):
    rows = [_row("Loose Case", "9800000021", **{
        "Lead Source": "walk-in", "Executive": "  amit ", "Interested Model": "turbo max",
        "Variant": "maxx (pv)", "Finance Required": "yes", "Priority": "high"})]
    body = await _preview(client, rows)
    assert body["errorCount"] == 0, body["errors"]
    row = body["sample"][0]
    assert (row["leadSource"], row["executive"]) == ("Walk-in", "Amit")
    assert (row["interestedModel"], row["variant"]) == ("Turbo Max", "Maxx (PV)")
    assert (row["financeRequired"], row["priority"]) == ("Yes", "High")


@pytest.mark.asyncio
async def test_blank_optional_cells_use_new_lead_defaults(client):
    rows = [_row("Minimal Row", "9800000031", **{
        "Lead Date": "", "Lead Source": "", "Executive": "", "Interested Model": "",
        "Variant": "", "Current Status": "", "Priority": "", "Budget": "",
        "Finance Required": "", "Exchange Required": ""})]
    body = await _preview(client, rows)
    assert body["errorCount"] == 0, body["errors"]
    row = body["sample"][0]
    assert row["leadSource"] == "Walk-in"
    assert row["currentStatus"] == "New" and row["priority"] == "Normal"
    assert row["financeRequired"] == "No" and row["exchangeRequired"] == "No"
    assert row["createdDate"] == server.today()
    # The old import wrote leadSource="Import", which is not a Settings value.
    assert row["leadSource"] in server.seeder.MASTERS["leadSources"]


@pytest.mark.asyncio
async def test_duplicate_mobiles_blocked_in_file_and_against_crm(client):
    await client.post("/api/leads/import/commit",
                      files={"file": ("a.csv", _csv([_row("First In", "9800000041")]), "text/csv")})
    rows = [_row("Same File A", "9800000042"), _row("Same File B", "9800000042"),
            _row("Already There", "9800000041"), _row("No Mobile", "")]
    body = await _preview(client, rows)
    assert body["validCount"] == 1
    problems = {e["customerName"]: " ".join(e["errors"]) for e in body["errors"]}
    assert "Duplicate mobile" in problems["Same File B"]
    assert "already used by lead" in problems["Already There"]
    assert "Mobile is required" in problems["No Mobile"]


@pytest.mark.asyncio
async def test_dates_accept_indian_format_and_reject_junk(client):
    rows = [_row("DMY Date", "9800000051", **{"Lead Date": "10-08-2026"}),
            _row("Junk Date", "9800000052", **{"Lead Date": "next monday"})]
    body = await _preview(client, rows)
    assert body["validCount"] == 1
    assert body["sample"][0]["createdDate"] == "2026-08-10"
    assert "is not a date" in " ".join(body["errors"][0]["errors"])
